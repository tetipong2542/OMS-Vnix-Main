# app.py
from __future__ import annotations

import os, csv, json, re
import math
import time
import uuid
import tempfile
from datetime import datetime, date, timedelta, timezone
from io import BytesIO
from functools import wraps

import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file, jsonify, session, make_response
)
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import func, text, or_, not_, and_
from sqlalchemy.sql import bindparam
from sqlalchemy.exc import IntegrityError
from dotenv import load_dotenv

from utils import (
    now_thai, to_thai_be, to_be_date_str, TH_TZ, current_be_year,
    normalize_platform, sla_text, compute_due_date,
    parse_datetime_guess,
)
from models import (
    db, Shop, Product, Stock, Sales, OrderLine, User, UserPreference, PriceUserPreference,
    SkuPricing, BrandControl, PlatformFeeSetting, MarketItem, MarketPriceLog,
    PriceConfig,
    PriceImportBatch, PriceImportOp,
    BrandOwnerSetting,
    PriceExportSetting,
    SupplierSkuMaster, SupplierConfig, SupplierImportBatch,
)
from importers import (
    import_products, import_stock, import_sales, import_orders,
    import_sku_pricing, import_market_prices, import_brand_control,
    import_monthly_sales,
    import_supplier_sku_stock,
)
from allocation import compute_allocation

# โหลด environment variables จากไฟล์ .env (สำหรับ Local Development)
load_dotenv()

# Import sqlalchemy_libsql to register the libsql dialect
# This must be imported before creating any SQLAlchemy engines
# The import automatically registers "sqlite.libsql" and "sqlite.aiolibsql" dialects
try:
    import sqlalchemy_libsql
except ImportError:
    # If sqlalchemy_libsql is not installed, Turso features won't be available
    # The app will fall back to local SQLite mode
    pass

# Suppress deprecation warnings for legacy code
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning, message=".*datetime.*utcnow.*")
# Suppress SQLAlchemy legacy API warnings for Query.get()
try:
    from sqlalchemy.exc import LegacyAPIWarning
    warnings.filterwarnings("ignore", category=LegacyAPIWarning)
except ImportError:
    pass

APP_NAME = os.environ.get("APP_NAME", "VNIX ERP")


# ---------------------------------------------------------------------------
# Price Dashboard: cache rows for infinite scroll (in-process)
# NOTE: This cache is per-process (fine for single worker). If you run
# multiple workers, each worker will have its own cache.
# ---------------------------------------------------------------------------

PRICE_DASH_ROWS_CACHE: dict[str, dict] = {}  # key -> {"user_id":int, "ts":float, "rows":list[dict]}
PRICE_DASH_CACHE_TTL = 15 * 60  # 15 minutes


# ---------------------------------------------------------------------------
# Price Dashboard: Platform Import cache (in-process)
# Stores output workbook path for download after apply
# ---------------------------------------------------------------------------

PLATFORM_IMPORT_CACHE: dict[str, dict] = {}  # key -> {"user_id":int, "path":str, "ts":float, "stats":dict}
PLATFORM_IMPORT_TTL = 30 * 60  # 30 minutes


def _platform_import_gc(ttl_sec: int = PLATFORM_IMPORT_TTL):
    now = time.time()
    for k, v in list(PLATFORM_IMPORT_CACHE.items()):
        try:
            ts = float((v or {}).get("ts", 0) or 0)
        except Exception:
            ts = 0
        if now - ts > ttl_sec:
            path = (v or {}).get("path")
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass
            PLATFORM_IMPORT_CACHE.pop(k, None)


def _norm_sku(v) -> str:
    if v is None:
        return ""
    if isinstance(v, int):
        return str(v).strip()
    if isinstance(v, float):
        try:
            if v.is_integer():
                return str(int(v))
        except Exception:
            pass
        return str(v).strip()
    return str(v).strip()


# --- Detect real SKU values (avoid template notes/header rows) ---
_SKU_HEADER_WORDS = {
    "sku",
    "sellersku",
    "seller sku",
    "parent sku",
    "seller_sku",
    "seller-sku",
    "เลข sku",
    "ราคา",
    "คลัง",
    "stock",
}


def _is_real_sku_value(s: str) -> bool:
    s = (s or "").strip()
    if not s:
        return False

    s_low = s.lower()

    # header-like values
    if s_low in _SKU_HEADER_WORDS:
        return False

    # template notes often contain Thai text or spaces
    if re.search(r"[ก-๙]", s):
        return False
    if any(ch.isspace() for ch in s):
        return False

    # unusually long: likely a description sentence
    if len(s) > 80:
        return False

    return True


def _detect_header_row(ws, scan_rows: int = 50) -> int:
    """Heuristic: find likely header row by counting short non-empty strings."""
    best_row = 1
    best_score = -1
    max_col = min(getattr(ws, "max_column", 1) or 1, 200)
    max_row = min(getattr(ws, "max_row", 1) or 1, scan_rows)
    for r in range(1, max_row + 1):
        score = 0
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str):
                s = val.strip()
                if s and len(s) <= 60:
                    score += 1
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def _find_header_row_by_keywords(ws, keywords, scan_rows: int = 80, min_hits: int = 2):
    """Find header row by searching for multiple keyword hits in a row.

    Useful for templates (e.g. Shopee) that contain instruction rows.
    """
    max_col = min(getattr(ws, "max_column", 1) or 1, 200)
    max_row = min(getattr(ws, "max_row", 1) or 1, scan_rows)
    keys = [str(k).strip().lower() for k in (keywords or []) if str(k).strip()]

    best_row = None
    best_hits = 0

    for r in range(1, max_row + 1):
        parts: list[str] = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                s = v.strip().lower()
                if s:
                    parts.append(s)
        if not parts:
            continue

        row_text = " | ".join(parts)
        hits = sum(1 for k in keys if k in row_text)

        if hits > best_hits:
            best_hits = hits
            best_row = r

    if best_row and best_hits >= min_hits:
        return int(best_row)
    return None


def _extract_columns(ws, header_row: int) -> list[dict]:
    cols: list[dict] = []
    seen: dict[str, int] = {}
    max_col = min(getattr(ws, "max_column", 1) or 1, 200)
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        name = (str(v).strip() if v is not None else "")
        if not name:
            continue
        key = name.lower()
        if key in seen:
            seen[key] += 1
            name_show = f"{name} ({seen[key]})"
        else:
            seen[key] = 1
            name_show = name
        cols.append({"idx": c, "name": name_show})
    return cols


def _first_data_row(ws, header_row: int, *sku_cols: int) -> int:
    r = header_row + 1
    max_row = getattr(ws, "max_row", header_row + 1) or (header_row + 1)
    while r <= max_row:
        for col in sku_cols:
            if not col or int(col) <= 0:
                continue
            v = _norm_sku(ws.cell(row=r, column=int(col)).value)
            if _is_real_sku_value(v):
                return r
        r += 1
    return 0


def _dash_cache_gc():
    now = time.time()

    for k in list(PRICE_DASH_ROWS_CACHE.keys()):
        try:
            ts = float(PRICE_DASH_ROWS_CACHE[k].get("ts", 0) or 0)
        except Exception:
            ts = 0
        if now - ts > PRICE_DASH_CACHE_TTL:
            PRICE_DASH_ROWS_CACHE.pop(k, None)


# ---------------------------------------------------------------------------
# Supplier Stock Dashboard: cache rows for infinite scroll (in-process)
# NOTE: This cache is per-process (fine for single worker). If you run
# multiple workers, each worker will have its own cache.
# ---------------------------------------------------------------------------

SUPPLIER_DASH_ROWS_CACHE: dict[str, dict] = {}  # key -> {"user_id":int, "ts":float, "rows":list[dict]}
SUPPLIER_DASH_CACHE_TTL = 15 * 60  # 15 minutes


def _supplier_dash_cache_gc():
    now = time.time()

    for k in list(SUPPLIER_DASH_ROWS_CACHE.keys()):
        try:
            ts = float(SUPPLIER_DASH_ROWS_CACHE[k].get("ts", 0) or 0)
        except Exception:
            ts = 0
        if now - ts > SUPPLIER_DASH_CACHE_TTL:
            SUPPLIER_DASH_ROWS_CACHE.pop(k, None)

    MAX_ITEMS = 30
    if len(SUPPLIER_DASH_ROWS_CACHE) > MAX_ITEMS:
        items = sorted(SUPPLIER_DASH_ROWS_CACHE.items(), key=lambda kv: (kv[1] or {}).get("ts", 0))
        for k, _ in items[: max(0, len(items) - MAX_ITEMS)]:
            SUPPLIER_DASH_ROWS_CACHE.pop(k, None)

    # Prevent unbounded growth (e.g., multiple tabs)
    MAX_ITEMS = 30
    if len(PRICE_DASH_ROWS_CACHE) > MAX_ITEMS:
        items = sorted(PRICE_DASH_ROWS_CACHE.items(), key=lambda kv: (kv[1] or {}).get("ts", 0))
        for k, _ in items[: max(0, len(items) - MAX_ITEMS)]:
            PRICE_DASH_ROWS_CACHE.pop(k, None)


def get_engine(bind_key: str | None = None):
    """Return SQLAlchemy engine for a specific bind.

    - bind_key=None -> main (data.db)
    - bind_key='price' -> price.db
    """
    if not bind_key:
        return db.engine
    return db.engines[bind_key]


# ---------------------------------------------------------------------------
# Module-level helpers (used by multiple routes)
# Note: Some routes are defined inside create_app(), but static analysis and
# runtime may still require these names to exist at module scope.
# ---------------------------------------------------------------------------


def _ol_table_name() -> str:
    try:
        return OrderLine.__table__.name
    except Exception:
        return getattr(OrderLine, "__tablename__", "order_lines")


def _has_any_sales(r: dict) -> bool:
    sales_status = (str(r.get("sales_status") or "")).strip()
    po_no = (str(r.get("po_no") or "")).strip()
    return bool(sales_status or po_no)


def _orders_packed_set(rows: list[dict]) -> set[str]:
    packed: set[str] = set()
    for r in rows or []:
        oid = (r.get("order_id") or "").strip()
        if not oid:
            continue
        if bool(r.get("is_packed")) or bool(r.get("packed")):
            packed.add(oid)
            continue
        status = (r.get("allocation_status") or "").strip().upper()
        if status == "PACKED":
            packed.add(oid)
    return packed


def _orders_not_in_sbs_set(rows: list[dict]) -> set[str]:
    result: set[str] = set()
    for r in rows or []:
        oid = (r.get("order_id") or "").strip()
        if not oid:
            continue
        if not _has_any_sales(r):
            result.add(oid)
    return result


def _orders_no_sales_set(rows: list[dict]) -> set[str]:
    result: set[str] = set()
    for r in rows or []:
        oid = (r.get("order_id") or "").strip()
        if not oid:
            continue
        sales_status = (str(r.get("sales_status") or "")).strip()
        sales_status_u = sales_status.upper()
        if sales_status and ("ยังไม่มี" in sales_status or "NO_SALES" in sales_status_u):
            result.add(oid)
    return result


def normalize_platform_key(s: str) -> str:
    """Normalize user-provided platform key to a safe identifier.

    - Collapses whitespace to underscore
    - Keeps only A-Z, a-z, 0-9, underscore
    - Max length 50
    """
    s = (s or "").strip()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^A-Za-z0-9_]+", "", s)
    return s[:50]


def _round_to_10_thai(n: float | None) -> float | None:
    """Round to the nearest 10 using Thai rule (1–4 down, 5–9 up).

    Drops decimals first to match examples: 268.5 -> 268 -> 270.
    """
    if n is None:
        return None
    try:
        x = int(float(n))
    except Exception:
        return None
    base = (x // 10) * 10
    if (x % 10) >= 5:
        base += 10
    return float(base)


# ===== Rule r9 default tiers (percent is in % units) =====
_R9_DEFAULT_TIERS: list[tuple[float, float | None, float]] = [
    (5, 10, 200.0),
    (11, 20, 150.0),
    (21, 49, 120.0),
    (50, 99, 80.0),
    (100, 499, 50.0),
    (500, 699, 45.0),
    (700, 999, 40.0),
    (1000, None, 35.0),  # None = no upper bound
]


# ===== Rule r10/r11 default configs =====
_R10_DEFAULT_CFG = {"min_loss_pct": 5.0}
_R11_DEFAULT_CFG = {"min_loss_pct": 5.0, "max_loss_pct": 20.0}


def _abs_pct(v, default: float, lo: float = 0.0, hi: float = 100.0) -> float:
    try:
        x = abs(float(v))
    except Exception:
        x = float(default)
    return float(max(lo, min(hi, x)))


def _load_cfg_dict(raw, default: dict) -> dict:
    if not raw:
        return dict(default)
    if isinstance(raw, dict):
        return raw
    try:
        js = json.loads(raw)
        return js if isinstance(js, dict) else dict(default)
    except Exception:
        return dict(default)


def _norm_r10_cfg(cfg: dict) -> dict:
    cfg = cfg or {}
    return {
        "min_loss_pct": _abs_pct(cfg.get("min_loss_pct"), _R10_DEFAULT_CFG["min_loss_pct"], 0.0, 50.0)
    }


def _norm_r11_cfg(cfg: dict) -> dict:
    cfg = cfg or {}
    mn = _abs_pct(cfg.get("min_loss_pct"), _R11_DEFAULT_CFG["min_loss_pct"], 0.0, 50.0)
    mx = _abs_pct(cfg.get("max_loss_pct"), _R11_DEFAULT_CFG["max_loss_pct"], 0.0, 80.0)
    if mx < mn:
        mx = mn
    return {"min_loss_pct": mn, "max_loss_pct": mx}


def _loss_abs_pct(profit_match_pct) -> float | None:
    if profit_match_pct is None:
        return None
    try:
        return abs(float(profit_match_pct))
    except Exception:
        return None


def _r9_default_cfg() -> list[dict]:
    return [{"min": lo, "max": hi, "pct": pct} for (lo, hi, pct) in _R9_DEFAULT_TIERS]


def _r9_cfg_to_tiers(cfg_list) -> list[tuple[float, float | None, float]]:
    """Normalize r9 config to tiers.

    - cfg_list = [{"min": 5, "max": 10, "pct": 200}, ...]
    - pct clamped to [0, 500]
    - tiers sorted by min asc
    """
    tiers: list[tuple[float, float | None, float]] = []
    for it in (cfg_list or []):
        if not isinstance(it, dict):
            continue
        lo = it.get("min")
        hi = it.get("max")
        pct = it.get("pct")
        if lo is None or pct is None:
            continue
        try:
            lo_f = float(lo)
        except Exception:
            continue

        hi_f: float | None = None
        if hi not in (None, "", "null"):
            try:
                hi_f = float(hi)
            except Exception:
                hi_f = None

        try:
            pct_f = float(pct)
        except Exception:
            continue

        pct_f = max(0.0, min(pct_f, 500.0))
        tiers.append((lo_f, hi_f, pct_f))

    tiers.sort(key=lambda t: float(t[0]))
    return tiers or list(_R9_DEFAULT_TIERS)


def _clamp_float(v, lo: float, hi: float, default: float | None = None) -> float:
    try:
        x = float(v)
    except Exception:
        x = float(default if default is not None else lo)
    if x < lo:
        return float(lo)
    if x > hi:
        return float(hi)
    return float(x)


def _round_up_to_5(x: float | None) -> float | None:
    """บวกขึ้น: ปัดขึ้นเป็นเลขลงท้าย 0/5 (ceil to multiple of 5)."""
    if x is None:
        return None
    try:
        v = float(x)
    except Exception:
        return None
    return float(int(math.ceil(v / 5.0) * 5))


def _round_down_to_5(x: float | None) -> float | None:
    """ลดลง: ปัดลงเป็นเลขลงท้าย 0/5 (floor to multiple of 5)."""
    if x is None:
        return None
    try:
        v = float(x)
    except Exception:
        return None
    return float(int(math.floor(v / 5.0) * 5))


def adjust_our_price_export(our_price: float | None, adj_pct: float) -> float | None:
    """ปรับราคาเพื่อ Export เท่านั้น (ไม่แตะ DB)

    - adj_pct > 0: คูณแล้วปัดขึ้นเป็น 0/5
    - adj_pct < 0: คูณแล้วปัดลงเป็น 0/5
    ตัวอย่าง: 100 +4% = 104 => 105, 100 -4% = 96 => 95
    """
    if our_price is None:
        return None
    try:
        p = float(our_price)
    except Exception:
        return None
    if p <= 0:
        return None

    raw = p * (1.0 + float(adj_pct) / 100.0)
    if adj_pct > 0:
        return _round_up_to_5(raw)
    if adj_pct < 0:
        return _round_down_to_5(raw)
    return p


def _floor_int(x: float | None) -> int:
    if x is None:
        return 0
    try:
        return int(math.floor(float(x)))
    except Exception:
        return 0


def _ceil_to_5_int(x: float | None) -> int:
    """ปัดขึ้นเป็นเลขที่หาร 5 ลงตัว (ตามกติกา 1–4→5, 6–9→10)."""
    if x is None:
        return 0
    try:
        v = float(x)
    except Exception:
        return 0
    r = _round_up_to_5(v)
    try:
        return int(r or 0)
    except Exception:
        return 0


def build_sell_prices(
    our_price: float | None,
    cost: float | None,
    step_pct: float | None,
    min_profit_pct: float | None,
    loss_aging3_pct: float | None,
    loss_aging6_pct: float | None,
    loss_aging12_pct: float | None,
    aging_bucket: str | None,

) -> list[int]:
    """Build tiers Sell1..Sell5 + SellMin (6 values), floor decimals.

    Rules:
    - Reduce compound by step_pct each tier.
        Notes:
        - The `cost` parameter should be the *effective offline cost* you want to protect
            (e.g., cost + pack_cost + ship_subsidy). Do NOT include platform fees here.

        Rules:
        - Non-aging: enforce threshold = cost*(1 + min_profit_pct).
        - Aging bucket (aging3/aging6/aging12): allow loss: threshold = cost*(1 - max_loss_pct).
    - Any tier below threshold is raised to the lowest tier that still passes threshold.
    """

    step = _clamp_float(step_pct, 0.0, 10.0, default=0.0) / 100.0
    p = float(our_price or 0.0)

    tiers: list[int] = []
    for _ in range(6):
        p = p * (1.0 - step)
        tiers.append(_ceil_to_5_int(p))

    try:
        c = float(cost) if cost is not None else None
    except Exception:
        c = None

    if c is None or c <= 0:
        return tiers

    bucket = (aging_bucket or "").strip().lower() or None
    if bucket == "aging3":
        max_loss = _clamp_float(loss_aging3_pct, 0.0, 50.0, default=0.0)
        threshold = c * (1.0 - max_loss / 100.0)
    elif bucket == "aging6":
        max_loss = _clamp_float(loss_aging6_pct, 0.0, 50.0, default=0.0)
        threshold = c * (1.0 - max_loss / 100.0)
    elif bucket == "aging12":
        max_loss = _clamp_float(loss_aging12_pct, 0.0, 50.0, default=0.0)
        threshold = c * (1.0 - max_loss / 100.0)
    else:
        min_profit = _clamp_float(min_profit_pct, 0.0, 10.0, default=0.0)
        threshold = c * (1.0 + min_profit / 100.0)

    ok = [t for t in tiers if float(t) >= float(threshold)]
    floor_price = min(ok) if ok else _ceil_to_5_int(float(threshold))

    tiers2 = [t if float(t) >= float(threshold) else floor_price for t in tiers]

    # NEW: Sell1 special when "ติด Min Profit" แล้ว tier แบน (Sell1==Sell2)
    # Rule:
    #   Sell1 = floor_to_5(Sell2 * 1.03)
    #   if Sell1 > OurPrice -> Sell1 = floor_to_5(OurPrice * 0.99)
    #   and must keep Sell1 > Sell2
    try:
        our = float(our_price or 0.0)
    except Exception:
        our = 0.0

    # ใช้เฉพาะ non-aging (Min Profit) ตาม logic เดิม
    if (bucket is None) and (len(tiers2) >= 2) and (our > 0) and (our >= float(threshold)):
        sell2 = int(tiers2[1] or 0)

        # ทำเฉพาะกรณี “แบน” จริง ๆ (Sell1 เท่ากับ Sell2)
        if int(tiers2[0] or 0) == sell2 and sell2 > 0:

            # 1) Sell1 = Sell2 + 3% แล้วปัดลงเป็นเลขลงท้าย 0/5
            cand = _round_down_to_5(float(sell2) * 1.03)
            cand_int = int(cand or sell2)

            # 2) ถ้าเกิน Our Price -> ใช้ Our Price -1% แทน (แล้วปัดลง)
            if cand_int > our:
                cap = _round_down_to_5(our * 0.99)
                cand_int = int(cap or our)

                # 3) กันเคสแปลก: ถ้า cap <= Sell2 ให้บังคับอย่างน้อย > Sell2 แต่ไม่เกิน Our
                if cand_int <= sell2:
                    our_floor = int(_round_down_to_5(our) or our)
                    cand_int = min(our_floor, sell2 + 5)

            # 4) สุดท้ายต้องมากกว่า Sell2 เท่านั้นถึงจะเขียนทับ
            if cand_int > sell2:
                tiers2[0] = cand_int

    return tiers2


def calc_profit_offline(price, cost, pack_cost: float = 0.0, ship_subsidy: float = 0.0) -> float:
    """Offline profit (no platform fee deduction)."""
    price = float(price or 0.0)
    cost = float(cost or 0.0)
    return price - cost - float(pack_cost or 0.0) - float(ship_subsidy or 0.0)


def _auto_price_from_cost(cost: float | None, tiers=None) -> float | None:
    """Rule r9: when no market price, derive Our Price from Cost by configurable tiers.

    - tiers: list of (min_cost, max_cost_or_None, pct_in_percent_units)
    - rounding: Thai tens rule (1–4 down, 5–9 up), drop decimals first (same as before)
    """
    if cost is None:
        return None
    try:
        c = float(cost)
    except Exception:
        return None

    if c < 5:
        return None

    tiers = tiers or _R9_DEFAULT_TIERS

    pct: float | None = None
    for lo, hi, p in (tiers or []):
        try:
            lo_f = float(lo)
        except Exception:
            continue
        if c < lo_f:
            continue

        if hi is None:
            try:
                pct = float(p)
            except Exception:
                pct = None
            break

        try:
            hi_f = float(hi)
        except Exception:
            continue
        if lo_f <= c <= hi_f:
            try:
                pct = float(p)
            except Exception:
                pct = None
            break

    if pct is None:
        return None

    raw = c * (1.0 + (pct / 100.0))
    return _round_to_10_thai(raw)


def _auto_price_from_cost_plus_pct(cost: float | None, pct: float) -> float | None:
    """Rule r10: Our Price = Cost * (1 + pct) and round up to multiple of 5.

    Rounding rule: 1–4 → 5, 6–9 → 10 (ceil to multiple of 5).
    """
    if cost is None:
        return None
    try:
        c = float(cost)
    except Exception:
        return None
    if c <= 0:
        return None

    raw = c * (1.0 + float(pct))
    return _round_up_to_5(raw)


def _print_cols_for_kind(kind: str) -> tuple[str, str, str]:
    kind = (kind or "").strip().lower()
    if kind == "warehouse":
        return ("printed_warehouse", "printed_warehouse_at", "printed_warehouse_by")
    if kind == "picking":
        return ("printed_picking", "printed_picking_at", "printed_picking_by")
    if kind == "lowstock":
        return ("printed_lowstock", "printed_lowstock_at", "printed_lowstock_by")
    if kind == "nostock":
        return ("printed_nostock", "printed_nostock_at", "printed_nostock_by")
    if kind == "notenough":
        return ("printed_notenough", "printed_notenough_at", "printed_notenough_by")
    return ("printed_picking", "printed_picking_at", "printed_picking_by")


def _detect_already_printed(oids: list[str], kind: str) -> set[str]:
    if not oids:
        return set()
    tbl = _ol_table_name()
    col, _, _ = _print_cols_for_kind(kind)
    sql = text(
        f"SELECT DISTINCT order_id FROM {tbl} WHERE order_id IN :oids AND COALESCE({col},0) > 0"
    ).bindparams(bindparam("oids", expanding=True))
    rows_sql = db.session.execute(sql, {"oids": oids}).fetchall()
    return {str(r[0]) for r in rows_sql if r and r[0]}


def _mark_printed(
    oids: list[str],
    kind: str,
    user_id: int | None,
    when_iso: str,
    commit: bool = True,
) -> int:
    if not oids:
        return 0

    username = None
    if user_id is not None:
        try:
            u = User.query.get(int(user_id))
            username = (u.username if u else None)
        except Exception:
            username = None

    tbl = _ol_table_name()
    col, col_at, col_by = _print_cols_for_kind(kind)
    sql = text(
        f"""
        UPDATE {tbl}
           SET {col}=COALESCE({col},0)+1,
               {col_at}=:ts,
               {col_by}=:byu
         WHERE order_id IN :oids
        """
    ).bindparams(bindparam("oids", expanding=True))
    res = db.session.execute(sql, {"ts": when_iso, "byu": username, "oids": oids})
    if commit:
        db.session.commit()
    try:
        return int(getattr(res, "rowcount", 0) or 0)
    except Exception:
        return 0


def get_google_credentials():
    """
    โหลด Google Service Account Credentials จาก Environment Variables หรือไฟล์

    สำหรับ Production (Railway):
    - ตั้งค่า Environment Variables ใน Railway Dashboard
    - ใช้ GOOGLE_CREDENTIALS_JSON (JSON string ทั้งหมด) หรือ
    - ใช้ตัวแปรแยก: GOOGLE_PROJECT_ID, GOOGLE_PRIVATE_KEY, GOOGLE_CLIENT_EMAIL, ฯลฯ

    สำหรับ Local Development:
    - วางไฟล์ credentials.json ในโฟลเดอร์โปรเจกต์
    """
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']

    # ลองอ่านจาก Environment Variable (JSON string ทั้งก้อน)
    google_creds_json = os.environ.get('GOOGLE_CREDENTIALS_JSON')
    if google_creds_json:
        try:
            creds_dict = json.loads(google_creds_json)
            return ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        except json.JSONDecodeError as e:
            raise RuntimeError(f"GOOGLE_CREDENTIALS_JSON ไม่ถูกต้อง: {e}")

    # ลองสร้างจาก Environment Variables แยก
    if os.environ.get('GOOGLE_PRIVATE_KEY'):
        creds_dict = {
            "type": "service_account",
            "project_id": os.environ.get('GOOGLE_PROJECT_ID'),
            "private_key_id": os.environ.get('GOOGLE_PRIVATE_KEY_ID'),
            "private_key": os.environ.get('GOOGLE_PRIVATE_KEY', '').replace('\\n', '\n'),
            "client_email": os.environ.get('GOOGLE_CLIENT_EMAIL'),
            "client_id": os.environ.get('GOOGLE_CLIENT_ID'),
            "auth_uri": os.environ.get('GOOGLE_AUTH_URI', 'https://accounts.google.com/o/oauth2/auth'),
            "token_uri": os.environ.get('GOOGLE_TOKEN_URI', 'https://oauth2.googleapis.com/token'),
            "auth_provider_x509_cert_url": os.environ.get('GOOGLE_AUTH_PROVIDER_CERT_URL', 'https://www.googleapis.com/oauth2/v1/certs'),
            "client_x509_cert_url": os.environ.get('GOOGLE_CLIENT_CERT_URL'),
            "universe_domain": os.environ.get('GOOGLE_UNIVERSE_DOMAIN', 'googleapis.com')
        }
        return ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)

    # ลองอ่านจากไฟล์ credentials.json (สำหรับ Local Development)
    creds_path = os.path.join(os.path.dirname(__file__), 'credentials.json')
    if os.path.exists(creds_path):
        return ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)

    # ไม่พบ credentials ทั้งหมด
    raise RuntimeError(
        "ไม่พบ Google Service Account Credentials\n\n"
        "สำหรับ Production: ตั้งค่า Environment Variables ใน Railway:\n"
        "- GOOGLE_CREDENTIALS_JSON (แนะนำ) หรือ\n"
        "- GOOGLE_PRIVATE_KEY, GOOGLE_CLIENT_EMAIL, ฯลฯ\n\n"
        "สำหรับ Local: วางไฟล์ credentials.json ในโฟลเดอร์โปรเจกต์"
    )


# -----------------------------
# สร้างแอป + บูตระบบเบื้องต้น
# -----------------------------
def create_app():
    app = Flask(__name__)
    app.secret_key = os.environ.get("SECRET_KEY", "vnix-secret")

    # =========[ Database Configuration ]=========
    # Support 3 separate Turso databases with embedded replicas:
    # - data: Main database (orders, products, customers, etc.)
    # - price: Pricing database (SKU pricing, price history)
    # - supplier-stock: Supplier and stock database

    # NEW: Check for 3 separate Turso databases
    data_url = os.environ.get("DATA_DB_URL")
    data_token = os.environ.get("DATA_DB_AUTH_TOKEN")
    data_local = os.environ.get("DATA_DB_LOCAL")  # No default - must be explicitly set for embedded replica

    price_url = os.environ.get("PRICE_DB_URL")
    price_token = os.environ.get("PRICE_DB_AUTH_TOKEN")
    price_local = os.environ.get("PRICE_DB_LOCAL")  # No default - must be explicitly set for embedded replica

    supplier_url = os.environ.get("SUPPLIER_DB_URL")
    supplier_token = os.environ.get("SUPPLIER_DB_AUTH_TOKEN")
    supplier_local = os.environ.get("SUPPLIER_DB_LOCAL")  # No default - must be explicitly set for embedded replica

    # OLD: Support legacy single Turso database mode (fallback)
    turso_url = os.environ.get("TURSO_DATABASE_URL")
    turso_token = os.environ.get("TURSO_AUTH_TOKEN")
    enable_dual_db = os.environ.get("ENABLE_DUAL_DB_MODE", "").lower() in ("true", "1", "yes")

    if data_url and data_token and price_url and price_token and supplier_url and supplier_token:
        # ====================
        # 3 SEPARATE TURSO DATABASES
        # ====================
        # Support both Embedded Replica (with local files) and Remote-only (no local files)

        def build_turso_uri(sync_url, auth_token, local_file):
            """
            Build SQLAlchemy URI for Turso database with embedded replica

            Note: sqlalchemy-libsql REQUIRES embedded replica mode (cannot do remote-only)

            - If local_file is set: Use that path (e.g., /data/data.db for Volume)
            - If local_file is NOT set: Use in-container temp path (e.g., /tmp/data.db)
            """
            # Ensure sync_url uses libsql:// protocol
            if not sync_url.startswith("libsql://"):
                if sync_url.startswith("https://"):
                    sync_url = sync_url.replace("https://", "libsql://", 1)

            # If local_file is not set, use in-container temp path (no Volume needed)
            if not local_file or not local_file.strip():
                # Use /tmp for in-container storage (will be synced on startup, lost on restart)
                db_name = sync_url.split('/')[-1].split('.')[0]  # Extract db name from URL
                local_file = f"/tmp/{db_name}.db"
            else:
                # Clean up file: prefix if present
                if local_file.startswith("file:"):
                    local_file = local_file[5:]

            # ALWAYS use embedded replica mode (required by sqlalchemy-libsql)
            return f"sqlite+libsql:///{local_file}?sync_url={sync_url}&authToken={auth_token}"

        # Detect mode based on whether local files are configured
        use_volume = bool(data_local and data_local.strip())

        if use_volume:
            print("[INFO] 🚀 Using 3 separate Turso databases with EMBEDDED REPLICAS + VOLUME")
            print("[INFO] 💾 Local files persisted in Railway Volume (fast restarts)")
        else:
            print("[INFO] 🚀 Using 3 separate Turso databases with EMBEDDED REPLICAS (In-Container)")
            print("[INFO] 📦 Local files in container /tmp (will re-sync on restart)")
            print("[INFO] ⚠️  No Volume - data syncs from Turso on each startup (~10-30s)")

        # Build URIs for each database
        data_uri = build_turso_uri(data_url, data_token, data_local)
        price_uri = build_turso_uri(price_url, price_token, price_local)
        supplier_uri = build_turso_uri(supplier_url, supplier_token, supplier_local)

        # Main database: data DB
        app.config["SQLALCHEMY_DATABASE_URI"] = data_uri
        app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

        # Bind keys for price and supplier databases
        app.config["SQLALCHEMY_BINDS"] = {
            "price": price_uri,
            "supplier": supplier_uri
        }

        # Show the actual paths being used
        data_path = data_local if data_local else "/tmp/data-tetipong2542.db"
        price_path = price_local if price_local else "/tmp/price-tetipong2542.db"
        supplier_path = supplier_local if supplier_local else "/tmp/supplier-stock-tetipong2542.db"

        print(f"[DEBUG] ✅ Data DB: {data_url}")
        print(f"[DEBUG]    → Local file: {data_path}")
        print(f"[DEBUG] ✅ Price DB: {price_url}")
        print(f"[DEBUG]    → Local file: {price_path}")
        print(f"[DEBUG] ✅ Supplier DB: {supplier_url}")
        print(f"[DEBUG]    → Local file: {supplier_path}")

        if use_volume:
            print(f"[DEBUG] 💾 Using Railway Volume - files persist across restarts")
        else:
            print(f"[DEBUG] 📦 Using container /tmp - files re-sync on each restart")

    elif enable_dual_db and turso_url and turso_token:
        # ====================
        # DUAL DATABASE MODE
        # ====================
        # - Old databases (sqlite3): read-only for historical data
        # - New database (Turso/libSQL): read/write for new data
        print("[INFO] ⚙️  DUAL DATABASE MODE ENABLED")
        print("[INFO] Using Turso (libSQL) for NEW data (read/write)")
        print("[INFO] Using SQLite for OLD data (read-only)")

        # Use Embedded Replica Mode for better performance and reliability
        # Format: sqlite+libsql:///local-file.db?sync_url=libsql://...&authToken=...
        local_db_file = os.environ.get("LOCAL_DB", "vnix-erp.db")

        # Clean up file: prefix if present
        if local_db_file.startswith("file:"):
            local_db_file = local_db_file[5:]  # Remove "file:" prefix

        # Ensure turso_url uses libsql:// protocol
        if not turso_url.startswith("libsql://"):
            if turso_url.startswith("https://"):
                turso_url = turso_url.replace("https://", "libsql://", 1)

        db_uri = f"sqlite+libsql:///{local_db_file}?sync_url={turso_url}&authToken={turso_token}"

        # Primary database: Turso (for new data)
        app.config["SQLALCHEMY_DATABASE_URI"] = db_uri
        app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

        # Setup bind keys for both old and new databases
        base_dir = os.path.abspath(os.path.dirname(__file__))
        volume_path = os.environ.get("RAILWAY_VOLUME_MOUNT_PATH")

        # Old SQLite databases (read-only)
        if volume_path:
            old_data_path = os.path.join(volume_path, "data.db")
            old_price_path = os.path.join(volume_path, "price.db")
            old_supplier_path = os.path.join(volume_path, "supplier_stock.db")
        else:
            old_data_path = os.path.join(base_dir, "data.db")
            old_price_path = os.path.join(base_dir, "price.db")
            old_supplier_path = os.path.join(base_dir, "supplier_stock.db")

        binds = {
            # New databases on Turso (read/write)
            "price": db_uri,
            "supplier": db_uri,

            # Old databases on SQLite (read-only for historical data)
            "data_old": f"sqlite:///{old_data_path}",
            "price_old": f"sqlite:///{old_price_path}",
            "supplier_old": f"sqlite:///{old_supplier_path}"
        }
        app.config["SQLALCHEMY_BINDS"] = binds

        print(f"[DEBUG] ✅ Turso (NEW data): {turso_url}")
        print(f"[DEBUG] ✅ SQLite OLD data.db: {old_data_path}")
        print(f"[DEBUG] ✅ SQLite OLD price.db: {old_price_path}")
        print(f"[DEBUG] ✅ SQLite OLD supplier_stock.db: {old_supplier_path}")
        print(f"[DEBUG] Full Turso URI: {db_uri.replace(turso_token, '***TOKEN***')}")

    elif turso_url and turso_token:
        # Production: Use Turso (libSQL) - Serverless Database
        print("[INFO] Using Turso (libSQL) database")

        # Use Embedded Replica Mode for better performance and reliability
        # Format: sqlite+libsql:///local-file.db?sync_url=libsql://...&authToken=...
        local_db_file = os.environ.get("LOCAL_DB", "vnix-erp.db")

        # Clean up file: prefix if present
        if local_db_file.startswith("file:"):
            local_db_file = local_db_file[5:]  # Remove "file:" prefix

        # Ensure turso_url uses libsql:// protocol
        if not turso_url.startswith("libsql://"):
            if turso_url.startswith("https://"):
                turso_url = turso_url.replace("https://", "libsql://", 1)

        db_uri = f"sqlite+libsql:///{local_db_file}?sync_url={turso_url}&authToken={turso_token}"

        app.config["SQLALCHEMY_DATABASE_URI"] = db_uri
        app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

        # All 3 databases (data, price, supplier) are in the same Turso database
        # So we use the same connection URL for all binds
        binds = {
            "price": db_uri,
            "supplier": db_uri
        }
        app.config["SQLALCHEMY_BINDS"] = binds

        print(f"[DEBUG] Turso URL: {turso_url}")
        print(f"[DEBUG] Full DB URI: {db_uri.replace(turso_token, '***TOKEN***')}")
        print(f"[DEBUG] Using single Turso database for all binds")
    else:
        # Local Development: Use SQLite files
        print("[INFO] Using local SQLite database files")
        volume_path = os.environ.get("RAILWAY_VOLUME_MOUNT_PATH")

        if volume_path:
            db_path = os.path.join(volume_path, "data.db")
            price_db_path = os.path.join(volume_path, "price.db")
            supplier_db_path = os.path.join(volume_path, "supplier_stock.db")
        else:
            base_dir = os.path.abspath(os.path.dirname(__file__))
            db_path = os.path.join(base_dir, "data.db")
            price_db_path = os.path.join(base_dir, "price.db")
            supplier_db_path = os.path.join(base_dir, "supplier_stock.db")

        app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_path}"
        app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

        binds = {
            "price": f"sqlite:///{price_db_path}",
            "supplier": f"sqlite:///{supplier_db_path}"
        }
        app.config["SQLALCHEMY_BINDS"] = binds

        print(f"[DEBUG] Main DB path: {db_path}")
        print(f"[DEBUG] Price DB path: {price_db_path}")
        print(f"[DEBUG] Supplier DB path: {supplier_db_path}")

    db.init_app(app)

    # =========[ NEW ]=========
    # Model: ออเดอร์ที่ถูกทำเป็น "ยกเลิก"
    class CancelledOrder(db.Model):
        __tablename__ = "cancelled_orders"
        id = db.Column(db.Integer, primary_key=True)
        order_id = db.Column(db.String(128), unique=True, index=True, nullable=False)
        imported_at = db.Column(db.DateTime, default=datetime.utcnow, index=True, nullable=False)
        imported_by_user_id = db.Column(db.Integer, db.ForeignKey("users.id"))
        note = db.Column(db.String(255))

    # =========[ NEW ]=========  Order "จ่ายงานแล้ว"
    class IssuedOrder(db.Model):
        __tablename__ = "issued_orders"
        id = db.Column(db.Integer, primary_key=True)
        order_id = db.Column(db.String(128), unique=True, index=True, nullable=False)
        issued_at = db.Column(db.DateTime, default=datetime.utcnow, index=True, nullable=False)
        issued_by_user_id = db.Column(db.Integer, db.ForeignKey("users.id"))
        source = db.Column(db.String(32))  # 'import' | 'print:picking' | 'print:warehouse' | 'manual'
        note = db.Column(db.String(255))
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  Order ที่ถูกลบ (Soft Delete / Recycle Bin)
    class DeletedOrder(db.Model):
        __tablename__ = "deleted_orders"
        id = db.Column(db.Integer, primary_key=True)
        order_id = db.Column(db.String(128), unique=True, index=True, nullable=False)
        deleted_at = db.Column(db.DateTime, default=datetime.utcnow, index=True, nullable=False)
        deleted_by_user_id = db.Column(db.Integer, db.ForeignKey("users.id"))
        note = db.Column(db.String(255))
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  ตารางเก็บประวัติการนำเข้า Orders (สำหรับ Dashboard)
    class ImportLog(db.Model):
        __tablename__ = "import_logs"
        id = db.Column(db.Integer, primary_key=True)
        import_date = db.Column(db.Date, index=True, nullable=False)  # วันที่ import
        platform = db.Column(db.String(50))
        shop_name = db.Column(db.String(128))  # ชื่อร้านที่นำเข้า
        filename = db.Column(db.String(255))
        
        # เก็บยอดเฉพาะเหตุการณ์ในไฟล์นั้นๆ
        added_count = db.Column(db.Integer, default=0)
        duplicates_count = db.Column(db.Integer, default=0)
        duplicates_same_day = db.Column(db.Integer, default=0)  # ซ้ำในวันเดียวกัน (ไม่แสดงในการ์ด)
        failed_count = db.Column(db.Integer, default=0)
        
        # เก็บรายชื่อ Error (JSON String)
        error_details = db.Column(db.Text, nullable=True)
        
        # เก็บ Batch Data (IDs ที่เพิ่ม/ซ้ำ/ไม่สำเร็จ) JSON String
        batch_data = db.Column(db.Text, nullable=True)
        
        created_at = db.Column(db.DateTime, default=datetime.utcnow)
    # =========[ /NEW ]=========

    # ---------- Helper: Table name (OrderLine) ----------
    def _ol_table_name() -> str:
        try:
            return OrderLine.__table__.name
        except Exception:
            return getattr(OrderLine, "__tablename__", "order_lines")

    # ---------- Auto-migrate: ensure print columns exist ----------
    def _ensure_orderline_print_columns():
        """Auto-migrate: เพิ่มคอลัมน์สำหรับติดตามสถานะการพิมพ์ Warehouse และ Picking"""
        tbl = _ol_table_name()
        with db.engine.connect() as con:
            cols = {row[1] for row in con.execute(text(f"PRAGMA table_info({tbl})")).fetchall()}

            def add(col, ddl):
                if col not in cols:
                    con.execute(text(f"ALTER TABLE {tbl} ADD COLUMN {col} {ddl}"))

            # สำหรับ "ใบงานคลัง (Warehouse Job Sheet)"
            add("printed_warehouse", "INTEGER DEFAULT 0")  # จำนวนครั้งที่พิมพ์
            add("printed_warehouse_at", "TEXT")  # timestamp ครั้งล่าสุด
            add("printed_warehouse_by", "TEXT")  # username ผู้พิมพ์

            # สำหรับ "Picking List"
            add("printed_picking", "INTEGER DEFAULT 0")  # จำนวนครั้งที่พิมพ์
            add("printed_picking_at", "TEXT")  # timestamp ครั้งล่าสุด
            add("printed_picking_by", "TEXT")  # username ผู้พิมพ์

            # สำหรับ "จ่ายงาน(รอบที่)"
            add("dispatch_round", "INTEGER")

            # สำหรับ "รายงานสินค้าน้อย" (แยกจากคลัง/Picking)
            add("printed_lowstock", "INTEGER DEFAULT 0")
            add("printed_lowstock_at", "TEXT")
            add("printed_lowstock_by", "TEXT")
            add("lowstock_round", "INTEGER")

            # สำหรับ "รายงานไม่มีสินค้า" (แยกจาก lowstock)
            add("printed_nostock", "INTEGER DEFAULT 0")
            add("printed_nostock_at", "TEXT")
            add("printed_nostock_by", "TEXT")
            add("nostock_round", "INTEGER")

            # สำหรับ "รายงานสินค้าไม่พอส่ง" (NOT_ENOUGH)
            add("printed_notenough", "INTEGER DEFAULT 0")
            add("printed_notenough_at", "TEXT")
            add("printed_notenough_by", "TEXT")
            add("notenough_round", "INTEGER")

            # สำหรับ "Barcode Scan Check" (Warehouse)
            add("scanned_at", "TEXT")
            add("scanned_by", "TEXT")

            con.commit()

    # ========== [NEW] Auto-migrate shops unique: (platform, name) ==========
    def _has_unique_index_on(conn, table: str, columns_exact: list[str]) -> tuple[bool, str | None]:
        idx_list = conn.execute(text(f"PRAGMA index_list({table})")).fetchall()
        for row in idx_list:
            idx_name = row[1]
            is_unique = int(row[2]) == 1
            if not is_unique:
                continue
            cols = [r[2] for r in conn.execute(text(f"PRAGMA index_info('{idx_name}')")).fetchall()]
            if cols == columns_exact:
                return True, idx_name
        return False, None

    def _migrate_shops_unique_to_platform_name():
        """ย้าย unique จาก name เดี่ยว → เป็น (platform, name)"""
        with db.engine.begin() as con:
            has_composite, _ = _has_unique_index_on(con, "shops", ["platform", "name"])
            if has_composite:
                return
            has_name_unique, idx_name = _has_unique_index_on(con, "shops", ["name"])
            if has_name_unique:
                is_auto = idx_name.startswith("sqlite_autoindex")
                if is_auto:
                    cols_info = con.execute(text("PRAGMA table_info(shops)")).fetchall()
                    col_names = [c[1] for c in cols_info]
                    has_created_at = "created_at" in col_names
                    con.execute(text("ALTER TABLE shops RENAME TO shops_old"))
                    create_sql = """
                    CREATE TABLE shops (
                        id INTEGER PRIMARY KEY,
                        platform TEXT,
                        name TEXT NOT NULL,
                        created_at TEXT
                    )
                    """ if has_created_at else """
                    CREATE TABLE shops (
                        id INTEGER PRIMARY KEY,
                        platform TEXT,
                        name TEXT NOT NULL
                    )
                    """
                    con.execute(text(create_sql))
                    copy_cols = "id, platform, name" + (", created_at" if has_created_at else "")
                    con.execute(text(f"INSERT INTO shops ({copy_cols}) SELECT {copy_cols} FROM shops_old"))
                    con.execute(text("DROP TABLE shops_old"))
                else:
                    con.execute(text(f"DROP INDEX IF EXISTS {idx_name}"))
            con.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_shops_platform_name ON shops(platform, name)"))
    # ========== [/NEW] ==========

    # =========[ NEW ]=========
    def _ensure_issue_table():
        try:
            IssuedOrder.__table__.create(bind=db.engine, checkfirst=True)
        except Exception as e:
            app.logger.warning(f"[issued_orders] ensure table failed: {e}")
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  ตารางเก็บ Order ที่ถูกลบ
    def _ensure_deleted_table():
        try:
            DeletedOrder.__table__.create(bind=db.engine, checkfirst=True)
        except Exception as e:
            app.logger.warning(f"[deleted_orders] ensure table failed: {e}")
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  ตารางเก็บประวัติการนำเข้า (Import Log)
    def _ensure_import_log_table():
        try:
            ImportLog.__table__.create(bind=db.engine, checkfirst=True)
        except Exception as e:
            app.logger.warning(f"[import_logs] ensure table failed: {e}")
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  ตาราง dedupe กัน request ซ้ำ (Idempotency)
    def _ensure_action_dedupe_table():
        """Create a minimal dedupe table used to make actions idempotent (e.g., picking print)."""
        try:
            with db.engine.begin() as con:
                con.execute(
                    text(
                        """
                        CREATE TABLE IF NOT EXISTS action_dedupe (
                            token TEXT PRIMARY KEY,
                            kind TEXT NOT NULL,
                            created_at TEXT NOT NULL,
                            user_id INTEGER
                        )
                        """
                    )
                )
        except Exception as e:
            app.logger.warning(f"[action_dedupe] ensure table failed: {e}")
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  เพิ่มคอลัมน์ใหม่ให้ตาราง Shop และ ImportLog
    def _ensure_shop_url_and_log_batch_columns():
        """Auto-migrate: เพิ่มคอลัมน์ google_sheet_url ให้ Shop และ batch_data, shop_name, duplicates_same_day ให้ ImportLog"""
        with db.engine.connect() as con:
            # เพิ่ม google_sheet_url ให้ Shop
            cols_shop = {row[1] for row in con.execute(text("PRAGMA table_info(shops)")).fetchall()}
            if "google_sheet_url" not in cols_shop:
                con.execute(text("ALTER TABLE shops ADD COLUMN google_sheet_url TEXT"))
            
            # เพิ่มคอลัมน์ให้ ImportLog
            cols_log = {row[1] for row in con.execute(text("PRAGMA table_info(import_logs)")).fetchall()}
            if "batch_data" not in cols_log:
                con.execute(text("ALTER TABLE import_logs ADD COLUMN batch_data TEXT"))
            if "shop_name" not in cols_log:
                con.execute(text("ALTER TABLE import_logs ADD COLUMN shop_name TEXT"))
            if "duplicates_same_day" not in cols_log:
                con.execute(text("ALTER TABLE import_logs ADD COLUMN duplicates_same_day INTEGER DEFAULT 0"))
            con.commit()
    # =========[ /NEW ]=========

    with app.app_context():
        # Check if we're using Turso - if so, sync instead of create
        using_turso = bool(data_url and data_token and price_url and price_token and supplier_url and supplier_token)

        if using_turso:
            # For Turso embedded replica: sync data from cloud instead of creating empty tables
            print(f"[INFO] 🔄 Syncing data from Turso cloud databases...")
            print(f"[INFO] ⏳ Initial sync may take 20-30 seconds...")

            # CRITICAL FIX: Force sync with libsql client BEFORE SQLAlchemy
            # This downloads data from Turso cloud to local embedded replica files
            try:
                import libsql_experimental as libsql

                data_path = data_local if data_local else "/tmp/data-tetipong2542.db"
                price_path = price_local if price_local else "/tmp/price-tetipong2542.db"
                supplier_path = supplier_local if supplier_local else "/tmp/supplier-stock-tetipong2542.db"

                print(f"[DEBUG] 🔄 Pre-syncing data.db from Turso cloud...")
                conn_data = libsql.connect(data_path, sync_url=data_url, auth_token=data_token)
                conn_data.sync()
                cursor = conn_data.cursor()
                cursor.execute("SELECT COUNT(*) FROM products")
                count = cursor.fetchone()[0]
                print(f"[DEBUG] ✅ data.db synced: {count} products")
                conn_data.close()

                print(f"[DEBUG] 🔄 Pre-syncing price.db from Turso cloud...")
                conn_price = libsql.connect(price_path, sync_url=price_url, auth_token=price_token)
                conn_price.sync()
                cursor = conn_price.cursor()
                cursor.execute("SELECT COUNT(*) FROM sku_pricing")
                count = cursor.fetchone()[0]
                print(f"[DEBUG] ✅ price.db synced: {count} SKU pricings")
                conn_price.close()

                print(f"[DEBUG] 🔄 Pre-syncing supplier_stock.db from Turso cloud...")
                conn_supplier = libsql.connect(supplier_path, sync_url=supplier_url, auth_token=supplier_token)
                conn_supplier.sync()
                cursor = conn_supplier.cursor()
                cursor.execute("SELECT COUNT(*) FROM supplier_sku_master")
                count = cursor.fetchone()[0]
                print(f"[DEBUG] ✅ supplier_stock.db synced: {count} supplier SKUs")
                conn_supplier.close()

                print(f"[INFO] ✅ All databases pre-synced successfully!")

            except Exception as e:
                print(f"[ERROR] Failed to pre-sync with libsql: {e}")
                import traceback
                traceback.print_exc()
                print(f"[WARNING] Continuing anyway - databases may be empty...")

            try:
                print(f"[DEBUG] Verifying synced data with SQLAlchemy...")
                with db.engine.connect() as conn:
                    # Try to query existing tables to trigger sync
                    try:
                        result = conn.execute(db.text("SELECT COUNT(*) FROM products"))
                        count = result.scalar()
                        print(f"[DEBUG] ✅ Data DB synced: {count} products found")
                    except Exception as e:
                        # Table doesn't exist yet, create it
                        print(f"[DEBUG] Data DB is empty, creating tables...")
                        db.create_all()
                        print(f"[DEBUG] Data DB tables created")

                print(f"[DEBUG] Syncing price database...")
                with db.engines["price"].connect() as conn:
                    try:
                        result = conn.execute(db.text("SELECT COUNT(*) FROM sku_pricing"))
                        count = result.scalar()
                        print(f"[DEBUG] ✅ Price DB synced: {count} SKU pricings found")
                    except Exception as e:
                        print(f"[DEBUG] Price DB is empty, creating tables...")
                        db.create_all(bind_key="price")
                        print(f"[DEBUG] Price DB tables created")

                print(f"[DEBUG] Syncing supplier database...")
                with db.engines["supplier"].connect() as conn:
                    try:
                        result = conn.execute(db.text("SELECT COUNT(*) FROM supplier_sku_master"))
                        count = result.scalar()
                        print(f"[DEBUG] ✅ Supplier DB synced: {count} supplier SKUs found")
                    except Exception as e:
                        print(f"[DEBUG] Supplier DB is empty, creating tables...")
                        db.create_all(bind_key="supplier")
                        print(f"[DEBUG] Supplier DB tables created")

                print(f"[INFO] ✅ All databases synced successfully!")

            except Exception as e:
                print(f"[ERROR] Failed to sync Turso databases: {e}")
                import traceback
                traceback.print_exc()
                # Fallback to create_all
                print(f"[DEBUG] Falling back to create_all()...")
                db.create_all()
                db.create_all(bind_key="price")
                db.create_all(bind_key="supplier")
        else:
            # Local SQLite mode: create tables normally
            try:
                print(f"[DEBUG] Creating main database tables...")
                db.create_all()  # Main database (data.db)
                print(f"[DEBUG] Main database tables created successfully")
            except Exception as e:
                print(f"[ERROR] Failed to create main database tables: {e}")
                raise

            try:
                print(f"[DEBUG] Creating price database tables...")
                db.create_all(bind_key="price")  # Price database (price.db)
                print(f"[DEBUG] Price database tables created successfully")
            except Exception as e:
                print(f"[ERROR] Failed to create price database tables: {e}")
                raise

            try:
                print(f"[DEBUG] Creating supplier database tables...")
                db.create_all(bind_key="supplier")  # Supplier database (supplier_stock.db)
                print(f"[DEBUG] Supplier database tables created successfully")
            except Exception as e:
                print(f"[ERROR] Failed to create supplier database tables: {e}")
                raise

        # ---- Price DB auto-migrate (SQLite): ensure new columns exist ----
        def _ensure_price_sku_pricing_columns():
            try:
                eng = get_engine("price")
                with eng.begin() as con:
                    cols = {row[1] for row in con.execute(text("PRAGMA table_info(sku_pricing)")).fetchall()}
                    if "brand" not in cols:
                        con.execute(text("ALTER TABLE sku_pricing ADD COLUMN brand TEXT"))
                    if "name" not in cols:
                        con.execute(text("ALTER TABLE sku_pricing ADD COLUMN name TEXT"))
                    if "stock_qty" not in cols:
                        con.execute(text("ALTER TABLE sku_pricing ADD COLUMN stock_qty INTEGER"))
                    if "stock_internal_qty" not in cols:
                        con.execute(text("ALTER TABLE sku_pricing ADD COLUMN stock_internal_qty INTEGER DEFAULT 0"))
                        # Backfill: ค่าเดิมที่เคยเก็บใน stock_qty ให้เป็น internal ด้วย
                        con.execute(
                            text(
                                """
                                UPDATE sku_pricing
                                   SET stock_internal_qty = COALESCE(stock_qty, 0)
                                 WHERE stock_internal_qty IS NULL OR stock_internal_qty = 0
                                """
                            )
                        )
                    if "monthly_sales_qty" not in cols:
                        con.execute(text("ALTER TABLE sku_pricing ADD COLUMN monthly_sales_qty INTEGER DEFAULT 0"))
            except Exception as e:
                app.logger.warning(f"[price] ensure sku_pricing columns failed: {e}")

        def _ensure_price_configs_columns():
            try:
                eng = get_engine("price")
                with eng.begin() as con:
                    cols = {row[1] for row in con.execute(text("PRAGMA table_info(price_configs)")).fetchall()}
                    if "worksheet" not in cols:
                        con.execute(text("ALTER TABLE price_configs ADD COLUMN worksheet TEXT"))
            except Exception as e:
                app.logger.warning(f"[price] ensure price_configs columns failed: {e}")

        def _ensure_platform_fee_settings_columns():
            try:
                eng = get_engine("price")
                with eng.begin() as con:
                    cols = {row[1] for row in con.execute(text("PRAGMA table_info(platform_fee_settings)")).fetchall()}

                    if "label" not in cols:
                        con.execute(text("ALTER TABLE platform_fee_settings ADD COLUMN label TEXT"))
                    if "is_active" not in cols:
                        con.execute(text("ALTER TABLE platform_fee_settings ADD COLUMN is_active INTEGER DEFAULT 1"))
                    if "sort_order" not in cols:
                        con.execute(text("ALTER TABLE platform_fee_settings ADD COLUMN sort_order INTEGER DEFAULT 0"))

                    # backfill label ให้แถวเก่า
                    con.execute(text("""
                        UPDATE platform_fee_settings
                           SET label = COALESCE(label, platform)
                         WHERE label IS NULL OR TRIM(label) = ''
                    """))
            except Exception as e:
                app.logger.warning(f"[price] ensure platform_fee_settings columns failed: {e}")

        def _ensure_market_items_columns():
            """Auto-migrate: เพิ่มคอลัมน์ใหม่ให้ market_items ใน price.db"""
            try:
                eng = get_engine("price")
                with eng.begin() as con:
                    cols = {row[1] for row in con.execute(text("PRAGMA table_info(market_items)")).fetchall()}
                    if "is_mall" not in cols:
                        con.execute(text("ALTER TABLE market_items ADD COLUMN is_mall INTEGER DEFAULT 0"))
            except Exception as e:
                app.logger.warning(f"[price] ensure market_items columns failed: {e}")

        _ensure_price_sku_pricing_columns()
        _ensure_price_configs_columns()
        _ensure_platform_fee_settings_columns()
        _ensure_market_items_columns()

        _ensure_orderline_print_columns()
        _migrate_shops_unique_to_platform_name()
        _ensure_issue_table()  # <<< NEW
        _ensure_deleted_table()  # <<< NEW สำหรับ Soft Delete
        _ensure_import_log_table()  # <<< NEW สำหรับ Import Dashboard
        _ensure_action_dedupe_table()  # <<< NEW กันกด/ส่งซ้ำ
        _ensure_shop_url_and_log_batch_columns()  # <<< NEW สำหรับบันทึก URL และ Batch Data
        # bootstrap admin
        if User.query.count() == 0:
            admin = User(
                username="admin",
                password_hash=generate_password_hash("admin123"),
                role="admin",
                active=True
            )
            db.session.add(admin)
            db.session.commit()

    # -----------------
    # Jinja filters
    # -----------------
    @app.template_filter("thai_be")
    def thai_be_filter(dt):
        try:
            return to_thai_be(dt)
        except Exception:
            return ""

    @app.template_filter("be_date")
    def be_date_filter(d):
        try:
            return to_be_date_str(d)
        except Exception:
            return ""

    # -----------------
    # UI context
    # -----------------
    @app.context_processor
    def inject_globals():
        return {
            "APP_NAME": APP_NAME,
            "BE_YEAR": current_be_year(),
            "CURRENT_USER": current_user()
        }

    # ให้ template ตรวจ endpoint ได้ (กันพังค่า has_endpoint)
    @app.template_global()
    def has_endpoint(endpoint: str) -> bool:
        try:
            return endpoint in app.view_functions
        except Exception:
            return False

    # -----------------
    # Auth helpers
    # -----------------
    def current_user():
        uid = session.get("uid")
        if not uid:
            return None
        return db.session.get(User, uid)

    def get_user_pref(user_id: int, key: str, default=None):
        try:
            row = PriceUserPreference.query.filter_by(user_id=user_id, key=key).first()
            return row.value if row and row.value is not None else default
        except Exception:
            return default

    def set_user_pref(user_id: int, key: str, value: str | None):
        try:
            row = PriceUserPreference.query.filter_by(user_id=user_id, key=key).first()
            if not row:
                row = PriceUserPreference(user_id=user_id, key=key, value=value)
                db.session.add(row)
            else:
                row.value = value
            db.session.commit()
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass

    def get_user_pref_main(user_id: int, key: str, default=None):
        try:
            row = UserPreference.query.filter_by(user_id=user_id, key=key).first()
            return row.value if row and row.value is not None else default
        except Exception:
            return default

    def set_user_pref_main(user_id: int, key: str, value: str | None):
        try:
            row = UserPreference.query.filter_by(user_id=user_id, key=key).first()
            if not row:
                row = UserPreference(user_id=user_id, key=key, value=value)
                db.session.add(row)
            else:
                row.value = value
            db.session.commit()
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass

    def login_required(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user():
                return redirect(url_for("login", next=request.path))
            return fn(*args, **kwargs)
        return wrapper

    # =========[ Database Management Helpers ]=========
    def get_db_paths():
        """Get paths for all 3 databases"""
        volume_path = os.environ.get("RAILWAY_VOLUME_MOUNT_PATH")
        if volume_path:
            return {
                "data": os.path.join(volume_path, "data.db"),
                "price": os.path.join(volume_path, "price.db"),
                "supplier": os.path.join(volume_path, "supplier_stock.db"),
            }
        else:
            base_dir = os.path.dirname(__file__)
            return {
                "data": os.path.join(base_dir, "data.db"),
                "price": os.path.join(base_dir, "price.db"),
                "supplier": os.path.join(base_dir, "supplier_stock.db"),
            }

    def get_db_info(db_path):
        """Get database file information"""
        info = {
            "exists": False,
            "size": "N/A",
            "size_bytes": 0,
            "modified": None
        }
        try:
            if os.path.exists(db_path):
                info["exists"] = True
                size_bytes = os.path.getsize(db_path)
                info["size_bytes"] = size_bytes
                info["size"] = f"{size_bytes / (1024 * 1024):.2f} MB"

                import time
                mtime = os.path.getmtime(db_path)
                info["modified"] = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(mtime))
        except Exception:
            pass
        return info

    @app.route("/api/userpref/set", methods=["POST"])
    @login_required
    def api_userpref_set():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "unauthorized"}), 401

        data = request.get_json(silent=True) or {}
        key = (data.get("key") or "").strip()
        value_raw = data.get("value", None)
        value = None if value_raw is None else str(value_raw).strip()

        if not key.startswith("supplier_stock."):
            return jsonify({"success": False, "msg": "key not allowed"}), 400

        set_user_pref_main(int(cu.id), key, value)
        return jsonify({"success": True})

    @app.route("/api/userpref/clear", methods=["POST"])
    @login_required
    def api_userpref_clear():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "unauthorized"}), 401

        data = request.get_json(silent=True) or {}
        keys = data.get("keys") or data.get("key") or []
        if isinstance(keys, str):
            keys = [keys]

        keys = [str(k).strip() for k in (keys or []) if str(k).strip()]
        keys = [k for k in keys if k.startswith("supplier_stock.")]

        if not keys:
            return jsonify({"success": False, "msg": "key not allowed"}), 400

        try:
            for k in keys:
                UserPreference.query.filter_by(user_id=int(cu.id), key=k).delete(
                    synchronize_session=False
                )
            db.session.commit()
            return jsonify({"success": True})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "msg": str(e)}), 400

    # -----------------
    # Utilities (app)
    # -----------------
    def parse_date_any(s: str | None):
        if not s:
            return None
        s = s.strip()
        try:
            if "-" in s:
                y, m, d = s.split("-")
                return date(int(y), int(m), int(d))
            else:
                d, m, y = s.split("/")
                y = int(y)
                if y > 2400:
                    y -= 543
                return date(y, int(m), int(d))
        except Exception:
            return None

    def _get_line_sku(line) -> str:
        if hasattr(line, "sku") and line.sku:
            return str(line.sku).strip()
        try:
            prod = getattr(line, "product", None)
            if prod and getattr(prod, "sku", None):
                return str(prod.sku).strip()
        except Exception:
            pass
        return ""

    def _calc_stock_qty_for_line(line: OrderLine) -> int:
        sku = _get_line_sku(line)
        if not sku:
            return 0
        prod = Product.query.filter_by(sku=sku).first()
        if prod and hasattr(prod, "stock_qty"):
            try:
                return int(prod.stock_qty or 0)
            except Exception:
                pass
        st = Stock.query.filter_by(sku=sku).first()
        try:
            return int(st.qty) if st and st.qty is not None else 0
        except Exception:
            return 0

    def _build_allqty_map(rows: list[dict]) -> dict[str, int]:
        total_by_sku: dict[str, int] = {}
        for r in rows:
            sku = (r.get("sku") or "").strip()
            if not sku:
                continue
            total_by_sku[sku] = total_by_sku.get(sku, 0) + int(r.get("qty", 0) or 0)
        return total_by_sku

    # [DEPRECATED] ฟังก์ชันนี้ไม่ใช้แล้ว - ใช้ compute_allocation() จาก allocation.py แทน
    # เก็บไว้สำหรับ reference เท่านั้น
    def _recompute_allocation_row(r: dict) -> dict:
        stock_qty = int(r.get("stock_qty", 0) or 0)
        allqty = int(r.get("allqty", r.get("qty", 0)) or 0)
        sales_status = (r.get("sales_status") or "").upper()
        packed_flag = bool(r.get("packed", False))
        accepted = bool(r.get("accepted", False))
        order_time = r.get("order_time")
        platform = r.get("platform") or (r.get("shop_platform") if r.get("shop_platform") else "")

        if sales_status == "PACKED" or packed_flag:
            allocation_status = "PACKED"
        elif accepted:
            allocation_status = "ACCEPTED"
        elif stock_qty <= 0:
            allocation_status = "SHORTAGE"
        elif allqty > stock_qty:
            allocation_status = "NOT_ENOUGH"
        elif stock_qty <= 3:
            allocation_status = "LOW_STOCK"
        else:
            allocation_status = "READY_ACCEPT"

        if allocation_status == "PACKED":
            sla = ""
        else:
            try:
                sla = sla_text(platform, order_time) if order_time else ""
            except Exception:
                sla = ""
        try:
            due_date = compute_due_date(platform, order_time) if order_time else None
        except Exception:
            due_date = None

        r["allocation_status"] = allocation_status
        r["sla"] = sla
        r["due_date"] = due_date
        return r

    def _check_mixed_status(order_id: str, all_rows: list[dict]) -> set:
        """
        ตรวจสอบว่า Order นี้มีสินค้าที่มีสถานะต่างกันปนอยู่หรือไม่
        คืนค่าเป็น set ของสถานะทั้งหมดที่พบใน Order นี้
        """
        statuses = set()
        for r in all_rows:
            if (r.get("order_id") or "").strip() == order_id:
                status = r.get("allocation_status")
                if status:
                    statuses.add(status)
        return statuses

    def _annotate_order_spans(rows: list[dict]) -> list[dict]:
        seen = set()
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                r["show_order_id"] = True
                r["order_id_display"] = ""
                continue
            if oid in seen:
                r["show_order_id"] = False
                r["order_id_display"] = ""
            else:
                r["show_order_id"] = True
                r["order_id_display"] = oid
                seen.add(oid)
        return rows

    def _group_rows_for_report(rows: list[dict]) -> list[dict]:
        def _key(r):
            return (
                (r.get("order_id") or ""),
                (r.get("platform") or ""),
                (r.get("shop") or ""),
                (r.get("logistic") or ""),
                (r.get("sku") or "")
            )
        rows = sorted(rows, key=_key)
        rows = _annotate_order_spans(rows)

        counts: dict[str, int] = {}
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            counts[oid] = counts.get(oid, 0) + 1

        for r in rows:
            oid = (r.get("order_id") or "").strip()
            r["order_rowspan"] = counts.get(oid, 1) if r.get("show_order_id") else 0
            r["order_id_display"] = oid if r.get("show_order_id") else ""
        return rows

    def _group_rows_for_warehouse_report(rows: list[dict]) -> list[dict]:
        """Group rows by order_id to show only 1 row per order for warehouse report"""
        order_map = {}
        
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                continue
            
            if oid not in order_map:
                # First row for this order - keep it
                # ใช้ printed_warehouse_count หรือ printed_count ที่มาจาก DB (ไม่ใช่ printed_warehouse ที่เป็น 0 ตลอด)
                order_map[oid] = {
                    "order_id": oid,
                    "platform": r.get("platform", ""),
                    "shop": r.get("shop", ""),
                    "logistic": r.get("logistic", ""),
                    "accepted_by": r.get("accepted_by", ""),
                    "printed_count": r.get("printed_warehouse_count") or r.get("printed_count") or r.get("printed_warehouse") or 0,
                    "printed_warehouse": r.get("printed_warehouse_count") or r.get("printed_count") or r.get("printed_warehouse") or 0,
                    "printed_warehouse_at": r.get("printed_warehouse_at"),
                    "printed_warehouse_by": r.get("printed_warehouse_by"),
                    "dispatch_round": r.get("dispatch_round"),
                    "scanned_at": r.get("scanned_at"),
                }
        
        # Convert back to list and sort
        result = list(order_map.values())
        result.sort(key=lambda r: (r["platform"], r["shop"], r["order_id"]))
        return result

    # -----------------
    # สร้างเซ็ต Order พร้อมรับทั้งออเดอร์ / สินค้าน้อยทั้งออเดอร์
    # -----------------
    def _orders_ready_set(rows: list[dict]) -> set[str]:
        by_oid: dict[str, list[dict]] = {}
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                continue
            by_oid.setdefault(oid, []).append(r)

        ready = set()
        for oid, items in by_oid.items():
            if not items:
                continue
            all_ready = True
            for it in items:
                status = (it.get("allocation_status") or "").upper()
                accepted = bool(it.get("accepted", False))
                packed = (status == "PACKED") or bool(it.get("packed", False))
                is_issued = bool(it.get("is_issued", False))  # [NEW] เช็ค Order จ่ายงานแล้ว
                # [แก้ไข] ถ้าจ่ายงานแล้ว (is_issued) ถือว่าจบงาน ไม่ต้องนับเข้ากอง 1
                if not (status == "READY_ACCEPT" and not accepted and not packed and not is_issued):
                    all_ready = False
                    break
            if all_ready:
                ready.add(oid)
        return ready

    def _orders_lowstock_order_set(rows: list[dict]) -> set[str]:
        by_oid: dict[str, list[dict]] = {}
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                continue
            by_oid.setdefault(oid, []).append(r)

        result = set()
        for oid, items in by_oid.items():
            if not items:
                continue
            all_sendable = True
            has_low = False
            for it in items:
                status = (it.get("allocation_status") or "").upper()
                accepted = bool(it.get("accepted", False))
                packed = (status == "PACKED") or bool(it.get("packed", False))
                is_issued = bool(it.get("is_issued", False))  # [NEW] เช็ค Order จ่ายงานแล้ว
                # [แก้ไข] ถ้าจ่ายงานแล้ว (is_issued) ถือว่าจบงาน ไม่ต้องนับเข้ากอง 2
                if packed or accepted or is_issued:
                    all_sendable = False
                    break
                if status not in ("READY_ACCEPT", "LOW_STOCK"):
                    all_sendable = False
                    break
                if status == "LOW_STOCK":
                    has_low = True
            if all_sendable and has_low:
                result.add(oid)
        return result

    # ===================== NEW: Orders ที่ยังไม่มีการเปิดใบขาย =====================
    def _has_any_sales(r: dict) -> bool:
        """คืน True ถ้า row นี้ 'มีการเปิดใบขายแล้วบางส่วนหรือทั้งหมด'"""
        sales_status = (str(r.get("sales_status") or "")).strip()
        po_no = (str(r.get("po_no") or "")).strip()
        return bool(sales_status or po_no)

    def _orders_packed_set(rows: list[dict]) -> set[str]:
        """คืน set ของ order_id ที่ถือว่า PACKED (ระดับ Order)"""
        packed: set[str] = set()
        for r in rows or []:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                continue
            if bool(r.get("is_packed")) or bool(r.get("packed")):
                packed.add(oid)
                continue
            status = (r.get("allocation_status") or "").strip().upper()
            if status == "PACKED":
                packed.add(oid)
        return packed

    def _orders_not_in_sbs_set(rows: list[dict]) -> set[str]:
        """คืน set ของ order_id ที่ยังไม่เข้า SBS (ไม่มี po_no และไม่มี sales_status)"""
        result: set[str] = set()
        for r in rows or []:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                continue
            if not _has_any_sales(r):
                result.add(oid)
        return result

    def _orders_no_sales_set(rows: list[dict]) -> set[str]:
        """คืน set ของ order_id ที่สถานะใบขายเป็น 'ยังไม่มีการเปิดใบขาย'"""
        result: set[str] = set()
        for r in rows or []:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                continue
            sales_status = (str(r.get("sales_status") or "")).strip()
            sales_status_u = sales_status.upper()
            if sales_status and ("ยังไม่มี" in sales_status or "NO_SALES" in sales_status_u):
                result.add(oid)
        return result

    def _print_cols_for_kind(kind: str) -> tuple[str, str, str]:
        """Map kind -> (count_col, at_col, by_col) in order_lines."""
        kind = (kind or "").strip().lower()
        if kind == "warehouse":
            return ("printed_warehouse", "printed_warehouse_at", "printed_warehouse_by")
        if kind == "picking":
            return ("printed_picking", "printed_picking_at", "printed_picking_by")
        if kind == "lowstock":
            return ("printed_lowstock", "printed_lowstock_at", "printed_lowstock_by")
        if kind == "nostock":
            return ("printed_nostock", "printed_nostock_at", "printed_nostock_by")
        if kind == "notenough":
            return ("printed_notenough", "printed_notenough_at", "printed_notenough_by")
        # default
        return ("printed_picking", "printed_picking_at", "printed_picking_by")

    def _detect_already_printed(oids: list[str], kind: str) -> set[str]:
        """คืน set ของ order_id ที่เคยพิมพ์ kind นี้ไปแล้ว (count > 0)"""
        if not oids:
            return set()
        tbl = _ol_table_name()
        col, _, _ = _print_cols_for_kind(kind)
        sql = text(
            f"SELECT DISTINCT order_id FROM {tbl} WHERE order_id IN :oids AND COALESCE({col},0) > 0"
        ).bindparams(bindparam("oids", expanding=True))
        rows_sql = db.session.execute(sql, {"oids": oids}).fetchall()
        return {str(r[0]) for r in rows_sql if r and r[0]}

    def _mark_printed(
        oids: list[str],
        kind: str,
        user_id: int | None,
        when_iso: str,
        commit: bool = True,
    ) -> int:
        """บวกจำนวนครั้งพิมพ์ + อัปเดตเวลา/ผู้พิมพ์ (ระดับ order_id)"""
        if not oids:
            return 0

        username = None
        if user_id is not None:
            try:
                u = User.query.get(int(user_id))
                username = (u.username if u else None)
            except Exception:
                username = None

        tbl = _ol_table_name()
        col, col_at, col_by = _print_cols_for_kind(kind)
        sql = text(
            f"""
            UPDATE {tbl}
               SET {col}=COALESCE({col},0)+1,
                   {col_at}=:ts,
                   {col_by}=:byu
             WHERE order_id IN :oids
            """
        ).bindparams(bindparam("oids", expanding=True))

        res = db.session.execute(sql, {"ts": when_iso, "byu": username, "oids": oids})
        if commit:
            db.session.commit()
        try:
            return int(getattr(res, "rowcount", 0) or 0)
        except Exception:
            return 0

    def _get_print_counts_local(oids: list[str], kind: str) -> dict[str, int]:
        if not oids:
            return {}
        tbl = _ol_table_name()
        if kind == "warehouse":
            col = "printed_warehouse"
        elif kind == "picking":
            col = "printed_picking"
        elif kind == "lowstock":
            col = "printed_lowstock"
        elif kind == "nostock":
            col = "printed_nostock"
        elif kind == "notenough":
            col = "printed_notenough"
        else:
            col = "printed_picking"

        sql = text(
            f"SELECT order_id, COALESCE(MAX({col}),0) AS c FROM {tbl} WHERE order_id IN :oids GROUP BY order_id"
        ).bindparams(bindparam("oids", expanding=True))
        rows_sql = db.session.execute(sql, {"oids": oids}).fetchall()
        return {str(r0): int(c or 0) for (r0, c) in rows_sql if r0}

    def _mark_lowstock_printed(oids: list[str], username: str | None, when_iso: str):
        """อัปเดตการพิมพ์สำหรับรายงานสินค้าน้อย"""
        if not oids:
            return
        tbl = _ol_table_name()
        sql = text(
            f"""
            UPDATE {tbl}
               SET printed_lowstock=COALESCE(printed_lowstock,0)+1,
                   printed_lowstock_at=:ts,
                   printed_lowstock_by=:byu
             WHERE order_id IN :oids
            """
        ).bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"ts": when_iso, "byu": username, "oids": oids})
        db.session.commit()

    def _mark_nostock_printed(oids: list[str], username: str | None, when_iso: str):
        """อัปเดตการพิมพ์สำหรับรายงานไม่มีสินค้า"""
        if not oids:
            return
        tbl = _ol_table_name()
        sql = text(f"""
            UPDATE {tbl}
               SET printed_nostock=COALESCE(printed_nostock,0)+1,
                   printed_nostock_at=:ts,
                   printed_nostock_by=:byu
             WHERE order_id IN :oids
        """).bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"ts": when_iso, "byu": username, "oids": oids})
        db.session.commit()

    def _mark_notenough_printed(oids: list[str], username: str | None, when_iso: str):
        """อัปเดตการพิมพ์สำหรับรายงานสินค้าไม่พอส่ง"""
        if not oids:
            return
        tbl = _ol_table_name()
        sql = text(f"""
            UPDATE {tbl}
               SET printed_notenough=COALESCE(printed_notenough,0)+1,
                   printed_notenough_at=:ts,
                   printed_notenough_by=:byu
             WHERE order_id IN :oids
        """).bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"ts": when_iso, "byu": username, "oids": oids})
        db.session.commit()

    def _inject_scan_status(rows: list[dict]):
        """ดึงข้อมูลว่าออเดอร์ไหนสแกนแล้วบ้าง"""
        oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        if not oids:
            return
        
        tbl = _ol_table_name()
        sql = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
        sql = sql.bindparams(bindparam("oids", expanding=True))
        res = db.session.execute(sql, {"oids": oids}).fetchall()
        scan_map = {r[0]: r[1] for r in res if r[0]}
        
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            r["scanned_at"] = scan_map.get(oid)

    def _inject_print_counts_to_rows(rows: list[dict], kind: str):
        """ฝัง printed_*_count และ printed_*_at ลงในแต่ละแถว (ใช้กับ Warehouse report)"""
        oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        counts = _get_print_counts_local(oids, kind)
        
        # Also get the timestamp of last print
        if not oids:
            return
        
        tbl = _ol_table_name()
        col_at = "printed_warehouse_at" if kind == "warehouse" else "printed_picking_at"
        sql = text(f"SELECT order_id, MAX({col_at}) AS last_printed_at FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
        sql = sql.bindparams(bindparam("oids", expanding=True))
        rows_sql = db.session.execute(sql, {"oids": oids}).all()
        timestamps = {}
        
        # Convert ISO string to datetime object
        for r_sql in rows_sql:
            if r_sql and r_sql[0] and r_sql[1]:
                try:
                    # Parse ISO datetime string
                    dt = datetime.fromisoformat(r_sql[1])
                    if dt.tzinfo is None:
                        dt = TH_TZ.localize(dt)
                    timestamps[str(r_sql[0])] = dt
                except Exception:
                    pass
        
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            c = int(counts.get(oid, 0))
            r["printed_count"] = c
            if kind == "warehouse":
                r["printed_warehouse_count"] = c
                r["printed_warehouse"] = c  # <-- เพิ่มบรรทัดนี้เพื่อให้เทมเพลตอ่ยใช้ได้
                r["printed_warehouse_at"] = timestamps.get(oid)
            else:
                r["printed_picking_count"] = c
                r["printed_picking"] = c  # <-- และบรรทัดน้
                r["printed_picking_at"] = timestamps.get(oid)

    # =========[ NEW ]=========
    # ส่วนเสริมเพื่อ "Order ยกเลิก"
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        _OPENPYXL_OK = True
    except Exception:
        _OPENPYXL_OK = False
        ILLEGAL_CHARACTERS_RE = None

    def sanitize_excel_value(v):
        """Remove Excel-illegal control characters from strings.

        openpyxl raises IllegalCharacterError when these characters appear.
        """
        if ILLEGAL_CHARACTERS_RE is not None and isinstance(v, str):
            return ILLEGAL_CHARACTERS_RE.sub("", v)
        return v

    def sanitize_excel_df(df: pd.DataFrame) -> pd.DataFrame:
        """Sanitize dataframe cells before exporting to Excel."""
        if ILLEGAL_CHARACTERS_RE is None:
            return df

        try:
            obj_cols = df.select_dtypes(include=["object"]).columns
            if len(obj_cols) == 0:
                return df
            df[obj_cols] = df[obj_cols].applymap(sanitize_excel_value)
            return df
        except Exception:
            # Fallback: sanitize everything (slower but safe)
            try:
                return df.applymap(sanitize_excel_value)
            except Exception:
                return df

    def _ensure_cancel_table():
        try:
            CancelledOrder.__table__.create(bind=db.engine, checkfirst=True)
        except Exception as e:
            app.logger.warning(f"[cancelled_orders] ensure table failed: {e}")

    def _cancelled_oids_set() -> set[str]:
        """คืนค่า set ของ order_id ที่ถูกยกเลิก (สำหรับ backward compatibility)"""
        rows = db.session.query(CancelledOrder.order_id).all()
        return {r[0] for r in rows if r and r[0]}

    def _cancelled_oids_map() -> dict[str, dict]:
        """คืนค่า dict ของ {order_id: {'note': note, 'at': timestamp}}"""
        rows = db.session.query(
            CancelledOrder.order_id, 
            CancelledOrder.note, 
            CancelledOrder.imported_at
        ).all()
        # เก็บทั้ง Note และ เวลา
        return {r[0]: {'note': (r[1] or ""), 'at': r[2]} for r in rows if r and r[0]}

    def _filter_out_cancelled_rows(rows: list[dict]) -> list[dict]:
        canc = _cancelled_oids_set()
        if not canc:
            return rows
        res = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if oid and oid in canc:
                continue
            res.append(r)
        return res

    # ===== HELPER: Issued (จ่ายงานแล้ว) =====
    def _issued_oids_set() -> set[str]:
        rows = db.session.query(IssuedOrder.order_id).all()
        return {r[0] for r in rows if r and r[0]}

    def _filter_out_issued_rows(rows: list[dict]) -> list[dict]:
        issued = _issued_oids_set()
        if not issued:
            return rows
        res = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if oid and oid in issued:
                continue
            res.append(r)
        return res

    # ===== HELPER: Deleted Orders (ถูกลบ / ถังขยะ) =====
    def _deleted_oids_set() -> set[str]:
        """ดึง order_id ทั้งหมดที่ถูกลบ (Soft Delete)"""
        rows = db.session.query(DeletedOrder.order_id).all()
        return {r[0] for r in rows if r and r[0]}

    def _filter_out_deleted_rows(rows: list[dict]) -> list[dict]:
        """กรอง order ที่ถูกลบออกจากรายการ"""
        deleted = _deleted_oids_set()
        if not deleted:
            return rows
        res = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if oid and oid in deleted:
                continue
            res.append(r)
        return res

    # ===== HELPER: Low Stock Printed (พิมพ์รายงานสินค้าน้อยแล้ว) =====
    def _lowstock_printed_oids_set() -> set[str]:
        """ดึง order_id ที่เคยพิมพ์รายงานสินค้าน้อยแล้ว"""
        tbl = _ol_table_name()
        rows = db.session.execute(text(f"""
            SELECT DISTINCT order_id
            FROM {tbl}
            WHERE printed_lowstock > 0
        """)).fetchall()
        return {r[0] for r in rows if r and r[0]}

    def _filter_out_lowstock_printed_rows(rows: list[dict]) -> list[dict]:
        """กรองออเดอร์ที่พิมพ์รายงานสินค้าน้อยออกแล้ว (ข้อ 2)"""
        printed = _lowstock_printed_oids_set()
        if not printed:
            return rows
        res = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if oid and oid in printed:
                continue
            res.append(r)
        return res

    def _mark_issued(oids: list[str], user_id: int | None, source: str = "manual", when_dt=None, commit: bool = True):
        """ทำเครื่องหมาย 'จ่ายงานแล้ว' โดยไม่แก้ทับข้อมูลเก่า (ยึดเวลาเดิม)"""
        if not oids:
            return 0
        # ใช้เวลาที่ส่งมา (เช่น ตอน import) ถ้าไม่ส่งมาก็ใช้เวลาปัจจุบันโซนไทย
        when_dt = when_dt or now_thai()
        try:
            # เก็บแบบ naive เพื่อให้ SQLite รับได้
            if getattr(when_dt, "tzinfo", None) is not None:
                when_dt = when_dt.replace(tzinfo=None)
        except Exception:
            pass

        existing = {
            r[0] for r in db.session.query(IssuedOrder.order_id)
            .filter(IssuedOrder.order_id.in_(oids)).distinct().all()
        }
        inserted = 0
        for oid in oids:
            oid = (oid or "").strip()
            if not oid or oid in existing:
                # มีข้อมูลเก่าแล้ว (เช่นมาจากการพิมพ์) ก็ไม่แก้ทับ ⇒ ยึดเวลาเก่าไว้
                continue
            db.session.add(IssuedOrder(order_id=oid, issued_at=when_dt, issued_by_user_id=user_id, source=source))
            inserted += 1
        if commit:
            db.session.commit()
        return inserted

    def _unissue(oids: list[str]) -> int:
        if not oids:
            return 0
        n = db.session.query(IssuedOrder).filter(IssuedOrder.order_id.in_(oids)).delete(synchronize_session=False)
        db.session.commit()
        return n

    # ให้ import "จ่ายงานแล้ว" ตั้งค่า counter ขั้นต่ำเป็น 1
    def _ensure_min_print_count(oids: list[str], min_count: int = 1, user_id: int | None = None, when_iso: str | None = None):
        """บังคับให้ printed_picking_count >= min_count (เฉพาะ Picking เท่านั้น)"""
        if not oids:
            return
        tbl = _ol_table_name()
        when_iso = when_iso or now_thai().isoformat()

        # เซ็ตเฉพาะ Picking (ไม่แตะ Warehouse)
        sql = text(f"""
            UPDATE {tbl}
               SET printed_picking=1,
                   printed_picking_count = CASE WHEN COALESCE(printed_picking_count,0) < :mc THEN :mc ELSE printed_picking_count END,
                   printed_picking_by_user_id = COALESCE(printed_picking_by_user_id, :uid),
                   printed_picking_at = COALESCE(printed_picking_at, :ts)
             WHERE order_id IN :oids
        """).bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"mc": min_count, "uid": user_id, "ts": when_iso, "oids": oids})

        db.session.commit()

    def _ensure_shops_from_df(df, platform: str, default_shop_name: str = None):
        """สร้างหรือใช้ Shop ที่มีอยู่แล้ว ก่อนที่จะ import orders (กัน UNIQUE constraint พัง)"""
        from utils import normalize_platform
        platform = normalize_platform(platform)
        
        # รวบรวม shop names ที่มีใน df (ลองดูหลายคอลัมน์ที่อาจมีชื่อร้าน)
        shop_names = set()
        for col in df.columns:
            col_lower = str(col).lower()
            if "shop" in col_lower or "ร้าน" in col_lower:
                for val in df[col].dropna().unique():
                    name = str(val).strip()
                    if name:
                        shop_names.add(name)
        
        # ถ้าไม่เจอใน df ให้ใช้ default_shop_name
        if not shop_names and default_shop_name:
            shop_names.add(default_shop_name.strip())
        
        # สร้าง/ใช้ shop ที่มีอยู่แล้ว
        for name in shop_names:
            existing = Shop.query.filter_by(platform=platform, name=name).first()
            if not existing:
                new_shop = Shop(platform=platform, name=name)
                db.session.add(new_shop)
        db.session.commit()

    def _parse_order_ids_from_upload(file_storage) -> list[str]:
        filename = (file_storage.filename or "").lower()
        data = file_storage.read()
        file_storage.stream.seek(0)

        order_ids: list[str] = []

        # Excel
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            if not _OPENPYXL_OK:
                raise RuntimeError("ไม่พบไลบรารี openpyxl สำหรับอ่านไฟล์ Excel, ติดตั้งด้วย: pip install openpyxl")
            wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not row:
                    continue
                val = row[0]
                if i == 1 and isinstance(val, str) and val.strip().lower() in {"order_id", "order_no", "เลขออเดอร์"}:
                    continue
                if val is None:
                    continue
                s = str(val).strip()
                if s:
                    order_ids.append(s)
            return order_ids

        # CSV
        if filename.endswith(".csv"):
            text_data = data.decode("utf-8-sig", errors="ignore")
            reader = csv.reader(text_data.splitlines())
            for i, row in enumerate(reader, start=1):
                if not row:
                    continue
                val = row[0]
                if i == 1 and isinstance(val, str) and val.strip().lower() in {"order_id", "order_no", "เลขออเดอร์"}:
                    continue
                s = str(val).strip()
                if s:
                    order_ids.append(s)
            return order_ids

        raise RuntimeError("รองรับเฉพาะไฟล์ .xlsx .xls หรือ .csv เท่านั้น")
    # =========[ /NEW ]=========

    # -------------
    # Routes: Auth & Users
    # -------------

    # --------- Admin: Shops (เดิม) ---------
    @app.route("/admin/shops")
    @login_required
    def admin_shops():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ต้องเป็นผู้ดูแลระบบหรือพนักงานเท่านั้น", "danger")
            return redirect(url_for("dashboard"))
        # Filter ออก system config shops (CANCEL_SYSTEM, EMPTY_BILL_SYSTEM, etc.)
        shops = Shop.query.filter_by(is_system_config=False).order_by(Shop.platform.asc(), Shop.name.asc()).all()
        counts = {s.id: db.session.query(func.count(OrderLine.id)).filter_by(shop_id=s.id).scalar() for s in shops}
        return render_template("admin_shops.html", shops=shops, counts=counts)

    @app.route("/admin/shops/<int:shop_id>/delete", methods=["POST"])
    @login_required
    def delete_shop(shop_id):
        cu = current_user()
        if not cu or cu.role != "admin":
            flash("เฉพาะแอดมินเท่านั้นที่ลบได้", "danger")
            return redirect(url_for("admin_shops"))
        s = Shop.query.get(shop_id)
        if not s:
            flash("ไม่พบร้านนี้", "warning")
            return redirect(url_for("admin_shops"))
        cnt = db.session.query(func.count(OrderLine.id)).filter_by(shop_id=s.id).scalar()
        if cnt and cnt > 0:
            flash("ไม่สามารถลบได้: มีออเดอร์ผูกกับร้านนี้อยู่", "danger")
            return redirect(url_for("admin_shops"))
        db.session.delete(s)
        db.session.commit()
        flash(f"ลบร้าน '{s.name}' แล้ว", "success")
        return redirect(url_for("admin_shops"))
    # --------------------------------------

    # -----------------------
    # API: จัดการ Link Google Sheet ของร้าน
    # -----------------------
    @app.route("/api/shop/url", methods=["POST"])
    @login_required
    def api_shop_url():
        data = request.get_json() or {}
        shop_name = (data.get("shop_name") or "").strip()
        platform = normalize_platform(data.get("platform") or "")
        url = (data.get("url") or "").strip()
        action = data.get("action")  # 'save' or 'delete'
        
        # ถ้าไม่ได้ระบุชื่อร้าน ให้ใช้ชื่อ Platform เป็นชื่อร้านแทน (เพื่อเก็บ URL กลาง)
        if not shop_name and platform:
            shop_name = platform
        
        if not shop_name:
            return jsonify({"success": False, "msg": "กรุณาเลือกแพลตฟอร์ม"})
        
        # ค้นหาร้านตาม platform + name
        shop = Shop.query.filter_by(platform=platform, name=shop_name).first()
        if not shop:
            # ถ้าไม่มี platform ตรง ลองหาแค่ชื่อร้าน
            shop = Shop.query.filter_by(name=shop_name).first()
        
        if not shop:
            # สร้างร้านใหม่ถ้ายังไม่มี
            shop = Shop(platform=platform or "อื่นๆ", name=shop_name)
            db.session.add(shop)
            db.session.commit()  # Commit เพื่อให้ได้ ID มาใช้
        
        if action == "save":
            # [แก้ไข] ใช้ SQL Update ตรงๆ
            db.session.execute(
                text("UPDATE shops SET google_sheet_url = :u WHERE id = :id"),
                {"u": url, "id": shop.id}
            )
            db.session.commit()
            return jsonify({"success": True, "msg": "บันทึกลิงก์เรียบร้อย"})
        elif action == "delete":
            # [แก้ไข] ใช้ SQL Update ตรงๆ (Set เป็น NULL)
            db.session.execute(
                text("UPDATE shops SET google_sheet_url = NULL WHERE id = :id"),
                {"id": shop.id}
            )
            db.session.commit()
            return jsonify({"success": True, "msg": "ลบลิงก์เรียบร้อย"})
            
        return jsonify({"success": False, "msg": "ไม่ระบุ action"})
    # --------------------------------------

    # -----------------------
    # API: เก็บค่า Google Sheet URL (config) แบบระบุ platform+name ชัดเจน
    # ใช้กับหน้าที่ต้องการ "จำ URL" โดยไม่ไปชนกับร้านจริง
    # -----------------------
    @app.route("/api/price/config/gsheet_url", methods=["POST"])
    @login_required
    def api_price_config_gsheet_url():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            return jsonify({"success": False, "msg": "ไม่มีสิทธิ์ใช้งาน"}), 403

        data = request.get_json() or {}
        platform = (data.get("platform") or "").strip()
        name = (data.get("name") or "").strip()
        action = (data.get("action") or "").strip().lower()
        url = (data.get("url") or "").strip()
        worksheet = (data.get("worksheet") or "").strip()
        worksheet_in_payload = ("worksheet" in data)

        if not platform or not name:
            return jsonify({"success": False, "msg": "ข้อมูลไม่ครบ (platform/name)"}), 400

        try:
            cfg = PriceConfig.query.filter_by(platform=platform, name=name).first()

            if action == "save":
                if not cfg:
                    if not url:
                        return jsonify({"success": False, "msg": "กรุณาใส่ URL"}), 400
                    cfg = PriceConfig(platform=platform, name=name)
                    db.session.add(cfg)

                # url: ถ้าส่งมาให้ set, ถ้าไม่ส่งมาให้คงค่าเดิมไว้
                if url:
                    cfg.url = url
                elif not (cfg.url or ""):
                    return jsonify({"success": False, "msg": "กรุณาใส่ URL"}), 400

                # worksheet: ถ้า payload มี key worksheet ให้ set (ส่ง "" = ล้างชื่อแท็บ)
                if worksheet_in_payload:
                    cfg.worksheet = worksheet or None

                db.session.commit()
                return jsonify({"success": True, "msg": "บันทึกค่า Google Sheet (URL/Worksheet) เรียบร้อย"})

            if action == "delete":
                if cfg:
                    db.session.delete(cfg)
                    db.session.commit()
                return jsonify({"success": True, "msg": "ลบลิงก์ (PRICE) เรียบร้อย"})

            return jsonify({"success": False, "msg": "action ต้องเป็น save หรือ delete"}), 400

        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "msg": f"เกิดข้อผิดพลาด: {e}"}), 500

    @app.route("/api/config/gsheet_url", methods=["POST"])
    @login_required
    def api_config_gsheet_url():
        # alias เพื่อกันหน้าเก่า/โค้ดเก่าแตก: ให้ชี้ไปฝั่ง PRICE
        return api_price_config_gsheet_url()
    # --------------------------------------

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            u = User.query.filter_by(username=username, active=True).first()
            if u and check_password_hash(u.password_hash, password):
                session["uid"] = u.id
                flash("เข้าสู่ระบบสำเร็จ", "success")
                return redirect(request.args.get("next") or url_for("dashboard"))
            flash("ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง", "danger")
        return render_template("login.html")

    @app.route("/logout")
    def logout():
        session.clear()
        flash("ออกจากระบบแล้ว", "info")
        return redirect(url_for("login"))

    # -----------------
    # Price Dashboard: KPI multi-select helpers
    # -----------------
    def _parse_kpi_multi(kpi_raw: str | None, allowed: set[str]) -> list[str]:
        raw = (kpi_raw or "").strip()
        if not raw:
            return []
        arr = [s.strip() for s in raw.split(",") if s.strip()]
        # ไม่ให้ tracked เป็นเงื่อนไขกรอง
        arr = [s for s in arr if s != "tracked"]
        # allow เฉพาะที่รู้จัก
        arr = [s for s in arr if s in allowed]
        # unique + preserve order
        seen: set[str] = set()
        out: list[str] = []
        for s in arr:
            if s not in seen:
                seen.add(s)
                out.append(s)
        return out

    def _apply_kpi_filters(rows: list[dict], kpi_selected: list[str]) -> list[dict]:
        """AND filter: ต้องเข้าเงื่อนไขทุกข้อที่ติ๊ก"""
        out = rows
        for key in kpi_selected or []:
            if key == "no_sales":
                out = [r for r in out if r.get("no_sales")]
            elif key == "need_market":
                out = [r for r in out if r.get("need_market")]
            elif key == "brand_control":
                out = [
                    r for r in out
                    if (r.get("brand_control") is not None) and (float(r.get("brand_control") or 0) > 0)
                ]
            elif key in {"aging3", "aging6", "aging12"}:
                out = [r for r in out if r.get("aging_bucket") == key]
            else:
                out = [r for r in out if key in (r.get("rec_keys") or [])]
        return out

    def _compute_kpi_counts(rows: list[dict], allowed_keys: set[str]) -> dict[str, int]:
        """นับจำนวน KPI ต่อ key จากชุด rows ที่ส่งเข้ามา (ใช้กับ rows_master)."""
        counts: dict[str, int] = {k: 0 for k in (allowed_keys or set())}
        for r in rows or []:
            for k in (r.get("rec_keys") or []):
                if k in counts:
                    counts[k] += 1

            b = r.get("aging_bucket")
            if b in counts:
                counts[b] += 1

            if r.get("no_sales") and ("no_sales" in counts):
                counts["no_sales"] += 1

            if "brand_control" in counts:
                bc = r.get("brand_control")
                if (bc is not None) and (float(bc or 0) > 0):
                    counts["brand_control"] += 1

            if "need_market" in counts and r.get("need_market"):
                counts["need_market"] += 1

        return counts

    def _resolve_master(
        master_raw: str | None,
        kpi_selected: list[str],
        allowed_keys: set[str],
    ) -> tuple[str, list[str]]:
        """Resolve master KPI (single) and sub-KPIs (multi).

        Contract after 2025-12:
        - `master` is a single primary KPI (stored in URL as `master=...`)
        - `kpi_selected` is the list of *sub* KPIs (stored in URL as `kpi=a,b,c`)
        - master MUST NOT be duplicated in `kpi_selected`
        - If master is missing but sub-KPIs exist, promote the first sub-KPI to master.
        """
        master = (master_raw or "").strip()
        if master not in (allowed_keys or set()):
            master = ""

        # Ensure master is not duplicated in sub KPIs
        if master:
            kpi_selected = [k for k in (kpi_selected or []) if k != master]

        # If no master but there are sub KPIs, promote the first to master
        if (not master) and (kpi_selected or []):
            master = kpi_selected[0]
            kpi_selected = kpi_selected[1:]

        return master, (kpi_selected or [])

    @app.route("/price/dashboard", methods=["GET"])
    @login_required
    def price_dashboard():
        cu = current_user()
        uid = cu.id if cu else None

        # ---- Auto rule configs (per user) ----
        r10_cfg = _norm_r10_cfg(_load_cfg_dict(get_user_pref(int(uid), "pm_auto_r10_cfg", default="") if uid else "", _R10_DEFAULT_CFG))
        r11_cfg = _norm_r11_cfg(_load_cfg_dict(get_user_pref(int(uid), "pm_auto_r11_cfg", default="") if uid else "", _R11_DEFAULT_CFG))
        r10_min_loss = float(r10_cfg.get("min_loss_pct") or _R10_DEFAULT_CFG["min_loss_pct"])
        r11_min_loss = float(r11_cfg.get("min_loss_pct") or _R11_DEFAULT_CFG["min_loss_pct"])
        r11_max_loss = float(r11_cfg.get("max_loss_pct") or _R11_DEFAULT_CFG["max_loss_pct"])

        platform_rows = (
            PlatformFeeSetting.query
            .filter(PlatformFeeSetting.is_active == True)
            .order_by(PlatformFeeSetting.sort_order.asc(), PlatformFeeSetting.platform.asc())
            .all()
        )

        # seed ค่าเริ่มต้นกันหน้าแตก (กรณีเพิ่งติดตั้ง/ยังไม่เคยตั้งค่า)
        if not platform_rows:
            for p_key, p_label in [("Shopee", "Shopee"), ("Lazada", "Lazada"), ("TikTok", "TikTok")]:
                db.session.add(
                    PlatformFeeSetting(
                        platform=p_key,
                        label=p_label,
                        fee_pct=0.0,
                        fixed_fee=0.0,
                        is_active=True,
                        sort_order=0,
                    )
                )
            db.session.commit()
            platform_rows = (
                PlatformFeeSetting.query
                .filter(PlatformFeeSetting.is_active == True)
                .order_by(PlatformFeeSetting.sort_order.asc(), PlatformFeeSetting.platform.asc())
                .all()
            )

        platforms = [(p.platform, (p.label or p.platform)) for p in platform_rows]
        allowed = {p[0] for p in platforms}

        default_platform = platforms[0][0] if platforms else "Shopee"

        platform_arg = (request.args.get("platform") or "").strip()

        # ถ้า URL ไม่ส่ง platform มา ให้ใช้ค่าที่เคยเลือกไว้ (ผูกกับ user)
        if not platform_arg and uid:
            platform_arg = (get_user_pref(uid, "price_dash_platform", "") or "").strip()

        platform_norm = normalize_platform(platform_arg) if platform_arg else ""
        platform = platform_norm if platform_norm in allowed else default_platform

        # เซฟค่าที่ใช้งานจริงกลับไปเป็นค่า default ครั้งถัดไป
        if uid:
            prev = (get_user_pref(uid, "price_dash_platform", "") or "").strip()
            if prev != platform:
                set_user_pref(uid, "price_dash_platform", platform)

        q = (request.args.get("q") or "").strip().lower()

        owner_sel = (request.args.get("owner") or "").strip()
        limit_sel = (request.args.get("limit") or "").strip().lower()

        # -----------------------------
        # Market stale days (for need_market KPI)
        # -----------------------------
        stale_days_raw = (request.args.get("stale_days") or "").strip()
        if uid and ("stale_days" not in request.args):
            stale_days_raw = (get_user_pref(uid, "price_dash_stale_days", "14") or "14").strip()
        try:
            stale_days = int(stale_days_raw or 14)
        except Exception:
            stale_days = 14
        stale_days = max(1, min(stale_days, 365))
        stale_days_sel = str(stale_days)

        now0 = now_thai()
        try:
            if getattr(now0, "tzinfo", None) is not None:
                now0 = now0.replace(tzinfo=None)
        except Exception:
            pass
        cutoff = now0 - timedelta(days=stale_days)

        kpi_sel = (request.args.get("kpi") or "").strip()

        # -----------------------------
        # Sort (server-side) + remember per-user
        # -----------------------------
        sort_sel = (request.args.get("sort") or "").strip()
        sort_dir = (request.args.get("dir") or "").strip().lower() or "asc"
        clear_sort = (request.args.get("clear_sort") or "").strip() == "1"

        if sort_dir not in {"asc", "desc"}:
            sort_dir = "asc"

        ALLOWED_SORTS = {
            "sku", "brand", "name",
            "stock_internal", "stock", "monthly_sales",
            "cost", "our_price", "market_best", "voucher", "brand_control",
            "gap", "profit_our", "profit_match", "recommend",
            "shop", "mall", "url", "owner", "updated",
        }

        # ถ้า URL ไม่ส่ง owner/limit มา ให้ใช้ค่าที่ล็อกไว้ (ผูกกับ user)
        # หมายเหตุ: ต้องแยกกรณี "ไม่มีพารามิเตอร์ owner" ออกจาก "ส่ง owner='' (ทั้งหมด)"
        if uid:
            if "owner" not in request.args:
                owner_sel = (get_user_pref(uid, "price_dash_owner", "") or "").strip()
            if ("limit" not in request.args) or (not limit_sel):
                limit_sel = (get_user_pref(uid, "price_dash_limit", "200") or "200").strip().lower()

            if clear_sort:
                sort_sel = ""
                sort_dir = "asc"
                set_user_pref(uid, "price_dash_sort", "")
                set_user_pref(uid, "price_dash_dir", "asc")
            else:
                if "sort" not in request.args:
                    sort_sel = (get_user_pref(uid, "price_dash_sort", "") or "").strip()
                if "dir" not in request.args:
                    sort_dir = (get_user_pref(uid, "price_dash_dir", "asc") or "asc").strip().lower() or "asc"

                if "sort" in request.args:
                    set_user_pref(uid, "price_dash_sort", sort_sel)
                if "dir" in request.args:
                    set_user_pref(uid, "price_dash_dir", sort_dir)

        if sort_dir not in {"asc", "desc"}:
            sort_dir = "asc"
        if sort_sel not in ALLOWED_SORTS:
            sort_sel = ""
            sort_dir = "asc"

        if not limit_sel:
            limit_sel = "200"
        allowed_limits = {"100", "200", "300", "500", "1000", "all"}
        if limit_sel not in allowed_limits:
            limit_sel = "200"
        # Under infinite scroll, this dropdown is treated as "page size" (chunk size)
        # to avoid rendering thousands of rows at once.
        page_size = 500 if limit_sel == "all" else int(limit_sel)

        # fee setting (สร้าง default ถ้ายังไม่มี) -> อยู่ใน price.db
        fee = PlatformFeeSetting.query.get(platform)
        if not fee:
            fee = PlatformFeeSetting(
                platform=platform,
                label=platform,
                fee_pct=0.0,
                fixed_fee=0.0,
                is_active=True,
                sort_order=0,
            )
            db.session.add(fee)
            db.session.commit()

        def calc_profit(price, cost, fee_pct, fixed_fee, pack_cost, ship_subsidy):
            price = float(price or 0.0)
            cost = float(cost or 0.0)
            fee_amt = (price * (float(fee_pct or 0.0) / 100.0)) + float(fixed_fee or 0.0)
            return price - cost - fee_amt - float(pack_cost or 0.0) - float(ship_subsidy or 0.0)

        def is_close_price(a, b, tol: float = 0.01) -> bool:
            if a is None or b is None:
                return False
            try:
                return abs(float(a) - float(b)) <= float(tol)
            except Exception:
                return False

        # -----------------------------
        # ✅ price.db เท่านั้น (Option B)
        # -----------------------------
        pricing_map = {p.sku: p for p in SkuPricing.query.all()}
        bc_map = {b.sku: b for b in BrandControl.query.all()}

        # Brand list for autocomplete in Export Price & Stock Adj (Rule 4/5)
        brand_vals: set[str] = set()
        for p in (pricing_map.values() or []):
            b = (getattr(p, "brand", None) or "").strip()
            if b:
                brand_vals.add(b)
        for bc0 in (bc_map.values() or []):
            b = (getattr(bc0, "brand", None) or "").strip()
            if b:
                brand_vals.add(b)
        brands = sorted(brand_vals, key=lambda s: s.lower())

        # Brand -> Owner mapping (settings)
        owner_rows = BrandOwnerSetting.query.all()
        brand_owner_map = {str(r.brand or "").strip(): str(r.owner or "").strip() for r in owner_rows if r}
        owners_list = sorted({v for v in brand_owner_map.values() if v})

        items = (
            MarketItem.query
            .filter(MarketItem.platform == platform, MarketItem.is_active == True)
            .all()
        )

        # เปลี่ยนจาก "ราคาต่ำสุด" -> "ล่าสุด" (อิง last_updated)
        # หมายเหตุ: อนุญาตให้ latest_net_price เป็น None/0 ได้ เพื่อให้รีเฟรชแล้ว "ว่าง" ตามล่าสุดจริง
        from datetime import datetime

        def _ts_market(it):
            t = getattr(it, "last_updated", None)
            if not t:
                return datetime.min
            try:
                if getattr(t, "tzinfo", None) is not None:
                    t = t.replace(tzinfo=None)
            except Exception:
                pass
            return t

        latest_by_sku = {}
        for it in items:
            cur = latest_by_sku.get(it.sku)
            if cur is None:
                latest_by_sku[it.sku] = it
                continue
            if _ts_market(it) > _ts_market(cur):
                latest_by_sku[it.sku] = it
                continue
            if _ts_market(it) == _ts_market(cur) and (it.id or 0) > (cur.id or 0):
                latest_by_sku[it.sku] = it

        sku_set = set(pricing_map.keys()) | set(latest_by_sku.keys()) | set(bc_map.keys())

        if q:
            filtered = set()
            for sku in sku_set:
                pr = pricing_map.get(sku)
                mk = latest_by_sku.get(sku)
                hay = [
                    (sku or "").lower(),
                    (getattr(pr, "spec_text", "") or "").lower(),
                    (getattr(pr, "brand", "") or "").lower(),
                    (getattr(pr, "name", "") or "").lower(),
                    (getattr(mk, "shop_name", "") or "").lower(),
                ]
                if any(q in s for s in hay):
                    filtered.add(sku)
            sku_set = filtered

        # filter by owner (match from internal brand)
        if owner_sel:
            filtered = set()
            for sku in sku_set:
                pr = pricing_map.get(sku)
                brand = (getattr(pr, "brand", "") or "").strip() if pr else ""
                owner = brand_owner_map.get(brand, "") if brand else ""
                if owner == owner_sel:
                    filtered.add(sku)
            sku_set = filtered

        REC_DEFS = {
            "market_cheaper": {"label": "ตลาดถูกกว่า", "badge": "pm-badge pm-badge-info", "level": "info"},
            "equal_price": {"label": "ราคาเท่ากัน", "badge": "pm-badge pm-badge-neutral", "level": "secondary"},
            "follow_ok": {"label": "ตามได้", "badge": "pm-badge pm-badge-ok", "level": "success"},
            "loss_0_5": {"label": "ขาดทุน(0-5%)", "badge": "pm-badge pm-badge-orange-subtle", "level": "warning"},
            "loss_6_10": {"label": "ขาดทุน(6-10%)", "badge": "pm-badge pm-badge-orange", "level": "warning"},
            "loss_heavy": {"label": "ขาดทุนหนัก", "badge": "pm-badge pm-badge-red-dark", "level": "danger"},
            "no_market": {"label": "ไม่มีราคาตลาด", "badge": "pm-badge pm-badge-neutral", "level": "secondary"},
            "missing_internal": {"label": "ข้อมูลฝั่งเราไม่ครบ", "badge": "pm-badge pm-badge-neutral", "level": "secondary"},
        }
        LEVEL_RANK = {"danger": 4, "warning": 3, "info": 2, "success": 1, "secondary": 0}

        rows = []

        sku_list = sorted(sku_set)

        for sku in sku_list:
            pr = pricing_map.get(sku)
            mk = latest_by_sku.get(sku)
            bc = bc_map.get(sku)

            spec_text = pr.spec_text if pr else None
            cost = float(pr.cost) if (pr and pr.cost is not None) else None
            our_price = float(pr.our_price) if (pr and pr.our_price is not None) else None
            floor_price = float(pr.floor_price) if (pr and pr.floor_price is not None) else None
            min_margin_pct = float(pr.min_margin_pct) if (pr and pr.min_margin_pct is not None) else 0.0
            pack_cost = float(pr.pack_cost) if (pr and pr.pack_cost is not None) else 0.0
            ship_subsidy = float(pr.ship_subsidy) if (pr and pr.ship_subsidy is not None) else 0.0

            brand = (getattr(pr, "brand", "") or "").strip() if pr else ""
            name = (getattr(pr, "name", "") or "").strip() if pr else ""

            stock_qty = int(pr.stock_qty) if (pr and getattr(pr, "stock_qty", None) is not None) else None
            stock_internal = int(getattr(pr, "stock_internal_qty", 0) or 0) if (pr and getattr(pr, "stock_internal_qty", None) is not None) else None
            monthly_sales = int(getattr(pr, "monthly_sales_qty", 0) or 0) if pr else 0

            owner = brand_owner_map.get(brand, "") if brand else ""

            market_net = float(mk.latest_net_price) if (mk and mk.latest_net_price is not None) else None
            market_voucher = float(mk.latest_voucher_discount) if (mk and mk.latest_voucher_discount is not None) else None
            market_shop = mk.shop_name if mk else None
            market_url = mk.product_url if mk else None
            market_updated = mk.last_updated if mk else None

            market_is_mall = bool(getattr(mk, "is_mall", False)) if mk else False

            brand_control = float(bc.price_control) if (bc and bc.price_control is not None) else None

            # Aging logic (Exclusive): choose the highest bucket only (1ปี > 6เดือน > 3เดือน)
            aging_bucket = None
            aging_label = None
            if stock_internal is not None:
                if (monthly_sales * 12 - stock_internal) < 0:
                    aging_bucket = "aging12"
                    aging_label = "Aging(1ปีขึ้นไป)"
                elif (monthly_sales * 6 - stock_internal) < 0:
                    aging_bucket = "aging6"
                    aging_label = "Aging(6เดือนขึ้นไป)"
                elif (monthly_sales * 3 - stock_internal) < 0:
                    aging_bucket = "aging3"
                    aging_label = "Aging(3เดือนขึ้นไป)"

            aging3 = aging_bucket == "aging3"
            aging6 = aging_bucket == "aging6"
            aging12 = aging_bucket == "aging12"

            age_tags: list[str] = []

            # No sales: MonthlySales = 0 and StockInternal >= 1
            no_sales = (monthly_sales == 0) and (stock_internal is not None and stock_internal >= 1)
            if no_sales:
                age_tags.append("ไม่มียอดขาย")

            if aging_bucket and aging_label:
                age_tags.append(aging_label)

            gap = (our_price - market_net) if (our_price is not None and market_net is not None) else None
            gap_pct = None
            if our_price is not None and market_net is not None and market_net > 0:
                gap_pct = ((our_price - market_net) / market_net) * 100.0

            profit_now = None
            profit_match = None
            profit_now_pct = None
            profit_match_pct = None

            recs = []
            rec_keys = []

            def add_rec(key: str):
                if key in REC_DEFS and key not in rec_keys:
                    rec_keys.append(key)
                    recs.append({
                        "key": key,
                        "text": REC_DEFS[key]["label"],
                        "badge_class": REC_DEFS[key]["badge"],
                    })
            no_market_flag = (mk is None) or (market_net is None) or (market_net <= 0)
            missing_internal_flag = (pr is None) or (cost is None) or (our_price is None) or (our_price <= 0)

            if no_market_flag:
                add_rec("no_market")

            if missing_internal_flag:
                add_rec("missing_internal")

            if (our_price is not None and our_price > 0) and (market_net is not None and market_net > 0):
                if abs(our_price - market_net) < 0.01:
                    add_rec("equal_price")
                elif market_net < our_price:
                    add_rec("market_cheaper")

            if pr is not None and our_price is not None and our_price > 0 and cost is not None:
                profit_now = calc_profit(our_price, cost, fee.fee_pct, fee.fixed_fee, pack_cost, ship_subsidy)
                profit_now_pct = (profit_now / our_price) * 100.0 if our_price else None

            if (market_net is not None and market_net > 0) and (cost is not None):
                profit_match = calc_profit(market_net, cost, fee.fee_pct, fee.fixed_fee, pack_cost, ship_subsidy)
                profit_match_pct = (profit_match / market_net) * 100.0

                if profit_match_pct >= 0:
                    add_rec("follow_ok")
                elif profit_match_pct > -6:
                    add_rec("loss_0_5")
                elif profit_match_pct > -10:
                    add_rec("loss_6_10")
                else:
                    add_rec("loss_heavy")

            levels = [REC_DEFS[k]["level"] for k in rec_keys] or ["secondary"]
            rec_level = max(levels, key=lambda x: LEVEL_RANK.get(x, 0))

            base = (
                (market_net is not None and market_net > 0)
                and (our_price is not None and our_price > 0)
                and (market_net < our_price)
            )

            base_up = (
                (market_net is not None and market_net > 0)
                and (our_price is not None and our_price > 0)
                and (market_net > our_price)
            )

            profit_ok = (profit_match is not None and profit_match >= 0)
            profit_neg = (profit_match is not None and profit_match < 0)

            cost_zero = (cost is not None and abs(cost) < 1e-9)
            cost_pos = (cost is not None and cost > 0)

            stock_i = int(stock_internal or 0)
            stock_t = int(stock_qty or 0)
            stock_i_pos = stock_i > 0
            stock_ok = (stock_i > 0) or (stock_t > 0)

            bucket = aging_bucket or ""
            aging3p = bucket in ("aging3", "aging6", "aging12")
            aging6p = bucket in ("aging6", "aging12")
            aging12p = bucket in ("aging12",)

            not_aging = (aging_bucket is None) and (not no_sales)

            has_loss_0_5 = "loss_0_5" in (rec_keys or [])
            has_loss_6_10 = "loss_6_10" in (rec_keys or [])

            r1 = base and profit_ok and cost_pos
            r2 = base and profit_ok and cost_zero and stock_i_pos
            r3 = base and profit_neg and (aging3p or no_sales)
            r4 = base and profit_neg and (aging6p or no_sales)
            r5 = base and profit_neg and (aging12p or no_sales)
            r6 = base and profit_neg and has_loss_0_5
            r7 = base and profit_neg and has_loss_6_10 and stock_i_pos

            # ✅ Rule 8: Brand Control
            # เข้า Auto เฉพาะ Market(best) == Brand Control เท่านั้น และต้องเป็นการ "ลดราคา" (ไม่ขึ้นราคา)
            r8 = (
                base
                and (brand_control is not None and float(brand_control or 0) > 0)
                and is_close_price(market_net, brand_control, tol=0.01)
            )

            # ✅ Rule 9: No market -> price from Cost by tiers, rounded to tens
            r9 = no_market_flag and cost_pos and (float(cost) >= 5)

            # ✅ Rule 10 (NEW): market ถูกกว่าเรา + ขาดทุนหนัก (< -X%) + มีสต๊อก (internal/stock > 0) + cost > 0 + ไม่ใช่ aging
            r10 = (
                base
                and not_aging
                and cost_pos
                and stock_ok
                and (profit_match_pct is not None and float(profit_match_pct) < -float(r10_min_loss))
            )

            # ✅ Rule 11 (NEW): Priority สูงสุด
            # market ถูกกว่าเรา + ขาดทุน < -Min% + มีสต๊อก + cost > 0 + ไม่ใช่ aging
            r11 = (
                base
                and not_aging
                and cost_pos
                and stock_ok
                and (profit_match_pct is not None and float(profit_match_pct) < -float(r11_min_loss))
            )

            # ✅ Rule 12 (NEW): market แพงกว่าเรา + ตามตลาดแล้วยังมีกำไร + มี Cost + มีสต๊อก
            r12 = base_up and profit_ok and cost_pos and stock_ok

            auto_rules = []
            if r1:
                auto_rules.append("r1")
            if r2:
                auto_rules.append("r2")
            if r3:
                auto_rules.append("r3")
            if r4:
                auto_rules.append("r4")
            if r5:
                auto_rules.append("r5")
            if r6:
                auto_rules.append("r6")
            if r7:
                auto_rules.append("r7")
            if r8:
                auto_rules.append("r8")
            if r9:
                auto_rules.append("r9")
            if r10:
                auto_rules.append("r10")
            if r11:
                auto_rules.append("r11")
            if r12:
                auto_rules.append("r12")

            mu = market_updated
            try:
                if mu is not None and getattr(mu, "tzinfo", None) is not None:
                    mu = mu.replace(tzinfo=None)
            except Exception:
                pass

            has_stock = (int(stock_internal or 0) > 0) or (int(stock_qty or 0) > 0)
            is_stale = (mu is None) or (mu < cutoff)
            need_market = has_stock and is_stale

            rows.append({
                "sku": sku,
                "brand": brand,
                "name": name,
                "market_item_id": (mk.id if mk else None),
                "stock_internal": stock_internal,
                "spec_text": spec_text,
                "stock_qty": stock_qty,
                "monthly_sales": monthly_sales,
                "cost": cost,
                "our_price": our_price,
                "market_net": market_net,
                "market_voucher": market_voucher,
                "brand_control": brand_control,
                "market_shop": market_shop,
                "market_is_mall": market_is_mall,
                "market_url": market_url,
                "owner": owner,
                "gap": gap,
                "gap_pct": gap_pct,
                "profit_now": profit_now,
                "profit_now_pct": profit_now_pct,
                "profit_match": profit_match,
                "profit_match_pct": profit_match_pct,
                "recs": recs,
                "rec_keys": rec_keys,
                "rec_level": rec_level,
                "market_updated": market_updated,
                "need_market": need_market,
                "aging3": aging3,
                "aging6": aging6,
                "aging12": aging12,
                "aging_bucket": aging_bucket,
                "age_tags": age_tags,
                "no_sales": no_sales,
                "auto_rules": auto_rules,
            })

        allowed_kpis = set(REC_DEFS.keys()) | {"no_sales", "aging3", "aging6", "aging12", "brand_control", "need_market"}
        kpi_selected = _parse_kpi_multi(kpi_sel, allowed_kpis)
        master_raw = (request.args.get("master") or "").strip()
        master_sel, kpi_selected = _resolve_master(master_raw, kpi_selected, allowed_kpis)

        # ---- Canonicalize URL (กัน master ซ้ำใน kpi + ทำ URL ให้สะอาด) ----
        raw_kpi = (request.args.get("kpi") or "").strip()
        raw_master = (request.args.get("master") or "").strip()
        canon_kpi = ",".join(kpi_selected or [])
        canon_master = (master_sel or "").strip()
        if (raw_kpi != canon_kpi) or (raw_master != canon_master):
            args = request.args.to_dict(flat=True)
            if canon_kpi:
                args["kpi"] = canon_kpi
            else:
                args.pop("kpi", None)
            if canon_master:
                args["master"] = canon_master
            else:
                args.pop("master", None)
            return redirect(url_for("price_dashboard", **args))
        # ---- /Canonicalize URL ----

        rows_master = _apply_kpi_filters(rows, [master_sel]) if master_sel else rows
        display_rows = _apply_kpi_filters(rows_master, kpi_selected)
        tracked_after_filter = len(display_rows)

        base_counts = _compute_kpi_counts(rows_master, allowed_kpis)
        counts = dict(base_counts)
        counts["tracked"] = tracked_after_filter

        auto_counts = {f"r{i}": 0 for i in range(1, 13)}
        auto_change_counts = {f"r{i}": 0 for i in range(1, 13)}
        auto_any_count = 0

        # For "pending change" badges: compute each rule's target and compare with current Our Price
        r9_tiers = list(_R9_DEFAULT_TIERS)
        try:
            raw_pref = get_user_pref(int(cu.id), "pm_auto_r9_cfg", default="") if cu else ""
            r9_cfg = json.loads(raw_pref) if raw_pref else None
            if isinstance(r9_cfg, list):
                r9_tiers = _r9_cfg_to_tiers(r9_cfg)
        except Exception:
            r9_tiers = list(_R9_DEFAULT_TIERS)

        def _to_float(x):
            try:
                if x in (None, ""):
                    return None
                return float(x)
            except Exception:
                return None

        def _target_for_rule(rid: str, rr: dict) -> float | None:
            cost = _to_float(rr.get("cost"))
            market_net = _to_float(rr.get("market_net"))
            brand_ctrl = _to_float(rr.get("brand_control"))
            loss_abs = _loss_abs_pct(_to_float(rr.get("profit_match_pct")))

            if rid == "r12":
                return market_net

            if rid == "r11":
                if (market_net is None) or (market_net <= 0) or (loss_abs is None):
                    return None
                if float(loss_abs) <= float(r11_max_loss):
                    return market_net
                return _auto_price_from_cost_plus_pct(cost, float(loss_abs) / 100.0)

            if rid == "r10":
                if loss_abs is None:
                    return None
                return _auto_price_from_cost_plus_pct(cost, float(loss_abs) / 100.0)

            if rid == "r9":
                return _auto_price_from_cost(cost, tiers=r9_tiers)
            if rid == "r8":
                return brand_ctrl
            # r1..r7 (and any other): set Our Price = Market(best)
            return market_net

        for rr in display_rows:
            rules = rr.get("auto_rules") or []
            if rules:
                auto_any_count += 1
            before = _to_float(rr.get("our_price"))
            for rid in rules:
                if rid in auto_counts:
                    auto_counts[rid] += 1

                if rid in auto_change_counts:
                    target = _target_for_rule(rid, rr)
                    newv = _to_float(target)
                    if newv is None or newv <= 0:
                        continue
                    if (before is None) or (not is_close_price(before, newv, tol=0.01)):
                        auto_change_counts[rid] += 1

        kpi_cards = [
            {"key": "tracked", "label": "Tracked", "count": tracked_after_filter, "icon": "bi-bookmark-check", "tone": "primary"},
            {"key": "market_cheaper", "label": "ตลาดถูกกว่า", "count": base_counts.get("market_cheaper", 0), "icon": "bi-graph-down-arrow", "tone": "info"},
            {"key": "follow_ok", "label": "ตามได้", "count": base_counts.get("follow_ok", 0), "icon": "bi-check-circle", "tone": "success"},
            {"key": "loss_0_5", "label": "ขาดทุน(0-5%)", "count": base_counts.get("loss_0_5", 0), "icon": "bi-exclamation-circle", "tone": "orange-subtle"},
            {"key": "loss_6_10", "label": "ขาดทุน(6-10%)", "count": base_counts.get("loss_6_10", 0), "icon": "bi-exclamation-octagon", "tone": "orange"},
            {"key": "loss_heavy", "label": "ขาดทุนหนัก", "count": base_counts.get("loss_heavy", 0), "icon": "bi-x-circle", "tone": "danger"},
            {"key": "equal_price", "label": "ราคาเท่ากัน", "count": base_counts.get("equal_price", 0), "icon": "bi-arrow-left-right", "tone": "secondary"},
            {"key": "no_market", "label": "ไม่มีราคาตลาด", "count": base_counts.get("no_market", 0), "icon": "bi-exclamation-triangle", "tone": "secondary"},
            {"key": "no_sales", "label": "ไม่มียอดขาย", "count": base_counts.get("no_sales", 0), "icon": "bi-slash-circle", "tone": "secondary"},
            {"key": "aging3", "label": "Aging (3เดือนขึ้นไป)", "count": base_counts.get("aging3", 0), "icon": "bi-hourglass-split", "tone": "warning"},
            {"key": "aging6", "label": "Aging (6เดือนขึ้นไป)", "count": base_counts.get("aging6", 0), "icon": "bi-hourglass", "tone": "warning"},
            {"key": "aging12", "label": "Aging (1ปีขึ้นไป)", "count": base_counts.get("aging12", 0), "icon": "bi-hourglass-bottom", "tone": "danger"},
            {"key": "missing_internal", "label": "Missing Internal", "count": base_counts.get("missing_internal", 0), "icon": "bi-database-x", "tone": "secondary"},
            {"key": "brand_control", "label": "Brand Control", "count": base_counts.get("brand_control", 0), "icon": "bi-shield-lock", "tone": "danger"},
            {"key": "need_market", "label": "หาราคาตลาด", "count": base_counts.get("need_market", 0), "icon": "bi-search", "tone": "warning"},
        ]

        # -----------------------------
        # Apply server-side sort to display_rows (before caching)
        # -----------------------------
        if sort_sel:
            SORT_META: dict[str, tuple[str, callable]] = {
                "sku": ("text", lambda r: r.get("sku")),
                "brand": ("text", lambda r: r.get("brand")),
                "name": ("text", lambda r: r.get("name")),
                "owner": ("text", lambda r: r.get("owner")),
                "shop": ("text", lambda r: r.get("market_shop")),
                "url": ("text", lambda r: r.get("market_url")),

                "stock_internal": ("num", lambda r: r.get("stock_internal")),
                "stock": ("num", lambda r: r.get("stock_qty")),
                "monthly_sales": ("num", lambda r: r.get("monthly_sales")),

                "cost": ("num", lambda r: r.get("cost")),
                "our_price": ("num", lambda r: r.get("our_price")),
                "market_best": ("num", lambda r: r.get("market_net")),
                "voucher": ("num", lambda r: r.get("market_voucher")),
                "brand_control": ("num", lambda r: r.get("brand_control")),
                "gap": ("num", lambda r: r.get("gap")),
                "profit_our": ("num", lambda r: r.get("profit_now")),
                "profit_match": ("num", lambda r: r.get("profit_match")),

                # Recommend: sort by severity (secondary<success<info<warning<danger)
                "recommend": ("level", lambda r: LEVEL_RANK.get(r.get("rec_level"), 0)),

                "mall": ("bool", lambda r: r.get("market_is_mall")),
                "updated": ("dt", lambda r: r.get("market_updated")),
            }

            kind, getter = SORT_META.get(sort_sel, ("text", lambda r: ""))

            def _sku_tiebreak(r: dict) -> str:
                return (r.get("sku") or "").strip()

            def _norm_text(v) -> str | None:
                s = ("" if v is None else str(v)).strip()
                return (s.lower() if s else None)

            def _norm_num(v) -> float | None:
                if v in (None, ""):
                    return None
                try:
                    return float(v)
                except Exception:
                    return None

            def _norm_bool(v) -> int | None:
                if v is None:
                    return None
                return 1 if bool(v) else 0

            def _norm_dt(v) -> datetime | None:
                if not v:
                    return None
                if isinstance(v, datetime):
                    try:
                        if getattr(v, "tzinfo", None) is not None:
                            v = v.replace(tzinfo=None)
                    except Exception:
                        pass
                    return v
                return None

            present: list[tuple[object, str, dict]] = []
            missing: list[dict] = []
            for rr in display_rows:
                raw_v = getter(rr)
                if kind == "text":
                    vv = _norm_text(raw_v)
                elif kind in {"num", "level"}:
                    vv = _norm_num(raw_v)
                elif kind == "bool":
                    vv = _norm_bool(raw_v)
                elif kind == "dt":
                    vv = _norm_dt(raw_v)
                else:
                    vv = raw_v

                if vv is None:
                    missing.append(rr)
                else:
                    present.append((vv, _sku_tiebreak(rr), rr))

            reverse = (sort_dir == "desc")
            present_sorted = sorted(present, key=lambda t: (t[0], t[1]), reverse=reverse)
            missing_sorted = sorted(missing, key=_sku_tiebreak)
            display_rows = [t[2] for t in present_sorted] + missing_sorted

        # -----------------------------
        # Infinite scroll: cache all filtered rows, render only first page
        # -----------------------------
        _dash_cache_gc()
        dash_key = uuid.uuid4().hex
        PRICE_DASH_ROWS_CACHE[dash_key] = {
            "user_id": int(cu.id) if cu else 0,
            "ts": time.time(),
            "rows": display_rows,
        }

        initial_rows = display_rows[:page_size]
        total_rows = len(display_rows)

        resp = make_response(render_template(
            "price_dashboard.html",
            rows=initial_rows,
            dash_key=dash_key,
            dash_total_rows=total_rows,
            dash_page_size=page_size,
            sort_sel=sort_sel,
            sort_dir=sort_dir,
            platform=platform,
            platforms=platforms,
            q=(request.args.get("q") or ""),
            fee=fee,
            counts=counts,
            kpi_cards=kpi_cards,
            kpi_sel=",".join(kpi_selected),
            kpi_selected=kpi_selected,
            master_sel=master_sel,
            owners_list=owners_list,
            owner_sel=owner_sel,
            limit_sel=limit_sel,
            stale_days_sel=stale_days_sel,
            auto_counts=auto_counts,
            auto_change_counts=auto_change_counts,
            auto_any_count=auto_any_count,
            brands=brands,
        ))
        resp.headers["Cache-Control"] = "no-store, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        return resp

    @app.get("/api/price/dashboard/rows")
    @login_required
    def api_price_dashboard_rows():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        _dash_cache_gc()

        key = (request.args.get("key") or "").strip()
        try:
            offset = int(request.args.get("offset") or 0)
            limit = int(request.args.get("limit") or 200)
        except Exception:
            return jsonify({"success": False, "msg": "พารามิเตอร์ไม่ถูกต้อง"}), 400

        entry = PRICE_DASH_ROWS_CACHE.get(key)
        if not entry:
            return jsonify({"success": False, "msg": "หมดอายุ/ไม่พบ cache (ให้รีเฟรชหน้า)"}), 400

        if int(entry.get("user_id") or 0) != int(cu.id):
            return jsonify({"success": False, "msg": "ไม่มีสิทธิ์เข้าถึงชุดข้อมูลนี้"}), 403

        rows_all = entry.get("rows") or []
        total = len(rows_all)

        if offset < 0:
            offset = 0
        if limit <= 0:
            limit = 200
        if limit > 1000:
            limit = 1000

        slice_rows = rows_all[offset : offset + limit]
        html = render_template("_price_dashboard_rows.html", rows=slice_rows)

        next_offset = offset + len(slice_rows)
        has_more = next_offset < total

        return jsonify({
            "success": True,
            "html": html,
            "next_offset": next_offset,
            "total": total,
            "has_more": has_more,
        })

    @app.post("/api/price/dashboard/update_cell")
    @login_required
    def api_price_dashboard_update_cell():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            return jsonify({"success": False, "msg": "ไม่มีสิทธิ์ใช้งาน"}), 403

        data = request.get_json(silent=True) or {}
        sku = (data.get("sku") or "").strip()
        col = (data.get("col") or "").strip()
        value = data.get("value")
        platform_raw = (data.get("platform") or "").strip()
        platform = normalize_platform(platform_raw) or platform_raw
        market_item_id = data.get("market_item_id")

        if not sku:
            return jsonify({"success": False, "msg": "SKU ว่าง"}), 400

        allowed_cols = {
            "cost",
            "our_price",
            "brand_control",
            "market_best",
            "voucher",
            "shop",
            "mall",
            "url",
            "updated",
        }
        if col not in allowed_cols:
            return jsonify({"success": False, "msg": "คอลัมน์นี้ต้องแก้ผ่าน Import"}), 400

        def _as_float(v):
            if v in (None, ""):
                return None
            return float(v)

        def _as_bool(v) -> bool:
            if isinstance(v, bool):
                return v
            if v is None:
                return False
            s = str(v).strip().lower()
            return s in {"1", "true", "yes", "on"}

        try:
            # ---- Internal (SkuPricing) ----
            if col in {"cost", "our_price"}:
                pr = SkuPricing.query.get(sku)
                if not pr:
                    pr = SkuPricing(sku=sku)
                    db.session.add(pr)

                v = _as_float(value)
                setattr(pr, col, v)
                db.session.commit()

                display = "" if v is None else f"{v:,.2f}"
                return jsonify({"success": True, "display": display})

            # ---- Brand Control ----
            if col == "brand_control":
                bc = BrandControl.query.get(sku)
                if not bc:
                    bc = BrandControl(sku=sku)
                    db.session.add(bc)

                v = _as_float(value)
                bc.price_control = v
                db.session.commit()

                display = "" if v is None else f"{v:,.2f}"
                return jsonify({"success": True, "display": display})

            # ---- MarketItem ----
            if not platform:
                return jsonify({"success": False, "msg": "Platform ว่าง"}), 400

            mk = None
            if market_item_id not in (None, ""):
                try:
                    mk = MarketItem.query.get(int(market_item_id))
                except Exception:
                    mk = None

            # safety: ensure id matches sku/platform
            if mk and ((mk.sku or "") != sku or (mk.platform or "") != platform):
                mk = None

            if not mk:
                mk = (
                    MarketItem.query
                    .filter(
                        MarketItem.sku == sku,
                        MarketItem.platform == platform,
                        MarketItem.is_active == True,
                        MarketItem.latest_net_price.isnot(None),
                        MarketItem.latest_net_price > 0,
                    )
                    .order_by(MarketItem.latest_net_price.asc())
                    .first()
                )

            if not mk:
                # create manual record (shop_name is NOT NULL and unique with sku+platform)
                base_shop = "(manual)"
                shop_name = base_shop
                for i in range(0, 10):
                    mk = MarketItem(sku=sku, platform=platform, shop_name=shop_name, is_active=True)
                    db.session.add(mk)
                    try:
                        db.session.flush()
                        break
                    except IntegrityError:
                        db.session.rollback()
                        mk = None
                        shop_name = f"{base_shop} {i+2}"

            if not mk:
                return jsonify({"success": False, "msg": "ไม่พบ/สร้างข้อมูลตลาดไม่ได้"}), 400

            display = ""

            if col == "market_best":
                v = _as_float(value)
                mk.latest_net_price = v
                display = "" if v is None else f"{v:,.2f}"

            elif col == "voucher":
                v = _as_float(value)
                mk.latest_voucher_discount = v
                display = "" if v is None else f"{v:,.2f}"

            elif col == "shop":
                v = (value or "").strip()
                if not v:
                    return jsonify({"success": False, "msg": "Shop ห้ามว่าง"}), 400

                # กันชน UniqueConstraint (sku + platform + shop_name)
                conflict = (
                    MarketItem.query
                    .filter(
                        MarketItem.sku == sku,
                        MarketItem.platform == platform,
                        MarketItem.shop_name == v,
                        MarketItem.id != mk.id,
                    )
                    .first()
                )
                if conflict:
                    # Shop ซ้ำภายใน SKU เดียวกัน: ไม่ถือเป็น error แต่สลับไปใช้ record เดิมแทน
                    return jsonify({
                        "success": True,
                        "display": v,
                        "market_item_id": conflict.id,
                        "switched": True,
                    }), 200

                mk.shop_name = v
                display = v

            elif col == "mall":
                mk.is_mall = _as_bool(value)
                display = "MALL" if mk.is_mall else "-"

            elif col == "url":
                v = (value or "").strip()
                mk.product_url = v or None
                display = v or "-"

            elif col == "updated":
                # รองรับ "NOW" = เวลาปัจจุบัน (ไทย)
                dt_store = None
                dt_display = None
                if isinstance(value, str) and value.strip().upper() in {"NOW", "__NOW__", "CURRENT"}:
                    dt_display = now_thai()
                    dt_store = dt_display
                    # กัน SQLite/SQLAlchemy บางเคสไม่ชอบ tz-aware -> ทำเป็น naive สำหรับการบันทึก
                    try:
                        if getattr(dt_store, "tzinfo", None) is not None:
                            dt_store = dt_store.replace(tzinfo=None)
                    except Exception:
                        pass
                else:
                    dt_store = parse_datetime_guess(value) if value not in (None, "") else None
                    dt_display = dt_store

                mk.last_updated = dt_store
                db.session.commit()
                return jsonify({
                    "success": True,
                    "display": (to_thai_be(dt_display) if dt_display else "-"),
                    "iso": (dt_store.isoformat() if dt_store else ""),
                    "market_item_id": mk.id,
                })

            db.session.commit()
            return jsonify({"success": True, "display": display, "market_item_id": mk.id})

        except IntegrityError as e:
            db.session.rollback()

            raw = str(getattr(e, "orig", "")) or str(e)

            # SQLite มักเป็นข้อความแนว: UNIQUE constraint failed: market_items.sku, market_items.platform, market_items.shop_name
            # หรือบางครั้งมีชื่อ constraint เช่น uq_market_item_sku_plat_shop
            if (
                "uq_market_item_sku_plat_shop" in raw
                or "UNIQUE constraint failed: market_items.sku" in raw
                or "UNIQUE constraint failed: market_items.sku, market_items.platform, market_items.shop_name" in raw
            ):
                return jsonify({
                    "success": False,
                    "msg": "Shop ซ้ำ: ใน SKU + Platform เดียวกัน ห้ามใช้ชื่อร้านซ้ำกัน (กรุณาเปลี่ยนชื่อ Shop ให้ไม่ซ้ำ)",
                }), 400

            return jsonify({"success": False, "msg": f"DB constraint error: {raw}"}), 400
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "msg": str(e)}), 400


    @app.get("/api/price/export_preset")
    @login_required
    def api_price_export_preset():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        dest = (request.args.get("dest_platform") or "").strip()
        dest_norm = normalize_platform(dest) or dest
        dest_key = normalize_platform_key(dest_norm) or "DEFAULT"
        key = f"pm_export_preset::{dest_key}"

        raw = get_user_pref(cu.id, key, default="")
        if not raw:
            return jsonify({"success": True, "preset": None})

        try:
            return jsonify({"success": True, "preset": json.loads(raw)})
        except Exception:
            return jsonify({"success": True, "preset": None})


    @app.post("/api/price/export_preset/save")
    @login_required
    def api_price_export_preset_save():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        data = request.get_json(silent=True) or {}
        dest = (data.get("dest_platform") or "").strip()
        preset = data.get("preset") or {}

        dest_norm = normalize_platform(dest) or dest
        dest_key = normalize_platform_key(dest_norm) or "DEFAULT"
        key = f"pm_export_preset::{dest_key}"

        try:
            set_user_pref(cu.id, key, json.dumps(preset, ensure_ascii=False))
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "msg": str(e)}), 400


    @app.post("/api/price/export_preset/clear")
    @login_required
    def api_price_export_preset_clear():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        data = request.get_json(silent=True) or {}
        dest = (data.get("dest_platform") or "").strip()

        dest_norm = normalize_platform(dest) or dest
        dest_key = normalize_platform_key(dest_norm) or "DEFAULT"
        key = f"pm_export_preset::{dest_key}"

        try:
            PriceUserPreference.query.filter_by(user_id=cu.id, key=key).delete(synchronize_session=False)
            db.session.commit()
            return jsonify({"success": True})
        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            return jsonify({"success": False, "msg": str(e)}), 400


    @app.get("/api/price/auto_r9")
    @login_required
    def api_price_auto_r9_get():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        key = "pm_auto_r9_cfg"
        raw = get_user_pref(int(cu.id), key, default="")
        if not raw:
            return jsonify({"success": True, "cfg": _r9_default_cfg()})

        try:
            cfg = json.loads(raw)
            if not isinstance(cfg, list):
                raise ValueError("cfg not list")
            return jsonify({"success": True, "cfg": cfg})
        except Exception:
            return jsonify({"success": True, "cfg": _r9_default_cfg()})


    @app.post("/api/price/auto_r9/save")
    @login_required
    def api_price_auto_r9_save():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        data = request.get_json(silent=True) or {}
        cfg = data.get("cfg")
        if not isinstance(cfg, list):
            return jsonify({"success": False, "msg": "cfg ต้องเป็น list"}), 400

        tiers = _r9_cfg_to_tiers(cfg)
        cfg_norm = [{"min": lo, "max": hi, "pct": pct} for (lo, hi, pct) in tiers]

        key = "pm_auto_r9_cfg"
        set_user_pref(int(cu.id), key, json.dumps(cfg_norm, ensure_ascii=False))
        return jsonify({"success": True})


    @app.get("/api/price/auto_r10")
    @login_required
    def api_price_auto_r10_get():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        raw = get_user_pref(int(cu.id), "pm_auto_r10_cfg", default="")
        cfg = _norm_r10_cfg(_load_cfg_dict(raw, _R10_DEFAULT_CFG))
        return jsonify({"success": True, "cfg": cfg})


    @app.post("/api/price/auto_r10/save")
    @login_required
    def api_price_auto_r10_save():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        data = request.get_json(silent=True) or {}
        cfg_norm = _norm_r10_cfg(data.get("cfg") or {})
        set_user_pref(int(cu.id), "pm_auto_r10_cfg", json.dumps(cfg_norm, ensure_ascii=False))
        return jsonify({"success": True})


    @app.get("/api/price/auto_r11")
    @login_required
    def api_price_auto_r11_get():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        raw = get_user_pref(int(cu.id), "pm_auto_r11_cfg", default="")
        cfg = _norm_r11_cfg(_load_cfg_dict(raw, _R11_DEFAULT_CFG))
        return jsonify({"success": True, "cfg": cfg})


    @app.post("/api/price/auto_r11/save")
    @login_required
    def api_price_auto_r11_save():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        data = request.get_json(silent=True) or {}
        cfg_norm = _norm_r11_cfg(data.get("cfg") or {})
        set_user_pref(int(cu.id), "pm_auto_r11_cfg", json.dumps(cfg_norm, ensure_ascii=False))
        return jsonify({"success": True})


    @app.get("/api/price/export_price_cols")
    @login_required
    def api_export_price_cols_get():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        platform_arg = (request.args.get("platform") or "").strip()
        platform = normalize_platform(platform_arg) if platform_arg else "Shopee"
        plat_key = normalize_platform_key(platform) or "DEFAULT"
        key = f"pm_export_price_cols::{plat_key}"

        raw = get_user_pref(int(cu.id), key, default="")
        if not raw:
            return jsonify({"success": True, "cfg": None})

        try:
            cfg = json.loads(raw)
            return jsonify({"success": True, "cfg": cfg})
        except Exception:
            return jsonify({"success": True, "cfg": None})


    @app.post("/api/price/export_price_cols/save")
    @login_required
    def api_export_price_cols_save():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        data = request.get_json(silent=True) or {}
        platform_arg = (data.get("platform") or "").strip()
        platform = normalize_platform(platform_arg) if platform_arg else "Shopee"
        plat_key = normalize_platform_key(platform) or "DEFAULT"
        key = f"pm_export_price_cols::{plat_key}"

        cfg = data.get("cfg") or {}
        if not isinstance(cfg, dict):
            return jsonify({"success": False, "msg": "cfg ต้องเป็น dict"}), 400

        ORDERED_KEYS = [
            "sku","brand","name","spec","stock_internal","stock","monthly_sales","cost","our_price",
            "sell_1","sell_2","sell_3","sell_4","sell_5","sell_min",
            "market_best","voucher","brand_control","gap","profit_our","profit_match",
            "recommend","shop","mall","url","owner","updated"
        ]
        allowed = set(ORDERED_KEYS)

        cols = cfg.get("cols") or []
        if not isinstance(cols, list):
            cols = []
        cols = [c for c in cols if isinstance(c, str) and c in allowed]
        if "sku" not in cols:
            cols = ["sku"] + [c for c in cols if c != "sku"]

        cfg_norm = {
            "cols": cols,
            "sort_brand": bool(cfg.get("sort_brand", False)),
            "include_zero_stock": bool(cfg.get("include_zero_stock", False)),
        }

        set_user_pref(int(cu.id), key, json.dumps(cfg_norm, ensure_ascii=False))
        return jsonify({"success": True, "cfg": cfg_norm})


    @app.post("/api/price/export_price_cols/clear")
    @login_required
    def api_export_price_cols_clear():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        data = request.get_json(silent=True) or {}
        platform_arg = (data.get("platform") or "").strip()
        platform = normalize_platform(platform_arg) if platform_arg else "Shopee"
        plat_key = normalize_platform_key(platform) or "DEFAULT"
        key = f"pm_export_price_cols::{plat_key}"

        try:
            PriceUserPreference.query.filter_by(user_id=cu.id, key=key).delete(synchronize_session=False)
            db.session.commit()
            return jsonify({"success": True})
        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            return jsonify({"success": False, "msg": str(e)}), 400

    @app.post("/price/dashboard/set_platform")
    @login_required
    def price_dashboard_set_platform():
        cu = current_user()
        if not cu:
            return redirect(url_for("login", next=request.path))

        platform_raw = (request.form.get("platform") or "").strip()
        platform = normalize_platform(platform_raw) or platform_raw
        q = (request.form.get("q") or "").strip()
        owner_sel = (request.form.get("owner") or "").strip()
        limit_sel = (request.form.get("limit") or "").strip().lower()
        stale_days_sel = (request.form.get("stale_days") or "").strip()
        kpi_sel = (request.form.get("kpi") or "").strip()
        master_sel = (request.form.get("master") or "").strip()
        sort_sel = (request.form.get("sort") or "").strip()
        sort_dir = (request.form.get("dir") or "asc").strip().lower() or "asc"

        platform_rows = (
            PlatformFeeSetting.query
            .filter(PlatformFeeSetting.is_active == True)
            .all()
        )
        allowed = {p.platform for p in platform_rows}

        if platform not in allowed:
            flash("Platform ไม่ถูกต้องหรือถูกปิดการใช้งาน", "warning")
            return redirect(url_for("price_dashboard"))

        set_user_pref(cu.id, "price_dash_platform", platform)
        flash(f"ล็อก Platform เป็น {platform} เรียบร้อย", "success")
        return redirect(url_for(
            "price_dashboard",
            platform=platform,
            q=q,
            owner=owner_sel,
            limit=limit_sel,
            stale_days=stale_days_sel,
            kpi=kpi_sel,
            master=master_sel,
            sort=sort_sel,
            dir=sort_dir,
        ))

    @app.post("/price/dashboard/set_filters")
    @login_required
    def price_dashboard_set_filters():
        cu = current_user()
        if not cu:
            return redirect(url_for("login", next=request.path))

        owner_sel = (request.form.get("owner") or "").strip()  # "" = ทั้งหมด
        limit_sel = (request.form.get("limit") or "200").strip().lower()
        stale_days_raw = (request.form.get("stale_days") or "").strip()
        allowed_limits = {"100", "200", "300", "500", "1000", "all"}
        if limit_sel not in allowed_limits:
            limit_sel = "200"

        try:
            stale_days = int(stale_days_raw or 14)
        except Exception:
            stale_days = 14
        stale_days = max(1, min(stale_days, 365))
        stale_days_sel = str(stale_days)

        set_user_pref(cu.id, "price_dash_owner", owner_sel)
        set_user_pref(cu.id, "price_dash_limit", limit_sel)
        set_user_pref(cu.id, "price_dash_stale_days", stale_days_sel)

        # คง platform เดิมไว้ (ล็อกแยกอยู่แล้ว)
        platform = (get_user_pref(cu.id, "price_dash_platform", "") or "").strip()
        if not platform:
            platform = (request.form.get("platform") or "").strip()

        q = (request.form.get("q") or "").strip()
        kpi_sel = (request.form.get("kpi") or "").strip()
        master_sel = (request.form.get("master") or "").strip()
        sort_sel = (request.form.get("sort") or "").strip()
        sort_dir = (request.form.get("dir") or "asc").strip().lower() or "asc"

        flash("บันทึกฟิลเตอร์เรียบร้อย", "success")
        return redirect(url_for(
            "price_dashboard",
            platform=platform,
            q=q,
            owner=owner_sel,
            limit=limit_sel,
            stale_days=stale_days_sel,
            kpi=kpi_sel,
            master=master_sel,
            sort=sort_sel,
            dir=sort_dir,
        ))

    @app.get("/price/dashboard/export.xlsx")
    @login_required
    def price_dashboard_export():
        """Export Price Marketing dashboard rows to Excel (ตาม filter/limit ที่เลือก)."""

        # คำนวณ rows ด้วย logic เดียวกับหน้า dashboard (แบบย่อ)
        platform_arg = (request.args.get("platform") or "").strip()
        platform = normalize_platform(platform_arg) if platform_arg else "Shopee"

        q = (request.args.get("q") or "").strip().lower()
        owner_sel = (request.args.get("owner") or "").strip()
        limit_sel = (request.args.get("limit") or "200").strip().lower()
        stale_days_raw = (request.args.get("stale_days") or "").strip()
        kpi_sel = (request.args.get("kpi") or "").strip()
        master_raw = (request.args.get("master") or "").strip()
        allowed_limits = {"100", "200", "300", "500", "1000", "all"}
        if limit_sel not in allowed_limits:
            limit_sel = "200"
        limit_n = None if limit_sel == "all" else int(limit_sel)

        try:
            stale_days = int(stale_days_raw or 14)
        except Exception:
            stale_days = 14
        stale_days = max(1, min(stale_days, 365))
        now0 = now_thai()
        try:
            if getattr(now0, "tzinfo", None) is not None:
                now0 = now0.replace(tzinfo=None)
        except Exception:
            pass
        cutoff = now0 - timedelta(days=stale_days)

        fee = PlatformFeeSetting.query.get(platform)
        if not fee:
            fee = PlatformFeeSetting(platform=platform, label=platform, fee_pct=0.0, fixed_fee=0.0, is_active=True, sort_order=0)
            db.session.add(fee)
            db.session.commit()

        def calc_profit(price, cost, fee_pct, fixed_fee, pack_cost, ship_subsidy):
            price = float(price or 0.0)
            cost = float(cost or 0.0)
            fee_amt = (price * (float(fee_pct or 0.0) / 100.0)) + float(fixed_fee or 0.0)
            return price - cost - fee_amt - float(pack_cost or 0.0) - float(ship_subsidy or 0.0)

        pricing_map = {p.sku: p for p in SkuPricing.query.all()}
        bc_map = {b.sku: b for b in BrandControl.query.all()}

        owner_rows = BrandOwnerSetting.query.all()
        brand_owner_map = {str(r.brand or "").strip(): str(r.owner or "").strip() for r in owner_rows if r}

        items = (
            MarketItem.query
            .filter(MarketItem.platform == platform, MarketItem.is_active == True)
            .all()
        )
        from datetime import datetime

        def _ts_market(it):
            t = getattr(it, "last_updated", None)
            if not t:
                return datetime.min
            try:
                if getattr(t, "tzinfo", None) is not None:
                    t = t.replace(tzinfo=None)
            except Exception:
                pass
            return t

        latest_by_sku = {}
        for it in items:
            cur = latest_by_sku.get(it.sku)
            if cur is None:
                latest_by_sku[it.sku] = it
                continue
            if _ts_market(it) > _ts_market(cur):
                latest_by_sku[it.sku] = it
                continue
            if _ts_market(it) == _ts_market(cur) and (it.id or 0) > (cur.id or 0):
                latest_by_sku[it.sku] = it

        sku_set = set(pricing_map.keys()) | set(latest_by_sku.keys()) | set(bc_map.keys())

        if q:
            filtered = set()
            for sku in sku_set:
                pr = pricing_map.get(sku)
                mk = latest_by_sku.get(sku)
                hay = [
                    (sku or "").lower(),
                    (getattr(pr, "spec_text", "") or "").lower(),
                    (getattr(pr, "brand", "") or "").lower(),
                    (getattr(pr, "name", "") or "").lower(),
                    (getattr(mk, "shop_name", "") or "").lower(),
                ]
                if any(q in s for s in hay):
                    filtered.add(sku)
            sku_set = filtered

        if owner_sel:
            filtered = set()
            for sku in sku_set:
                pr = pricing_map.get(sku)
                brand = (getattr(pr, "brand", "") or "").strip() if pr else ""
                owner = brand_owner_map.get(brand, "") if brand else ""
                if owner == owner_sel:
                    filtered.add(sku)
            sku_set = filtered

        sku_list = sorted(sku_set)
        if limit_n is not None:
            sku_list = sku_list[:limit_n]

        export_rows = []

        REC_DEFS = {
            "market_cheaper": {"label": "ตลาดถูกกว่า"},
            "equal_price": {"label": "ราคาเท่ากัน"},
            "follow_ok": {"label": "ตามได้"},
            "loss_0_5": {"label": "ขาดทุน(0-5%)"},
            "loss_6_10": {"label": "ขาดทุน(6-10%)"},
            "loss_heavy": {"label": "ขาดทุนหนัก"},
            "no_market": {"label": "ไม่มีราคาตลาด"},
            "missing_internal": {"label": "ข้อมูลฝั่งเราไม่ครบ"},
        }

        allowed_kpis = set(REC_DEFS.keys()) | {"no_sales", "aging3", "aging6", "aging12", "brand_control", "need_market"}
        kpi_selected = _parse_kpi_multi(kpi_sel, allowed_kpis)
        master_sel, kpi_selected = _resolve_master(master_raw, kpi_selected, allowed_kpis)

        # ให้หัวคอลัมน์ “มีเสมอ” และเรียงเหมือนหน้า Dashboard
        cols = [
            "SKU", "Brand", "Name", "Stock Internal", "Stock", "Monthly Sales", "Cost", "Our Price",
            "Market (best)", "Voucher", "Brand Control", "Gap",
            "Profit@Our", "Profit@Match", "Recommend",
            "Shop", "MALL", "URL", "ผู้ดูแล", "Updated",
        ]
        for sku in sku_list:
            pr = pricing_map.get(sku)
            mk = latest_by_sku.get(sku)
            bc = bc_map.get(sku)

            cost = float(pr.cost) if (pr and pr.cost is not None) else None
            our_price = float(pr.our_price) if (pr and pr.our_price is not None) else None
            floor_price = float(pr.floor_price) if (pr and pr.floor_price is not None) else None
            min_margin_pct = float(pr.min_margin_pct) if (pr and pr.min_margin_pct is not None) else 0.0
            pack_cost = float(pr.pack_cost) if (pr and pr.pack_cost is not None) else 0.0
            ship_subsidy = float(pr.ship_subsidy) if (pr and pr.ship_subsidy is not None) else 0.0

            brand = (getattr(pr, "brand", "") or "").strip() if pr else ""
            name = (getattr(pr, "name", "") or "").strip() if pr else ""
            owner = brand_owner_map.get(brand, "") if brand else ""

            stock_internal = int(getattr(pr, "stock_internal_qty", 0) or 0) if (pr and getattr(pr, "stock_internal_qty", None) is not None) else None
            stock_total = int(getattr(pr, "stock_qty", 0) or 0) if (pr and getattr(pr, "stock_qty", None) is not None) else None
            monthly_sales = int(getattr(pr, "monthly_sales_qty", 0) or 0) if pr else 0

            market_net = float(mk.latest_net_price) if (mk and mk.latest_net_price is not None) else None
            market_voucher = float(mk.latest_voucher_discount) if (mk and mk.latest_voucher_discount is not None) else None
            market_shop = mk.shop_name if mk else None
            market_is_mall = bool(getattr(mk, "is_mall", False)) if mk else False
            market_url = mk.product_url if mk else None
            market_updated = mk.last_updated if mk else None

            brand_control = float(bc.price_control) if (bc and bc.price_control is not None) else None

            gap = (our_price - market_net) if (our_price is not None and market_net is not None) else None
            profit_now = None
            profit_match = None

            rec_keys = []

            def add_rec(key: str):
                if key in REC_DEFS and key not in rec_keys:
                    rec_keys.append(key)

            no_market_flag = (mk is None) or (market_net is None) or (market_net <= 0)
            missing_internal_flag = (pr is None) or (cost is None) or (our_price is None) or (our_price <= 0)

            if no_market_flag:
                add_rec("no_market")
            if missing_internal_flag:
                add_rec("missing_internal")

            if (our_price is not None and our_price > 0) and (market_net is not None and market_net > 0):
                if abs(our_price - market_net) < 0.01:
                    add_rec("equal_price")
                elif market_net < our_price:
                    add_rec("market_cheaper")

            if our_price is not None and our_price > 0 and cost is not None:
                profit_now = calc_profit(our_price, cost, fee.fee_pct, fee.fixed_fee, pack_cost, ship_subsidy)

            if market_net is not None and market_net > 0 and cost is not None:
                profit_match = calc_profit(market_net, cost, fee.fee_pct, fee.fixed_fee, pack_cost, ship_subsidy)

                profit_match_pct = (profit_match / market_net) * 100.0
                if profit_match_pct >= 0:
                    add_rec("follow_ok")
                elif profit_match_pct > -6:
                    add_rec("loss_0_5")
                elif profit_match_pct > -10:
                    add_rec("loss_6_10")
                else:
                    add_rec("loss_heavy")

            # Aging logic (Exclusive): choose the highest bucket only (1ปี > 6เดือน > 3เดือน)
            aging_bucket = None
            aging_label = None
            if stock_internal is not None:
                if (monthly_sales * 12 - stock_internal) < 0:
                    aging_bucket = "aging12"
                    aging_label = "Aging(1ปีขึ้นไป)"
                elif (monthly_sales * 6 - stock_internal) < 0:
                    aging_bucket = "aging6"
                    aging_label = "Aging(6เดือนขึ้นไป)"
                elif (monthly_sales * 3 - stock_internal) < 0:
                    aging_bucket = "aging3"
                    aging_label = "Aging(3เดือนขึ้นไป)"

            no_sales = (monthly_sales == 0) and (stock_internal is not None and stock_internal >= 1)

            mu = market_updated
            try:
                if mu is not None and getattr(mu, "tzinfo", None) is not None:
                    mu = mu.replace(tzinfo=None)
            except Exception:
                pass
            has_stock = (int(stock_internal or 0) > 0) or (int(stock_total or 0) > 0)
            is_stale = (mu is None) or (mu < cutoff)
            need_market = has_stock and is_stale

            age_tags: list[str] = []
            if no_sales:
                age_tags.append("ไม่มียอดขาย")
            if aging_label:
                age_tags.append(aging_label)

            rec_labels = [REC_DEFS[k]["label"] for k in rec_keys]
            rec_text = ", ".join(rec_labels) if rec_labels else "-"
            if age_tags:
                rec_text = (rec_text + " | " + ", ".join(age_tags)) if rec_text and rec_text != "-" else ", ".join(age_tags)

            export_rows.append({
                "SKU": sku,
                "Brand": brand,
                "Name": name,
                "Stock Internal": stock_internal,
                "Stock": stock_total,
                "Monthly Sales": monthly_sales,
                "Cost": cost,
                "Our Price": our_price,
                "Market (best)": market_net,
                "Voucher": market_voucher,
                "Brand Control": brand_control,
                "Gap": gap,
                "Profit@Our": profit_now,
                "Profit@Match": profit_match,
                "Recommend": rec_text,
                "Shop": market_shop,
                "MALL": "MALL" if market_is_mall else "",
                "URL": market_url,
                "ผู้ดูแล": owner,
                "Updated": to_thai_be(market_updated) if market_updated else "",
                # for KPI filtering
                "brand_control": brand_control,
                "need_market": need_market,
                # fields for KPI filtering
                "rec_keys": rec_keys,
                "aging_bucket": aging_bucket,
                "no_sales": no_sales,
            })

        rows_master = _apply_kpi_filters(export_rows, [master_sel]) if master_sel else export_rows
        export_rows = _apply_kpi_filters(rows_master, kpi_selected)

        # ✅ สำคัญ: ต่อให้ export_rows ว่าง ก็ยังมีหัวคอลัมน์ครบ
        df = pd.DataFrame(export_rows, columns=cols)
        df = sanitize_excel_df(df)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="PriceDashboard")
        bio.seek(0)

        filename = f"price_dashboard_{platform}_{now_thai().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/price/dashboard/export_price.xlsx")
    @login_required
    def price_dashboard_export_price():
        """Export Price with selectable columns + computed sell tiers."""

        platform_arg = (request.args.get("platform") or "").strip()
        platform = normalize_platform(platform_arg) if platform_arg else "Shopee"

        q = (request.args.get("q") or "").strip().lower()
        owner_sel = (request.args.get("owner") or "").strip()
        limit_sel = (request.args.get("limit") or "200").strip().lower()
        stale_days_raw = (request.args.get("stale_days") or "").strip()
        kpi_sel = (request.args.get("kpi") or "").strip()
        master_raw = (request.args.get("master") or "").strip()

        sort_brand_param = request.args.get("sort_brand")
        include_zero_param = request.args.get("include_zero_stock")

        cfg = {}
        try:
            cu = current_user()
            plat_key = normalize_platform_key(platform) or "DEFAULT"
            pref_key = f"pm_export_price_cols::{plat_key}"
            raw = get_user_pref(int(cu.id), pref_key, default="") if cu else ""
            cfg = json.loads(raw) if raw else {}
        except Exception:
            cfg = {}
        if not isinstance(cfg, dict):
            cfg = {}

        def _as_bool(v) -> bool:
            return str(v).strip().lower() in {"1", "true", "on", "yes"}

        if sort_brand_param is None:
            sort_brand = bool(cfg.get("sort_brand", False))
        else:
            sort_brand = _as_bool(sort_brand_param)

        if include_zero_param is None:
            include_zero_stock = bool(cfg.get("include_zero_stock", False))
        else:
            include_zero_stock = _as_bool(include_zero_param)

        allowed_limits = {"100", "200", "300", "500", "1000", "all"}
        if limit_sel not in allowed_limits:
            limit_sel = "200"
        limit_n = None if limit_sel == "all" else int(limit_sel)

        try:
            stale_days = int(stale_days_raw or 14)
        except Exception:
            stale_days = 14
        stale_days = max(1, min(stale_days, 365))

        now0 = now_thai()
        try:
            if getattr(now0, "tzinfo", None) is not None:
                now0 = now0.replace(tzinfo=None)
        except Exception:
            pass
        cutoff = now0 - timedelta(days=stale_days)

        fee = PlatformFeeSetting.query.get(platform)
        if not fee:
            fee = PlatformFeeSetting(platform=platform, label=platform, fee_pct=0.0, fixed_fee=0.0, is_active=True, sort_order=0)
            db.session.add(fee)
            db.session.commit()

        try:
            export_setting = PriceExportSetting.query.get(1)
        except Exception:
            export_setting = None
        if not export_setting:
            export_setting = PriceExportSetting(
                id=1,
                step_pct=5.0,
                min_profit_pct=5.0,
                loss_aging3_pct=5.0,
                loss_aging6_pct=10.0,
                loss_aging12_pct=20.0,
            )

        def calc_profit(price, cost, fee_pct, fixed_fee, pack_cost, ship_subsidy):
            price = float(price or 0.0)
            cost = float(cost or 0.0)
            fee_amt = (price * (float(fee_pct or 0.0) / 100.0)) + float(fixed_fee or 0.0)
            return price - cost - fee_amt - float(pack_cost or 0.0) - float(ship_subsidy or 0.0)

        COL_DEFS = [
            ("sku", "SKU"),
            ("brand", "Brand"),
            ("name", "Name"),
            ("spec", "Spec"),
            ("stock_internal", "Stock Internal"),
            ("stock", "Stock"),
            ("monthly_sales", "Monthly Sales"),
            ("cost", "Cost"),
            ("our_price", "Our Price"),
            ("sell_1", "ราคาขาย 1"),
            ("sell_2", "ราคาขาย 2"),
            ("sell_3", "ราคาขาย 3"),
            ("sell_4", "ราคาขาย 4"),
            ("sell_5", "ราคาขาย 5"),
            ("sell_min", "ราคาขายต่ำสุด"),
            ("market_best", "Market (best)"),
            ("voucher", "Voucher"),
            ("brand_control", "Brand Control"),
            ("gap", "Gap"),
            ("profit_our", "Profit@Our"),
            ("profit_match", "Profit@Match"),
            ("recommend", "Recommend"),
            ("shop", "Shop"),
            ("mall", "MALL"),
            ("url", "URL"),
            ("owner", "ผู้ดูแล"),
            ("updated", "Updated"),
        ]
        LABEL = {k: v for k, v in COL_DEFS}
        ORDERED_KEYS = [k for k, _ in COL_DEFS]

        selected = request.args.getlist("cols")

        # ✅ ถ้า user ไม่ส่ง cols มาเลย ให้ใช้ค่าที่บันทึกไว้ก่อน
        if not selected:
            try:
                saved_cols = cfg.get("cols")
                if isinstance(saved_cols, list) and saved_cols:
                    selected = saved_cols
            except Exception:
                pass
        if not selected:
            selected_keys = ORDERED_KEYS
        else:
            allowed = set(ORDERED_KEYS)
            selected_set = {c for c in selected if c in allowed}
            selected_keys = [k for k in ORDERED_KEYS if k in selected_set]

        pricing_map = {p.sku: p for p in SkuPricing.query.all()}
        bc_map = {b.sku: b for b in BrandControl.query.all()}

        owner_rows = BrandOwnerSetting.query.all()
        brand_owner_map = {str(r.brand or "").strip(): str(r.owner or "").strip() for r in owner_rows if r}

        items = (
            MarketItem.query
            .filter(MarketItem.platform == platform, MarketItem.is_active == True)
            .all()
        )
        from datetime import datetime

        def _ts_market(it):
            t = getattr(it, "last_updated", None)
            if not t:
                return datetime.min
            try:
                if getattr(t, "tzinfo", None) is not None:
                    t = t.replace(tzinfo=None)
            except Exception:
                pass
            return t

        latest_by_sku = {}
        for it in items:
            cur = latest_by_sku.get(it.sku)
            if cur is None:
                latest_by_sku[it.sku] = it
                continue
            if _ts_market(it) > _ts_market(cur):
                latest_by_sku[it.sku] = it
                continue
            if _ts_market(it) == _ts_market(cur) and (it.id or 0) > (cur.id or 0):
                latest_by_sku[it.sku] = it

        sku_set = set(pricing_map.keys()) | set(latest_by_sku.keys()) | set(bc_map.keys())

        if q:
            filtered = set()
            for sku in sku_set:
                pr = pricing_map.get(sku)
                mk = latest_by_sku.get(sku)
                hay = [
                    (sku or "").lower(),
                    (getattr(pr, "spec_text", "") or "").lower(),
                    (getattr(pr, "brand", "") or "").lower(),
                    (getattr(pr, "name", "") or "").lower(),
                    (getattr(mk, "shop_name", "") or "").lower(),
                ]
                if any(q in s for s in hay):
                    filtered.add(sku)
            sku_set = filtered

        if owner_sel:
            filtered = set()
            for sku in sku_set:
                pr = pricing_map.get(sku)
                brand = (getattr(pr, "brand", "") or "").strip() if pr else ""
                owner = brand_owner_map.get(brand, "") if brand else ""
                if owner == owner_sel:
                    filtered.add(sku)
            sku_set = filtered

        sku_list = sorted(sku_set)
        if limit_n is not None:
            sku_list = sku_list[:limit_n]

        REC_DEFS = {
            "market_cheaper": {"label": "ตลาดถูกกว่า"},
            "equal_price": {"label": "ราคาเท่ากัน"},
            "follow_ok": {"label": "ตามได้"},
            "loss_0_5": {"label": "ขาดทุน(0-5%)"},
            "loss_6_10": {"label": "ขาดทุน(6-10%)"},
            "loss_heavy": {"label": "ขาดทุนหนัก"},
            "no_market": {"label": "ไม่มีราคาตลาด"},
            "missing_internal": {"label": "ข้อมูลฝั่งเราไม่ครบ"},
        }

        allowed_kpis = set(REC_DEFS.keys()) | {"no_sales", "aging3", "aging6", "aging12", "brand_control", "need_market"}
        kpi_selected = _parse_kpi_multi(kpi_sel, allowed_kpis)
        master_sel, kpi_selected = _resolve_master(master_raw, kpi_selected, allowed_kpis)

        export_rows: list[dict] = []
        for sku in sku_list:
            pr = pricing_map.get(sku)
            mk = latest_by_sku.get(sku)
            bc = bc_map.get(sku)

            stock_total_for_filter = int(getattr(pr, "stock_qty", 0) or 0) if pr else 0
            if (not include_zero_stock) and (stock_total_for_filter <= 0):
                continue

            cost = float(pr.cost) if (pr and pr.cost is not None) else None
            our_price = float(pr.our_price) if (pr and pr.our_price is not None) else None
            pack_cost = float(pr.pack_cost) if (pr and pr.pack_cost is not None) else 0.0
            ship_subsidy = float(pr.ship_subsidy) if (pr and pr.ship_subsidy is not None) else 0.0

            # Offline effective cost for sell tiers (no platform fee)
            effective_cost = None
            if cost is not None:
                try:
                    effective_cost = float(cost) + float(pack_cost or 0.0) + float(ship_subsidy or 0.0)
                except Exception:
                    effective_cost = cost

            brand = (getattr(pr, "brand", "") or "").strip() if pr else ""
            name = (getattr(pr, "name", "") or "").strip() if pr else ""
            spec = (getattr(pr, "spec_text", "") or "").strip() if pr else ""
            owner = brand_owner_map.get(brand, "") if brand else ""

            stock_internal = int(getattr(pr, "stock_internal_qty", 0) or 0) if (pr and getattr(pr, "stock_internal_qty", None) is not None) else None
            stock_total = int(getattr(pr, "stock_qty", 0) or 0) if (pr and getattr(pr, "stock_qty", None) is not None) else None
            monthly_sales = int(getattr(pr, "monthly_sales_qty", 0) or 0) if pr else 0

            market_net = float(mk.latest_net_price) if (mk and mk.latest_net_price is not None) else None
            market_voucher = float(mk.latest_voucher_discount) if (mk and mk.latest_voucher_discount is not None) else None
            market_shop = mk.shop_name if mk else None
            market_is_mall = bool(getattr(mk, "is_mall", False)) if mk else False
            market_url = mk.product_url if mk else None
            market_updated = mk.last_updated if mk else None

            brand_control = float(bc.price_control) if (bc and bc.price_control is not None) else None

            gap = (our_price - market_net) if (our_price is not None and market_net is not None) else None
            profit_now = None
            profit_match = None

            rec_keys: list[str] = []

            def add_rec(key: str):
                if key in REC_DEFS and key not in rec_keys:
                    rec_keys.append(key)

            no_market_flag = (mk is None) or (market_net is None) or (market_net <= 0)
            missing_internal_flag = (pr is None) or (cost is None) or (our_price is None) or (our_price <= 0)
            if no_market_flag:
                add_rec("no_market")
            if missing_internal_flag:
                add_rec("missing_internal")

            if (our_price is not None and our_price > 0) and (market_net is not None and market_net > 0):
                if abs(our_price - market_net) < 0.01:
                    add_rec("equal_price")
                elif market_net < our_price:
                    add_rec("market_cheaper")

            if our_price is not None and our_price > 0 and cost is not None:
                profit_now = calc_profit(our_price, cost, fee.fee_pct, fee.fixed_fee, pack_cost, ship_subsidy)

            if market_net is not None and market_net > 0 and cost is not None:
                profit_match = calc_profit(market_net, cost, fee.fee_pct, fee.fixed_fee, pack_cost, ship_subsidy)
                profit_match_pct = (profit_match / market_net) * 100.0
                if profit_match_pct >= 0:
                    add_rec("follow_ok")
                elif profit_match_pct > -6:
                    add_rec("loss_0_5")
                elif profit_match_pct > -10:
                    add_rec("loss_6_10")
                else:
                    add_rec("loss_heavy")

            aging_bucket = None
            aging_label = None
            if stock_internal is not None:
                if (monthly_sales * 12 - stock_internal) < 0:
                    aging_bucket = "aging12"
                    aging_label = "Aging(1ปีขึ้นไป)"
                elif (monthly_sales * 6 - stock_internal) < 0:
                    aging_bucket = "aging6"
                    aging_label = "Aging(6เดือนขึ้นไป)"
                elif (monthly_sales * 3 - stock_internal) < 0:
                    aging_bucket = "aging3"
                    aging_label = "Aging(3เดือนขึ้นไป)"

            no_sales = (monthly_sales == 0) and (stock_internal is not None and stock_internal >= 1)

            mu = market_updated
            try:
                if mu is not None and getattr(mu, "tzinfo", None) is not None:
                    mu = mu.replace(tzinfo=None)
            except Exception:
                pass
            has_stock = (int(stock_internal or 0) > 0) or (int(stock_total or 0) > 0)
            is_stale = (mu is None) or (mu < cutoff)
            need_market = has_stock and is_stale

            age_tags: list[str] = []
            if no_sales:
                age_tags.append("ไม่มียอดขาย")
            if aging_label:
                age_tags.append(aging_label)

            rec_labels = [REC_DEFS[k]["label"] for k in rec_keys]
            rec_text = ", ".join(rec_labels) if rec_labels else "-"
            if age_tags:
                rec_text = (rec_text + " | " + ", ".join(age_tags)) if rec_text and rec_text != "-" else ", ".join(age_tags)

            sell = None
            if our_price is not None and float(our_price or 0) > 0:
                aging_for_loss = "aging12" if no_sales else aging_bucket
                sell = build_sell_prices(
                    our_price=our_price,
                    cost=effective_cost,
                    step_pct=export_setting.step_pct,
                    min_profit_pct=export_setting.min_profit_pct,
                    loss_aging3_pct=export_setting.loss_aging3_pct,
                    loss_aging6_pct=export_setting.loss_aging6_pct,
                    loss_aging12_pct=export_setting.loss_aging12_pct,
                    aging_bucket=aging_for_loss,
                )

            export_rows.append({
                "sku": sku,
                "brand": brand,
                "name": name,
                "spec": spec,
                "stock_internal": stock_internal,
                "stock": stock_total,
                "monthly_sales": monthly_sales,
                "cost": cost,
                "our_price": our_price,
                "sell_1": (sell[0] if sell else None),
                "sell_2": (sell[1] if sell else None),
                "sell_3": (sell[2] if sell else None),
                "sell_4": (sell[3] if sell else None),
                "sell_5": (sell[4] if sell else None),
                "sell_min": (sell[5] if sell else None),
                "market_best": market_net,
                "voucher": market_voucher,
                "brand_control": brand_control,
                "gap": gap,
                "profit_our": profit_now,
                "profit_match": profit_match,
                "recommend": rec_text,
                "shop": market_shop,
                "mall": "MALL" if market_is_mall else "",
                "url": market_url,
                "owner": owner,
                "updated": to_thai_be(market_updated) if market_updated else "",
                # for KPI filtering
                "need_market": need_market,
                "rec_keys": rec_keys,
                "aging_bucket": aging_bucket,
                "no_sales": no_sales,
            })

        rows_master = _apply_kpi_filters(export_rows, [master_sel]) if master_sel else export_rows
        export_rows = _apply_kpi_filters(rows_master, kpi_selected)

        if sort_brand:
            export_rows.sort(
                key=lambda r: (
                    1 if not (str(r.get("brand") or "").strip()) else 0,
                    str(r.get("brand") or "").strip().lower(),
                    str(r.get("sku") or "").strip().lower(),
                )
            )

        final_rows = [{LABEL[k]: r.get(k) for k in selected_keys} for r in export_rows]
        df = pd.DataFrame(final_rows, columns=[LABEL[k] for k in selected_keys])
        df = sanitize_excel_df(df)

        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="ExportPrice")
        bio.seek(0)

        ts = now_thai().strftime("%Y%m%d_%H%M")
        filename = f"price_export_sale_{ts}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/price/dashboard/export_stock_adj.xlsx")
    @login_required
    def price_dashboard_export_stock_adj():
        """Export Price Marketing dashboard rows to Excel with export-only Our Price adjustment."""

        # --- 0) Filters (keep same behavior as dashboard export) ---
        platform_arg = (request.args.get("platform") or "").strip()
        platform = normalize_platform(platform_arg) if platform_arg else "Shopee"

        q = (request.args.get("q") or "").strip().lower()
        owner_sel = (request.args.get("owner") or "").strip()
        limit_sel = (request.args.get("limit") or "200").strip().lower()
        stale_days_raw = (request.args.get("stale_days") or "").strip()
        kpi_sel = (request.args.get("kpi") or "").strip()
        master_raw = (request.args.get("master") or "").strip()

        allowed_limits = {"100", "200", "300", "500", "1000", "all"}
        if limit_sel not in allowed_limits:
            limit_sel = "200"
        limit_n = None if limit_sel == "all" else int(limit_sel)

        try:
            stale_days = int(stale_days_raw or 14)
        except Exception:
            stale_days = 14
        stale_days = max(1, min(stale_days, 365))

        now0 = now_thai()
        try:
            if getattr(now0, "tzinfo", None) is not None:
                now0 = now0.replace(tzinfo=None)
        except Exception:
            pass
        cutoff = now0 - timedelta(days=stale_days)

        # --- 1) Modal inputs ---
        adj_pct = _clamp_float(request.args.get("adj_pct"), -50.0, 50.0, default=0.0)

        preview = (request.args.get("preview") or "").strip() == "1"

        # New: allow multi-rule selection (checkbox)
        stock_rules = request.args.getlist("stock_rule")  # ['1','2','4','5',...]
        stock_rules = [str(x).strip() for x in stock_rules if str(x).strip()]
        allowed_rules = {"1", "2", "3", "4", "5"}
        stock_rules_set = {r for r in stock_rules if r in allowed_rules}

        # Backward-compat: map legacy single-mode into new rule set if no stock_rule provided
        if not stock_rules_set:
            stock_mode = (request.args.get("stock_mode") or "").strip()
            legacy_map = {
                "r11": {"1"},
                "r12": {"2"},
                "div": {"3"},
                "brand0": {"4"},
                "brand_div": {"5"},
            }
            if stock_mode in legacy_map:
                stock_rules_set = set(legacy_map[stock_mode])

        # Enforce constraints: rule 3 cannot be used with 1-2
        if "3" in stock_rules_set:
            stock_rules_set.discard("1")
            stock_rules_set.discard("2")
        elif "1" in stock_rules_set or "2" in stock_rules_set:
            stock_rules_set.discard("3")

        try:
            stock_divisor = int((request.args.get("stock_divisor") or "").strip() or 3)
        except Exception:
            stock_divisor = 3

        # Rule 3 divisor
        if "3" in stock_rules_set:
            stock_divisor = max(1, int(stock_divisor or 0))

        # Rule 5 divisor (separate; backward-compat: fallback to stock_divisor if not provided)
        stock_divisor_brand_raw = (request.args.get("stock_divisor_brand") or "").strip()
        if stock_divisor_brand_raw:
            try:
                stock_divisor_brand = int(stock_divisor_brand_raw)
            except Exception:
                stock_divisor_brand = 2
        else:
            # legacy UI used stock_divisor for rule 5
            stock_divisor_brand = int(stock_divisor or 3)
        if "5" in stock_rules_set:
            stock_divisor_brand = max(1, int(stock_divisor_brand or 0))

        # Brand lists
        stock_brands_4_raw = (request.args.get("stock_brands_4") or "").strip()
        stock_brands_5_raw = (request.args.get("stock_brands_5") or "").strip()
        # Backward-compat: old UI used a single stock_brands for both (but only one mode selectable)
        stock_brands_raw = (request.args.get("stock_brands") or "").strip()

        def _parse_brands(s: str) -> set[str]:
            if not s:
                return set()
            return {b.strip().lower() for b in re.split(r"[,\n]+", s) if b and b.strip()}

        brands_4_set = _parse_brands(stock_brands_4_raw)
        brands_5_set = _parse_brands(stock_brands_5_raw)
        legacy_brands_set = _parse_brands(stock_brands_raw)

        if ("4" in stock_rules_set) and (not brands_4_set) and legacy_brands_set:
            brands_4_set = set(legacy_brands_set)
        if ("5" in stock_rules_set) and (not brands_5_set) and legacy_brands_set:
            brands_5_set = set(legacy_brands_set)

        def _as_bool_last_arg(name: str, default: bool = True) -> bool:
            vals = request.args.getlist(name)
            if not vals:
                return default
            return str(vals[-1]).strip().lower() in {"1", "true", "on", "yes"}

        # Rule 5 toggle: floor 1-2 result to 0 (default: True for backward compat)
        brand_floor12_to0 = _as_bool_last_arg("stock_divide_brand_floor12_to0", default=True)

        # Rule 5 toggle: min 1 when internal>0 but division yields 0 (default: False)
        brand_min1_when_zero = _as_bool_last_arg("stock_divide_brand_min1_when_zero", default=False)

        # NEW: price base for export (Our Price vs Sell 1)
        use_sell1 = _as_bool_last_arg("use_sell1", default=False)

        # NEW: Skip price adjustment when Brand Control > 0 (use Our Price only)
        skip_adj_when_brand_control = _as_bool_last_arg("skip_adj_when_brand_control", default=False)

        # NEW: Skip price adjustment when Profit@Our% >= X (use Our Price only)
        # Note: Applies only when adj_pct > 0 (to avoid lowering price when negative).
        skip_adj_when_profit_our = _as_bool_last_arg("skip_adj_when_profit_our", default=False)
        profit_our_min_pct = _clamp_float(request.args.get("profit_our_min_pct"), 0.0, 50.0, default=10.0)

        adj_platform = (request.args.get("adj_platform") or "").strip()
        adj_platform_other = (request.args.get("adj_platform_other") or "").strip()
        if adj_platform.lower() in {"other", "อื่นๆ"} and adj_platform_other:
            adj_platform = adj_platform_other
        adj_platform = adj_platform or platform

        base_txt = "Sell1" if use_sell1 else "OurPrice"
        adj_note = f"Platform={adj_platform} | Base={base_txt} | Adj={adj_pct:+.2f}%"

        # Export Price Settings (for Sell1 computation)
        export_setting = None
        if use_sell1:
            try:
                export_setting = PriceExportSetting.query.get(1)
            except Exception:
                export_setting = None
            if not export_setting:
                export_setting = PriceExportSetting(
                    id=1,
                    step_pct=5.0,
                    min_profit_pct=5.0,
                    loss_aging3_pct=5.0,
                    loss_aging6_pct=10.0,
                    loss_aging12_pct=20.0,
                )

        # --- 2) Selectable columns + 4 audit columns ---
        COL_DEFS = [
            ("sku", "SKU"),
            ("brand", "Brand"),
            ("name", "Name"),
            ("stock_internal", "Stock Internal"),
            ("stock", "Stock"),
            ("stock_adj", "Stock Adj"),
            ("monthly_sales", "Monthly Sales"),
            ("cost", "Cost"),
            ("our_price", "Our Price"),
            ("market_best", "Market (best)"),
            ("voucher", "Voucher"),
            ("brand_control", "Brand Control"),
            ("gap", "Gap"),
            ("profit_our", "Profit@Our"),
            ("profit_match", "Profit@Match"),
            ("recommend", "Recommend"),
            ("shop", "Shop"),
            ("mall", "MALL"),
            ("url", "URL"),
            ("owner", "ผู้ดูแล"),
            ("updated", "Updated"),
            # === audit columns (ท้ายไฟล์) ===
            ("adj_note", "หมายเหตุการแก้ไข Our Price"),
            ("our_before", "Our Price (ก่อนปรับ)"),
            ("our_after", "Our Price (หลังปรับ)"),
            ("adj_pct", "ปรับกี่ %"),
        ]
        LABEL = {k: v for k, v in COL_DEFS}
        ORDERED_KEYS = [k for k, _ in COL_DEFS]

        selected = request.args.getlist("cols")
        if not selected:
            selected_keys = ORDERED_KEYS
        else:
            allowed = set(ORDERED_KEYS)
            selected_set = {c for c in selected if c in allowed}
            selected_keys = [k for k in ORDERED_KEYS if k in selected_set]

        # Force audit columns to always be included (match UI note: ท้ายไฟล์มีให้เสมอ)
        for k in ["adj_note", "our_before", "our_after", "adj_pct"]:
            if k not in selected_keys:
                selected_keys.append(k)

        # --- 3) Load base data (same sources as dashboard export) ---
        fee = PlatformFeeSetting.query.get(platform)
        if not fee:
            fee = PlatformFeeSetting(platform=platform, label=platform, fee_pct=0.0, fixed_fee=0.0, is_active=True, sort_order=0)
            db.session.add(fee)
            db.session.commit()

        def calc_profit(price, cost, fee_pct, fixed_fee, pack_cost, ship_subsidy):
            price = float(price or 0.0)
            cost = float(cost or 0.0)
            fee_amt = (price * (float(fee_pct or 0.0) / 100.0)) + float(fixed_fee or 0.0)
            return price - cost - fee_amt - float(pack_cost or 0.0) - float(ship_subsidy or 0.0)

        pricing_map = {p.sku: p for p in SkuPricing.query.all()}
        bc_map = {b.sku: b for b in BrandControl.query.all()}

        owner_rows = BrandOwnerSetting.query.all()
        brand_owner_map = {str(r.brand or "").strip(): str(r.owner or "").strip() for r in owner_rows if r}

        items = (
            MarketItem.query
            .filter(MarketItem.platform == platform, MarketItem.is_active == True)
            .all()
        )
        from datetime import datetime

        def _ts_market(it):
            t = getattr(it, "last_updated", None)
            if not t:
                return datetime.min
            try:
                if getattr(t, "tzinfo", None) is not None:
                    t = t.replace(tzinfo=None)
            except Exception:
                pass
            return t

        latest_by_sku = {}
        for it in items:
            cur = latest_by_sku.get(it.sku)
            if cur is None:
                latest_by_sku[it.sku] = it
                continue
            if _ts_market(it) > _ts_market(cur):
                latest_by_sku[it.sku] = it
                continue
            if _ts_market(it) == _ts_market(cur) and (it.id or 0) > (cur.id or 0):
                latest_by_sku[it.sku] = it

        sku_set = set(pricing_map.keys()) | set(latest_by_sku.keys()) | set(bc_map.keys())

        if q:
            filtered = set()
            for sku in sku_set:
                pr = pricing_map.get(sku)
                mk = latest_by_sku.get(sku)
                hay = [
                    (sku or "").lower(),
                    (getattr(pr, "spec_text", "") or "").lower(),
                    (getattr(pr, "brand", "") or "").lower(),
                    (getattr(pr, "name", "") or "").lower(),
                    (getattr(mk, "shop_name", "") or "").lower(),
                ]
                if any(q in s for s in hay):
                    filtered.add(sku)
            sku_set = filtered

        if owner_sel:
            filtered = set()
            for sku in sku_set:
                pr = pricing_map.get(sku)
                brand = (getattr(pr, "brand", "") or "").strip() if pr else ""
                owner = brand_owner_map.get(brand, "") if brand else ""
                if owner == owner_sel:
                    filtered.add(sku)
            sku_set = filtered

        sku_list = sorted(sku_set)
        if limit_n is not None:
            sku_list = sku_list[:limit_n]

        REC_DEFS = {
            "market_cheaper": {"label": "ตลาดถูกกว่า"},
            "equal_price": {"label": "ราคาเท่ากัน"},
            "follow_ok": {"label": "ตามได้"},
            "loss_0_5": {"label": "ขาดทุน(0-5%)"},
            "loss_6_10": {"label": "ขาดทุน(6-10%)"},
            "loss_heavy": {"label": "ขาดทุนหนัก"},
            "no_market": {"label": "ไม่มีราคาตลาด"},
            "missing_internal": {"label": "ข้อมูลฝั่งเราไม่ครบ"},
        }

        allowed_kpis = set(REC_DEFS.keys()) | {"no_sales", "aging3", "aging6", "aging12", "brand_control", "need_market"}
        kpi_selected = _parse_kpi_multi(kpi_sel, allowed_kpis)
        master_sel, kpi_selected = _resolve_master(master_raw, kpi_selected, allowed_kpis)

        def _as_int0(x) -> int:
            try:
                return int(x or 0)
            except Exception:
                return 0

        def compute_stock_adj(stock_total, stock_internal, brand: str) -> int:
            s = _as_int0(stock_total)
            i = _as_int0(stock_internal)
            b = (brand or "").strip().lower()

            # (1.7) ถ้าทั้งคู่ 0/ว่าง -> 0 เสมอ
            if s <= 0 and i <= 0:
                return 0

            # ถ้าไม่เลือกกฎ -> ไม่ปรับ
            if not stock_rules_set:
                return s

            # ---------- Priority 1: Brand override (ข้อ 4/5) ----------
            if "5" in stock_rules_set and b and (b in brands_5_set):
                if i <= 0:
                    return 0
                v = i // stock_divisor_brand
                from_min1 = False
                if brand_min1_when_zero and i > 0 and v == 0:
                    v = 1
                    from_min1 = True

                if brand_floor12_to0 and (not from_min1) and (1 <= v <= 2):
                    v = 0
                return int(v)

            if "4" in stock_rules_set and b and (b in brands_4_set):
                return 0

            # ---------- Base rule (ข้อ 1/2 หรือ 3) ----------
            # (3) โหมด global: ใช้ internal ÷ divisor (ไม่ใช้ stock)
            if "3" in stock_rules_set:
                v = i // stock_divisor
                if 1 <= v <= 2:
                    v = 0
                return int(v)

            # (1)(2) ใช้ stock เป็นฐาน
            v = s
            if "1" in stock_rules_set and (1 <= s <= 3) and (i == 0):
                v = 0
            if "2" in stock_rules_set and (1 <= s <= 5) and (1 <= i <= 5):
                v = 1
            return int(v)

        export_rows: list[dict] = []
        for sku in sku_list:
            pr = pricing_map.get(sku)
            mk = latest_by_sku.get(sku)
            bc = bc_map.get(sku)

            cost = float(pr.cost) if (pr and pr.cost is not None) else None
            our_price = float(pr.our_price) if (pr and pr.our_price is not None) else None
            pack_cost = float(pr.pack_cost) if (pr and pr.pack_cost is not None) else 0.0
            ship_subsidy = float(pr.ship_subsidy) if (pr and pr.ship_subsidy is not None) else 0.0

            # Offline effective cost for sell tiers (no platform fee)
            effective_cost = None
            if use_sell1 and cost is not None:
                try:
                    effective_cost = float(cost) + float(pack_cost or 0.0) + float(ship_subsidy or 0.0)
                except Exception:
                    effective_cost = cost

            brand = (getattr(pr, "brand", "") or "").strip() if pr else ""
            name = (getattr(pr, "name", "") or "").strip() if pr else ""
            owner = brand_owner_map.get(brand, "") if brand else ""

            stock_internal = int(getattr(pr, "stock_internal_qty", 0) or 0) if (pr and getattr(pr, "stock_internal_qty", None) is not None) else None
            stock_total = int(getattr(pr, "stock_qty", 0) or 0) if (pr and getattr(pr, "stock_qty", None) is not None) else None
            monthly_sales = int(getattr(pr, "monthly_sales_qty", 0) or 0) if pr else 0

            stock_adj = compute_stock_adj(stock_total, stock_internal, brand)

            market_net = float(mk.latest_net_price) if (mk and mk.latest_net_price is not None) else None
            market_voucher = float(mk.latest_voucher_discount) if (mk and mk.latest_voucher_discount is not None) else None
            market_shop = mk.shop_name if mk else None
            market_is_mall = bool(getattr(mk, "is_mall", False)) if mk else False
            market_url = mk.product_url if mk else None
            market_updated = mk.last_updated if mk else None

            brand_control = float(bc.price_control) if (bc and bc.price_control is not None) else None

            # KPI/filter logic should match the dashboard (use original our_price)
            rec_keys: list[str] = []

            def add_rec(key: str):
                if key in REC_DEFS and key not in rec_keys:
                    rec_keys.append(key)

            no_market_flag = (mk is None) or (market_net is None) or (market_net <= 0)
            missing_internal_flag = (pr is None) or (cost is None) or (our_price is None) or (our_price <= 0)
            if no_market_flag:
                add_rec("no_market")
            if missing_internal_flag:
                add_rec("missing_internal")

            if (our_price is not None and our_price > 0) and (market_net is not None and market_net > 0):
                if abs(our_price - market_net) < 0.01:
                    add_rec("equal_price")
                elif market_net < our_price:
                    add_rec("market_cheaper")

            profit_match = None
            if market_net is not None and market_net > 0 and cost is not None:
                profit_match = calc_profit(market_net, cost, fee.fee_pct, fee.fixed_fee, pack_cost, ship_subsidy)
                profit_match_pct = (profit_match / market_net) * 100.0
                if profit_match_pct >= 0:
                    add_rec("follow_ok")
                elif profit_match_pct > -6:
                    add_rec("loss_0_5")
                elif profit_match_pct > -10:
                    add_rec("loss_6_10")
                else:
                    add_rec("loss_heavy")

            aging_bucket = None
            aging_label = None
            if stock_internal is not None:
                if (monthly_sales * 12 - stock_internal) < 0:
                    aging_bucket = "aging12"
                    aging_label = "Aging(1ปีขึ้นไป)"
                elif (monthly_sales * 6 - stock_internal) < 0:
                    aging_bucket = "aging6"
                    aging_label = "Aging(6เดือนขึ้นไป)"
                elif (monthly_sales * 3 - stock_internal) < 0:
                    aging_bucket = "aging3"
                    aging_label = "Aging(3เดือนขึ้นไป)"

            no_sales = (monthly_sales == 0) and (stock_internal is not None and stock_internal >= 1)

            mu = market_updated
            try:
                if mu is not None and getattr(mu, "tzinfo", None) is not None:
                    mu = mu.replace(tzinfo=None)
            except Exception:
                pass
            has_stock = (int(stock_internal or 0) > 0) or (int(stock_total or 0) > 0)
            is_stale = (mu is None) or (mu < cutoff)
            need_market = has_stock and is_stale

            # --- export-only price adjust (does not affect KPI filtering) ---
            sell1 = None
            if use_sell1 and (our_price is not None) and float(our_price or 0) > 0:
                aging_for_loss = "aging12" if no_sales else aging_bucket
                sell = build_sell_prices(
                    our_price=our_price,
                    cost=effective_cost,
                    step_pct=export_setting.step_pct,
                    min_profit_pct=export_setting.min_profit_pct,
                    loss_aging3_pct=export_setting.loss_aging3_pct,
                    loss_aging6_pct=export_setting.loss_aging6_pct,
                    loss_aging12_pct=export_setting.loss_aging12_pct,
                    aging_bucket=aging_for_loss,
                )
                sell1 = (sell[0] if sell else None)

            base_price = sell1 if (use_sell1 and sell1 is not None) else our_price

            has_bc = (brand_control is not None) and (float(brand_control or 0) > 0)

            # Profit@Our% (computed from original Our Price)
            profit_base_pct = None
            if (our_price is not None and float(our_price or 0) > 0) and (cost is not None):
                try:
                    p0 = calc_profit(our_price, cost, fee.fee_pct, fee.fixed_fee, pack_cost, ship_subsidy)
                    profit_base_pct = (p0 / float(our_price)) * 100.0 if float(our_price) else None
                except Exception:
                    profit_base_pct = None

            skip_bc = bool(skip_adj_when_brand_control and has_bc)
            skip_profit = (
                bool(skip_adj_when_profit_our)
                and (float(adj_pct or 0.0) > 0.0)
                and (profit_base_pct is not None)
                and (float(profit_base_pct) >= float(profit_our_min_pct))
            )

            if skip_bc:
                # Keep Our Price as-is. Do not apply adj_pct and do not use Sell1.
                base_price = our_price
                our_before = our_price
                our_export = our_price
                row_adj_pct = 0.0
                adj_note_row = adj_note + " | SkipAdj(Bc>0)"

            elif skip_profit:
                # Keep Our Price as-is. Do not apply adj_pct and do not use Sell1.
                base_price = our_price
                our_before = our_price
                our_export = our_price
                row_adj_pct = 0.0
                adj_note_row = adj_note + f" | SkipAdj(ProfitOur≥{profit_our_min_pct:g}%)"
            else:
                our_before = base_price
                our_after = adjust_our_price_export(base_price, adj_pct) if base_price is not None else None
                our_export = our_after if our_after is not None else our_before
                row_adj_pct = adj_pct
                adj_note_row = adj_note

            # Recalc gap/profit to match exported Our Price
            gap = (our_export - market_net) if (our_export is not None and market_net is not None) else None
            profit_now = None
            if our_export is not None and float(our_export or 0) > 0 and cost is not None:
                profit_now = calc_profit(our_export, cost, fee.fee_pct, fee.fixed_fee, pack_cost, ship_subsidy)

            # Recommend text (display) based on exported Our Price for readability
            rec_keys_disp: list[str] = []

            def add_rec_disp(key: str):
                if key in REC_DEFS and key not in rec_keys_disp:
                    rec_keys_disp.append(key)

            if (our_export is not None and our_export > 0) and (market_net is not None and market_net > 0):
                if abs(our_export - market_net) < 0.01:
                    add_rec_disp("equal_price")
                elif market_net < our_export:
                    add_rec_disp("market_cheaper")
            if no_market_flag:
                add_rec_disp("no_market")
            if missing_internal_flag:
                add_rec_disp("missing_internal")
            if profit_match is not None and market_net is not None and market_net > 0:
                profit_match_pct = (profit_match / market_net) * 100.0
                if profit_match_pct >= 0:
                    add_rec_disp("follow_ok")
                elif profit_match_pct > -6:
                    add_rec_disp("loss_0_5")
                elif profit_match_pct > -10:
                    add_rec_disp("loss_6_10")
                else:
                    add_rec_disp("loss_heavy")

            age_tags: list[str] = []
            if no_sales:
                age_tags.append("ไม่มียอดขาย")
            if aging_label:
                age_tags.append(aging_label)

            rec_labels_disp = [REC_DEFS[k]["label"] for k in rec_keys_disp]
            rec_text = ", ".join(rec_labels_disp) if rec_labels_disp else "-"
            if age_tags:
                rec_text = (rec_text + " | " + ", ".join(age_tags)) if rec_text and rec_text != "-" else ", ".join(age_tags)

            export_rows.append({
                "sku": sku,
                "brand": brand,
                "name": name,
                "stock_internal": stock_internal,
                "stock": stock_total,
                "stock_adj": stock_adj,
                "monthly_sales": monthly_sales,
                "cost": cost,
                "our_price": our_export,
                "market_best": market_net,
                "voucher": market_voucher,
                "brand_control": brand_control,
                "gap": gap,
                "profit_our": profit_now,
                "profit_match": profit_match,
                "recommend": rec_text,
                "shop": market_shop,
                "mall": "MALL" if market_is_mall else "",
                "url": market_url,
                "owner": owner,
                "updated": to_thai_be(market_updated) if market_updated else "",
                "adj_note": adj_note_row,
                "our_before": our_before,
                "our_after": our_export,
                "adj_pct": row_adj_pct,
                # for KPI filtering
                "brand_control": brand_control,
                "need_market": need_market,
                "rec_keys": rec_keys,
                "aging_bucket": aging_bucket,
                "no_sales": no_sales,
            })

        rows_master = _apply_kpi_filters(export_rows, [master_sel]) if master_sel else export_rows
        export_rows = _apply_kpi_filters(rows_master, kpi_selected)

        if preview:
            total_rows = len(export_rows)
            adjusted_rows = sum(
                1
                for r in export_rows
                if _as_int0(r.get("stock_adj")) != _as_int0(r.get("stock"))
            )
            return jsonify({"total": total_rows, "adjusted": adjusted_rows})

        final_rows = [{LABEL[k]: r.get(k) for k in selected_keys} for r in export_rows]
        df = pd.DataFrame(final_rows, columns=[LABEL[k] for k in selected_keys])
        df = sanitize_excel_df(df)

        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="StockAdj")
        bio.seek(0)

        filename = f"stock_adj_{adj_platform}_{now_thai().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.post("/api/price/platform_import/inspect")
    @login_required
    def api_platform_import_inspect():
        cu = current_user()
        if not cu:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

        _platform_import_gc()

        f = request.files.get("file")
        if not f:
            return jsonify({"ok": False, "error": "missing file"}), 400

        platform_hint = (request.form.get("platform") or "").strip().lower()
        sheet_name = (request.form.get("sheet_name") or "").strip()
        header_row_raw = (request.form.get("header_row") or "").strip()

        try:
            f.stream.seek(0)
        except Exception:
            pass
        wb = load_workbook(f, data_only=False)
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

        header_row = None
        if header_row_raw.isdigit():
            header_row = int(header_row_raw)
        if not header_row or header_row <= 0:
            if platform_hint == "shopee":
                header_row = (
                    _find_header_row_by_keywords(
                        ws,
                        keywords=["parent sku", "เลข sku", "ราคา", "คลัง", "stock"],
                        scan_rows=80,
                        min_hits=2,
                    )
                    or _detect_header_row(ws)
                )
            else:
                header_row = _detect_header_row(ws)

        cols = _extract_columns(ws, header_row)

        def _guess(keywords: list[str]):
            for col in cols:
                n = (col.get("name") or "").lower()
                if any(k in n for k in keywords):
                    return col.get("idx")
            return None

        sugg = {
            "sku_col_idx": _guess(["parent sku", "sellersku", "seller sku", "shop sku", "sku ของผู้ขาย", "sku"]),
            "sku_col_idx_alt": _guess(["เลข sku", "sku id", "item sku", "model sku"]),
            # Lazada commonly has both "Special Price" and "Price"
            "price_col_idx": _guess(["special price", "specialprice", "ราคาพิเศษ"]),
            "price_col_idx2": _guess(["price", "ราคาปกติ", "ราคาขาย", "ราคา"]),
            "stock_col_idx": _guess(["คลัง", "stock", "ปริมาณ", "จำนวน", "qty", "quantity"]),
        }

        return jsonify({
            "ok": True,
            "sheet_names": wb.sheetnames,
            "sheet_name": ws.title,
            "header_row": header_row,
            "columns": cols,
            "suggest": sugg,
        })


    # -------------------------------------------------------------------
    # Price Dashboard: Platform Import mapping presets (per-user/per-platform)
    # Stored in price.db (PriceUserPreference)
    # key format: plat_imp_map.<platform>
    # -------------------------------------------------------------------

    def _plat_key(p: str) -> str:
        p = (p or "").strip().lower()
        p = re.sub(r"[^a-z0-9_]+", "_", p)
        return p[:30] or "platform"


    @app.get("/api/price/platform_import/mapping")
    @login_required
    def api_platform_import_mapping_get():
        cu = current_user()
        if not cu:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

        plat = _plat_key(request.args.get("platform") or "")
        key = f"plat_imp_map.{plat}"

        raw = get_user_pref(int(cu.id), key, default="")
        if not raw:
            return jsonify({"ok": True, "mapping": None})

        try:
            return jsonify({"ok": True, "mapping": json.loads(raw)})
        except Exception:
            return jsonify({"ok": True, "mapping": None})


    @app.post("/api/price/platform_import/mapping/save")
    @login_required
    def api_platform_import_mapping_save():
        cu = current_user()
        if not cu:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

        data = request.get_json(silent=True) or {}
        plat = _plat_key(data.get("platform") or "")
        key = f"plat_imp_map.{plat}"

        mapping = data.get("mapping") or {}
        payload = {
            "v": 1,
            "sku": mapping.get("sku"),
            "sku2": mapping.get("sku2"),
            "p1": mapping.get("p1"),
            "p2": mapping.get("p2"),
            "st": mapping.get("st"),
        }
        raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        if len(raw) > 240:
            return jsonify({"ok": False, "error": "mapping too long"}), 400

        set_user_pref(int(cu.id), key, raw)
        return jsonify({"ok": True})


    @app.post("/api/price/platform_import/mapping/clear")
    @login_required
    def api_platform_import_mapping_clear():
        cu = current_user()
        if not cu:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

        data = request.get_json(silent=True) or {}
        plat = _plat_key(data.get("platform") or "")
        key = f"plat_imp_map.{plat}"

        set_user_pref(int(cu.id), key, None)
        return jsonify({"ok": True})

    @app.post("/api/price/platform_import/apply")
    @login_required
    def api_platform_import_apply():
        cu = current_user()
        if not cu:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

        _platform_import_gc()

        f = request.files.get("file")
        if not f:
            return jsonify({"ok": False, "error": "missing file"}), 400

        # ===== 1) Mapping =====
        sheet_name = (request.form.get("sheet_name") or "").strip()
        try:
            header_row = int((request.form.get("header_row") or "").strip() or 1)
        except Exception:
            header_row = 1

        def _as_int(v, default: int = 0) -> int:
            try:
                return int(str(v).strip() or default)
            except Exception:
                return int(default)

        sku_col_idx = _as_int(request.form.get("sku_col_idx"), 0)
        sku_col_idx_alt = _as_int(request.form.get("sku_col_idx_alt"), 0)
        price_col_idx = _as_int(request.form.get("price_col_idx"), 0)  # 0 = no update
        price_col_idx2 = _as_int(request.form.get("price_col_idx2"), 0)  # 0 = no update
        stock_col_idx = _as_int(request.form.get("stock_col_idx"), 0)  # 0 = no update

        if sku_col_idx <= 0:
            return jsonify({"ok": False, "error": "SKU column is required"}), 400

        # ===== 2) Settings (names aligned with Export Price & Stock Adj.) =====
        platform = normalize_platform(request.form.get("platform") or "Shopee")
        adj_pct = _clamp_float(request.form.get("adj_pct"), -50.0, 50.0, default=0.0)

        stock_rules = request.form.getlist("stock_rule")
        stock_rules = [str(x).strip() for x in stock_rules if str(x).strip()]
        allowed_rules = {"1", "2", "3", "4", "5"}
        stock_rules_set = {r for r in stock_rules if r in allowed_rules}

        # Enforce constraints: rule 3 cannot be used with 1-2
        if "3" in stock_rules_set:
            stock_rules_set.discard("1")
            stock_rules_set.discard("2")
        elif "1" in stock_rules_set or "2" in stock_rules_set:
            stock_rules_set.discard("3")

        stock_divisor = _as_int(request.form.get("stock_divisor"), 3)
        if "3" in stock_rules_set:
            stock_divisor = max(1, int(stock_divisor or 0))

        stock_divisor_brand = _as_int(request.form.get("stock_divisor_brand"), 2)
        if "5" in stock_rules_set:
            stock_divisor_brand = max(1, int(stock_divisor_brand or 0))

        stock_brands_4_raw = (request.form.get("stock_brands_4") or "").strip()
        stock_brands_5_raw = (request.form.get("stock_brands_5") or "").strip()

        def _parse_brands(s: str) -> set[str]:
            if not s:
                return set()
            return {b.strip().lower() for b in re.split(r"[,\n]+", s) if b and b.strip()}

        brands_4_set = _parse_brands(stock_brands_4_raw)
        brands_5_set = _parse_brands(stock_brands_5_raw)

        def _as_bool(name: str, default: bool = False) -> bool:
            v = request.form.get(name)
            if v is None:
                return default
            return str(v).strip().lower() in {"1", "true", "on", "yes"}

        brand_floor12_to0 = _as_bool("stock_divide_brand_floor12_to0", default=True)
        brand_min1_when_zero = _as_bool("stock_divide_brand_min1_when_zero", default=False)
        use_sell1 = _as_bool("use_sell1", default=False)
        skip_adj_when_brand_control = _as_bool("skip_adj_when_brand_control", default=False)

        # NEW: Skip price adjustment when Profit@Our% >= X (use Our Price only)
        # Note: Applies only when adj_pct > 0 (to avoid lowering price when negative).
        skip_adj_when_profit_our = _as_bool("skip_adj_when_profit_our", default=False)
        profit_our_min_pct = _clamp_float(request.form.get("profit_our_min_pct"), 0.0, 50.0, default=10.0)

        # Fee settings (used for Profit@Our% calculation)
        fee = PlatformFeeSetting.query.get(platform)
        if not fee:
            fee = PlatformFeeSetting(platform=platform, label=platform, fee_pct=0.0, fixed_fee=0.0, is_active=True, sort_order=0)
            db.session.add(fee)
            db.session.commit()

        def calc_profit(price, cost, fee_pct, fixed_fee, pack_cost, ship_subsidy):
            price = float(price or 0.0)
            cost = float(cost or 0.0)
            fee_amt = (price * (float(fee_pct or 0.0) / 100.0)) + float(fixed_fee or 0.0)
            return price - cost - fee_amt - float(pack_cost or 0.0) - float(ship_subsidy or 0.0)

        export_setting = None
        if use_sell1:
            try:
                export_setting = PriceExportSetting.query.get(1)
            except Exception:
                export_setting = None
            if not export_setting:
                export_setting = PriceExportSetting(
                    id=1,
                    step_pct=5.0,
                    min_profit_pct=5.0,
                    loss_aging3_pct=5.0,
                    loss_aging6_pct=10.0,
                    loss_aging12_pct=20.0,
                )

        # BrandControl map: sku(lower) -> price_control float
        bc_val_map: dict[str, float] = {}
        try:
            for bc in (BrandControl.query.all() or []):
                sku0 = (getattr(bc, "sku", "") or "").strip().lower()
                if not sku0:
                    continue
                try:
                    v = float(getattr(bc, "price_control", 0) or 0)
                except Exception:
                    v = 0.0
                bc_val_map[sku0] = v
        except Exception:
            bc_val_map = {}

        def _as_int0(x) -> int:
            try:
                return int(x or 0)
            except Exception:
                return 0

        def compute_stock_adj(stock_total, stock_internal, brand: str) -> int:
            s = _as_int0(stock_total)
            i = _as_int0(stock_internal)
            b = (brand or "").strip().lower()

            if s <= 0 and i <= 0:
                return 0

            if not stock_rules_set:
                return s

            if "5" in stock_rules_set and b and (b in brands_5_set):
                if i <= 0:
                    return 0
                v = i // stock_divisor_brand
                from_min1 = False
                if brand_min1_when_zero and i > 0 and v == 0:
                    v = 1
                    from_min1 = True
                if brand_floor12_to0 and (not from_min1) and (1 <= v <= 2):
                    v = 0
                return int(v)

            if "4" in stock_rules_set and b and (b in brands_4_set):
                return 0

            if "3" in stock_rules_set:
                v = i // stock_divisor
                if 1 <= v <= 2:
                    v = 0
                return int(v)

            v = s
            if "1" in stock_rules_set and (1 <= s <= 3) and (i == 0):
                v = 0
            if "2" in stock_rules_set and (1 <= s <= 5) and (1 <= i <= 5):
                v = 1
            return int(v)

        # ===== 3) Build sku -> {price, stock} map from price.db =====
        sku_adj_map: dict[str, dict] = {}
        for pr in (SkuPricing.query.all() or []):
            sku = (getattr(pr, "sku", "") or "").strip()
            if not sku:
                continue

            brand = (getattr(pr, "brand", "") or "").strip()
            stock_internal = int(getattr(pr, "stock_internal_qty", 0) or 0)
            stock_total = int(getattr(pr, "stock_qty", 0) or 0)
            monthly_sales = int(getattr(pr, "monthly_sales_qty", 0) or 0)

            our_price = float(pr.our_price) if getattr(pr, "our_price", None) is not None else None
            cost = float(pr.cost) if getattr(pr, "cost", None) is not None else None
            pack_cost = float(getattr(pr, "pack_cost", 0) or 0)
            ship_subsidy = float(getattr(pr, "ship_subsidy", 0) or 0)

            effective_cost = None
            if use_sell1 and cost is not None:
                try:
                    effective_cost = float(cost) + float(pack_cost or 0.0) + float(ship_subsidy or 0.0)
                except Exception:
                    effective_cost = cost

            stock_adj = compute_stock_adj(stock_total, stock_internal, brand)

            aging_bucket = None
            if stock_internal is not None:
                if (monthly_sales * 12 - stock_internal) < 0:
                    aging_bucket = "aging12"
                elif (monthly_sales * 6 - stock_internal) < 0:
                    aging_bucket = "aging6"
                elif (monthly_sales * 3 - stock_internal) < 0:
                    aging_bucket = "aging3"

            no_sales = (monthly_sales == 0) and (stock_internal is not None and stock_internal >= 1)
            sell1 = None
            if use_sell1 and our_price is not None and float(our_price or 0) > 0:
                aging_for_loss = "aging12" if no_sales else aging_bucket
                sell = build_sell_prices(
                    our_price=our_price,
                    cost=effective_cost,
                    step_pct=export_setting.step_pct,
                    min_profit_pct=export_setting.min_profit_pct,
                    loss_aging3_pct=export_setting.loss_aging3_pct,
                    loss_aging6_pct=export_setting.loss_aging6_pct,
                    loss_aging12_pct=export_setting.loss_aging12_pct,
                    aging_bucket=aging_for_loss,
                )
                sell1 = (sell[0] if sell else None)

            bc_val = 0.0
            try:
                bc_val = float(bc_val_map.get(sku.lower(), 0) or 0)
            except Exception:
                bc_val = 0.0
            has_bc = (bc_val > 0)

            # Profit@Our% (computed from original Our Price)
            profit_base_pct = None
            if (our_price is not None and float(our_price or 0) > 0) and (cost is not None):
                try:
                    p0 = calc_profit(our_price, cost, fee.fee_pct, fee.fixed_fee, pack_cost, ship_subsidy)
                    profit_base_pct = (p0 / float(our_price)) * 100.0 if float(our_price) else None
                except Exception:
                    profit_base_pct = None

            skip_profit = (
                bool(skip_adj_when_profit_our)
                and (float(adj_pct or 0.0) > 0.0)
                and (profit_base_pct is not None)
                and (float(profit_base_pct) >= float(profit_our_min_pct))
            )

            # Priority: BrandControl skip first, then Profit@Our skip
            if skip_adj_when_brand_control and has_bc:
                base_price = our_price
                new_price = our_price
            elif skip_profit:
                base_price = our_price
                new_price = our_price
            else:
                base_price = sell1 if (use_sell1 and sell1 is not None) else our_price
                new_price = adjust_our_price_export(base_price, adj_pct) if base_price is not None else None

            sku_adj_map[sku.lower()] = {"price": new_price, "stock": stock_adj}

        # ===== 4) Load workbook + overwrite mapped columns only =====
        try:
            f.stream.seek(0)
        except Exception:
            pass
        wb = load_workbook(f, data_only=False)
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

        header_row = max(1, int(header_row or 1))
        start_row = _first_data_row(ws, header_row, sku_col_idx, sku_col_idx_alt)
        if not start_row:
            start_row = header_row + 1

        total_rows = 0
        matched = 0
        updated_price = 0
        updated_stock = 0

        max_row = getattr(ws, "max_row", 0) or 0
        for r in range(start_row, max_row + 1):
            sku_val = _norm_sku(ws.cell(row=r, column=sku_col_idx).value)
            if (not sku_val) and sku_col_idx_alt > 0:
                sku_val = _norm_sku(ws.cell(row=r, column=sku_col_idx_alt).value)
            if not sku_val:
                continue

            # skip template note rows / repeated header rows
            if not _is_real_sku_value(sku_val):
                continue

            total_rows += 1

            adj = sku_adj_map.get(sku_val.strip().lower())
            has_adj = bool(adj)
            if has_adj:
                matched += 1

            # ===== PRICE: update only when there is a valid new price (>0) =====
            new_price = (adj or {}).get("price") if has_adj else None
            should_update_price = False
            try:
                should_update_price = (new_price is not None) and (float(new_price or 0) > 0)
            except Exception:
                should_update_price = False

            if should_update_price:
                if price_col_idx > 0:
                    ws.cell(row=r, column=price_col_idx).value = float(new_price)
                    updated_price += 1
                if price_col_idx2 > 0:
                    ws.cell(row=r, column=price_col_idx2).value = float(new_price)
                    updated_price += 1

            # ===== STOCK: always write Stock Adj; if SKU not found -> set 0 =====
            if stock_col_idx > 0:
                new_stock = int((adj or {}).get("stock")) if has_adj else 0
                ws.cell(row=r, column=stock_col_idx).value = new_stock
                updated_stock += 1

        # ===== 5) Save temp + return download key =====
        out_key = uuid.uuid4().hex
        out_path = os.path.join(tempfile.gettempdir(), f"platform_adjusted_{out_key}.xlsx")
        wb.save(out_path)

        exported_at = now_thai()
        try:
            if getattr(exported_at, "tzinfo", None) is not None:
                exported_at = exported_at.replace(tzinfo=None)
        except Exception:
            pass
        ts_str = exported_at.strftime("%Y%m%d_%H%M")

        plat_safe = normalize_platform_key(platform) or "platform"

        stats = {
            "header_row": header_row,
            "start_row": start_row,
            "total_rows": total_rows,
            "matched": matched,
            "updated_price": updated_price,
            "updated_stock": updated_stock,
            "match_pct": (matched * 100.0 / total_rows) if total_rows else 0.0,
        }

        PLATFORM_IMPORT_CACHE[out_key] = {
            "user_id": cu.id,
            "path": out_path,
            "ts": time.time(),
            "stats": stats,
            "platform": plat_safe,
            "ts_str": ts_str,
        }

        return jsonify({
            "ok": True,
            "download_key": out_key,
            "download_url": f"/api/price/platform_import/download?key={out_key}",
            "stats": stats,
        })

    @app.get("/api/price/platform_import/download")
    @login_required
    def api_platform_import_download():
        cu = current_user()
        if not cu:
            return "unauthorized", 401

        _platform_import_gc()

        key = (request.args.get("key") or "").strip()
        info = PLATFORM_IMPORT_CACHE.get(key)
        if not info or info.get("user_id") != cu.id:
            return "not found", 404

        path = info.get("path")
        if not path or not os.path.exists(path):
            return "file missing", 404

        plat = (info.get("platform") or "platform").strip() or "platform"
        ts_str = (info.get("ts_str") or now_thai().strftime("%Y%m%d_%H%M")).strip()
        download_name = f"{plat}_{ts_str}.xlsx"

        return send_file(
            path,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.post("/price/dashboard/auto_price_apply")
    @login_required
    def price_dashboard_auto_price_apply():
        cu = current_user()
        if not cu:
            return redirect(url_for("login", next=request.path))

        platform_raw = (request.form.get("platform") or "").strip()
        platform = normalize_platform(platform_raw) or platform_raw

        q_raw = (request.form.get("q") or "").strip()
        q = q_raw.lower()

        owner_sel = (request.form.get("owner") or "").strip()
        limit_sel = (request.form.get("limit") or "200").strip().lower()
        stale_days_raw = (request.form.get("stale_days") or "").strip()
        kpi_sel = (request.form.get("kpi") or "").strip()
        master_raw = (request.form.get("master") or "").strip()

        rules_raw = (request.form.get("rules") or "").strip()
        selected_rules = {r.strip() for r in rules_raw.split(",") if r.strip()}
        allowed_rules = {f"r{i}" for i in range(1, 13)}
        selected_rules = selected_rules & allowed_rules

        # --- Load r10/r11 configs (priority: form -> saved pref -> default) ---
        r10_cfg_src = None
        raw_r10 = (request.form.get("r10_cfg") or "").strip()
        if raw_r10:
            try:
                r10_cfg_src = json.loads(raw_r10)
            except Exception:
                r10_cfg_src = None
        if not isinstance(r10_cfg_src, dict):
            r10_cfg_src = get_user_pref(int(cu.id), "pm_auto_r10_cfg", default="")
        r10_cfg = _norm_r10_cfg(_load_cfg_dict(r10_cfg_src, _R10_DEFAULT_CFG))
        r10_min_loss = float(r10_cfg.get("min_loss_pct") or _R10_DEFAULT_CFG["min_loss_pct"])

        r11_cfg_src = None
        raw_r11 = (request.form.get("r11_cfg") or "").strip()
        if raw_r11:
            try:
                r11_cfg_src = json.loads(raw_r11)
            except Exception:
                r11_cfg_src = None
        if not isinstance(r11_cfg_src, dict):
            r11_cfg_src = get_user_pref(int(cu.id), "pm_auto_r11_cfg", default="")
        r11_cfg = _norm_r11_cfg(_load_cfg_dict(r11_cfg_src, _R11_DEFAULT_CFG))
        r11_min_loss = float(r11_cfg.get("min_loss_pct") or _R11_DEFAULT_CFG["min_loss_pct"])
        r11_max_loss = float(r11_cfg.get("max_loss_pct") or _R11_DEFAULT_CFG["max_loss_pct"])

        if not selected_rules:
            flash("กรุณาเลือกอย่างน้อย 1 เงื่อนไขสำหรับปรับราคา", "warning")
            return redirect(url_for("price_dashboard", platform=platform, q=q_raw, owner=owner_sel, limit=limit_sel, kpi=kpi_sel, master=master_raw))

        platform_rows = (
            PlatformFeeSetting.query
            .filter(PlatformFeeSetting.is_active == True)
            .all()
        )
        allowed_platforms = {p.platform for p in platform_rows}
        if platform not in allowed_platforms:
            flash("Platform ไม่ถูกต้องหรือถูกปิดการใช้งาน", "warning")
            return redirect(url_for("price_dashboard"))

        allowed_limits = {"100", "200", "300", "500", "1000", "all"}
        if limit_sel not in allowed_limits:
            limit_sel = "200"
        limit_n = None if limit_sel == "all" else int(limit_sel)

        try:
            stale_days = int(stale_days_raw or 14)
        except Exception:
            stale_days = 14
        stale_days = max(1, min(stale_days, 365))

        now0 = now_thai()
        try:
            if getattr(now0, "tzinfo", None) is not None:
                now0 = now0.replace(tzinfo=None)
        except Exception:
            pass
        cutoff = now0 - timedelta(days=stale_days)

        fee = PlatformFeeSetting.query.get(platform)
        if not fee:
            fee = PlatformFeeSetting(platform=platform, label=platform, fee_pct=0.0, fixed_fee=0.0, is_active=True, sort_order=0)
            db.session.add(fee)
            db.session.commit()

        def calc_profit(price, cost, fee_pct, fixed_fee, pack_cost, ship_subsidy):
            price = float(price or 0.0)
            cost = float(cost or 0.0)
            fee_amt = (price * (float(fee_pct or 0.0) / 100.0)) + float(fixed_fee or 0.0)
            return price - cost - fee_amt - float(pack_cost or 0.0) - float(ship_subsidy or 0.0)

        def is_close_price(a, b, tol: float = 0.01) -> bool:
            if a is None or b is None:
                return False
            try:
                return abs(float(a) - float(b)) <= float(tol)
            except Exception:
                return False

        pricing_map = {p.sku: p for p in SkuPricing.query.all()}
        bc_map = {b.sku: b for b in BrandControl.query.all()}

        owner_rows = BrandOwnerSetting.query.all()
        brand_owner_map = {str(r.brand or "").strip(): str(r.owner or "").strip() for r in owner_rows if r}

        items = (
            MarketItem.query
            .filter(MarketItem.platform == platform, MarketItem.is_active == True)
            .all()
        )
        from datetime import datetime

        def _ts_market(it):
            t = getattr(it, "last_updated", None)
            if not t:
                return datetime.min
            try:
                if getattr(t, "tzinfo", None) is not None:
                    t = t.replace(tzinfo=None)
            except Exception:
                pass
            return t

        latest_by_sku = {}
        for it in items:
            cur = latest_by_sku.get(it.sku)
            if cur is None:
                latest_by_sku[it.sku] = it
                continue
            if _ts_market(it) > _ts_market(cur):
                latest_by_sku[it.sku] = it
                continue
            if _ts_market(it) == _ts_market(cur) and (it.id or 0) > (cur.id or 0):
                latest_by_sku[it.sku] = it

        sku_set = set(pricing_map.keys()) | set(latest_by_sku.keys()) | set(bc_map.keys())

        if q:
            filtered = set()
            for sku in sku_set:
                pr = pricing_map.get(sku)
                mk = latest_by_sku.get(sku)
                hay = [
                    (sku or "").lower(),
                    (getattr(pr, "spec_text", "") or "").lower(),
                    (getattr(pr, "brand", "") or "").lower(),
                    (getattr(pr, "name", "") or "").lower(),
                    (getattr(mk, "shop_name", "") or "").lower(),
                ]
                if any(q in s for s in hay):
                    filtered.add(sku)
            sku_set = filtered

        if owner_sel:
            filtered = set()
            for sku in sku_set:
                pr = pricing_map.get(sku)
                brand = (getattr(pr, "brand", "") or "").strip() if pr else ""
                owner = brand_owner_map.get(brand, "") if brand else ""
                if owner == owner_sel:
                    filtered.add(sku)
            sku_set = filtered

        sku_list = sorted(sku_set)
        if limit_n is not None:
            sku_list = sku_list[:limit_n]

        rows = []
        for sku in sku_list:
            pr = pricing_map.get(sku)
            mk = latest_by_sku.get(sku)
            bc = bc_map.get(sku)

            cost = float(pr.cost) if (pr and pr.cost is not None) else None
            our_price = float(pr.our_price) if (pr and pr.our_price is not None) else None

            pack_cost = float(pr.pack_cost) if (pr and pr.pack_cost is not None) else 0.0
            ship_subsidy = float(pr.ship_subsidy) if (pr and pr.ship_subsidy is not None) else 0.0

            stock_internal = int(getattr(pr, "stock_internal_qty", 0) or 0) if (pr and getattr(pr, "stock_internal_qty", None) is not None) else None
            stock_qty = int(getattr(pr, "stock_qty", 0) or 0) if (pr and getattr(pr, "stock_qty", None) is not None) else None
            monthly_sales = int(getattr(pr, "monthly_sales_qty", 0) or 0) if pr else 0

            market_net = float(mk.latest_net_price) if (mk and mk.latest_net_price is not None) else None
            brand_control = float(bc.price_control) if (bc and bc.price_control is not None) else None
            market_updated = mk.last_updated if mk else None

            # Aging logic (Exclusive): choose the highest bucket only (1ปี > 6เดือน > 3เดือน)
            aging_bucket = None
            if stock_internal is not None:
                if (monthly_sales * 12 - stock_internal) < 0:
                    aging_bucket = "aging12"
                elif (monthly_sales * 6 - stock_internal) < 0:
                    aging_bucket = "aging6"
                elif (monthly_sales * 3 - stock_internal) < 0:
                    aging_bucket = "aging3"

            no_sales = (monthly_sales == 0) and (stock_internal is not None and stock_internal >= 1)

            mu = market_updated
            try:
                if mu is not None and getattr(mu, "tzinfo", None) is not None:
                    mu = mu.replace(tzinfo=None)
            except Exception:
                pass
            has_stock = (int(stock_internal or 0) > 0) or (int(stock_qty or 0) > 0)
            is_stale = (mu is None) or (mu < cutoff)
            need_market = has_stock and is_stale

            rec_keys = []

            no_market_flag = (mk is None) or (market_net is None) or (market_net <= 0)
            missing_internal_flag = (pr is None) or (cost is None) or (our_price is None) or (our_price <= 0)
            if no_market_flag:
                rec_keys.append("no_market")
            if missing_internal_flag:
                rec_keys.append("missing_internal")
            if (our_price is not None and our_price > 0) and (market_net is not None and market_net > 0):
                if abs(our_price - market_net) < 0.01:
                    rec_keys.append("equal_price")
                elif market_net < our_price:
                    rec_keys.append("market_cheaper")

            profit_match = None
            profit_match_pct = None
            if (market_net is not None and market_net > 0) and (cost is not None):
                profit_match = calc_profit(market_net, cost, fee.fee_pct, fee.fixed_fee, pack_cost, ship_subsidy)
                profit_match_pct = (profit_match / market_net) * 100.0
                if profit_match_pct >= 0:
                    rec_keys.append("follow_ok")
                elif profit_match_pct > -6:
                    rec_keys.append("loss_0_5")
                elif profit_match_pct > -10:
                    rec_keys.append("loss_6_10")
                else:
                    rec_keys.append("loss_heavy")

            base = (
                (market_net is not None and market_net > 0)
                and (our_price is not None and our_price > 0)
                and (market_net < our_price)
            )

            base_up = (
                (market_net is not None and market_net > 0)
                and (our_price is not None and our_price > 0)
                and (market_net > our_price)
            )

            profit_ok = (profit_match is not None and profit_match >= 0)
            profit_neg = (profit_match is not None and profit_match < 0)

            cost_zero = (cost is not None and abs(cost) < 1e-9)
            cost_pos = (cost is not None and cost > 0)

            stock_i = int(stock_internal or 0)
            stock_t = int(stock_qty or 0)
            stock_i_pos = stock_i > 0
            stock_ok = (stock_i > 0) or (stock_t > 0)

            bucket = aging_bucket or ""
            aging3p = bucket in ("aging3", "aging6", "aging12")
            aging6p = bucket in ("aging6", "aging12")
            aging12p = bucket in ("aging12",)

            not_aging = (aging_bucket is None) and (not no_sales)

            has_loss_0_5 = "loss_0_5" in (rec_keys or [])
            has_loss_6_10 = "loss_6_10" in (rec_keys or [])

            r1 = base and profit_ok and cost_pos
            r2 = base and profit_ok and cost_zero and stock_i_pos
            r3 = base and profit_neg and (aging3p or no_sales)
            r4 = base and profit_neg and (aging6p or no_sales)
            r5 = base and profit_neg and (aging12p or no_sales)
            r6 = base and profit_neg and has_loss_0_5
            r7 = base and profit_neg and has_loss_6_10 and stock_i_pos

            r8 = (
                base
                and (brand_control is not None and float(brand_control or 0) > 0)
                and is_close_price(market_net, brand_control, tol=0.01)
            )

            r9 = no_market_flag and cost_pos and (float(cost) >= 5)

            r10 = (
                base
                and not_aging
                and cost_pos
                and stock_ok
                and (profit_match_pct is not None and float(profit_match_pct) < -float(r10_min_loss))
            )

            r11 = (
                base
                and not_aging
                and cost_pos
                and stock_ok
                and (profit_match_pct is not None and float(profit_match_pct) < -float(r11_min_loss))
            )

            r12 = base_up and profit_ok and cost_pos and stock_ok

            auto_rules = []
            if r1:
                auto_rules.append("r1")
            if r2:
                auto_rules.append("r2")
            if r3:
                auto_rules.append("r3")
            if r4:
                auto_rules.append("r4")
            if r5:
                auto_rules.append("r5")
            if r6:
                auto_rules.append("r6")
            if r7:
                auto_rules.append("r7")
            if r8:
                auto_rules.append("r8")
            if r9:
                auto_rules.append("r9")
            if r10:
                auto_rules.append("r10")
            if r11:
                auto_rules.append("r11")
            if r12:
                auto_rules.append("r12")

            rows.append({
                "sku": sku,
                "market_net": market_net,
                "brand_control": brand_control,
                "cost": cost,
                "profit_match_pct": profit_match_pct,
                "stock_internal": stock_internal,
                "stock_qty": stock_qty,
                "rec_keys": rec_keys,
                "aging_bucket": aging_bucket,
                "no_sales": no_sales,
                "need_market": need_market,
                "auto_rules": auto_rules,
            })

        allowed_kpis = {
            "market_cheaper",
            "equal_price",
            "follow_ok",
            "loss_0_5",
            "loss_6_10",
            "loss_heavy",
            "no_market",
            "missing_internal",
            "no_sales",
            "aging3",
            "aging6",
            "aging12",
            "brand_control",
            "need_market",
        }
        kpi_selected = _parse_kpi_multi(kpi_sel, allowed_kpis)
        master_sel, kpi_selected = _resolve_master(master_raw, kpi_selected, allowed_kpis)
        rows_master = _apply_kpi_filters(rows, [master_sel]) if master_sel else rows
        display_rows = _apply_kpi_filters(rows_master, kpi_selected)

        updated = 0
        skipped = 0

        # breakdown (debuggable UX)
        matched = 0
        skip_no_target = 0
        skip_no_pr = 0
        skip_same = 0

        # --- Load r9 config (priority: form -> saved pref -> default) ---
        r9_cfg = None
        raw_form = (request.form.get("r9_cfg") or "").strip()
        if raw_form:
            try:
                r9_cfg = json.loads(raw_form)
            except Exception:
                r9_cfg = None

        if not isinstance(r9_cfg, list):
            raw_pref = get_user_pref(int(cu.id), "pm_auto_r9_cfg", default="")
            if raw_pref:
                try:
                    r9_cfg = json.loads(raw_pref)
                except Exception:
                    r9_cfg = None

        r9_tiers = _r9_cfg_to_tiers(r9_cfg if isinstance(r9_cfg, list) else _r9_default_cfg())

        for r in display_rows:
            rules = r.get("auto_rules") or []
            if not any(x in selected_rules for x in rules):
                continue

            matched += 1

            sku = r.get("sku")
            market_net = r.get("market_net")
            brand_control = r.get("brand_control")
            cost = r.get("cost")
            profit_match_pct = r.get("profit_match_pct")
            loss_abs = _loss_abs_pct(profit_match_pct)

            # Priority: r11 > r9 > r10 > r8 > market_best
            if ("r11" in selected_rules) and ("r11" in rules):
                if (market_net is None) or (float(market_net or 0) <= 0) or (loss_abs is None):
                    target_price = None
                elif float(loss_abs) <= float(r11_max_loss):
                    target_price = market_net
                else:
                    target_price = _auto_price_from_cost_plus_pct(cost, float(loss_abs) / 100.0)

            elif ("r9" in selected_rules) and ("r9" in rules):
                target_price = _auto_price_from_cost(cost, tiers=r9_tiers)

            elif ("r10" in selected_rules) and ("r10" in rules):
                target_price = _auto_price_from_cost_plus_pct(cost, float(loss_abs or 0.0) / 100.0) if (loss_abs is not None) else None
            elif ("r8" in selected_rules) and ("r8" in rules):
                target_price = brand_control
            else:
                target_price = market_net

            if (not sku) or (target_price is None):
                skipped += 1
                skip_no_target += 1
                continue

            try:
                new_price = float(target_price)
            except Exception:
                skipped += 1
                skip_no_target += 1
                continue

            if new_price <= 0:
                skipped += 1
                skip_no_target += 1
                continue

            pr = pricing_map.get(sku) or SkuPricing.query.get(sku)
            if not pr:
                skipped += 1
                skip_no_pr += 1
                continue

            before = float(pr.our_price) if pr.our_price is not None else None

            # Treat tiny diffs as same (avoid float noise)
            if before is not None and is_close_price(before, new_price, tol=0.01):
                skipped += 1
                skip_same += 1
                continue

            pr.our_price = new_price
            updated += 1

        if updated > 0:
            db.session.commit()
            flash(
                f"ปรับราคา Auto สำเร็จ: {updated} SKU (เข้าเงื่อนไข {matched} | เท่าเดิม {skip_same} | ราคาใหม่ไม่พร้อม {skip_no_target} | ไม่พบ SKU {skip_no_pr})",
                "success",
            )
        else:
            db.session.rollback()
            flash(
                f"เข้าเงื่อนไข {matched} SKU แต่เปลี่ยนจริง 0 (เท่าเดิม {skip_same} | ราคาใหม่ไม่พร้อม {skip_no_target} | ไม่พบ SKU {skip_no_pr})",
                "warning",
            )

        return redirect(url_for("price_dashboard", platform=platform, q=q_raw, owner=owner_sel, limit=limit_sel, stale_days=str(stale_days), kpi=kpi_sel, master=master_sel))

    @app.route("/price/settings", methods=["GET", "POST"])
    @login_required
    def price_settings():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("price_dashboard"))

        if request.method == "POST":
            rows = (
                PlatformFeeSetting.query
                .order_by(PlatformFeeSetting.sort_order.asc(), PlatformFeeSetting.platform.asc())
                .all()
            )

            for s in rows:
                key = s.platform

                fee_pct = request.form.get(f"fee_pct_{key}")
                fixed_fee = request.form.get(f"fixed_fee_{key}")
                label = request.form.get(f"label_{key}")
                sort_order = request.form.get(f"sort_{key}")
                is_active = request.form.get(f"active_{key}")

                s.label = (label or s.label or key).strip()
                try:
                    s.sort_order = int(sort_order) if sort_order not in (None, "") else (s.sort_order or 0)
                except Exception:
                    pass
                s.is_active = True if is_active in ("on", "true", "1", "yes") else False

                try:
                    s.fee_pct = float(fee_pct) if fee_pct not in (None, "") else 0.0
                except Exception:
                    s.fee_pct = 0.0

                try:
                    s.fixed_fee = float(fixed_fee) if fixed_fee not in (None, "") else 0.0
                except Exception:
                    s.fixed_fee = 0.0

            db.session.commit()
            flash("บันทึกค่าธรรมเนียมแพลตฟอร์มเรียบร้อย", "success")
            return redirect(url_for("price_settings"))

        platform_rows = (
            PlatformFeeSetting.query
            .order_by(PlatformFeeSetting.sort_order.asc(), PlatformFeeSetting.platform.asc())
            .all()
        )
        if not platform_rows:
            for p_key, p_label in [("Shopee", "Shopee"), ("Lazada", "Lazada"), ("TikTok", "TikTok")]:
                db.session.add(
                    PlatformFeeSetting(
                        platform=p_key,
                        label=p_label,
                        fee_pct=0.0,
                        fixed_fee=0.0,
                        is_active=True,
                        sort_order=0,
                    )
                )
            db.session.commit()
            platform_rows = (
                PlatformFeeSetting.query
                .order_by(PlatformFeeSetting.sort_order.asc(), PlatformFeeSetting.platform.asc())
                .all()
            )

        platforms = [(p.platform, (p.label or p.platform)) for p in platform_rows]
        settings = {p.platform: p for p in platform_rows}

        # brand list สำหรับผูกผู้ดูแล (ดึงจาก internal/brand control ที่มีอยู่ใน price.db)
        brand_vals = set()
        try:
            for (b,) in db.session.query(SkuPricing.brand).filter(SkuPricing.brand.isnot(None)).distinct().all():
                if b and str(b).strip():
                    brand_vals.add(str(b).strip())
        except Exception:
            pass
        try:
            for (b,) in db.session.query(BrandControl.brand).filter(BrandControl.brand.isnot(None)).distinct().all():
                if b and str(b).strip():
                    brand_vals.add(str(b).strip())
        except Exception:
            pass

        brands = sorted(brand_vals)

        owner_rows = BrandOwnerSetting.query.all()
        owner_map = {str(r.brand or "").strip(): str(r.owner or "").strip() for r in owner_rows if r}
        owners_list = sorted({v for v in owner_map.values() if v})

        export_setting = None
        try:
            export_setting = PriceExportSetting.query.get(1)
        except Exception:
            export_setting = None

        if not export_setting:
            export_setting = PriceExportSetting(
                id=1,
                step_pct=5.0,
                min_profit_pct=5.0,
                loss_aging3_pct=5.0,
                loss_aging6_pct=10.0,
                loss_aging12_pct=20.0,
            )

        # คำนวณมูลค่าขาย/เดือน (ประมาณการ) ต่อ Brand = SUM(Our Price * Monthly Sales)
        # ใช้ข้อมูลจาก price.db (SkuPricing) และเป็นค่าคำนวณ ไม่กระทบ schema
        brand_monthly_value_map: dict[str, float] = {}
        try:
            rows_val = (
                db.session.query(
                    SkuPricing.brand,
                    func.sum(
                        func.coalesce(SkuPricing.our_price, 0) * func.coalesce(SkuPricing.monthly_sales_qty, 0)
                    ),
                )
                .filter(SkuPricing.brand.isnot(None))
                .group_by(SkuPricing.brand)
                .all()
            )
            for b, v in rows_val:
                brand = (str(b or "").strip())
                if not brand:
                    continue
                brand_monthly_value_map[brand] = float(v or 0.0)
        except Exception:
            brand_monthly_value_map = {}

        # SKU count ต่อ Brand: นับเฉพาะ SKU ที่มี stock_internal_qty > 0 หรือ stock_qty > 0
        brand_sku_count_map: dict[str, int] = {}
        total_sku_instock = 0
        try:
            rows_sku = (
                db.session.query(
                    SkuPricing.brand,
                    func.count(func.distinct(SkuPricing.sku)),
                )
                .filter(SkuPricing.brand.isnot(None))
                .filter(
                    or_(
                        func.coalesce(SkuPricing.stock_internal_qty, 0) > 0,
                        func.coalesce(SkuPricing.stock_qty, 0) > 0,
                    )
                )
                .group_by(SkuPricing.brand)
                .all()
            )
            for b, c in rows_sku:
                brand = (str(b or "").strip())
                if not brand:
                    continue
                brand_sku_count_map[brand] = int(c or 0)

            total_sku_instock = int(
                db.session.query(func.count(func.distinct(SkuPricing.sku)))
                .filter(
                    or_(
                        func.coalesce(SkuPricing.stock_internal_qty, 0) > 0,
                        func.coalesce(SkuPricing.stock_qty, 0) > 0,
                    )
                )
                .scalar()
                or 0
            )
        except Exception:
            brand_sku_count_map = {}
            total_sku_instock = 0

        return render_template(
            "price_settings.html",
            platforms=platforms,
            settings=settings,
            brands=brands,
            owner_map=owner_map,
            owners_list=owners_list,
            export_setting=export_setting,
            brand_monthly_value_map=brand_monthly_value_map,
            brand_sku_count_map=brand_sku_count_map,
            total_sku_instock=total_sku_instock,
        )

    @app.post("/price/settings/export_price")
    @login_required
    def price_settings_export_price():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("price_settings"))

        step_pct = _clamp_float(request.form.get("step_pct"), 0.0, 10.0, default=5.0)
        min_profit_pct = _clamp_float(request.form.get("min_profit_pct"), 0.0, 10.0, default=5.0)
        loss_aging3_pct = _clamp_float(request.form.get("loss_aging3_pct"), 0.0, 50.0, default=5.0)
        loss_aging6_pct = _clamp_float(request.form.get("loss_aging6_pct"), 0.0, 50.0, default=10.0)
        loss_aging12_pct = _clamp_float(request.form.get("loss_aging12_pct"), 0.0, 50.0, default=20.0)

        s = None
        try:
            s = PriceExportSetting.query.get(1)
        except Exception:
            s = None

        if not s:
            s = PriceExportSetting(id=1)
            db.session.add(s)

        s.step_pct = step_pct
        s.min_profit_pct = min_profit_pct
        s.loss_aging3_pct = loss_aging3_pct
        s.loss_aging6_pct = loss_aging6_pct
        s.loss_aging12_pct = loss_aging12_pct

        db.session.commit()
        flash("บันทึกค่าตั้งค่า Export Price เรียบร้อย", "success")
        return redirect(url_for("price_settings"))

    @app.post("/price/settings/brand_owners")
    @login_required
    def price_settings_brand_owners():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("price_settings"))

        brands = request.form.getlist("brand")
        owners = request.form.getlist("owner")

        for b, o in zip(brands, owners):
            brand = (b or "").strip()
            owner = (o or "").strip()
            if not brand:
                continue

            row = BrandOwnerSetting.query.get(brand)
            if not owner:
                if row:
                    db.session.delete(row)
                continue

            if not row:
                db.session.add(BrandOwnerSetting(brand=brand, owner=owner))
            else:
                row.owner = owner

        db.session.commit()
        flash("บันทึกผู้ดูแลตาม Brand เรียบร้อย", "success")
        return redirect(url_for("price_settings"))

    @app.post("/price/settings/add_platform")
    @login_required
    def price_settings_add_platform():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("price_settings"))

        platform_key = normalize_platform_key(request.form.get("new_platform_key") or "")
        label = (request.form.get("new_platform_label") or "").strip()
        fee_pct = request.form.get("new_fee_pct")
        fixed_fee = request.form.get("new_fixed_fee")

        if not platform_key:
            flash("กรุณากรอก Platform Key", "warning")
            return redirect(url_for("price_settings"))

        exists = PlatformFeeSetting.query.get(platform_key)
        if exists:
            flash("มี Platform Key นี้อยู่แล้ว", "warning")
            return redirect(url_for("price_settings"))

        try:
            fee_pct_f = float(fee_pct or 0)
        except Exception:
            fee_pct_f = 0.0
        try:
            fixed_fee_f = float(fixed_fee or 0)
        except Exception:
            fixed_fee_f = 0.0

        s = PlatformFeeSetting(
            platform=platform_key,
            label=(label or platform_key),
            is_active=True,
            sort_order=0,
            fee_pct=fee_pct_f,
            fixed_fee=fixed_fee_f,
        )
        db.session.add(s)
        db.session.commit()
        flash(f"เพิ่มแพลตฟอร์ม {s.label} เรียบร้อย", "success")
        return redirect(url_for("price_settings"))

    @app.route("/import/price/internal", methods=["GET", "POST"])
    @login_required
    def import_price_internal_view():
        if request.method == "POST":
            mode = (request.form.get("mode") or "file").strip().lower()

            try:
                df = None
                source_text = ""
                sheet_url = ""
                worksheet = ""

                if mode == "gsheet":
                    sheet_url = (request.form.get("sheet_url") or "").strip()
                    worksheet = (request.form.get("worksheet") or "").strip()

                    cfg = None
                    try:
                        cfg = PriceConfig.query.filter_by(
                            platform="PRICE_INTERNAL_SYSTEM",
                            name="GoogleSheet_Price_Internal",
                        ).first()
                    except Exception:
                        cfg = None

                    # fallback worksheet/sheet_url จาก config (กันกรณี UI ส่งมาไม่ครบ)
                    if (not worksheet) and cfg and cfg.worksheet:
                        worksheet = (cfg.worksheet or "").strip()
                    if (not sheet_url) and cfg and cfg.url:
                        sheet_url = (cfg.url or "").strip()

                    if not sheet_url:
                        flash("กรุณาระบุ Google Sheet URL", "danger")
                        return redirect(url_for("import_price_internal_view"))

                    creds = get_google_credentials()
                    client = gspread.authorize(creds)
                    sh = client.open_by_url(sheet_url)
                    try:
                        ws = sh.worksheet(worksheet) if worksheet else sh.get_worksheet(0)
                    except gspread.WorksheetNotFound:
                        flash(f"ไม่พบ Worksheet: {worksheet}", "danger")
                        return redirect(url_for("import_price_internal_view"))

                    records = ws.get_all_records()
                    df = pd.DataFrame(records)
                    source_text = "Google Sheet"

                else:
                    f = request.files.get("file")
                    if not f or not f.filename:
                        flash("กรุณาเลือกไฟล์ Excel/CSV", "warning")
                        return redirect(url_for("import_price_internal_view"))

                    df = pd.read_csv(f) if f.filename.lower().endswith(".csv") else pd.read_excel(f)
                    source_text = f.filename

                if df is None or df.empty:
                    flash("ไม่พบข้อมูลในไฟล์/ชีท", "warning")
                    return redirect(url_for("import_price_internal_view"))

                batch = PriceImportBatch(
                    kind="internal",
                    source=(mode or "file"),
                    source_name=(sheet_url if mode == "gsheet" else source_text),
                    worksheet=((worksheet or "").strip() or None) if mode == "gsheet" else None,
                    created_by=(current_user().username if current_user() else "unknown"),
                )
                db.session.add(batch)
                db.session.flush()

                result = import_sku_pricing(df, batch_id=batch.id)
                batch.ok_rows = int(result.get("ok", 0) or 0)
                batch.skip_rows = int(result.get("skip", 0) or 0)
                db.session.commit()
                flash(
                    f"Import ข้อมูลฝั่งเราเรียบร้อย: สำเร็จ {result['ok']} แถว | ข้าม {result['skip']} | สร้างสินค้าใหม่ {result['new_products']} [จาก {source_text}]",
                    "success",
                )
                return redirect(url_for("price_dashboard"))

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import internal price failed")
                flash(f"Import ไม่สำเร็จ: {e}", "danger")
                return redirect(url_for("import_price_internal_view"))

        saved_url = ""
        saved_worksheet = ""
        try:
            cfg = PriceConfig.query.filter_by(
                platform="PRICE_INTERNAL_SYSTEM",
                name="GoogleSheet_Price_Internal",
            ).first()
            if cfg:
                saved_url = cfg.url or ""
                saved_worksheet = cfg.worksheet or ""
        except Exception:
            pass

        last_batch = None
        try:
            last_batch = (
                PriceImportBatch.query.filter_by(kind="internal", undone=False)
                .order_by(PriceImportBatch.id.desc())
                .first()
            )
        except Exception:
            last_batch = None

        return render_template(
            "import_price_internal.html",
            saved_url=saved_url,
            saved_worksheet=saved_worksheet,
            last_batch=last_batch,
        )

    @app.route("/import/price/internal/template", methods=["GET"])
    @login_required
    def download_price_internal_template():
        # Use Dashboard-friendly headers (Importer still supports legacy snake_case)
        df = pd.DataFrame(
            [
                {
                    "SKU": "",
                    "Brand": "",
                    "Name": "",
                    "Stock Internal": "",
                    "Stock": "",
                    "Cost": "",
                    "Our Price": "",
                    "Spec": "",
                    "Floor Price": "",
                    "Min Margin %": "",
                    "Pack Cost": "",
                    "Ship Subsidy": "",
                }
            ],
            columns=[
                "SKU",
                "Brand",
                "Name",
                "Stock Internal",
                "Stock",
                "Cost",
                "Our Price",
                "Spec",
                "Floor Price",
                "Min Margin %",
                "Pack Cost",
                "Ship Subsidy",
            ],
        )

        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Internal")
        bio.seek(0)

        return send_file(
            bio,
            as_attachment=True,
            download_name="template_price_internal.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/import/price/market", methods=["GET", "POST"])
    @login_required
    def import_market_prices_view():
        cu = current_user()
        checked_by = (cu.username if cu else "admin")

        if request.method == "POST":
            mode = (request.form.get("mode") or "file").strip().lower()
            default_platform = (request.form.get("default_platform") or "").strip().lower()
            default_platform = normalize_platform(default_platform) if default_platform else None

            try:
                df = None
                source_text = ""
                sheet_url = ""
                worksheet = ""

                if mode == "gsheet":
                    sheet_url = (request.form.get("sheet_url") or "").strip()
                    worksheet = (request.form.get("worksheet") or "").strip()

                    cfg = None
                    try:
                        cfg = PriceConfig.query.filter_by(
                            platform="PRICE_MARKET_SYSTEM",
                            name="GoogleSheet_Price_Market",
                        ).first()
                    except Exception:
                        cfg = None

                    # fallback worksheet/sheet_url จาก config (กันกรณี UI ส่งมาไม่ครบ)
                    if (not worksheet) and cfg and cfg.worksheet:
                        worksheet = (cfg.worksheet or "").strip()
                    if (not sheet_url) and cfg and cfg.url:
                        sheet_url = (cfg.url or "").strip()

                    if not sheet_url:
                        flash("กรุณาระบุ Google Sheet URL", "danger")
                        return redirect(url_for("import_market_prices_view"))

                    creds = get_google_credentials()
                    client = gspread.authorize(creds)
                    sh = client.open_by_url(sheet_url)
                    try:
                        ws = sh.worksheet(worksheet) if worksheet else sh.get_worksheet(0)
                    except gspread.WorksheetNotFound:
                        flash(f"ไม่พบ Worksheet: {worksheet}", "danger")
                        return redirect(url_for("import_market_prices_view"))
                    records = ws.get_all_records()
                    df = pd.DataFrame(records)
                    source_text = "Google Sheet"

                else:
                    f = request.files.get("file")
                    if not f or not f.filename:
                        flash("กรุณาเลือกไฟล์ Excel/CSV", "danger")
                        return redirect(url_for("import_market_prices_view"))

                    df = pd.read_csv(f) if f.filename.lower().endswith(".csv") else pd.read_excel(f)
                    source_text = f.filename

                if df is None or df.empty:
                    flash("ไม่พบข้อมูลในไฟล์/ชีท", "warning")
                    return redirect(url_for("import_market_prices_view"))

                batch = PriceImportBatch(
                    kind="market",
                    source=(mode or "file"),
                    source_name=(sheet_url if mode == "gsheet" else source_text),
                    worksheet=((worksheet or "").strip() or None) if mode == "gsheet" else None,
                    default_platform=(default_platform or None),
                    created_by=(checked_by or "unknown"),
                )
                db.session.add(batch)
                db.session.flush()

                result = import_market_prices(
                    df,
                    default_platform=default_platform,
                    checked_by=checked_by,
                    batch_id=batch.id,
                )
                batch.ok_rows = int(result.get("ok", 0) or 0)
                batch.skip_rows = int(result.get("skip", 0) or 0)
                db.session.commit()
                flash(
                    f"✅ นำเข้าราคาตลาดสำเร็จ {result.get('ok', 0)} แถว (ข้าม {result.get('skip', 0)}) [จาก {source_text}]",
                    "success",
                )
                return redirect(url_for("price_dashboard"))

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import market prices failed")
                flash(f"เกิดข้อผิดพลาด: {e}", "danger")
                return redirect(url_for("import_market_prices_view"))

        saved_url = ""
        saved_worksheet = ""
        try:
            cfg = PriceConfig.query.filter_by(
                platform="PRICE_MARKET_SYSTEM",
                name="GoogleSheet_Price_Market",
            ).first()
            if cfg:
                saved_url = cfg.url or ""
                saved_worksheet = cfg.worksheet or ""
        except Exception:
            pass

        last_batch = None
        try:
            last_batch = (
                PriceImportBatch.query.filter_by(kind="market", undone=False)
                .order_by(PriceImportBatch.id.desc())
                .first()
            )
        except Exception:
            last_batch = None

        return render_template(
            "import_market_prices.html",
            saved_url=saved_url,
            saved_worksheet=saved_worksheet,
            last_batch=last_batch,
        )

    @app.route("/import/price/market/template", methods=["GET"])
    @login_required
    def download_price_market_template():
        # Use Dashboard-friendly headers (Importer still supports legacy snake_case)
        df = pd.DataFrame(
            [
                {
                    "SKU": "",
                    "Platform": "",
                    "Shop": "",
                    "Market (best)": "",
                    "Voucher": "",
                    "MALL": "",
                    "URL": "",
                    "Updated": "",
                }
            ]
        )

        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Market")
        bio.seek(0)

        return send_file(
            bio,
            as_attachment=True,
            download_name="template_price_market.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/import/price/brand_control", methods=["GET", "POST"])
    @login_required
    def import_brand_control_view():
        if request.method == "POST":
            mode = (request.form.get("mode") or "file").strip().lower()

            try:
                df = None
                source_text = ""
                sheet_url = ""
                worksheet = ""

                if mode == "gsheet":
                    sheet_url = (request.form.get("sheet_url") or "").strip()
                    worksheet = (request.form.get("worksheet") or "").strip()

                    cfg = None
                    try:
                        cfg = PriceConfig.query.filter_by(
                            platform="PRICE_BRAND_CONTROL_SYSTEM",
                            name="GoogleSheet_Brand_Control",
                        ).first()
                    except Exception:
                        cfg = None

                    # fallback worksheet/sheet_url จาก config (กันกรณี UI ส่งมาไม่ครบ)
                    if (not worksheet) and cfg and cfg.worksheet:
                        worksheet = (cfg.worksheet or "").strip()
                    if (not sheet_url) and cfg and cfg.url:
                        sheet_url = (cfg.url or "").strip()

                    if not sheet_url:
                        flash("กรุณาระบุ Google Sheet URL", "danger")
                        return redirect(url_for("import_brand_control_view"))

                    creds = get_google_credentials()
                    client = gspread.authorize(creds)
                    sh = client.open_by_url(sheet_url)
                    try:
                        ws = sh.worksheet(worksheet) if worksheet else sh.get_worksheet(0)
                    except gspread.WorksheetNotFound:
                        flash(f"ไม่พบ Worksheet: {worksheet}", "danger")
                        return redirect(url_for("import_brand_control_view"))

                    records = ws.get_all_records()
                    df = pd.DataFrame(records)
                    source_text = "Google Sheet"

                else:
                    f = request.files.get("file")
                    if not f or not f.filename:
                        flash("กรุณาเลือกไฟล์ Excel/CSV", "danger")
                        return redirect(url_for("import_brand_control_view"))

                    df = pd.read_csv(f) if f.filename.lower().endswith(".csv") else pd.read_excel(f)
                    source_text = f.filename

                if df is None or df.empty:
                    flash("ไม่พบข้อมูลในไฟล์/ชีท", "warning")
                    return redirect(url_for("import_brand_control_view"))

                batch = PriceImportBatch(
                    kind="brand_control",
                    source=(mode or "file"),
                    source_name=(sheet_url if mode == "gsheet" else source_text),
                    worksheet=((worksheet or "").strip() or None) if mode == "gsheet" else None,
                    created_by=(current_user().username if current_user() else "unknown"),
                )
                db.session.add(batch)
                db.session.flush()

                result = import_brand_control(df, batch_id=batch.id)
                batch.ok_rows = int(result.get("ok", 0) or 0)
                batch.skip_rows = int(result.get("skip", 0) or 0)
                db.session.commit()
                flash(
                    f"✅ นำเข้า Brand Control สำเร็จ {result.get('ok', 0)} แถว (ข้าม {result.get('skip', 0)}) [จาก {source_text}]",
                    "success",
                )
                return redirect(url_for("price_dashboard"))

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import brand control failed")
                flash(f"เกิดข้อผิดพลาด: {e}", "danger")
                return redirect(url_for("import_brand_control_view"))

        saved_url = ""
        saved_worksheet = ""
        try:
            cfg = PriceConfig.query.filter_by(
                platform="PRICE_BRAND_CONTROL_SYSTEM",
                name="GoogleSheet_Brand_Control",
            ).first()
            if cfg:
                saved_url = cfg.url or ""
                saved_worksheet = cfg.worksheet or ""
        except Exception:
            pass

        last_batch = None
        try:
            last_batch = (
                PriceImportBatch.query.filter_by(kind="brand_control", undone=False)
                .order_by(PriceImportBatch.id.desc())
                .first()
            )
        except Exception:
            last_batch = None

        return render_template(
            "import_brand_control.html",
            saved_url=saved_url,
            saved_worksheet=saved_worksheet,
            last_batch=last_batch,
        )

    @app.route("/import/price/monthly_sales", methods=["GET", "POST"])
    @login_required
    def import_monthly_sales_view():
        if request.method == "POST":
            mode = (request.form.get("mode") or "file").strip().lower()

            try:
                df = None
                source_text = ""
                sheet_url = ""
                worksheet = ""

                if mode == "gsheet":
                    sheet_url = (request.form.get("sheet_url") or "").strip()
                    worksheet = (request.form.get("worksheet") or "").strip() or "Sheet1"

                    cfg = None
                    try:
                        cfg = PriceConfig.query.filter_by(
                            platform="PRICE_MONTHLY_SALES_SYSTEM",
                            name="GoogleSheet_Monthly_Sales",
                        ).first()
                    except Exception:
                        cfg = None

                    # fallback sheet_url/worksheet จาก config (กันกรณี UI ส่งมาไม่ครบ)
                    if (not sheet_url) and cfg and cfg.url:
                        sheet_url = (cfg.url or "").strip()
                    if (not worksheet) and cfg and cfg.worksheet:
                        worksheet = (cfg.worksheet or "").strip() or "Sheet1"

                    if not sheet_url:
                        flash("กรุณาระบุ Google Sheet URL", "danger")
                        return redirect(url_for("import_monthly_sales_view"))

                    creds = get_google_credentials()
                    client = gspread.authorize(creds)
                    sh = client.open_by_url(sheet_url)
                    try:
                        ws = sh.worksheet(worksheet) if worksheet else sh.get_worksheet(0)
                    except gspread.WorksheetNotFound:
                        flash(f"ไม่พบ Worksheet: {worksheet}", "danger")
                        return redirect(url_for("import_monthly_sales_view"))

                    records = ws.get_all_records()
                    df = pd.DataFrame(records)
                    source_text = "Google Sheet"

                    # save config for next time
                    try:
                        if not cfg:
                            cfg = PriceConfig(platform="PRICE_MONTHLY_SALES_SYSTEM", name="GoogleSheet_Monthly_Sales")
                            db.session.add(cfg)
                        cfg.url = sheet_url
                        cfg.worksheet = worksheet
                        db.session.commit()
                    except Exception:
                        db.session.rollback()

                else:
                    f = request.files.get("file")
                    if not f or not f.filename:
                        flash("กรุณาเลือกไฟล์ Excel/CSV", "danger")
                        return redirect(url_for("import_monthly_sales_view"))

                    df = pd.read_csv(f) if f.filename.lower().endswith(".csv") else pd.read_excel(f)
                    source_text = f.filename

                if df is None or df.empty:
                    flash("ไม่พบข้อมูลในไฟล์/ชีท", "warning")
                    return redirect(url_for("import_monthly_sales_view"))

                batch = PriceImportBatch(
                    kind="monthly_sales",
                    source=(mode or "file"),
                    source_name=(sheet_url if mode == "gsheet" else source_text),
                    worksheet=((worksheet or "").strip() or None) if mode == "gsheet" else None,
                    created_by=(current_user().username if current_user() else "unknown"),
                )
                db.session.add(batch)
                db.session.flush()

                result = import_monthly_sales(df, batch_id=batch.id)
                batch.ok_rows = int(result.get("ok", 0) or 0)
                batch.skip_rows = int(result.get("skip", 0) or 0)
                db.session.commit()
                flash(
                    f"✅ นำเข้า Monthly Sales สำเร็จ {result.get('ok', 0)} แถว (ข้าม {result.get('skip', 0)}) [จาก {source_text}]",
                    "success",
                )
                return redirect(url_for("price_dashboard"))

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import monthly sales failed")
                flash(f"เกิดข้อผิดพลาด: {e}", "danger")
                return redirect(url_for("import_monthly_sales_view"))

        saved_url = ""
        saved_ws = ""
        try:
            cfg = PriceConfig.query.filter_by(
                platform="PRICE_MONTHLY_SALES_SYSTEM",
                name="GoogleSheet_Monthly_Sales",
            ).first()
            if cfg:
                saved_url = cfg.url or ""
                saved_ws = cfg.worksheet or ""
        except Exception:
            pass

        last_batch = None
        try:
            last_batch = (
                PriceImportBatch.query.filter_by(kind="monthly_sales", undone=False)
                .order_by(PriceImportBatch.id.desc())
                .first()
            )
        except Exception:
            last_batch = None

        return render_template(
            "import_monthly_sales.html",
            saved_url=saved_url,
            saved_ws=saved_ws,
            last_batch=last_batch,
        )

    @app.route("/import/price/monthly_sales/template", methods=["GET"])
    @login_required
    def download_price_monthly_sales_template():
        df = pd.DataFrame(
            [
                {
                    "SKU": "",
                    "Quantity": "",
                }
            ]
        )

        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="MonthlySales")
        bio.seek(0)

        return send_file(
            bio,
            as_attachment=True,
            download_name="monthly_sales_template.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/import/price/monthly_sales/clear_all", methods=["POST"])
    @login_required
    def import_monthly_sales_clear_all():
        cu = current_user()
        if not cu or cu.role != "admin":
            flash("ต้องเป็น admin เท่านั้น", "danger")
            return redirect(url_for("import_monthly_sales_view"))

        confirm = (request.form.get("confirm") or "").strip()
        if confirm != "CLEAR SALES":
            flash('พิมพ์ "CLEAR SALES" เพื่อยืนยัน', "warning")
            return redirect(url_for("import_monthly_sales_view"))

        try:
            SkuPricing.query.update({SkuPricing.monthly_sales_qty: 0})
            _delete_price_import_history("monthly_sales")
            db.session.commit()
            flash("ล้างข้อมูล Monthly Sales แล้ว ✅", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"ล้างข้อมูลไม่สำเร็จ: {e}", "danger")

        return redirect(url_for("import_monthly_sales_view"))


    # -------------------------------------------------------------------
    # Supplier Stock: Import (file/gsheet) + Template + Clear + Dashboard
    # -------------------------------------------------------------------

    @app.route("/import/price/supplier_sku_stock", methods=["GET", "POST"])
    @login_required
    def import_supplier_sku_stock_view():
        if request.method == "POST":
            mode = (request.form.get("mode") or "file").strip().lower()

            try:
                df = None
                source_text = ""
                sheet_url = ""
                worksheet = ""

                if mode == "gsheet":
                    sheet_url = (request.form.get("sheet_url") or "").strip()
                    worksheet = (request.form.get("worksheet") or "").strip() or "Sheet1"

                    cfg = SupplierConfig.query.filter_by(name="GoogleSheet_SupplierSkuStock").first()

                    if (not sheet_url) and cfg and cfg.url:
                        sheet_url = (cfg.url or "").strip()
                    if (not worksheet) and cfg and cfg.worksheet:
                        worksheet = (cfg.worksheet or "").strip() or "Sheet1"

                    if not sheet_url:
                        flash("กรุณาระบุ Google Sheet URL", "danger")
                        return redirect(url_for("import_supplier_sku_stock_view"))

                    creds = get_google_credentials()
                    client = gspread.authorize(creds)
                    sh = client.open_by_url(sheet_url)
                    try:
                        ws = sh.worksheet(worksheet) if worksheet else sh.get_worksheet(0)
                    except gspread.WorksheetNotFound:
                        flash(f"ไม่พบ Worksheet: {worksheet}", "danger")
                        return redirect(url_for("import_supplier_sku_stock_view"))

                    df = pd.DataFrame(ws.get_all_records())
                    source_text = "Google Sheet"

                    # save config for next time
                    try:
                        if not cfg:
                            cfg = SupplierConfig(name="GoogleSheet_SupplierSkuStock")
                            db.session.add(cfg)
                        cfg.url = sheet_url
                        cfg.worksheet = worksheet
                        db.session.commit()
                    except Exception:
                        db.session.rollback()

                else:
                    f = request.files.get("file")
                    if not f or not f.filename:
                        flash("กรุณาเลือกไฟล์ Excel/CSV", "danger")
                        return redirect(url_for("import_supplier_sku_stock_view"))

                    df = pd.read_csv(f) if f.filename.lower().endswith(".csv") else pd.read_excel(f)
                    source_text = f.filename

                if df is None or df.empty:
                    flash("ไม่พบข้อมูลในไฟล์/ชีท", "warning")
                    return redirect(url_for("import_supplier_sku_stock_view"))

                batch = SupplierImportBatch(
                    kind="supplier_sku_stock",
                    source=(mode or "file"),
                    source_name=(sheet_url if mode == "gsheet" else source_text),
                    worksheet=((worksheet or "").strip() or None) if mode == "gsheet" else None,
                    created_by=(current_user().username if current_user() else "unknown"),
                )
                db.session.add(batch)
                db.session.flush()

                result = import_supplier_sku_stock(df)
                batch.ok_rows = int(result.get("ok", 0) or 0)
                batch.skip_rows = int(result.get("skip", 0) or 0)
                db.session.commit()

                flash(
                    f"✅ Import Supplier Stock สำเร็จ ok={result.get('ok',0)} | insert={result.get('insert',0)} | update={result.get('update',0)} | conflict={result.get('conflict',0)}",
                    "success",
                )
                return redirect(url_for("supplier_stock_dashboard"))

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import supplier sku stock failed")
                flash(f"เกิดข้อผิดพลาด: {e}", "danger")
                return redirect(url_for("import_supplier_sku_stock_view"))

        saved_url = ""
        saved_ws = ""
        cfg = SupplierConfig.query.filter_by(name="GoogleSheet_SupplierSkuStock").first()
        if cfg:
            saved_url = cfg.url or ""
            saved_ws = cfg.worksheet or ""

        last_batch = SupplierImportBatch.query.filter_by(kind="supplier_sku_stock").order_by(SupplierImportBatch.id.desc()).first()

        return render_template(
            "import_supplier_sku_stock.html",
            saved_url=saved_url,
            saved_ws=saved_ws,
            last_batch=last_batch,
        )


    @app.route("/import/price/supplier_sku_stock/template", methods=["GET"])
    @login_required
    def download_supplier_sku_stock_template():
        df = pd.DataFrame(
            [
                {
                    "SKU": "V-HS-SSD-E100-512G",
                    "SKU SUP": "HS-SSD-E100 512G",
                    "Supplier": "SIS",
                    "Brand": "HIKSEMI",
                    "Name": "HIKSEMI SSD E100 512GB",
                    "Stock": 1,
                }
            ],
            columns=["SKU", "SKU SUP", "Supplier", "Brand", "Name", "Stock"],
        )

        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Import_Supplier")
        bio.seek(0)

        return send_file(
            bio,
            as_attachment=True,
            download_name="template_import_sku_stock_supplier.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


    @app.route("/import/price/supplier_sku_stock/clear_stock", methods=["POST"])
    @login_required
    def clear_supplier_stock_only():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("import_supplier_sku_stock_view"))

        if (request.form.get("confirm") or "").strip() != "CLEAR SUP STOCK":
            flash('พิมพ์คำยืนยันไม่ถูกต้อง (ต้องเป็น "CLEAR SUP STOCK")', "warning")
            return redirect(url_for("import_supplier_sku_stock_view"))

        try:
            SupplierSkuMaster.query.update(
                {
                    SupplierSkuMaster.stock_sup_qty: 0,
                    SupplierSkuMaster.stock_updated_at: None,
                }
            )
            db.session.commit()
            flash("✅ ล้างเฉพาะ Stock Sup แล้ว (Mapping ยังอยู่)", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"ล้างไม่สำเร็จ: {e}", "danger")

        return redirect(url_for("import_supplier_sku_stock_view"))


    @app.route("/import/price/supplier_sku_stock/clear_all", methods=["POST"])
    @login_required
    def clear_supplier_all():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("import_supplier_sku_stock_view"))

        if (request.form.get("confirm") or "").strip() != "CLEAR SUP ALL":
            flash('พิมพ์คำยืนยันไม่ถูกต้อง (ต้องเป็น "CLEAR SUP ALL")', "warning")
            return redirect(url_for("import_supplier_sku_stock_view"))

        try:
            SupplierSkuMaster.query.delete(synchronize_session=False)
            SupplierImportBatch.query.delete(synchronize_session=False)
            db.session.commit()
            flash("✅ ล้างข้อมูล Supplier ทั้งหมดแล้ว", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"ล้างไม่สำเร็จ: {e}", "danger")

        return redirect(url_for("import_supplier_sku_stock_view"))


    @app.route("/supplier-stock", methods=["GET"])
    @login_required
    def supplier_stock_dashboard():
        cu = current_user()

        selected_supplier = (request.args.get("supplier") or "").strip()

        # ---- Read prefs as default ----
        raw_len = (get_user_pref_main(int(cu.id), "supplier_stock.page_length", "100") or "100").strip()
        try:
            pref_len = int(raw_len)
        except Exception:
            pref_len = 100
        if pref_len not in {100, 300, 500, 1000, -1}:
            pref_len = 100

        pref_stock = (get_user_pref_main(int(cu.id), "supplier_stock.stock_filter", "all") or "all").strip().lower()
        if pref_stock not in {"all", "nonzero", "zero"}:
            pref_stock = "all"

        # ---- Query args override prefs ----
        q_raw = (request.args.get("q") or "").strip()
        q = q_raw.lower()
        stock_mode = (request.args.get("stock") or pref_stock).strip().lower()
        brand_sel = (request.args.get("brand") or "").strip()
        sort_code = (request.args.get("sort") or "sku_asc").strip()
        limit_sel_raw = (request.args.get("limit") or str(pref_len)).strip()

        try:
            limit_sel = int(limit_sel_raw)
        except Exception:
            limit_sel = pref_len
        if limit_sel not in {100, 300, 500, 1000, -1}:
            limit_sel = 100

        if stock_mode not in {"all", "nonzero", "zero"}:
            stock_mode = "all"

        # ---- Save prefs when user changes ----
        if "stock" in request.args:
            set_user_pref_main(int(cu.id), "supplier_stock.stock_filter", stock_mode)
        if "limit" in request.args:
            set_user_pref_main(int(cu.id), "supplier_stock.page_length", str(limit_sel))

        # chunk size for infinite scroll
        page_size = 500 if limit_sel == -1 else int(limit_sel)

        # ---- Load supplier rows ----
        rows_raw = (
            SupplierSkuMaster.query.filter(SupplierSkuMaster.is_active == True)
            .order_by(
                SupplierSkuMaster.sku.asc(),
                SupplierSkuMaster.supplier.asc(),
                SupplierSkuMaster.sku_sup.asc(),
            )
            .all()
        )

        sup_all: set[str] = set()
        sup_skus_nonzero: dict[str, set[str]] = {}
        sup_qty_nonzero: dict[str, int] = {}
        all_sku_nonzero_set: set[str] = set()
        all_qty_total = 0

        for r in rows_raw:
            sup = (r.supplier or "").strip()
            sku = (r.sku or "").strip()
            qty = int(r.stock_sup_qty or 0)

            if sup:
                sup_all.add(sup)

            all_qty_total += qty

            if sup and sku and qty > 0:
                sup_skus_nonzero.setdefault(sup, set()).add(sku)
                sup_qty_nonzero[sup] = sup_qty_nonzero.get(sup, 0) + qty
                all_sku_nonzero_set.add(sku)

        supplier_cards = []
        for sup in sorted(sup_all):
            skus = sup_skus_nonzero.get(sup, set())
            supplier_cards.append(
                {
                    "supplier": sup,
                    "sku_nonzero": len(skus),
                    "qty_nonzero": int(sup_qty_nonzero.get(sup, 0)),
                }
            )

        supplier_cards.sort(
            key=lambda x: (x.get("sku_nonzero", 0), x.get("qty_nonzero", 0)),
            reverse=True,
        )

        all_suppliers_count = len(sup_all)
        all_sku_nonzero_count = len(all_sku_nonzero_set)

        # ---- Aggregate by SKU ----
        agg: dict[str, dict] = {}
        for r in rows_raw:
            sku = (r.sku or "").strip()
            if not sku:
                continue

            if sku not in agg:
                agg[sku] = {
                    "sku": sku,
                    "brand": r.brand or "",
                    "name": r.name or "",
                    "items": [],
                    "total_stock": 0,
                    "last_updated": None,
                }

            agg[sku]["items"].append(
                {
                    "supplier": r.supplier or "",
                    "sku_sup": r.sku_sup or "",
                    "stock": int(r.stock_sup_qty or 0),
                }
            )
            agg[sku]["total_stock"] += int(r.stock_sup_qty or 0)

            if r.stock_updated_at and (
                agg[sku]["last_updated"] is None or r.stock_updated_at > agg[sku]["last_updated"]
            ):
                agg[sku]["last_updated"] = r.stock_updated_at

            if (not agg[sku]["brand"]) and r.brand:
                agg[sku]["brand"] = r.brand
            if (not agg[sku]["name"]) and r.name:
                agg[sku]["name"] = r.name

        data = list(agg.values())

        def _total_stock(d: dict) -> int:
            return int(d.get("total_stock") or 0)

        def _safe_text(s: str | None) -> str:
            return (s or "").strip().lower()

        def _hay(d: dict) -> str:
            parts = [d.get("sku") or "", d.get("brand") or "", d.get("name") or ""]
            for it in (d.get("items") or []):
                parts += [it.get("supplier") or "", it.get("sku_sup") or ""]
            return " ".join(parts).lower()

        # ---- Supplier filter: keep only selected supplier items (stock>0) to reduce HTML ----
        if selected_supplier:
            for d in data:
                d["items"] = [
                    it
                    for it in (d.get("items") or [])
                    if (it.get("supplier") == selected_supplier and int(it.get("stock") or 0) > 0)
                ]
                d["total_stock"] = sum(int(it.get("stock") or 0) for it in (d.get("items") or []))
            data = [d for d in data if d.get("items")]

        # ---- Stock mode filter ----
        if stock_mode == "nonzero":
            data = [d for d in data if _total_stock(d) > 0]
        elif stock_mode == "zero":
            data = [d for d in data if _total_stock(d) == 0]

        # brand list after supplier/stock mode for more relevant dropdown
        brands = sorted({(d.get("brand") or "").strip() for d in data if (d.get("brand") or "").strip()})

        # ---- Search + Brand filter ----
        if q:
            data = [d for d in data if q in _hay(d)]
        if brand_sel:
            data = [d for d in data if (d.get("brand") or "") == brand_sel]

        # ---- Sort ----
        if sort_code == "sku_asc":
            data.sort(key=lambda d: _safe_text(d.get("sku")))
        elif sort_code == "sku_desc":
            data.sort(key=lambda d: _safe_text(d.get("sku")), reverse=True)
        elif sort_code == "brand_asc":
            data.sort(key=lambda d: (_safe_text(d.get("brand")), _safe_text(d.get("sku"))))
        elif sort_code == "brand_desc":
            data.sort(key=lambda d: (_safe_text(d.get("brand")), _safe_text(d.get("sku"))), reverse=True)
        elif sort_code == "name_asc":
            data.sort(key=lambda d: (_safe_text(d.get("name")), _safe_text(d.get("sku"))))
        elif sort_code == "name_desc":
            data.sort(key=lambda d: (_safe_text(d.get("name")), _safe_text(d.get("sku"))), reverse=True)
        elif sort_code == "stock_desc":
            data.sort(key=lambda d: (_total_stock(d), _safe_text(d.get("sku"))), reverse=True)
        elif sort_code == "stock_asc":
            data.sort(key=lambda d: (_total_stock(d), _safe_text(d.get("sku"))))
        elif sort_code == "upd_desc":
            data.sort(key=lambda d: ((d.get("last_updated") or datetime.min), _safe_text(d.get("sku"))), reverse=True)
        elif sort_code == "upd_asc":
            data.sort(key=lambda d: ((d.get("last_updated") or datetime.min), _safe_text(d.get("sku"))))
        else:
            data.sort(key=lambda d: (_total_stock(d), _safe_text(d.get("sku"))), reverse=True)

        # ---- KPI from FULL filtered data ----
        kpi_total_sku = len(data)
        kpi_total_stock = sum(int(d.get("total_stock") or 0) for d in data)

        # ---- Cache + render first chunk ----
        _supplier_dash_cache_gc()
        dash_key = uuid.uuid4().hex
        SUPPLIER_DASH_ROWS_CACHE[dash_key] = {
            "user_id": int(cu.id),
            "ts": time.time(),
            "rows": data,
        }

        initial_rows = data[:page_size]
        total_rows = len(data)

        return render_template(
            "supplier_stock_dashboard.html",
            rows=initial_rows,
            dash_key=dash_key,
            dash_total_rows=total_rows,
            dash_page_size=page_size,

            # filters (for UI)
            q=q_raw,
            stock_mode=stock_mode,
            brand_sel=brand_sel,
            sort_code=sort_code,
            limit_sel=limit_sel,

            # KPI
            kpi_total_sku=kpi_total_sku,
            kpi_total_stock=kpi_total_stock,

            supplier_cards=supplier_cards,
            selected_supplier=selected_supplier,
            brands=brands,
            all_suppliers_count=all_suppliers_count,
            all_sku_nonzero_count=all_sku_nonzero_count,
            all_qty_total=int(all_qty_total),
        )


    @app.get("/api/supplier-stock/rows")
    @login_required
    def api_supplier_stock_rows():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "msg": "กรุณาเข้าสู่ระบบ"}), 401

        _supplier_dash_cache_gc()

        key = (request.args.get("key") or "").strip()
        try:
            offset = int(request.args.get("offset") or 0)
            limit = int(request.args.get("limit") or 200)
        except Exception:
            return jsonify({"success": False, "msg": "พารามิเตอร์ไม่ถูกต้อง"}), 400

        entry = SUPPLIER_DASH_ROWS_CACHE.get(key)
        if not entry:
            return jsonify({"success": False, "msg": "หมดอายุ/ไม่พบ cache (ให้รีเฟรชหน้า)"}), 400

        if int(entry.get("user_id") or 0) != int(cu.id):
            return jsonify({"success": False, "msg": "ไม่มีสิทธิ์เข้าถึงชุดข้อมูลนี้"}), 403

        rows_all = entry.get("rows") or []
        total = len(rows_all)

        if offset < 0:
            offset = 0
        if limit <= 0:
            limit = 200
        if limit > 1000:
            limit = 1000

        slice_rows = rows_all[offset: offset + limit]
        html = render_template("_supplier_stock_rows.html", rows=slice_rows)

        next_offset = offset + len(slice_rows)
        has_more = next_offset < total

        return jsonify({
            "success": True,
            "html": html,
            "next_offset": next_offset,
            "total": total,
            "has_more": has_more,
        })


    @app.get("/supplier-stock/export.xlsx")
    @login_required
    def supplier_stock_export_xlsx():
        cu = current_user()
        if not cu:
            return redirect(url_for("login"))

        layout = (request.args.get("layout") or "template").strip().lower()  # template | horizontal
        selected_supplier = (request.args.get("supplier") or "").strip()
        q = (request.args.get("q") or "").strip().lower()
        stock_mode = (request.args.get("stock") or "all").strip().lower()  # all | nonzero | zero
        brand_sel = (request.args.get("brand") or "").strip()
        sort_code = (request.args.get("sort") or "sku_asc").strip()
        export_all = (request.args.get("export_all") or "").strip().lower() in {"1", "true", "yes", "on"}

        if stock_mode not in {"all", "nonzero", "zero"}:
            stock_mode = "all"
        if layout not in {"template", "horizontal"}:
            layout = "template"

        if export_all:
            q = ""
            stock_mode = "all"
            brand_sel = ""

        rows = (
            SupplierSkuMaster.query.filter(SupplierSkuMaster.is_active == True)
            .order_by(SupplierSkuMaster.sku.asc(), SupplierSkuMaster.supplier.asc())
            .all()
        )

        agg: dict[str, dict] = {}
        for r in rows:
            sku = (r.sku or "").strip()
            if not sku:
                continue

            if sku not in agg:
                agg[sku] = {
                    "sku": sku,
                    "brand": (r.brand or "").strip(),
                    "name": (r.name or "").strip(),
                    "items": [],
                }

            if (not agg[sku]["brand"]) and r.brand:
                agg[sku]["brand"] = (r.brand or "").strip()
            if (not agg[sku]["name"]) and r.name:
                agg[sku]["name"] = (r.name or "").strip()

            dt = r.stock_updated_at
            updated_str = dt.strftime("%d/%m/%Y %H:%M") if dt else ""

            agg[sku]["items"].append(
                {
                    "supplier": (r.supplier or "").strip(),
                    "sku_sup": (r.sku_sup or "").strip(),
                    "stock": int(r.stock_sup_qty or 0),
                    "updated": updated_str,
                    "updated_dt": dt,
                }
            )

        data = list(agg.values())

        def _total_stock(d: dict) -> int:
            return sum(int(it.get("stock") or 0) for it in (d.get("items") or []))

        def _safe_text(s: str | None) -> str:
            return (s or "").strip().lower()

        def _last_upd_dt(d: dict):
            dts = [it.get("updated_dt") for it in (d.get("items") or []) if it.get("updated_dt")]
            return max(dts) if dts else None

        if selected_supplier:
            for d in data:
                d["items"] = [
                    it
                    for it in (d.get("items") or [])
                    if (it.get("supplier") == selected_supplier and int(it.get("stock") or 0) > 0)
                ]
            data = [d for d in data if d.get("items")]

        if stock_mode == "nonzero":
            data = [d for d in data if _total_stock(d) > 0]
        elif stock_mode == "zero":
            data = [d for d in data if _total_stock(d) == 0]

        if brand_sel:
            data = [d for d in data if (d.get("brand") or "") == brand_sel]

        if q:
            def _hay(d: dict) -> str:
                parts = [d.get("sku") or "", d.get("brand") or "", d.get("name") or ""]
                for it in (d.get("items") or []):
                    parts += [it.get("supplier") or "", it.get("sku_sup") or ""]
                return " ".join(parts).lower()

            data = [d for d in data if q in _hay(d)]

        if sort_code == "sku_asc":
            data.sort(key=lambda d: _safe_text(d.get("sku")))
        elif sort_code == "sku_desc":
            data.sort(key=lambda d: _safe_text(d.get("sku")), reverse=True)
        elif sort_code == "brand_asc":
            data.sort(key=lambda d: (_safe_text(d.get("brand")), _safe_text(d.get("sku"))))
        elif sort_code == "brand_desc":
            data.sort(key=lambda d: (_safe_text(d.get("brand")), _safe_text(d.get("sku"))), reverse=True)
        elif sort_code == "name_asc":
            data.sort(key=lambda d: (_safe_text(d.get("name")), _safe_text(d.get("sku"))))
        elif sort_code == "name_desc":
            data.sort(key=lambda d: (_safe_text(d.get("name")), _safe_text(d.get("sku"))), reverse=True)
        elif sort_code == "stock_desc":
            data.sort(key=lambda d: (_total_stock(d), _safe_text(d.get("sku"))), reverse=True)
        elif sort_code == "stock_asc":
            data.sort(key=lambda d: (_total_stock(d), _safe_text(d.get("sku"))))
        elif sort_code == "upd_desc":
            data.sort(key=lambda d: (_last_upd_dt(d) or datetime.min, _safe_text(d.get("sku"))), reverse=True)
        elif sort_code == "upd_asc":
            data.sort(key=lambda d: (_last_upd_dt(d) or datetime.min, _safe_text(d.get("sku"))))

        if layout == "template":
            out_rows = []
            for d in data:
                sku = d.get("sku") or ""
                brand = d.get("brand") or ""
                name = d.get("name") or ""
                items = (d.get("items") or [])
                total_stock = sum(int(it.get("stock") or 0) for it in items)

                for idx, it in enumerate(items):
                    out_rows.append(
                        {
                            "SKU": sku,
                            "SKU SUP": it.get("sku_sup") or "",
                            "Supplier": it.get("supplier") or "",
                            "Brand": brand,
                            "Name": name,
                            "Stock": int(it.get("stock") or 0),
                            "Total Stock": (total_stock if idx == 0 else ""),
                            "Updated": it.get("updated") or "",
                        }
                    )

            df = pd.DataFrame(
                out_rows,
                columns=["SKU", "SKU SUP", "Supplier", "Brand", "Name", "Stock", "Total Stock", "Updated"],
            )
            sheet_name = "SupplierStock_Template"

        else:
            for d in data:
                d["items"] = sorted(d.get("items") or [], key=lambda it: _safe_text(it.get("supplier")))

            max_n = 0
            for d in data:
                max_n = max(max_n, len(d.get("items") or []))

            cols = ["SKU", "Brand", "Name"]
            for i in range(1, max_n + 1):
                cols += [f"Supplier {i}", f"SKU SUP {i}", f"Stock {i}", f"Updated {i}"]

            out_rows = []
            for d in data:
                row = {
                    "SKU": d.get("sku") or "",
                    "Brand": d.get("brand") or "",
                    "Name": d.get("name") or "",
                }

                items = d.get("items") or []
                for i in range(max_n):
                    it = items[i] if i < len(items) else {}
                    row[f"Supplier {i + 1}"] = it.get("supplier") or ""
                    row[f"SKU SUP {i + 1}"] = it.get("sku_sup") or ""
                    row[f"Stock {i + 1}"] = int(it.get("stock") or 0) if it else ""
                    row[f"Updated {i + 1}"] = it.get("updated") or ""

                out_rows.append(row)

            df = pd.DataFrame(out_rows, columns=cols)
            sheet_name = "SupplierStock_Horizontal"

        df = sanitize_excel_df(df)

        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        bio.seek(0)

        ts = now_thai().strftime("%Y%m%d_%H%M")
        filename = f"supplier_stock_{layout}_{ts}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # -------------------------------------------------------------------
    # Price Import: Undo latest batch + Clear all (danger zone)
    # -------------------------------------------------------------------

    def _parse_dt(s):
        if not s:
            return None
        try:
            return pd.to_datetime(s).to_pydatetime()
        except Exception:
            return None

    def undo_last_price_import(kind: str, undone_by: str):
        batch = (
            PriceImportBatch.query.filter_by(kind=kind, undone=False)
            .order_by(PriceImportBatch.id.desc())
            .first()
        )
        if not batch:
            return None

        ops = (
            PriceImportOp.query.filter_by(batch_id=batch.id)
            .order_by(PriceImportOp.seq.desc())
            .all()
        )

        for op in ops:
            t = op.table_name
            act = op.action
            pk = op.pk

            if t == "market_price_logs" and act == "insert":
                log = MarketPriceLog.query.get(int(pk))
                if log:
                    db.session.delete(log)

            elif t == "market_items":
                item_id = int(pk)
                if act == "insert":
                    item = MarketItem.query.get(item_id)
                    if item:
                        db.session.delete(item)
                elif act == "update":
                    before = json.loads(op.before_json or "{}")
                    item = MarketItem.query.get(item_id)
                    if not item:
                        item = MarketItem(
                            id=item_id,
                            sku=before.get("sku") or "",
                            platform=before.get("platform") or "",
                            shop_name=before.get("shop_name") or "",
                        )
                        db.session.add(item)

                    item.sku = before.get("sku") or item.sku
                    item.platform = before.get("platform") or item.platform
                    item.shop_name = before.get("shop_name") or item.shop_name

                    item.product_url = before.get("product_url")
                    item.is_active = bool(before.get("is_active", True))
                    item.latest_listed_price = before.get("latest_listed_price")
                    item.latest_shipping_fee = before.get("latest_shipping_fee")
                    item.latest_voucher_discount = before.get("latest_voucher_discount")
                    item.latest_coin_discount = before.get("latest_coin_discount")
                    item.latest_net_price = before.get("latest_net_price")
                    item.last_updated = _parse_dt(before.get("last_updated"))
                    item.note = before.get("note")

            elif t == "sku_pricing":
                sku = pk
                if act == "insert":
                    pr = SkuPricing.query.get(sku)
                    if pr:
                        db.session.delete(pr)
                elif act == "update":
                    before = json.loads(op.before_json or "{}")
                    pr = SkuPricing.query.get(sku)
                    if not pr:
                        pr = SkuPricing(sku=sku)
                        db.session.add(pr)

                    pr.brand = before.get("brand")
                    pr.name = before.get("name")
                    pr.spec_text = before.get("spec_text")
                    pr.stock_qty = before.get("stock_qty")
                    pr.stock_internal_qty = before.get("stock_internal_qty")
                    pr.monthly_sales_qty = before.get("monthly_sales_qty")
                    pr.cost = before.get("cost")
                    pr.our_price = before.get("our_price")
                    pr.floor_price = before.get("floor_price")
                    pr.min_margin_pct = before.get("min_margin_pct")
                    pr.pack_cost = before.get("pack_cost")
                    pr.ship_subsidy = before.get("ship_subsidy")

            elif t == "brand_controls":
                sku = pk
                if act == "insert":
                    bc = BrandControl.query.get(sku)
                    if bc:
                        db.session.delete(bc)
                elif act == "update":
                    before = json.loads(op.before_json or "{}")
                    bc = BrandControl.query.get(sku)
                    if not bc:
                        bc = BrandControl(sku=sku)
                        db.session.add(bc)
                    bc.brand = before.get("brand")
                    bc.name = before.get("name")
                    bc.price_control = before.get("price_control")

        batch.undone = True
        batch.undone_at = now_thai()
        batch.undone_by = undone_by
        db.session.commit()
        return batch

    @app.route("/import/price/internal/undo_last", methods=["POST"])
    @login_required
    def undo_price_internal_last():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("import_price_internal_view"))

        try:
            b = undo_last_price_import("internal", cu.username)
            flash(
                "✅ Undo สำเร็จ" if b else "ไม่พบรอบนำเข้าที่ Undo ได้",
                "success" if b else "warning",
            )
        except Exception as e:
            db.session.rollback()
            flash(f"Undo ไม่สำเร็จ: {e}", "danger")

        return redirect(url_for("import_price_internal_view"))

    @app.route("/import/price/market/undo_last", methods=["POST"])
    @login_required
    def undo_price_market_last():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("import_market_prices_view"))

        try:
            b = undo_last_price_import("market", cu.username)
            flash(
                "✅ Undo สำเร็จ" if b else "ไม่พบรอบนำเข้าที่ Undo ได้",
                "success" if b else "warning",
            )
        except Exception as e:
            db.session.rollback()
            flash(f"Undo ไม่สำเร็จ: {e}", "danger")

        return redirect(url_for("import_market_prices_view"))

    @app.route("/import/price/brand_control/undo_last", methods=["POST"])
    @login_required
    def undo_price_brand_last():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("import_brand_control_view"))

        try:
            b = undo_last_price_import("brand_control", cu.username)
            flash(
                "✅ Undo สำเร็จ" if b else "ไม่พบรอบนำเข้าที่ Undo ได้",
                "success" if b else "warning",
            )
        except Exception as e:
            db.session.rollback()
            flash(f"Undo ไม่สำเร็จ: {e}", "danger")

        return redirect(url_for("import_brand_control_view"))

    def _delete_price_import_history(kind: str):
        batch_ids = [b.id for b in PriceImportBatch.query.filter_by(kind=kind).all()]
        if batch_ids:
            PriceImportOp.query.filter(PriceImportOp.batch_id.in_(batch_ids)).delete(
                synchronize_session=False
            )
        PriceImportBatch.query.filter_by(kind=kind).delete(synchronize_session=False)

    @app.route("/import/price/internal/clear_all", methods=["POST"])
    @login_required
    def clear_price_internal_all():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("import_price_internal_view"))

        if (request.form.get("confirm_text") or "").strip() != "CLEAR INTERNAL":
            flash("พิมพ์คำยืนยันไม่ถูกต้อง", "warning")
            return redirect(url_for("import_price_internal_view"))

        try:
            _delete_price_import_history("internal")
            SkuPricing.query.delete(synchronize_session=False)
            db.session.commit()
            flash("✅ ล้าง Internal ทั้งหมดแล้ว", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"ล้างไม่สำเร็จ: {e}", "danger")

        return redirect(url_for("import_price_internal_view"))

    @app.route("/import/price/brand_control/clear_all", methods=["POST"])
    @login_required
    def clear_price_brand_all():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("import_brand_control_view"))

        if (request.form.get("confirm_text") or "").strip() != "CLEAR BRAND":
            flash("พิมพ์คำยืนยันไม่ถูกต้อง", "warning")
            return redirect(url_for("import_brand_control_view"))

        try:
            _delete_price_import_history("brand_control")
            BrandControl.query.delete(synchronize_session=False)
            db.session.commit()
            flash("✅ ล้าง Brand Control ทั้งหมดแล้ว", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"ล้างไม่สำเร็จ: {e}", "danger")

        return redirect(url_for("import_brand_control_view"))

    @app.route("/import/price/market/clear_all", methods=["POST"])
    @login_required
    def clear_price_market_all():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ใช้งาน", "danger")
            return redirect(url_for("import_market_prices_view"))

        if (request.form.get("confirm_text") or "").strip() != "CLEAR MARKET":
            flash("พิมพ์คำยืนยันไม่ถูกต้อง", "warning")
            return redirect(url_for("import_market_prices_view"))

        try:
            _delete_price_import_history("market")
            MarketPriceLog.query.delete(synchronize_session=False)
            MarketItem.query.delete(synchronize_session=False)
            db.session.commit()
            flash("✅ ล้าง Market ทั้งหมดแล้ว (ทั้ง Snapshot + Logs)", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"ล้างไม่สำเร็จ: {e}", "danger")

        return redirect(url_for("import_market_prices_view"))

    @app.route("/import/price/brand_control/template", methods=["GET"])
    @login_required
    def download_price_brand_control_template():
        # Use Dashboard-friendly headers (Importer still supports legacy snake_case)
        df = pd.DataFrame(
            [
                {
                    "SKU": "",
                    "Brand": "",
                    "Name": "",
                    "Brand Control": "",
                }
            ]
        )

        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="BrandControl")
        bio.seek(0)

        return send_file(
            bio,
            as_attachment=True,
            download_name="template_brand_control.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/admin/users", methods=["GET", "POST"])
    @login_required
    def admin_users():
        cu = current_user()
        if cu.role != "admin":
            flash("ต้องเป็นผู้ดูแลระบบเท่านั้น", "danger")
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            action = request.form.get("action")
            if action == "create":
                username = request.form.get("username").strip()
                password = request.form.get("password")
                role = request.form.get("role", "user")
                if not username or not password:
                    flash("กรุณากรอกชื่อผู้ใช้/รหัสผ่าน", "danger")
                elif User.query.filter_by(username=username).first():
                    flash("มีชื่อผู้ใช้นี้อยู่แล้ว", "warning")
                else:
                    u = User(
                        username=username,
                        password_hash=generate_password_hash(password),
                        role=role,
                        active=True
                    )
                    db.session.add(u)
                    db.session.commit()
                    flash(f"สร้างผู้ใช้ {username} แล้ว", "success")
            elif action == "delete":
                uid = int(request.form.get("uid"))
                if uid == cu.id:
                    flash("ลบตัวเองไม่ได้", "warning")
                else:
                    User.query.filter_by(id=uid).delete()
                    db.session.commit()
                    flash("ลบผู้ใช้แล้ว", "success")
        users = User.query.order_by(User.created_at.desc()).all() if hasattr(User, "created_at") else User.query.all()
        return render_template("users.html", users=users)

    # -------------
    # Dashboard
    # -------------
    @app.route("/")
    @login_required
    def dashboard():
        platform = normalize_platform(request.args.get("platform"))
        shop_id = request.args.get("shop_id")
        show_change = (request.args.get("show_change") or "").strip().upper()  # [NEW] Filter: สถานะเปลี่ยน (Warehouse Receive)
        
        # [แก้ไข] เปลี่ยนจาก import_date เดี่ยว เป็น Range
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        status = request.args.get("status")
        q = (request.args.get("q") or "").strip()  # รับค่าคำค้นหา Global Search
        all_time = request.args.get("all_time")  # Flag สำหรับดูทั้งหมด
        mode = request.args.get("mode")  # [NEW] โหมด Order ปัจจุบัน (today)

        shops = Shop.query.order_by(Shop.name.asc()).all()

        # แปลงวันที่
        def _p(s): return parse_date_any(s)
        
        imp_from = _p(import_from_str)
        imp_to = _p(import_to_str)
        d_from = datetime.combine(_p(date_from), datetime.min.time(), tzinfo=TH_TZ) if date_from else None
        d_to = datetime.combine(_p(date_to) + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if date_to else None

        # ตรวจสอบว่ามี Filter วันที่หรือไม่
        has_date_filter = bool(imp_from or imp_to or d_from or d_to)
        
        # ตรวจสอบโหมด All Time (Lock)
        is_all_time = bool(all_time)

        # กำหนด Strategy การดึงข้อมูล
        use_default_view = False
        rows = []

        # Base Filters (Platform/Shop)
        base_filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
        }

        if is_all_time:
            # CASE 1: All Time -> ดึงข้อมูลทั้งหมด ไม่สนวันที่
            # active_only=False เพื่อให้ดึง Packed/Cancelled ด้วย
            filters = base_filters.copy()
            filters["active_only"] = False 
            filters["all_time"] = True
            rows, _ = compute_allocation(db.session, filters)

        elif mode == 'today':
            # [NEW] CASE 1.5: Order ปัจจุบัน (วันนี้)
            # กรองเฉพาะ Import Date = วันนี้ + Order ที่ยกเลิกวันนี้
            today = now_thai().date()
            
            # 1. ดึง Order ที่นำเข้าวันนี้
            filters = base_filters.copy()
            filters["active_only"] = False
            filters["import_from"] = today
            filters["import_to"] = today
            rows_import, _ = compute_allocation(db.session, filters)
            
            # 2. ดึง Order ที่ "ยกเลิกวันนี้" (บวก 7 ชม. เพื่อให้ตรงกับเวลาไทย)
            cancel_today_oids = [
                r[0] for r in db.session.query(CancelledOrder.order_id)
                .filter(func.date(CancelledOrder.imported_at, '+7 hours') == today).all()
            ]
            
            rows_cancel = []
            if cancel_today_oids:
                # ดึงข้อมูลของ Order ที่ cancel วันนี้ (ใช้ all_time แล้ว filter เอาเฉพาะ ID)
                f_cancel = base_filters.copy()
                f_cancel["all_time"] = True
                f_cancel["active_only"] = False
                temp_rows, _ = compute_allocation(db.session, f_cancel)
                rows_cancel = [r for r in temp_rows if r.get("order_id") in cancel_today_oids]
            
            # 3. รวมรายการ (ตัดตัวซ้ำด้วย id)
            seen_ids = set()
            rows = []
            for r in (rows_import + rows_cancel):
                rid = r.get("id")
                if rid not in seen_ids:
                    rows.append(r)
                    seen_ids.add(rid)

        elif has_date_filter:
            # CASE 2: มีการเลือกช่วงเวลา (Import Date หรือ Order Date) -> ดึงตามช่วงเวลานั้น
            filters = base_filters.copy()
            filters["active_only"] = False
            filters["import_from"] = imp_from
            filters["import_to"] = imp_to
            filters["date_from"] = d_from
            filters["date_to"] = d_to
            rows, _ = compute_allocation(db.session, filters)
            
        else:
            # CASE 3: Default View (ไม่มี Filter วันที่ และไม่ใช่ All Time)
            use_default_view = True
            
            # 3.1 ดึง Order ค้างทั้งหมด (Active Orders - All Time)
            f_active = base_filters.copy()
            f_active["active_only"] = True
            rows_active, _ = compute_allocation(db.session, f_active)
            
            # 3.2 ดึง Order จบแล้ว (Packed/Cancelled) ของ "วันนี้" เท่านั้น
            today = now_thai().date()
            f_inactive = base_filters.copy()
            f_inactive["active_only"] = False
            f_inactive["import_from"] = today  # เฉพาะวันนี้
            f_inactive["import_to"] = today
            
            rows_today_all, _ = compute_allocation(db.session, f_inactive)
            
            # คัดเฉพาะ Packed/Cancelled จากของวันนี้
            existing_ids = set(r["id"] for r in rows_active)
            rows = list(rows_active)
            
            for r in rows_today_all:
                if r["id"] not in existing_ids:
                    # ถ้าไม่อยู่ใน Active แสดงว่าเป็น Packed หรือ Cancelled
                    if r.get("is_packed") or r.get("is_cancelled"):
                         rows.append(r)

        # --- Post-Processing Rows ---
        # ดึงเซ็ต/แมป Order ยกเลิก/จ่ายแล้ว/แพ็คแล้ว
        cancelled_map = _cancelled_oids_map()  # dict: order_id -> note
        packed_oids = _orders_packed_set(rows)
        orders_not_in_sbs = _orders_not_in_sbs_set(rows)
        orders_no_sales = _orders_no_sales_set(rows)
        
        # [NEW] ดึง Order ที่ถูกลบ (Soft Delete)
        # หมายเหตุ: สำหรับ Order ที่ "จ่ายงานแล้ว" เราจะยังคงแสดงใน Dashboard ได้
        # เพื่อให้ Warehouse เห็นว่า "สถานะเปลี่ยน" (เช่น ถูกลบ/ยกเลิก) หลังจ่ายงาน
        deleted_oids = _deleted_oids_set()
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            r["is_deleted"] = bool(oid and oid in deleted_oids)
        # กรองออกเฉพาะ Order ที่ถูกลบและยังไม่ได้จ่ายงาน (ไม่ต้องโชว์ในตารางหลัก)
        rows = [
            r for r in rows
            if not (r.get("is_deleted") and not r.get("is_issued"))
        ]

        # Inject scan status (scanned_at) at order level
        _inject_scan_status(rows)
        for r in rows:
            r["is_scanned"] = bool(r.get("scanned_at"))
        # Process Row Attributes
        totals = _build_allqty_map(rows)
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            
            # เติม stock
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty

            r["allqty"] = int(totals.get((r.get("sku") or "").strip(), r.get("qty", 0)) or 0)
            r["accepted"] = bool(r.get("accepted", False))
            r["sales_status"] = r.get("sales_status", None)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            
            # Set Flags
            r["is_cancelled"] = False
            r["is_not_in_sbs"] = False
            r["packed"] = False
            r["cancel_reason"] = ""  # [NEW] เหตุผลการยกเลิก
            r["cancel_at"] = ""      # [NEW] เวลาที่ยกเลิก (สำหรับแสดงใน HTML)
            r["cancel_str"] = ""     # [NEW] ข้อความรวมสำหรับ Excel
            
            # [NEW] เช็คว่า Order นี้เคยแพ็คแล้วหรือยัง (ก่อนถูกยกเลิก)
            r["was_packed"] = (oid in packed_oids)

            if oid in cancelled_map:
                r["allocation_status"] = "CANCELLED"
                r["is_cancelled"] = True
                
                # [NEW] แกะข้อมูล Note และ Time จาก dict ซ้อน
                c_info = cancelled_map[oid]
                note_txt = c_info.get('note', '')
                time_obj = c_info.get('at')
                
                # จัด Format เวลา (แปลงเป็น พ.ศ.)
                time_str = ""
                if time_obj:
                    try:
                        # แปลงเป็น พ.ศ. ถ้าปียังเป็น ค.ศ.
                        if time_obj.year < 2400:
                            time_obj_be = time_obj.replace(year=time_obj.year + 543)
                        else:
                            time_obj_be = time_obj
                        time_str = time_obj_be.strftime("%d/%m/%Y %H:%M")
                    except Exception:
                        pass
                
                r["cancel_reason"] = note_txt
                r["cancel_at"] = time_str  # ส่งไปโชว์ใน HTML
                r["cancel_str"] = f"{note_txt} [เมื่อ: {time_str}]" if time_str else note_txt  # ส่งไป Excel
                
                r["actions_disabled"] = True
            elif oid in packed_oids:
                r["allocation_status"] = "PACKED"
                r["packed"] = True
                r["actions_disabled"] = True
            else:
                r["actions_disabled"] = False
                if oid in orders_not_in_sbs:
                    r["is_not_in_sbs"] = True

        # --- STEP 3: คำนวณ KPI จากข้อมูลทั้งหมดใน Scope (ก่อนถูกซ่อนจากตาราง) ---
        # [แก้ไขจุดที่ 2] ใช้ rows (ซึ่งคือข้อมูลทั้งหมดใน Scope นี้) คำนวณ KPI เลย
        # เพื่อให้ปุ่ม Packed/Cancelled แสดงยอดได้ถูกต้อง แม้ตารางจะไม่ได้โชว์
        
        scope_rows = list(rows)  # สำรองข้อมูลไว้คำนวณ KPI
        
        # Helper lists for KPI counts from scope
        kpi_orders_ready = _orders_ready_set(scope_rows)
        
        # +++ [เพิ่ม] เก็บรายชื่อ Order ที่พร้อมจริงๆ (จากข้อมูลทั้งหมด) เอาไว้ใช้คุมปุ่มกดรับ
        # เพราะเดี๋ยว kpi_orders_ready จะถูกเขียนทับถ้ามีการ Search
        global_ready_oids = kpi_orders_ready.copy()
        
        kpi_orders_low = _orders_lowstock_order_set(scope_rows)
        kpi_orders_nosales = _orders_no_sales_set(scope_rows)
        kpi_orders_not_in_sbs = _orders_not_in_sbs_set(scope_rows)
        
        # [แก้ไข] ลบ Order ที่ยกเลิกออกจาก KPI "ยังไม่มีใบขาย" และ "ยังไม่เข้า SBS"
        # เพื่อไม่ให้ยอดเด้งทั้งที่ยกเลิกไปแล้ว
        cancelled_all_ids = set(cancelled_map.keys())
        kpi_orders_nosales = kpi_orders_nosales - cancelled_all_ids
        kpi_orders_not_in_sbs = kpi_orders_not_in_sbs - cancelled_all_ids
        
        # [NEW] คำนวณ Set ของ Order ที่เป็น "ไม่มีสินค้า" หรือ "สินค้าไม่พอส่ง"
        # ใช้ Set เพื่อให้เลข Order ไม่ซ้ำกัน
        kpi_orders_problem = set()
        for r in scope_rows:
            # [แก้ไข] เพิ่มเงื่อนไข: ต้องยังไม่จ่ายงาน (is_issued) ด้วย ถึงจะนับเข้ากอง 3
            if not r.get("packed") and not r.get("is_cancelled") and not r.get("is_issued"):
                status_alloc = (r.get("allocation_status") or "").strip().upper()
                if status_alloc in ("SHORTAGE", "NOT_ENOUGH"):
                    oid = (r.get("order_id") or "").strip()
                    if oid:
                        kpi_orders_problem.add(oid)

        # [NEW] คำนวณ Set ของ Order ที่เป็น "บิลเปล่า" (BILL_EMPTY)
        # [สำคัญ] บิลเปล่านับแยกต่างหาก ไม่กรอง packed/cancelled เพราะต้องการติดตามทั้งหมด
        kpi_orders_bill_empty = set()
        bill_empty_count_debug = 0
        for r in scope_rows:
            status_alloc = (r.get("allocation_status") or "").strip().upper()
            if status_alloc == "BILL_EMPTY":
                bill_empty_count_debug += 1
                oid = (r.get("order_id") or "").strip()
                if oid:
                    kpi_orders_bill_empty.add(oid)

        # [DEBUG] แสดงจำนวนแถวที่เจอ BILL_EMPTY
        app.logger.info(f"[BILL_EMPTY DEBUG] พบ {bill_empty_count_debug} แถวที่มี allocation_status='BILL_EMPTY' ใน scope_rows (total={len(scope_rows)} rows)")
        app.logger.info(f"[BILL_EMPTY DEBUG] Order IDs ที่เป็น BILL_EMPTY: {sorted(kpi_orders_bill_empty)}")

        # ===== Scan (Barcode) KPI Sets =====
        def _active_oids(source_rows: list[dict]) -> set[str]:
            return {
                (r.get("order_id") or "").strip()
                for r in source_rows
                if r.get("order_id") and not r.get("packed") and not r.get("is_cancelled")
            }

        kpi_active_oids = _active_oids(scope_rows)
        kpi_orders_scanned = {
            (r.get("order_id") or "").strip()
            for r in scope_rows
            if r.get("order_id")
            and not r.get("packed")
            and not r.get("is_cancelled")
            and r.get("scanned_at")
        }
        kpi_orders_not_scanned = kpi_active_oids - kpi_orders_scanned

        # ===== Warehouse Receive (Issued but Not Packed) KPI Sets =====
        def _compute_wh_receive_sets(source_rows: list[dict]):
            issued_active_oids = {
                (r.get("order_id") or "").strip()
                for r in source_rows
                if r.get("order_id")
                and r.get("is_issued")
                and not r.get("packed")
            }
            if not issued_active_oids:
                return {
                    "total": set(),
                    "g1": set(),
                    "g2": set(),
                    "g3": set(),
                    "issued_date": {},
                    "src": {},
                }

            issued_rows = (
                db.session.query(IssuedOrder.order_id, IssuedOrder.source, IssuedOrder.issued_at)
                .filter(IssuedOrder.order_id.in_(issued_active_oids))
                .all()
            )
            src_map: dict[str, str | None] = {str(r[0]): (r[1] if r and r[0] else None) for r in issued_rows}
            date_map: dict[str, date] = {}
            for r in issued_rows:
                if not r or not r[0] or not r[2]:
                    continue
                try:
                    dt = r[2]
                    if isinstance(dt, datetime):
                        date_map[str(r[0])] = dt.date()
                except Exception:
                    pass

            g1, g2, g3 = set(), set(), set()
            for oid in issued_active_oids:
                src = (src_map.get(oid) or "").strip().lower()
                # กอง 2: สินค้าน้อย
                if src == "print:lowstock":
                    g2.add(oid)
                # กอง 3: ไม่มีสินค้า + ไม่พอส่ง
                elif src in {"print:nostock", "print:notenough"}:
                    g3.add(oid)
                # กอง 1: ใบงานคลัง + Picking list (รวม manual/import/unknown)
                else:
                    g1.add(oid)

            total = set().union(g1, g2, g3)
            return {"total": total, "g1": g1, "g2": g2, "g3": g3, "issued_date": date_map, "src": src_map}

        wh_sets = _compute_wh_receive_sets(scope_rows)
        wh_total_oids = wh_sets["total"]
        wh_g1_oids = wh_sets["g1"]
        wh_g2_oids = wh_sets["g2"]
        wh_g3_oids = wh_sets["g3"]
        wh_issued_date_map = wh_sets["issued_date"]
        wh_src_map: dict[str, str | None] = wh_sets.get("src", {})

        # ===== [NEW] Status Change (Issued -> Current) =====
        # เป้าหมาย: งานที่จ่ายแล้ว (Issued) ยังอยู่กองเดิม (ตาม source)
        # แต่ต้องการเห็นว่า "ตอนนี้" มีการเปลี่ยนแปลงจากกองตอนจ่ายหรือไม่
        def _compute_wh_status_change(
            source_rows: list[dict],
            wh_total: set[str],
            wh_g1: set[str],
            wh_g2: set[str],
            wh_g3: set[str],
            src_map: dict[str, str | None],
        ):
            def _src_to_group(src: str | None) -> str:
                s = (src or "").strip().lower()
                if s == "print:lowstock":
                    return "G2"
                if s in {"print:nostock", "print:notenough"}:
                    return "G3"
                return "G1"

            def _current_group(agg: dict) -> str:
                if agg.get("is_deleted"):
                    return "DELETED"
                if agg.get("is_cancelled"):
                    return "CANCELLED"
                if agg.get("has_g3"):
                    return "G3"
                if agg.get("has_g2"):
                    return "G2"
                return "G1"

            wh_oid_set = set(wh_total)

            # Aggregate current status at ORDER level (worst-case among lines)
            wh_agg: dict[str, dict] = {}
            for r in source_rows:
                oid = (r.get("order_id") or "").strip()
                if not oid or oid not in wh_oid_set:
                    continue

                a = wh_agg.get(oid)
                if a is None:
                    a = {
                        "is_cancelled": False,
                        "is_deleted": False,
                        "has_g2": False,
                        "has_g3": False,
                    }
                    wh_agg[oid] = a

                if r.get("is_cancelled"):
                    a["is_cancelled"] = True
                if r.get("is_deleted"):
                    a["is_deleted"] = True

                st = (r.get("allocation_status") or "").strip().upper()
                try:
                    stock_qty = int(r.get("stock_qty") or 0)
                except Exception:
                    stock_qty = 0
                try:
                    qty = int(r.get("qty") or 0)
                except Exception:
                    qty = 0

                # G3: ของหมด/ไม่พอส่ง (priority สูงกว่า)
                if st in {"SHORTAGE", "NOT_ENOUGH"} or stock_qty <= 0:
                    a["has_g3"] = True
                # G2: สินค้าน้อย
                elif st == "LOW_STOCK" or (qty > 0 and stock_qty < qty):
                    a["has_g2"] = True

            status_change_map: dict[str, dict] = {}
            changed_total: set[str] = set()
            changed_g1: set[str] = set()
            changed_g2: set[str] = set()
            changed_g3: set[str] = set()

            for oid in wh_oid_set:
                src_group = _src_to_group(src_map.get(oid))
                cur_group = _current_group(wh_agg.get(oid) or {})

                change_info = None
                if cur_group == "DELETED":
                    change_info = {"label": "ถูกลบแล้ว", "cls": "bg-dark", "icon": "bi-trash"}
                elif cur_group == "CANCELLED":
                    change_info = {"label": "ยกเลิกแล้ว", "cls": "bg-danger", "icon": "bi-x-circle"}
                elif cur_group != src_group:
                    # improved (good news)
                    if src_group == "G3" and cur_group in {"G1", "G2"}:
                        change_info = {"label": f"ของเข้าแล้ว ({src_group}→{cur_group})", "cls": "bg-success", "icon": "bi-arrow-up"}
                    # worsened (bad news)
                    elif cur_group == "G3" and src_group in {"G1", "G2"}:
                        change_info = {"label": f"ของหมด/ไม่พอ ({src_group}→{cur_group})", "cls": "bg-danger", "icon": "bi-arrow-down"}
                    else:
                        change_info = {"label": f"สถานะเปลี่ยน ({src_group}→{cur_group})", "cls": "bg-warning text-dark", "icon": "bi-arrow-left-right"}

                if change_info:
                    status_change_map[oid] = change_info
                    changed_total.add(oid)
                    if oid in wh_g1:
                        changed_g1.add(oid)
                    if oid in wh_g2:
                        changed_g2.add(oid)
                    if oid in wh_g3:
                        changed_g3.add(oid)

            return status_change_map, changed_total, changed_g1, changed_g2, changed_g3

        # คำนวณครั้งแรก (ก่อนจะมีการ recalculates ตอน Search)
        status_change_map, changed_oids_total, changed_oids_g1, changed_oids_g2, changed_oids_g3 = _compute_wh_status_change(
            scope_rows,
            wh_total_oids,
            wh_g1_oids,
            wh_g2_oids,
            wh_g3_oids,
            wh_src_map,
        )

        # Inject order-level status_change into each line row (initial)
        wh_oid_set = set(wh_total_oids)
        for r in scope_rows:
            oid = (r.get("order_id") or "").strip()
            if oid and oid in wh_oid_set:
                r["status_change"] = status_change_map.get(oid)
        
        # Packed Sets จาก Scope
        kpi_packed_oids = set(r.get("order_id") for r in scope_rows if r.get("packed"))
        
        # [NEW] แยก KPI Order ยกเลิก เป็น 2 กลุ่ม
        # 1. ยกเลิกก่อนแพ็ค (ไม่เคยแพ็ค)
        # 2. ยกเลิกหลังแพ็ค (เคยแพ็คแล้ว)
        kpi_cancel_nopack = set()  # ยกเลิก (ก่อนแพ็ค)
        kpi_cancel_packed = set()  # ยกเลิก (หลังแพ็ค)
        
        for r in scope_rows:
            if r.get("is_cancelled"):
                oid = (r.get("order_id") or "").strip()
                if oid:
                    if r.get("was_packed"):
                        kpi_cancel_packed.add(oid)
                    else:
                        kpi_cancel_nopack.add(oid)

        # --- STEP 4: กรองข้อมูลเพื่อแสดงผลในตาราง (Filtering View) ---
        status_norm = (status or "").strip().upper()
        
        # กรณีที่ 1: มีการค้นหา (Global Search)
        if q:
            q_lower = q.lower()
            filtered_rows = []
            for r in rows:
                search_text = (
                    str(r.get("order_id") or "") + " " +
                    str(r.get("sku") or "") + " " +
                    str(r.get("brand") or "") + " " +
                    str(r.get("model") or "") + " " +
                    str(r.get("shop") or "") + " " +
                    str(r.get("sales_status") or "")
                ).lower()
                if q_lower in search_text:
                    filtered_rows.append(r)
            rows = filtered_rows
            
            # [เพิ่มเติม] ถ้ามีการค้นหา ให้ KPI นับตามผลการค้นหาด้วย
            scope_rows = rows
            
            # Recalculate sets for filtered scope (กรณี search)
            kpi_orders_ready = _orders_ready_set(scope_rows)
            kpi_orders_low = _orders_lowstock_order_set(scope_rows)
            kpi_orders_nosales = _orders_no_sales_set(scope_rows)
            kpi_orders_not_in_sbs = _orders_not_in_sbs_set(scope_rows)
            
            # [แก้ไข] ลบ Order ที่ยกเลิกออกจาก KPI "ยังไม่มีใบขาย" และ "ยังไม่เข้า SBS" (กรณี search)
            cancelled_all_ids = set(cancelled_map.keys())
            kpi_orders_nosales = kpi_orders_nosales - cancelled_all_ids
            kpi_orders_not_in_sbs = kpi_orders_not_in_sbs - cancelled_all_ids
            
            kpi_packed_oids = set(r.get("order_id") for r in scope_rows if r.get("packed"))
            
            # Recalculate kpi_cancel_nopack / kpi_cancel_packed for search
            kpi_cancel_nopack = set()
            kpi_cancel_packed = set()
            for r in scope_rows:
                if r.get("is_cancelled"):
                    oid = (r.get("order_id") or "").strip()
                    if oid:
                        if r.get("was_packed"):
                            kpi_cancel_packed.add(oid)
                        else:
                            kpi_cancel_nopack.add(oid)
            
            # Recalculate kpi_orders_problem for search
            kpi_orders_problem = set()
            for r in scope_rows:
                # [แก้ไข] เพิ่มเงื่อนไข: ต้องยังไม่จ่ายงาน (is_issued) ถึงจะนับเป็นงานค้างกอง 3
                if not r.get("packed") and not r.get("is_cancelled") and not r.get("is_issued"):
                    status_alloc = (r.get("allocation_status") or "").strip().upper()
                    if status_alloc in ("SHORTAGE", "NOT_ENOUGH"):
                        oid = (r.get("order_id") or "").strip()
                        if oid:
                            kpi_orders_problem.add(oid)

            # Recalculate kpi_orders_bill_empty for search
            # [สำคัญ] บิลเปล่านับแยกต่างหาก ไม่กรอง packed/cancelled เพราะต้องการติดตามทั้งหมด
            kpi_orders_bill_empty = set()
            for r in scope_rows:
                status_alloc = (r.get("allocation_status") or "").strip().upper()
                if status_alloc == "BILL_EMPTY":
                    oid = (r.get("order_id") or "").strip()
                    if oid:
                        kpi_orders_bill_empty.add(oid)

            # Recalculate scan sets for search scope
            kpi_active_oids = _active_oids(scope_rows)
            kpi_orders_scanned = {
                (r.get("order_id") or "").strip()
                for r in scope_rows
                if r.get("order_id")
                and not r.get("packed")
                and not r.get("is_cancelled")
                and r.get("scanned_at")
            }
            kpi_orders_not_scanned = kpi_active_oids - kpi_orders_scanned

            # Recalculate warehouse-receive sets for search scope
            wh_sets = _compute_wh_receive_sets(scope_rows)
            wh_total_oids = wh_sets["total"]
            wh_g1_oids = wh_sets["g1"]
            wh_g2_oids = wh_sets["g2"]
            wh_g3_oids = wh_sets["g3"]
            wh_issued_date_map = wh_sets["issued_date"]
            wh_src_map = wh_sets.get("src", {})

            # Recalculate status-change for search scope
            status_change_map, changed_oids_total, changed_oids_g1, changed_oids_g2, changed_oids_g3 = _compute_wh_status_change(
                scope_rows,
                wh_total_oids,
                wh_g1_oids,
                wh_g2_oids,
                wh_g3_oids,
                wh_src_map,
            )

            wh_oid_set = set(wh_total_oids)
            for r in scope_rows:
                oid = (r.get("order_id") or "").strip()
                if oid and oid in wh_oid_set:
                    r["status_change"] = status_change_map.get(oid)

        # กรณีที่ 2: ไม่ได้ค้นหา -> ใช้ Logic การกรองตามสถานะ
        else:
            if status_norm == "TOTAL":
                # [NEW] ถ้าเลือก "รวม Order" ให้แสดงทุกอย่างใน Scope (ไม่ซ่อน Packed/Cancelled)
                pass
            elif status_norm == "ORDER_CANCELLED":
                # [แก้ไข] กรองเฉพาะยกเลิกที่ยังไม่เคยแพ็ค (ก่อนแพ็ค)
                rows = [r for r in rows if r.get("is_cancelled") and not r.get("was_packed")]
            elif status_norm == "ORDER_CANCELLED_PACKED":
                # [NEW] กรองเฉพาะยกเลิกหลังแพ็ค (เคยแพ็คแล้ว)
                rows = [r for r in rows if r.get("is_cancelled") and r.get("was_packed")]
            elif status_norm == "ORDER_NOT_IN_SBS":
                rows = [r for r in rows if r.get("is_not_in_sbs")]
            elif status_norm in {"ORDER_NOT_SCANNED", "ORDER_SCAN_BARCODE"}:
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_not_scanned]
            elif status_norm == "WH_RECEIVE_TOTAL":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in wh_total_oids]
            elif status_norm == "WH_RECEIVE_G1":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in wh_g1_oids]
            elif status_norm == "WH_RECEIVE_G2":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in wh_g2_oids]
            elif status_norm == "WH_RECEIVE_G3":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in wh_g3_oids]
            elif status_norm == "ORDER_PROBLEM":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_problem]
            elif status_norm == "PACKED":
                rows = [r for r in rows if r.get("packed")]
            elif status_norm == "ORDER_READY":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_ready]
            elif status_norm in {"ORDER_LOW_STOCK", "ORDER_LOW"}:
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_low]
            elif status_norm == "ORDER_NO_SALES":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_nosales]
            elif status_norm == "BILL_EMPTY":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_bill_empty]
            elif status_norm:
                # กรองสถานะรายบรรทัด
                rows = [r for r in rows if (r.get("allocation_status") or "").strip().upper() == status_norm]
            else:
                # Default Table View: ซ่อน Packed/Cancelled ออกจากตารางหลัก
                # แต่ข้อมูลใน scope_rows ยังอยู่ครบ ทำให้ KPI ไม่เป็น 0
                if not status:
                     rows = [r for r in rows if not r.get("packed") and not r.get("is_cancelled")]

        # --- STEP 4.5: ใส่ "สถานะตามการ์ด (KPI Cards)" ลงในแต่ละแถว เพื่อให้คอลัมน์สถานะอ่านเข้าใจ ---
        # หมายเหตุ: การ์ดด้านบนเป็นการจัดกลุ่ม "ระดับ Order" แต่ตารางแสดง "ระดับบรรทัดสินค้า (SKU)"
        # ดังนั้นเราจะใส่ badge เพิ่มให้เห็นว่าบรรทัดนี้อยู่ในกลุ่มไหนของการ์ดบ้าง (เช่น กอง 1/2/3, ไม่มีใบขาย, ไม่เข้า SBS)
        def _build_card_tags(oid: str) -> list:
            tags = []
            if not oid:
                return tags

            # กลุ่มงานค้าง (Pending Tasks)
            if oid in kpi_orders_ready:
                tags.append({"code": "ORDER_READY", "label": "กอง 1", "cls": "bg-success"})
            if oid in kpi_orders_low:
                tags.append({"code": "ORDER_LOW_STOCK", "label": "กอง 2", "cls": "bg-warning text-dark"})
            if oid in kpi_orders_problem:
                tags.append({"code": "ORDER_PROBLEM", "label": "กอง 3", "cls": "bg-danger"})
            if oid in kpi_orders_nosales:
                tags.append({"code": "ORDER_NO_SALES", "label": "ยังไม่แพ็ค", "cls": "bg-light text-dark border border-secondary"})
            if oid in kpi_orders_not_in_sbs:
                tags.append({"code": "ORDER_NOT_IN_SBS", "label": "ยังไม่นำเข้าSBS", "cls": "bg-light text-dark border border-secondary"})

            # กลุ่มงานจบ (Completed Today) — เผื่อหน้า Total/ค้นหา ทำให้ Packed/Cancelled โผล่มาในตาราง
            if oid in kpi_packed_oids:
                tags.append({"code": "PACKED", "label": "แพ็คแล้ว", "cls": "bg-dark"})
            if oid in kpi_cancel_nopack:
                tags.append({"code": "ORDER_CANCELLED", "label": "ยกเลิก(ก่อนแพ็ค)", "cls": "bg-secondary"})
            if oid in kpi_cancel_packed:
                tags.append({"code": "ORDER_CANCELLED_PACKED", "label": "ยกเลิก(หลังแพ็ค)", "cls": "bg-secondary"})

            return tags

        for r in rows:
            oid = (r.get("order_id") or "").strip()
            r["card_tags"] = _build_card_tags(oid)

        # --- STEP 5: สร้าง Dict KPI ---
        
        # --- [NEW LOGIC] แยกนับยอด เก่า (Old) vs วันนี้ (Today) ---
        today_date = now_thai().date()

        def _count_split(oid_set, source_rows):
            """ฟังก์ชันช่วยนับ: คืนค่า (total, old_count, today_count)"""
            total = len(oid_set)
            old_c = 0
            today_c = 0
            
            # สร้าง Dict เพื่อ map order_id -> import_date จาก source_rows
            oid_date_map = {}
            for r in source_rows:
                if r.get("order_id"):
                    d = r.get("import_date")
                    # Fallback: ถ้าไม่มี import_date ให้ใช้วันที่สั่ง
                    if not d and r.get("order_time"):
                        if isinstance(r["order_time"], datetime):
                            d = r["order_time"].date()
                    oid_date_map[r["order_id"]] = d
            
            for oid in oid_set:
                d = oid_date_map.get(oid)
                # ตรวจสอบว่าเป็นเก่าหรือใหม่
                is_old = True
                if d:
                    # แปลงเป็น date object ถ้าจำเป็น
                    if isinstance(d, datetime):
                        d = d.date()
                    elif isinstance(d, str):
                        try:
                            d = datetime.strptime(d, "%Y-%m-%d").date()
                        except:
                            d = today_date

                    if d >= today_date:
                        is_old = False
                
                if is_old:
                    old_c += 1
                else:
                    today_c += 1
                    
            return total, old_c, today_c

        def _count_split_by_issued_at(oid_set: set[str], oid_issued_date_map: dict[str, date]):
            total = len(oid_set)
            old_c = 0
            today_c = 0
            for oid in oid_set:
                d = oid_issued_date_map.get(oid)
                if d and d >= today_date:
                    today_c += 1
                else:
                    old_c += 1
            return total, old_c, today_c

        # คำนวณยอดแยกของแต่ละกอง
        c_ready, c_ready_old, c_ready_new = _count_split(kpi_orders_ready, scope_rows)
        c_low, c_low_old, c_low_new = _count_split(kpi_orders_low, scope_rows)
        c_prob, c_prob_old, c_prob_new = _count_split(kpi_orders_problem, scope_rows)
        c_bill_empty, c_bill_empty_old, c_bill_empty_new = _count_split(kpi_orders_bill_empty, scope_rows)

        # [DEBUG] แสดงข้อมูล KPI บิลเปล่าใน console
        app.logger.info(f"[BILL_EMPTY DEBUG] kpi_orders_bill_empty = {kpi_orders_bill_empty}")
        app.logger.info(f"[BILL_EMPTY DEBUG] c_bill_empty={c_bill_empty}, old={c_bill_empty_old}, new={c_bill_empty_new}")

        # [NEW] แยกเก่า/ใหม่ ให้กับ "ไม่มีใบขาย" และ "ไม่เข้า SBS"
        c_nosale, c_nosale_old, c_nosale_new = _count_split(kpi_orders_nosales, scope_rows)
        c_nosbs, c_nosbs_old, c_nosbs_new = _count_split(kpi_orders_not_in_sbs, scope_rows)
        
        # รวม Order ค้างทั้งหมด (Unique) - กรองเฉพาะ active
        active_oids_set = set(kpi_active_oids)
        c_pending, c_pending_old, c_pending_new = _count_split(active_oids_set, scope_rows)

        c_not_scanned, c_not_scanned_old, c_not_scanned_new = _count_split(kpi_orders_not_scanned, scope_rows)

        # Warehouse Receive split (based on issued_at)
        c_wh_total, c_wh_total_old, c_wh_total_new = _count_split_by_issued_at(wh_total_oids, wh_issued_date_map)
        c_wh_g1, c_wh_g1_old, c_wh_g1_new = _count_split_by_issued_at(wh_g1_oids, wh_issued_date_map)
        c_wh_g2, c_wh_g2_old, c_wh_g2_new = _count_split_by_issued_at(wh_g2_oids, wh_issued_date_map)
        c_wh_g3, c_wh_g3_old, c_wh_g3_new = _count_split_by_issued_at(wh_g3_oids, wh_issued_date_map)
        
        # [NEW] นับถังขยะ (Deleted) เฉพาะของวันนี้ (ถ้าเป็น Default View)
        # ใช้ func.date เพื่อเทียบแค่วันที่ (ตัดเวลาออก)
        deleted_today_count = db.session.query(func.count(DeletedOrder.id)).filter(
            func.date(DeletedOrder.deleted_at) == today_date
        ).scalar() or 0

        kpis = {
            "total_items": len(scope_rows),
            "total_qty": sum(int(r.get("qty", 0) or 0) for r in scope_rows),
            
            # [แก้ไข] ปรับสูตรนับ "รวม Order":
            # - ถ้ามีการเลือกสถานะ (status ไม่ว่าง) -> นับจาก rows (ตามที่กรอง)
            # - ถ้าไม่มีการเลือกสถานะ (หน้า All Time/ปกติ) -> นับจาก scope_rows (รวม Packed/Cancel)
            "orders_total": len(set(
                r.get("order_id") for r in (rows if status else scope_rows) 
                if r.get("order_id")
            )),
            
            # --- กลุ่มงานค้าง (ใช้ยอดแยก เก่า/ใหม่) ---
            "orders_unique": c_pending,
            "orders_unique_old": c_pending_old,
            "orders_unique_new": c_pending_new,

            "orders_ready": c_ready,
            "orders_ready_old": c_ready_old,
            "orders_ready_new": c_ready_new,

            "orders_low": c_low,
            "orders_low_old": c_low_old,
            "orders_low_new": c_low_new,

            "orders_problem": c_prob,
            "orders_problem_old": c_prob_old,
            "orders_problem_new": c_prob_new,

            "orders_nosales": c_nosale,
            "orders_nosales_old": c_nosale_old,
            "orders_nosales_new": c_nosale_new,

            "orders_not_in_sbs": c_nosbs,
            "orders_not_in_sbs_old": c_nosbs_old,
            "orders_not_in_sbs_new": c_nosbs_new,

            "orders_bill_empty": c_bill_empty,
            "orders_bill_empty_old": c_bill_empty_old,
            "orders_bill_empty_new": c_bill_empty_new,

            # --- Scan Barcode (นับเฉพาะงานค้างใน Scope) ---
            "orders_not_scanned": c_not_scanned,
            "orders_not_scanned_old": c_not_scanned_old,
            "orders_not_scanned_new": c_not_scanned_new,

            # --- คลังรับงาน (Issued but Not Packed) ---
            "wh_receive_total": c_wh_total,
            "wh_receive_total_old": c_wh_total_old,
            "wh_receive_total_new": c_wh_total_new,
            "wh_change_total": len(changed_oids_total),
            "wh_receive_g1": c_wh_g1,
            "wh_receive_g1_old": c_wh_g1_old,
            "wh_receive_g1_new": c_wh_g1_new,
            "wh_change_g1": len(changed_oids_g1),
            "wh_receive_g2": c_wh_g2,
            "wh_receive_g2_old": c_wh_g2_old,
            "wh_receive_g2_new": c_wh_g2_new,
            "wh_change_g2": len(changed_oids_g2),
            "wh_receive_g3": c_wh_g3,
            "wh_receive_g3_old": c_wh_g3_old,
            "wh_receive_g3_new": c_wh_g3_new,
            "wh_change_g3": len(changed_oids_g3),
            
            # --- กลุ่มงานจบ (ใช้ยอดเดิม) ---
            # นับจาก Scope (ไม่ว่าจะซ่อนในตารางหรือไม่ ก็จะโชว์ตัวเลข)
            "ready": sum(1 for r in scope_rows if r.get("allocation_status") == "READY_ACCEPT" and not r.get("packed") and not r.get("is_cancelled")),
            "accepted": sum(1 for r in scope_rows if r.get("allocation_status") == "ACCEPTED"),
            "low": sum(1 for r in scope_rows if r.get("allocation_status") == "LOW_STOCK" and not r.get("packed") and not r.get("is_cancelled")),
            "nostock": sum(1 for r in scope_rows if r.get("allocation_status") == "SHORTAGE" and not r.get("packed") and not r.get("is_cancelled")),
            "notenough": sum(1 for r in scope_rows if r.get("allocation_status") == "NOT_ENOUGH" and not r.get("packed") and not r.get("is_cancelled")),
            
            "packed": len(kpi_packed_oids),
            
            # [แก้ไข] แยกเป็น 2 ยอด: ยกเลิกก่อนแพ็ค / ยกเลิกหลังแพ็ค
            "orders_cancelled": len(kpi_cancel_nopack),
            "orders_cancelled_packed": len(kpi_cancel_packed),
            
            # [NEW] จำนวน Order ที่ถูกลบ (Soft Delete) - เฉพาะวันนี้
            "orders_deleted": deleted_today_count,
        }

        # Sort
        def _sort_key(r):
            return ((r.get("order_id") or ""), (r.get("platform") or ""), (r.get("shop") or ""), (r.get("sku") or ""))
        rows = sorted(rows, key=_sort_key)

        # --- STEP 4.9: Apply "สถานะเปลี่ยน" filter (Warehouse Receive) ---
        if show_change:
            if show_change == "TOTAL":
                target = set(wh_total_oids)
            elif show_change == "G1":
                target = set(wh_g1_oids)
            elif show_change == "G2":
                target = set(wh_g2_oids)
            elif show_change == "G3":
                target = set(wh_g3_oids)
            else:
                target = set()

            # ใช้ scope_rows เพื่อให้ฟิลเตอร์นี้ทำงานได้แม้ก่อนหน้าจะกดการ์ด/สถานะอื่น
            rows = [
                r for r in scope_rows
                if (r.get("order_id") or "").strip() in target and r.get("status_change")
            ]

        # --- [แก้ไขจุดที่ 1] คำนวณยอด "Order จ่ายแล้ว" (Issued) ให้ขยับตามฟิลเตอร์ ---
        iq = db.session.query(IssuedOrder.order_id)\
               .join(OrderLine, OrderLine.order_id == IssuedOrder.order_id)\
               .join(Shop, Shop.id == OrderLine.shop_id)

        # 1. กรอง Platform / Shop
        if platform:
            iq = iq.filter(Shop.platform == platform)
        if shop_id:
            iq = iq.filter(Shop.id == int(shop_id))

        # 2. [เพิ่ม] กรองตามคำค้นหา (Global Search)
        if q:
            iq = iq.filter(IssuedOrder.order_id.contains(q))

        # 3. กรองวันที่สำหรับ "Order จ่ายแล้ว" (Issued Count)
        if is_all_time:
            # All Time -> ไม่กรองวันที่ (นับสะสม)
            pass
        elif mode == 'today':
            # โหมด Today -> กรองเฉพาะออเดอร์ที่นำเข้าวันนี้
            iq = iq.filter(OrderLine.import_date == now_thai().date())
        elif has_date_filter:
            # ถ้ามีการเลือกช่วงเวลา -> กรองตามวันที่นำเข้า
            if imp_from: iq = iq.filter(OrderLine.import_date >= imp_from)
            if imp_to:   iq = iq.filter(OrderLine.import_date <= imp_to)
            if d_from:   iq = iq.filter(OrderLine.order_time >= d_from)
            if d_to:     iq = iq.filter(OrderLine.order_time < d_to)
        else:
            # [แก้ไขสำคัญ] Default View (หน้าปกติ) 
            # ให้กรอง "เวลาจ่ายงาน (Issued At)" เป็น "วันนี้" เท่านั้น 
            # เพื่อให้ยอดรีเซ็ตเป็น 0 ทุกวัน
            iq = iq.filter(func.date(IssuedOrder.issued_at) == now_thai().date())

        # ใช้ distinct เพราะ 1 Order มีหลาย Line
        issued_count = iq.distinct().count()

        return render_template(
            "dashboard.html",
            rows=rows,
            shops=shops,
            platform_sel=platform,
            shop_sel=shop_id,
            import_from_sel=import_from_str,  # ส่งกลับไปแสดงผล
            import_to_sel=import_to_str,      # ส่งกลับไปแสดงผล
            status_sel=status,
            date_from_sel=date_from,
            date_to_sel=date_to,
            kpis=kpis,
            packed_oids=packed_oids,
            issued_count=issued_count,
            all_time=all_time,
            use_default_view=use_default_view,
            q=q,
            mode=mode,  # [แก้ไข] ส่งค่า mode ไปยัง template เพื่อให้ปุ่ม "ดูเฉพาะงานวันนี้" ทำงานได้
            ready_oids=global_ready_oids,  # [แก้ไข] เปลี่ยนจาก kpi_orders_ready เป็น global_ready_oids เพื่อให้ปุ่มกดรับอ้างอิงจากความพร้อมจริงๆ เท่านั้น (ไม่เพี้ยนตอน Search)
            change_filter=show_change,
        )

    # =========[ NEW ]=========  กดรับ Order ในหน้า Dashboard
    @app.post("/dashboard/accept_order")
    @login_required
    def dashboard_accept_order():
        cu = current_user()
        if not cu:
            flash("กรุณาเข้าสู่ระบบก่อน", "danger")
            return redirect(url_for("login"))

        order_id = request.form.get("order_id")
        sku = request.form.get("sku")
        platform = request.form.get("platform")
        shop_id = request.form.get("shop_id")

        if not order_id or not sku:
            flash("ข้อมูลไม่ครบถ้วน", "danger")
            return redirect(url_for("dashboard"))

        # อัปเดท OrderLine ให้เป็น accepted
        try:
            ol = OrderLine.query.filter_by(order_id=order_id, sku=sku).first()
            if ol:
                ol.accepted = True
                ol.accepted_at = now_thai()
                ol.accepted_by_user_id = cu.id
                ol.accepted_by_username = cu.username
                db.session.commit()
                flash(f"รับออเดอร์ {order_id} (SKU: {sku}) สำเร็จ", "success")
            else:
                flash("ไม่พบรายการที่ต้องการรับ", "warning")
        except Exception as e:
            db.session.rollback()
            app.logger.exception("Accept order failed")
            flash(f"เกิดข้อผิดพลาด: {e}", "danger")

        # redirect กลับไปหน้าเดิมพร้อมฟิลเตอร์
        return redirect(url_for("dashboard", platform=platform, shop_id=shop_id))
    # =========[ /NEW ]=========

    # -----------------------
    # Import endpoints
    # -----------------------
    @app.route("/import/orders", methods=["GET", "POST"])
    @login_required
    def import_orders_view():
        # 1. จัดการวันที่แบบ Range (จาก...ถึง...)
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        # Fallback: รองรับ filter_date เดิมสำหรับ backwards compatibility
        filter_date_str = request.args.get("filter_date")
        
        today_str = now_thai().date().isoformat()
        
        # ถ้ามี filter_date เดิม ให้ใช้เป็น date_from และ date_to เดียวกัน
        if filter_date_str and not date_from_str:
            date_from_str = filter_date_str
            date_to_str = filter_date_str
        
        if not date_from_str:
            date_from_str = today_str
        if not date_to_str:
            date_to_str = today_str
            
        try:
            view_date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
        except Exception:
            view_date_from = now_thai().date()
        try:
            view_date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except Exception:
            view_date_to = now_thai().date()

        # สร้างวันที่แบบไทยสำหรับแสดงผล (วัน/เดือน/ปี พ.ศ.)
        def _th_date(d):
            return f"{d.day:02d}/{d.month:02d}/{d.year + 543}"
        
        view_date_range_thai = _th_date(view_date_from)
        if view_date_from != view_date_to:
            view_date_range_thai += f" - {_th_date(view_date_to)}"

        # 2. Handle POST (การนำเข้า)
        if request.method == "POST":
            platform = request.form.get("platform")
            shop_name = request.form.get("shop_name")
            f = request.files.get("file")
            
            # วันที่นำเข้า = วันปัจจุบันเสมอ
            current_import_date = now_thai().date()

            if not platform or not f:
                flash("กรุณาเลือกแพลตฟอร์ม และเลือกไฟล์", "danger")
                return redirect(url_for("import_orders_view"))
            try:
                df = pd.read_excel(f)
                # >>> สร้าง/ใช้ร้านเดิมก่อนเสมอ (กัน UNIQUE พัง)
                _ensure_shops_from_df(df, platform=platform, default_shop_name=shop_name)
                
                # เรียก Importer ใหม่
                stats = import_orders(
                    df, platform=platform, shop_name=shop_name, import_date=current_import_date
                )
                
                # เก็บ Batch Data (IDs) ลง Log
                batch_data = json.dumps({
                    "added_ids": stats.get("added_ids", []),
                    "duplicate_ids": stats.get("duplicate_ids", []),
                    "duplicate_old_ids": stats.get("duplicate_old_ids", []),
                    "duplicate_today_ids": stats.get("duplicate_today_ids", []),
                    "failed_ids": stats.get("failed_ids", [])
                }, ensure_ascii=False)
                
                # บันทึก Log ลง DB
                log_entry = ImportLog(
                    import_date=current_import_date,
                    platform=platform,
                    filename=f.filename or "uploaded_file.xlsx",
                    added_count=stats["added"],
                    duplicates_count=stats["duplicates"],
                    failed_count=stats["failed"],
                    error_details=json.dumps(stats["errors"], ensure_ascii=False) if stats["errors"] else "[]"
                )
                # เพิ่ม batch_data และ shop_name ถ้าคอลัมน์มีอยู่
                if hasattr(log_entry, 'batch_data'):
                    log_entry.batch_data = batch_data
                if hasattr(log_entry, 'shop_name'):
                    log_entry.shop_name = shop_name or ""
                if hasattr(log_entry, 'duplicates_same_day'):
                    log_entry.duplicates_same_day = stats.get("duplicates_today", 0)
                db.session.add(log_entry)
                db.session.commit()

                # สร้างข้อความแจ้งเตือนแยกประเภทซ้ำ
                dup_old = stats.get('duplicates_old', 0)
                dup_today = stats.get('duplicates_today', 0)
                dup_msg = f"ซ้ำข้ามวัน {dup_old}"
                if dup_today > 0:
                    dup_msg += f" (ซ้ำวันนี้ {dup_today} - ไม่นับ)"
                
                flash(
                    f"นำเข้า: เพิ่ม {stats['added']} | {dup_msg} | ไม่สำเร็จ {stats['failed']}", 
                    "success" if stats['failed'] == 0 else "warning"
                )
                # Redirect กลับมาหน้า Dashboard ของวันที่นำเข้า
                return redirect(url_for('import_orders_view', date_from=current_import_date.isoformat(), date_to=current_import_date.isoformat()))

            except Exception as e:
                db.session.rollback()
                flash(f"เกิดข้อผิดพลาดในการนำเข้าออเดอร์: {e}", "danger")
                return redirect(url_for("import_orders_view"))

        # 3. คำนวณ Dashboard (นับ Unique Order IDs) - รองรับ Date Range
        
        # A. ยอดสำเร็จ (Success): นับ Order ID ไม่ซ้ำจาก OrderLine จริงๆ ตามช่วงวันที่เลือก
        success_count = db.session.query(func.count(func.distinct(OrderLine.order_id)))\
            .filter(OrderLine.import_date >= view_date_from)\
            .filter(OrderLine.import_date <= view_date_to).scalar() or 0

        # B. ดึง Logs ของช่วงวันนั้นเพื่อนับ Duplicate และ Failed (ตาม Order ID ไม่ซ้ำ)
        logs = ImportLog.query.filter(
            ImportLog.import_date >= view_date_from,
            ImportLog.import_date <= view_date_to
        ).order_by(ImportLog.created_at.desc()).all()
        
        # [แก้ไขใหม่] ใช้ Set เก็บ ID เพื่อตัดตัวซ้ำ (Unique Count)
        # ต่อให้นำเข้าไฟล์เดิม 10 รอบ ID เดิมก็จะถูกนับแค่ 1 ครั้ง
        log_dup_old_ids: set[str] = set()  # ซ้ำข้ามวัน (แสดงในการ์ด)
        log_dup_today_ids: set[str] = set()  # ซ้ำในวัน (ไม่แสดงในการ์ด)
        log_failed_ids: set[str] = set()  # Failed IDs (Unique) - เก็บ Order ID ที่ Failed
        anon_error_set: set[str] = set()  # เก็บข้อความ Error ที่ไม่มี ID (เช่น "แถว 12...") (ไม่ซ้ำ)
        grouped_errors: list[dict] = []  # เก็บ Error แยกตาม Log
        
        for l in logs:
            # ดึง Batch Data (IDs)
            batch_data_str = getattr(l, 'batch_data', None)
            batch_data = {}
            if batch_data_str:
                try:
                    batch_data = json.loads(batch_data_str)
                    log_dup_old_ids.update(batch_data.get("duplicate_old_ids", []))
                    log_dup_today_ids.update(batch_data.get("duplicate_today_ids", []))
                    
                    # Failed: เก็บ ID ที่ระบุได้ลง Set (ตัดซ้ำอัตโนมัติ)
                    current_failed_ids = batch_data.get("failed_ids", [])
                    for fid in current_failed_ids:
                        if fid:
                            log_failed_ids.add(str(fid).strip())
                except Exception:
                    pass
                
            # ดึง Error Details และจัดกลุ่ม
            if l.error_details and l.error_details != "[]":
                try:
                    errs = json.loads(l.error_details)
                    if errs:
                        # [สำคัญ] วนลูปเช็ค Error แต่ละบรรทัดเพื่อตัดซ้ำ
                        for err_msg in errs:
                            err_msg = str(err_msg).strip()
                            # ถ้าเป็น Error ที่มีคำว่า "Order " แสดงว่ามี ID แล้ว -> ข้าม (เพราะถูกนับใน log_failed_ids แล้ว)
                            # แต่ถ้าเป็น "แถว 12..." หรือ Error อื่นๆ -> เก็บลง Set
                            if not err_msg.startswith("Order "):
                                anon_error_set.add(err_msg)
                        
                        # แปลงเวลา Log เป็นไทย (UTC+7)
                        ts = l.created_at
                        time_str = "-"
                        if ts:
                            try:
                                ts_thai = ts + timedelta(hours=7)
                                year_be = ts_thai.year + 543
                                time_str = f"{ts_thai.day:02d}/{ts_thai.month:02d}/{year_be} ({ts_thai.hour:02d}:{ts_thai.minute:02d})"
                            except Exception:
                                pass
                        
                        grouped_errors.append({
                            "platform": l.platform or "-",
                            "shop_name": getattr(l, 'shop_name', '') or l.filename or "-",
                            "filename": l.filename or "-",
                            "time": time_str,
                            "errors": errs
                        })
                except Exception:
                    pass
        
        # C. คำนวณยอดสรุป (นับจาก Set ที่ตัดซ้ำแล้ว - ไม่ใช้ Fallback บวกสะสมอีกต่อไป)
        # ยอด Failed = (จำนวน Order ID ที่ไม่ซ้ำ) + (จำนวนข้อความ Error แถวที่ไม่ซ้ำ)
        real_fail_count = len(log_failed_ids) + len(anon_error_set)
        dup_old_count = len(log_dup_old_ids)  # ซ้ำข้ามวัน (แสดงในการ์ด) - Unique เท่านั้น
        dup_today_count = len(log_dup_today_ids)  # ซ้ำในวัน (ไม่แสดงในการ์ด)
        
        # Total = Success + Failed (ไม่รวม Duplicate เพราะคือออเดอร์เดิม)
        total_count = success_count + real_fail_count
        
        # D. ดึงข้อมูลร้านและ URL
        shops = Shop.query.order_by(Shop.name.asc()).all()
        
        # [แก้ไข] ใช้ SQL ดึง URL โดยตรง (แก้ปัญหารีเฟรชแล้วหาย)
        shop_urls = {}
        try:
            # ดึง name และ google_sheet_url จากตาราง shops ตรงๆ
            rows_url = db.session.execute(text("SELECT name, google_sheet_url FROM shops")).fetchall()
            for r_name, r_url in rows_url:
                shop_urls[r_name] = r_url or ""
        except Exception as e:
            app.logger.warning(f"Fetch shop urls failed: {e}")
        
        return render_template(
            "import_orders.html", 
            shops=shops,
            shop_urls=shop_urls,
            date_from=view_date_from.isoformat(),
            date_to=view_date_to.isoformat(),
            view_date_range_thai=view_date_range_thai,
            dash={
                "total": total_count,
                "success": success_count,
                "duplicate": dup_old_count,        # แสดงเฉพาะซ้ำข้ามวัน
                "duplicate_today": dup_today_count, # ซ้ำในวัน (ไม่แสดงในการ์ด แต่เก็บไว้อ้างอิง)
                "failed": real_fail_count,
                "grouped_errors": grouped_errors   # ส่งแบบกลุ่มไป
            }
        )

    # =========[ NEW ]=========
    # Import Orders จาก Google Sheet
    @app.route("/import/orders/gsheet", methods=["POST"])
    @login_required
    def import_orders_gsheet():
        platform = request.form.get("platform")
        shop_name = request.form.get("shop_name")
        sheet_url = request.form.get("sheet_url")

        if not platform or not sheet_url:
            flash("กรุณาระบุแพลตฟอร์มและลิงก์ Google Sheet", "danger")
            return redirect(url_for("import_orders_view"))

        # [NEW] อัปเดต URL ล่าสุดให้ร้านอัตโนมัติเมื่อกดดึงข้อมูล
        # Logic ใหม่: ถ้าไม่ได้ระบุชื่อร้าน ให้บันทึกเข้า Platform Name (เป็น URL กลาง)
        platform_std = normalize_platform(platform)
        
        # ชื่อที่จะใช้บันทึก URL (ถ้ามีชื่อร้านใช้ชื่อร้าน ถ้าไม่มีใช้ชื่อ Platform)
        target_save_name = shop_name.strip() if shop_name and shop_name.strip() else platform_std
        
        if sheet_url:
            # ค้นหาร้าน หรือ สร้างใหม่ถ้าไม่เจอ (เพื่อเก็บ URL)
            s = Shop.query.filter_by(platform=platform_std, name=target_save_name).first()
            if not s:
                # ถ้าไม่เจอ ลองหาจากชื่ออย่างเดียว (กรณีชื่อ Platform)
                s = Shop.query.filter_by(name=target_save_name).first()
            
            if not s:
                s = Shop(platform=platform_std, name=target_save_name)
                db.session.add(s)
                db.session.commit()  # Commit เพื่อให้ได้ ID มาใช้
            
            # บันทึก URL (ใช้ SQL ตรงๆ)
            if sheet_url:
                db.session.execute(
                    text("UPDATE shops SET google_sheet_url = :u WHERE id = :id"),
                    {"u": sheet_url, "id": s.id}
                )
                db.session.commit()

        # กำหนดชื่อ Tab ตามแพลตฟอร์ม
        target_tab_name = ""
        if platform == "Shopee":
            target_tab_name = "Import_Shopee"
        elif platform == "Lazada":
            target_tab_name = "Import_Lazada"
        elif platform == "TikTok":
            target_tab_name = "Import_Tiktok"
        else:
            target_tab_name = "Import_Order_other"

        try:
            # 1. เชื่อมต่อ Google API
            creds = get_google_credentials()
            client = gspread.authorize(creds)

            # 2. เปิด Google Sheet
            sheet = client.open_by_url(sheet_url)
            
            # 3. เลือก Tab ตามชื่อ
            try:
                worksheet = sheet.worksheet(target_tab_name)
            except gspread.WorksheetNotFound:
                flash(f"❌ ไม่พบ Tab ชื่อ '{target_tab_name}' ใน Google Sheet นี้", "danger")
                return redirect(url_for("import_orders_view"))
            
            # 4. ดึงข้อมูล
            data = worksheet.get_all_records()
            if not data:
                flash(f"Tab '{target_tab_name}' ไม่มีข้อมูล", "warning")
                return redirect(url_for("import_orders_view"))

            # 5. แปลงเป็น DataFrame และนำเข้า
            df = pd.DataFrame(data)
            
            # สร้าง/เช็คชื่อร้าน
            _ensure_shops_from_df(df, platform=platform, default_shop_name=shop_name)
            
            # วันที่นำเข้า = วันปัจจุบันเสมอ
            current_import_date = now_thai().date()
            
            # เรียก Importer ใหม่
            stats = import_orders(
                df, platform=platform, shop_name=shop_name, import_date=current_import_date
            )
            
            # เก็บ Batch Data (IDs) ลง Log
            batch_data = json.dumps({
                "added_ids": stats.get("added_ids", []),
                "duplicate_ids": stats.get("duplicate_ids", []),
                "duplicate_old_ids": stats.get("duplicate_old_ids", []),
                "duplicate_today_ids": stats.get("duplicate_today_ids", []),
                "failed_ids": stats.get("failed_ids", [])
            }, ensure_ascii=False)
            
            # บันทึก Log ลง DB
            log_entry = ImportLog(
                import_date=current_import_date,
                platform=platform,
                filename=f"Google Sheet ({target_tab_name})",
                added_count=stats["added"],
                duplicates_count=stats["duplicates"],
                failed_count=stats["failed"],
                error_details=json.dumps(stats["errors"], ensure_ascii=False) if stats["errors"] else "[]"
            )
            # เพิ่ม batch_data, shop_name และ duplicates_same_day ถ้าคอลัมน์มีอยู่
            if hasattr(log_entry, 'batch_data'):
                log_entry.batch_data = batch_data
            if hasattr(log_entry, 'shop_name'):
                log_entry.shop_name = shop_name or ""
            if hasattr(log_entry, 'duplicates_same_day'):
                log_entry.duplicates_same_day = stats.get("duplicates_today", 0)
            db.session.add(log_entry)
            db.session.commit()
            
            # สร้างข้อความแจ้งเตือนแยกประเภทซ้ำ
            dup_old = stats.get('duplicates_old', 0)
            dup_today = stats.get('duplicates_today', 0)
            dup_msg = f"ซ้ำข้ามวัน {dup_old}"
            if dup_today > 0:
                dup_msg += f" (ซ้ำวันนี้ {dup_today} - ไม่นับ)"
            
            flash(
                f"✅ ดึงข้อมูลจาก {target_tab_name}: เพิ่ม {stats['added']} | {dup_msg} | ไม่สำเร็จ {stats['failed']}", 
                "success" if stats['failed'] == 0 else "warning"
            )
            return redirect(url_for('import_orders_view', date_from=current_import_date.isoformat(), date_to=current_import_date.isoformat()))

        except Exception as e:
            db.session.rollback()
            if "PERMISSION_DENIED" in str(e):
                flash("❌ บอทเข้าถึงไฟล์ไม่ได้! กรุณาตรวจสอบสิทธิ์การแชร์ (Share) ของ Google Sheet", "danger")
            else:
                app.logger.exception("Google Sheet Import Error")
                flash(f"เกิดข้อผิดพลาด: {str(e)}", "danger")
            return redirect(url_for("import_orders_view"))

    # =========[ NEW ]=========
    # ล้างประวัติ Import Log (พร้อมออปชั่นลบข้อมูลออเดอร์จริง)
    @app.route("/import/orders/clear_log", methods=["POST"])
    @login_required
    def clear_import_log():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ลบประวัติ", "danger")
            return redirect(url_for("import_orders_view"))
            
        mode = request.form.get("mode")  # 'range' or 'all'
        delete_data = request.form.get("delete_data")  # 'yes' ถ้าติ๊ก checkbox
        
        try:
            order_deleted_count = 0
            log_deleted_count = 0
            
            if mode == 'all':
                # 1. ถ้าเลือกติ๊กลบข้อมูล -> ลบออเดอร์ทั้งหมด
                if delete_data == 'yes':
                    # ลบ OrderLines ทั้งหมด
                    order_deleted_count = OrderLine.query.delete()
                    # ลบ DeletedOrder ถังขยะด้วย
                    try:
                        db.session.query(DeletedOrder).delete()
                    except Exception:
                        pass
                
                # 2. ลบ Log ทั้งหมด
                log_deleted_count = ImportLog.query.delete()
                
                if order_deleted_count > 0:
                    msg = f"ล้างเกลี้ยง! (Log {log_deleted_count} รายการ, ออเดอร์ {order_deleted_count} รายการ)"
                else:
                    msg = f"ล้างประวัติการนำเข้าทั้งหมดเรียบร้อย ({log_deleted_count} รายการ)"

            else:
                # ลบตามช่วงวันที่
                d_from_str = request.form.get("date_from")
                d_to_str = request.form.get("date_to")
                
                if not d_from_str or not d_to_str:
                    flash("ระบุวันที่ไม่ถูกต้อง", "warning")
                    return redirect(url_for("import_orders_view"))
                    
                d_from = datetime.strptime(d_from_str, "%Y-%m-%d").date()
                d_to = datetime.strptime(d_to_str, "%Y-%m-%d").date()
                
                # 1. ถ้าเลือกติ๊กลบข้อมูล -> ลบออเดอร์ในช่วงวันที่นำเข้านั้น
                if delete_data == 'yes':
                    order_deleted_count = OrderLine.query.filter(
                        OrderLine.import_date >= d_from,
                        OrderLine.import_date <= d_to
                    ).delete(synchronize_session=False)
                
                # 2. ลบ Log ในช่วงวันที่
                log_deleted_count = ImportLog.query.filter(
                    ImportLog.import_date >= d_from,
                    ImportLog.import_date <= d_to
                ).delete(synchronize_session=False)
                
                if order_deleted_count > 0:
                    msg = f"ล้างข้อมูลช่วง {to_be_date_str(d_from)} - {to_be_date_str(d_to)} เรียบร้อย (Log {log_deleted_count}, ออเดอร์ {order_deleted_count})"
                else:
                    msg = f"ล้างประวัติช่วง {to_be_date_str(d_from)} - {to_be_date_str(d_to)} เรียบร้อย ({log_deleted_count} รายการ)"
                
            db.session.commit()
            flash(msg, "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"เกิดข้อผิดพลาด: {e}", "danger")
            
        return redirect(url_for("import_orders_view"))
    # =========[ /NEW ]=========

    # =========[ NEW ]=========
    # Import Orders ยกเลิก + Template
    @app.route("/import/cancel/template")
    @login_required
    def import_cancel_template():
        fmt = (request.args.get("format") or "xlsx").lower()
        sample = ["ORDER-001", "ORDER-002", "ORDER-ABC-003"]

        if fmt == "xlsx" and _OPENPYXL_OK:
            wb = Workbook()
            ws = wb.active
            ws.title = "cancelled_orders"
            ws["A1"] = "order_id"
            for i, no in enumerate(sample, start=2):
                ws[f"A{i}"] = no
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(
                bio,
                as_attachment=True,
                download_name="template_import_orders_cancel.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # Fallback CSV
        csv_io = BytesIO()
        csv_io.write(("order_id\n" + "\n".join(sample)).encode("utf-8-sig"))
        csv_io.seek(0)
        return send_file(
            csv_io,
            as_attachment=True,
            download_name="template_import_orders_cancel.csv",
            mimetype="text/csv",
        )

    # =========================================================
    #  Helper: Unique Daily Stats สำหรับ Cancel Orders
    # =========================================================
    def _get_cancel_daily_stats(view_date):
        """
        คำนวณสถิติ Dashboard ตามโจทย์:
        1. Success = นับจำนวน Order ที่ถูกสร้างใน DB 'วันนี้' จริงๆ (New Success)
        2. Duplicate = (จำนวน Unique ID ที่ตรวจสอบวันนี้) - (Success)
           * วิธีนี้ช่วยให้ถ้านำเข้าไฟล์เดิมซ้ำ 3 รอบ ยอด Duplicate จะคงที่ ไม่เบิ้ลตามจำนวนรอบ
        """
        # 1. ยอด Success: นับจาก DB โดยตรง (บวก 7 ชม. เพื่อให้ตรงกับเวลาไทย)
        success_count = CancelledOrder.query.filter(
            func.date(CancelledOrder.imported_at, '+7 hours') == view_date
        ).count()

        # 2. ยอด Unique Input: ดึง ID ทั้งหมดที่เคยนำเข้าวันนี้จาก Log มาทำ Unique Set
        logs = ImportLog.query.filter(
            ImportLog.import_date == view_date,
            ImportLog.platform == 'CANCEL_SYSTEM'
        ).all()
        
        all_attempted_ids = set()
        failed_total = 0
        
        for log in logs:
            failed_total += (log.failed_count or 0)
            # ดึง ID ที่บันทึกไว้ใน batch_data มารวมกัน
            if log.batch_data:
                try:
                    data = json.loads(log.batch_data)
                    if "ids" in data:
                        all_attempted_ids.update(data["ids"])
                except:
                    pass
                    
        # 3. ยอด Duplicate: (ที่เช็คทั้งหมด - ที่เข้า DB ได้)
        # หมายถึงรายการที่มีใน DB อยู่แล้ว (ไม่ว่าจะเก่าหรือใหม่)
        duplicate_count = max(0, len(all_attempted_ids) - success_count)

        return {
            "success": success_count,
            "duplicate": duplicate_count,
            "failed": failed_total
        }

    def _process_cancel_import(order_ids: list, source_name: str, user_id: int):
        """ประมวลผลการนำเข้า order ยกเลิก"""
        # 1. คลีนข้อมูลและตัดตัวซ้ำในไฟล์ (Internal Deduplicate)
        unique_input_ids = set()
        for oid in order_ids:
            s = str(oid).strip()
            if s: unique_input_ids.add(s)
        
        if not unique_input_ids:
            return 0, 0

        # 2. หา ID ที่มีอยู่แล้วใน DB (เช็คซ้ำทั้งหมด ไม่สนวันที่)
        existing_query = db.session.query(CancelledOrder.order_id).filter(
            CancelledOrder.order_id.in_(unique_input_ids)
        ).all()
        existing_ids = {r[0] for r in existing_query}

        # 3. แยกรายการ ใหม่ vs ซ้ำ
        new_ids = unique_input_ids - existing_ids
        
        # 4. บันทึกรายการใหม่ลง DB
        if new_ids:
            timestamp = datetime.now(timezone.utc)
            new_entries = []
            for oid in new_ids:
                new_entries.append(CancelledOrder(
                    order_id=oid,
                    imported_at=timestamp,
                    imported_by_user_id=user_id,
                    note=f"Import via {source_name}"
                ))
            db.session.bulk_save_objects(new_entries)
            db.session.commit()

        # 5. บันทึก Log พร้อม batch_data (สำคัญสำหรับคำนวณหน้าเว็บ)
        log = ImportLog(
            import_date=now_thai().date(),
            platform="CANCEL_SYSTEM",
            shop_name="-",
            filename=source_name,
            added_count=len(new_ids),
            duplicates_count=len(existing_ids),
            failed_count=0,
            # เก็บ ID ทั้งหมดเพื่อไปทำ Union ที่ Dashboard
            batch_data=json.dumps({"ids": list(unique_input_ids)})
        )
        db.session.add(log)
        db.session.commit()

        return len(new_ids), len(existing_ids)

    @app.route("/import/cancel", methods=["GET"])
    @login_required
    def import_cancel_view():
        _ensure_cancel_table()
        
        # ตรวจสอบให้ ImportLog table มีอยู่
        try:
            ImportLog.__table__.create(bind=db.engine, checkfirst=True)
        except Exception:
            pass

        # [แก้ไข] รับค่าเป็นช่วงวันที่
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        
        # Default: วันนี้
        if not date_from_str:
            date_from_str = now_thai().date().isoformat()
        if not date_to_str:
            date_to_str = now_thai().date().isoformat()

        d_from = parse_date_any(date_from_str)
        d_to = parse_date_any(date_to_str)

        # 1. ดึง Log ตามช่วงเวลา
        logs = ImportLog.query.filter(
            ImportLog.import_date >= d_from,
            ImportLog.import_date <= d_to,
            ImportLog.platform == 'CANCEL_SYSTEM'
        ).order_by(ImportLog.created_at.desc()).all()

        # 2. คำนวณ Success (นับจาก DB จริงที่สร้างในช่วงเวลานั้น + ปรับ Timezone ไทย UTC+7)
        success_count = CancelledOrder.query.filter(
            func.date(CancelledOrder.imported_at, '+7 hours') >= d_from,
            func.date(CancelledOrder.imported_at, '+7 hours') <= d_to
        ).count()

        # 3. [แก้ตรงนี้] คำนวณ Duplicate แบบไม่นับเบิ้ล (Unique ID)
        all_attempted_ids = set()  # ใช้ Set เพื่อตัดตัวซ้ำอัตโนมัติ
        failed_sum = 0

        for log in logs:
            failed_sum += (log.failed_count or 0)
            # ดึงรายชื่อ ID ที่เคยยิงเข้ามาทั้งหมดจาก batch_data
            if log.batch_data:
                try:
                    data = json.loads(log.batch_data)
                    # ในฟังก์ชัน _process_cancel_import เราบันทึก key "ids" เอาไว้
                    if "ids" in data:
                        all_attempted_ids.update(data["ids"])
                except Exception:
                    pass
        
        # สูตร: จำนวน ID ทั้งหมดที่ไม่ซ้ำที่ยิงเข้ามา - จำนวนที่สำเร็จจริง = จำนวนที่ซ้ำ
        unique_duplicate_count = max(0, len(all_attempted_ids) - success_count)

        stats = {
            "success": success_count,
            "duplicate": unique_duplicate_count,  # ใช้ค่าใหม่ที่คำนวณแบบ Unique
            "failed": failed_sum
        }

        # ดึง URL ที่บันทึกไว้ (Config) จาก Shop
        saved_url = ""
        try:
            config_row = db.session.execute(
                text("SELECT google_sheet_url FROM shops WHERE platform = 'CANCEL_SYSTEM' AND name = 'GoogleSheet' LIMIT 1")
            ).fetchone()
            if config_row and config_row[0]:
                saved_url = config_row[0]
        except Exception:
            pass
            
        # เตรียมวันที่ภาษาไทยสำหรับหัวข้อ Modal
        date_from_thai = to_be_date_str(d_from) if d_from else ""
        date_to_thai = to_be_date_str(d_to) if d_to else ""

        return render_template(
            "import_cancel.html",
            date_from=date_from_str,
            date_to=date_to_str,
            date_from_thai=date_from_thai,
            date_to_thai=date_to_thai,
            stats=stats,
            logs=logs,
            saved_url=saved_url
        )

    @app.route("/import/cancel/action", methods=["POST"])
    @login_required
    def import_cancel_action():
        _ensure_cancel_table()
        
        # ตรวจสอบให้ ImportLog table มีอยู่
        try:
            ImportLog.__table__.create(bind=db.engine, checkfirst=True)
        except Exception:
            pass

        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ต้องเป็นผู้ดูแลระบบหรือพนักงานเท่านั้น", "danger")
            return redirect(url_for("dashboard"))

        mode = request.form.get("mode")  # 'file' or 'gsheet'
        
        try:
            order_ids = []
            source_name = "File"
            
            # กรณี 1: Google Sheet
            if mode == "gsheet":
                url = request.form.get("sheet_url", "").strip()
                if not url:
                    flash("กรุณาระบุ URL", "warning")
                    return redirect(url_for("import_cancel_view"))
                    
                # Connect Google Sheet
                creds = get_google_credentials()
                client = gspread.authorize(creds)
                sh = client.open_by_url(url)
                source_name = f"GSheet: {sh.title}"
                
                # จับ Tab ชื่อ Import_Cancel
                try:
                    ws = sh.worksheet("Import_Cancel")
                except gspread.WorksheetNotFound:
                    flash("ไม่พบ Tab ชื่อ 'Import_Cancel' ใน Google Sheet นี้", "danger")
                    return redirect(url_for("import_cancel_view"))
                    
                # อ่านข้อมูล (หาคอลัมน์ Order ID หรือ อ่านคอลัมน์แรก)
                rows = ws.get_all_values()
                if rows:
                    header = [str(h).lower().strip() for h in rows[0]]
                    col_idx = 0  # Default คอลัมน์แรก
                    for idx, h in enumerate(header):
                        if h in ["order_id", "order id", "order_no", "เลขคำสั่งซื้อ", "เลขออเดอร์"]:
                            col_idx = idx
                            break
                    
                    # เก็บ ID (ข้าม Header)
                    for r in rows[1:]:
                        if len(r) > col_idx:
                            val = str(r[col_idx]).strip()
                            if val: order_ids.append(val)

            # กรณี 2: Upload File
            elif mode == "file":
                f = request.files.get("file")
                if f and f.filename:
                    order_ids = _parse_order_ids_from_upload(f)
                    source_name = f.filename
                else:
                    flash("โปรดเลือกไฟล์ Excel/CSV", "warning")
                    return redirect(url_for("import_cancel_view"))

            # ประมวลผล
            if order_ids:
                added, dups = _process_cancel_import(order_ids, source_name, cu.id)
                flash(f"✅ นำเข้าสำเร็จ: เพิ่มใหม่ {added}, ซ้ำ {dups} รายการ", "success")
            else:
                flash("ไม่พบข้อมูล Order ID", "warning")

        except Exception as e:
            db.session.rollback()
            app.logger.exception("Import cancelled orders failed")
            flash(f"เกิดข้อผิดพลาด: {e}", "danger")

        return redirect(url_for("import_cancel_view"))

    # =========[ NEW ]=========
    # ล้างประวัติ Import Cancel Log
    @app.route("/import/cancel/clear_log", methods=["POST"])
    @login_required
    def clear_cancel_log():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ลบประวัติ", "danger")
            return redirect(url_for("import_cancel_view"))
            
        mode = request.form.get("mode")  # 'range' or 'all'
        delete_data = request.form.get("delete_data")  # 'yes' ถ้าติ๊ก checkbox
        
        try:
            data_deleted_count = 0
            log_deleted_count = 0
            
            if mode == 'all':
                # 1. ถ้าเลือกติ๊กลบข้อมูล -> ลบข้อมูลใน CancelledOrder ทั้งหมด
                if delete_data == 'yes':
                    data_deleted_count = db.session.query(CancelledOrder).delete()
                
                # 2. ลบ Log ทั้งหมดที่เป็นของ CANCEL_SYSTEM
                log_deleted_count = ImportLog.query.filter_by(platform='CANCEL_SYSTEM').delete()
                
                if data_deleted_count > 0:
                    msg = f"ล้างเกลี้ยง! (Log {log_deleted_count} รายการ, ข้อมูล {data_deleted_count} รายการ)"
                else:
                    msg = f"ล้างประวัติทั้งหมดเรียบร้อย ({log_deleted_count} รายการ)"

            else:
                # ลบตามช่วงวันที่
                d_from_str = request.form.get("date_from")
                d_to_str = request.form.get("date_to")
                
                if not d_from_str or not d_to_str:
                    flash("ระบุวันที่ไม่ถูกต้อง", "warning")
                    return redirect(url_for("import_cancel_view"))
                    
                d_from = datetime.strptime(d_from_str, "%Y-%m-%d").date()
                d_to = datetime.strptime(d_to_str, "%Y-%m-%d").date()
                
                # 1. ถ้าเลือกติ๊กลบข้อมูล -> ลบข้อมูลในช่วงวันที่
                if delete_data == 'yes':
                    # สร้าง timestamp ครอบคลุมทั้งวัน
                    dt_start = datetime.combine(d_from, datetime.min.time())
                    dt_end = datetime.combine(d_to, datetime.max.time())
                    
                    data_deleted_count = CancelledOrder.query.filter(
                        CancelledOrder.imported_at >= dt_start,
                        CancelledOrder.imported_at <= dt_end
                    ).delete(synchronize_session=False)
                
                # 2. ลบ Log ในช่วงวันที่ (เฉพาะ CANCEL_SYSTEM)
                log_deleted_count = ImportLog.query.filter(
                    ImportLog.platform == 'CANCEL_SYSTEM',
                    ImportLog.import_date >= d_from,
                    ImportLog.import_date <= d_to
                ).delete(synchronize_session=False)
                
                if data_deleted_count > 0:
                    msg = f"ล้างข้อมูลช่วง {to_be_date_str(d_from)} - {to_be_date_str(d_to)} เรียบร้อย (Log {log_deleted_count}, ข้อมูล {data_deleted_count})"
                else:
                    msg = f"ล้างประวัติช่วง {to_be_date_str(d_from)} - {to_be_date_str(d_to)} เรียบร้อย ({log_deleted_count} รายการ)"
                
            db.session.commit()
            flash(msg, "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"เกิดข้อผิดพลาด: {e}", "danger")
            
        return redirect(url_for("import_cancel_view"))
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  Import Orders (จ่ายงานแล้ว)
    @app.route("/import/issued/template")
    @login_required
    def import_issued_template():
        # ใช้ logic เดียวกับ template ของ cancel (คืนไฟล์คอลัมน์ order_id)
        sample = ["ORDER-001", "ORDER-002", "ORDER-003"]
        try:
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title = "issued_orders"; ws["A1"] = "order_id"
            for i, no in enumerate(sample, start=2): ws[f"A{i}"] = no
            bio = BytesIO(); wb.save(bio); bio.seek(0)
            return send_file(bio, as_attachment=True, download_name="template_import_orders_issued.xlsx",
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception:
            csv_io = BytesIO()
            csv_io.write(("order_id\n" + "\n".join(sample)).encode("utf-8-sig"))
            csv_io.seek(0)
            return send_file(csv_io, as_attachment=True, download_name="template_import_orders_issued.csv", mimetype="text/csv")

    @app.route("/import/issued", methods=["GET", "POST"])
    @login_required
    def import_issued_orders():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ต้องเป็นผู้ดูแลระบบหรือพนักงานเท่านั้น", "danger")
            return redirect(url_for("dashboard"))

        result = None
        if request.method == "POST":
            f = request.files.get("file")
            if not f or (f.filename or "").strip() == "":
                flash("โปรดเลือกไฟล์ Excel/CSV ก่อน", "warning")
                return redirect(url_for("import_issued_orders"))
            try:
                order_ids_raw = _parse_order_ids_from_upload(f)
                order_ids = [s.strip() for s in order_ids_raw if s and s.strip()]
                order_ids = list(dict.fromkeys(order_ids))  # unique + preserve order

                # มีอยู่จริงในระบบ?
                exists_set = {
                    r[0] for r in db.session.query(OrderLine.order_id)
                    .filter(OrderLine.order_id.in_(order_ids)).distinct().all()
                }
                not_found = [s for s in order_ids if s not in exists_set]

                # mark เป็น "จ่ายงานแล้ว" พร้อมบันทึกเวลา import
                imported_at = now_thai()
                inserted = _mark_issued(list(exists_set), user_id=cu.id, source="import", when_dt=imported_at)

                # ตาม requirement: ถ้ายังไม่เคยนับพิมพ์ ให้ตั้งเป็น 1
                if exists_set:
                    _ensure_min_print_count(list(exists_set), min_count=1, user_id=cu.id, when_iso=now_thai().isoformat())

                result = {
                    "total_in_file": len(order_ids),
                    "matched_in_system": len(exists_set),
                    "inserted_issued": inserted,
                    "not_found": not_found[:50],
                }
                flash(f"ทำเครื่องหมาย 'จ่ายงานแล้ว' {inserted} ออเดอร์", "success")

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import issued orders failed")
                flash(f"เกิดข้อผิดพลาด: {e}", "danger")
                result = None

        return render_template("import_issued.html", result=result)
    # =========[ /NEW ]=========

    # -----------------------
    # Import บิลเปล่า (Empty Bill)
    # -----------------------
    def _update_bill_empty_status_from_df(df: pd.DataFrame) -> tuple[int, int, list[str], list[str], list[str]]:
        """
        อัพเดตสถานะ allocation_status เป็น BILL_EMPTY สำหรับ Order ID ที่ระบุใน DataFrame
        ค้นหา order_id จากคอลัมน์ต่างๆ ที่เป็นไปได้
        แยกนับ Order ที่เป็น BILL_EMPTY ใหม่ กับที่เป็นอยู่แล้ว (Duplicate)

        Returns: (new_count, duplicate_count, new_order_ids, duplicate_order_ids, failed_order_ids)
        """
        from importers import first_existing, COMMON_ORDER_ID

        # หาคอลัมน์ order_id
        order_col = first_existing(df, COMMON_ORDER_ID)
        if not order_col:
            raise ValueError("ไม่พบคอลัมน์ Order ID ในไฟล์")

        new_count = 0
        duplicate_count = 0
        new_order_ids = []
        duplicate_order_ids = []
        failed_order_ids = []

        for idx, row in df.iterrows():
            order_id = str(row.get(order_col, "")).strip()
            if not order_id or order_id.lower() in ("nan", "none", ""):
                failed_order_ids.append(f"แถว {idx+2}: ไม่มี Order ID")
                continue

            try:
                # อัพเดต OrderLine ทั้งหมดที่มี order_id นี้
                # เนื่องจาก allocation_status อาจไม่มีคอลัมน์ใน DB
                # เราจะใช้วิธีเพิ่มคอลัมน์ถ้ายังไม่มี (SQLite รองรับ)
                lines = OrderLine.query.filter_by(order_id=order_id).all()
                if lines:
                    # ลองเพิ่มคอลัมน์ถ้ายังไม่มี (SQLite จะ ignore ถ้ามีอยู่แล้ว)
                    try:
                        db.session.execute(text("ALTER TABLE order_lines ADD COLUMN allocation_status TEXT"))
                        db.session.commit()
                    except Exception:
                        db.session.rollback()
                        pass  # คอลัมน์มีอยู่แล้ว หรือ error อื่น

                    # ตรวจสอบว่า Order นี้เป็น BILL_EMPTY อยู่แล้วหรือไม่
                    # ถ้า OrderLine ใดๆ ใน Order นี้มี allocation_status = 'BILL_EMPTY' อยู่แล้ว = Duplicate
                    is_already_bill_empty = any(
                        hasattr(line, 'allocation_status') and line.allocation_status == 'BILL_EMPTY'
                        for line in lines
                    )

                    # อัพเดตสถานะเป็น BILL_EMPTY พร้อมอัพเดต import_date เป็นวันที่ Import
                    current_import_date = now_thai().date()
                    for line in lines:
                        line.allocation_status = 'BILL_EMPTY'
                        line.import_date = current_import_date

                    # นับแยก new กับ duplicate
                    if is_already_bill_empty:
                        if order_id not in duplicate_order_ids:
                            duplicate_order_ids.append(order_id)
                            duplicate_count += 1
                    else:
                        if order_id not in new_order_ids:
                            new_order_ids.append(order_id)
                            new_count += 1
                else:
                    failed_order_ids.append(f"Order {order_id}: ไม่พบในระบบ")
            except Exception as e:
                failed_order_ids.append(f"Order {order_id}: {str(e)}")

        db.session.commit()
        return new_count, duplicate_count, new_order_ids, duplicate_order_ids, failed_order_ids

    @app.route("/import/bill_empty", methods=["GET", "POST"])
    @login_required
    def import_bill_empty_view():
        """
        หน้าสำหรับนำเข้าบิลเปล่า (Empty Bill)
        ใช้สำหรับกรณีที่ต้องการนำเข้าข้อมูลออเดอร์ที่ไม่มีสินค้า
        """
        # ตรวจสอบให้ ImportLog table มีอยู่
        try:
            ImportLog.__table__.create(bind=db.engine, checkfirst=True)
        except Exception:
            pass

        if request.method == "POST":
            mode = request.form.get("mode")  # "excel" or "gsheet"
            platform = request.form.get("platform")
            shop_name = request.form.get("shop_name")
            f = request.files.get("file")
            sheet_url = request.form.get("sheet_url")

            if mode == "excel":
                if not f:
                    flash("กรุณาเลือกไฟล์ Excel", "danger")
                    return redirect(url_for("import_bill_empty_view"))

                try:
                    df = pd.read_excel(f)
                    new_count, duplicate_count, new_ids, duplicate_ids, failed_ids = _update_bill_empty_status_from_df(df)

                    # บันทึก ImportLog (นับเฉพาะ new orders, ไม่นับ duplicate)
                    current_import_date = now_thai().date()
                    filename = f.filename if hasattr(f, 'filename') else "Excel File"

                    log_entry = ImportLog(
                        import_date=current_import_date,
                        platform='EMPTY_BILL_SYSTEM',
                        filename=filename,
                        added_count=new_count,
                        duplicates_count=duplicate_count,
                        failed_count=len(failed_ids),
                        error_details=json.dumps(failed_ids[:10], ensure_ascii=False) if failed_ids else "[]"
                    )
                    if hasattr(log_entry, 'shop_name'):
                        log_entry.shop_name = ""
                    if hasattr(log_entry, 'batch_data'):
                        log_entry.batch_data = json.dumps({
                            "new_ids": new_ids,
                            "duplicate_ids": duplicate_ids,
                            "failed_ids": failed_ids[:50]  # จำกัดจำนวน
                        }, ensure_ascii=False)
                    db.session.add(log_entry)
                    db.session.commit()

                    # Flash messages แสดงผลแยก new/duplicate/failed
                    if new_count > 0:
                        flash(f"✅ อัพเดตสถานะบิลเปล่าสำเร็จ: {new_count} Order (ใหม่)", "success")
                    if duplicate_count > 0:
                        flash(f"<strong>Order</strong> ที่เป็นบิลเปล่าอยู่แล้ว (ซ้ำ): {duplicate_count} Order", "info")
                    if new_count == 0 and duplicate_count == 0:
                        flash("⚠️ ไม่พบ Order ID ที่สามารถอัพเดตได้", "warning")
                    if failed_ids:
                        flash(f"⚠️ ไม่สำเร็จ: {len(failed_ids)} Order", "warning")

                except Exception as e:
                    db.session.rollback()
                    flash(f"เกิดข้อผิดพลาด: {str(e)}", "danger")
                    app.logger.exception("Bill Empty Import Error")

                return redirect(url_for("import_bill_empty_view"))

            elif mode == "gsheet":
                if not sheet_url:
                    flash("กรุณาระบุ Google Sheet URL", "danger")
                    return redirect(url_for("import_bill_empty_view"))

                try:
                    # 1. เชื่อมต่อ Google API
                    creds = get_google_credentials()
                    client = gspread.authorize(creds)

                    # 2. เปิด Google Sheet
                    sheet = client.open_by_url(sheet_url)

                    # 3. เลือก Tab "Import_Scan_Bill_Empty"
                    try:
                        worksheet = sheet.worksheet("Import_Scan_Bill_Empty")
                    except gspread.WorksheetNotFound:
                        flash("❌ ไม่พบ Tab ชื่อ 'Import_Scan_Bill_Empty' ใน Google Sheet นี้", "danger")
                        return redirect(url_for("import_bill_empty_view"))

                    # 4. ดึงข้อมูล
                    data = worksheet.get_all_records()
                    if not data:
                        flash("Tab 'Import_Scan_Bill_Empty' ไม่มีข้อมูล", "warning")
                        return redirect(url_for("import_bill_empty_view"))

                    # 5. แปลงเป็น DataFrame และอัพเดตสถานะ
                    df = pd.DataFrame(data)
                    new_count, duplicate_count, new_ids, duplicate_ids, failed_ids = _update_bill_empty_status_from_df(df)

                    # 6. บันทึก ImportLog (นับเฉพาะ new orders, ไม่นับ duplicate)
                    current_import_date = now_thai().date()
                    log_entry = ImportLog(
                        import_date=current_import_date,
                        platform='EMPTY_BILL_SYSTEM',
                        filename=f"Google Sheet (Import_Scan_Bill_Empty)",
                        added_count=new_count,
                        duplicates_count=duplicate_count,
                        failed_count=len(failed_ids),
                        error_details=json.dumps(failed_ids[:10], ensure_ascii=False) if failed_ids else "[]"
                    )
                    if hasattr(log_entry, 'shop_name'):
                        log_entry.shop_name = ""
                    if hasattr(log_entry, 'batch_data'):
                        log_entry.batch_data = json.dumps({
                            "new_ids": new_ids,
                            "duplicate_ids": duplicate_ids,
                            "failed_ids": failed_ids[:50]  # จำกัดจำนวน
                        }, ensure_ascii=False)
                    db.session.add(log_entry)

                    # 7. บันทึก URL (ถ้ามี)
                    if sheet_url:
                        try:
                            config_shop = Shop.query.filter_by(platform='EMPTY_BILL_SYSTEM', name='GoogleSheet').first()
                            if not config_shop:
                                config_shop = Shop(platform='EMPTY_BILL_SYSTEM', name='GoogleSheet', is_system_config=True)
                                db.session.add(config_shop)
                                db.session.flush()  # ใช้ flush แทน commit เพื่อให้ได้ ID

                            db.session.execute(
                                text("UPDATE shops SET google_sheet_url = :u WHERE id = :id"),
                                {"u": sheet_url, "id": config_shop.id}
                            )
                        except Exception:
                            pass

                    db.session.commit()

                    # Flash messages แสดงผลแยก new/duplicate/failed
                    if new_count > 0:
                        flash(f"✅ ดึงข้อมูลจาก Google Sheet สำเร็จ: อัพเดตสถานะบิลเปล่า {new_count} Order (ใหม่)", "success")
                    if duplicate_count > 0:
                        flash(f"<strong>Order</strong> ที่เป็นบิลเปล่าอยู่แล้ว (ซ้ำ): {duplicate_count} Order", "info")
                    if new_count == 0 and duplicate_count == 0:
                        flash("⚠️ ไม่พบ Order ID ที่สามารถอัพเดตได้", "warning")
                    if failed_ids:
                        flash(f"⚠️ ไม่สำเร็จ: {len(failed_ids)} Order", "warning")

                except Exception as e:
                    db.session.rollback()
                    if "PERMISSION_DENIED" in str(e):
                        flash("❌ บอทเข้าถึงไฟล์ไม่ได้! กรุณาตรวจสอบสิทธิ์การแชร์ (Share) ของ Google Sheet", "danger")
                    else:
                        app.logger.exception("Google Sheet Import Error")
                        flash(f"เกิดข้อผิดพลาด: {str(e)}", "danger")

                return redirect(url_for("import_bill_empty_view"))

        # GET request - แสดงฟอร์ม
        # รับค่าเป็นช่วงวันที่
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")

        # Default: วันนี้
        if not date_from_str:
            date_from_str = now_thai().date().isoformat()
        if not date_to_str:
            date_to_str = now_thai().date().isoformat()

        d_from = parse_date_any(date_from_str)
        d_to = parse_date_any(date_to_str)

        # ดึง Log ตามช่วงเวลา (ใช้ platform เฉพาะสำหรับ Empty Bill)
        logs = ImportLog.query.filter(
            ImportLog.import_date >= d_from,
            ImportLog.import_date <= d_to,
            ImportLog.platform == 'EMPTY_BILL_SYSTEM'
        ).order_by(ImportLog.created_at.desc()).all()

        # คำนวณสถิติจาก Logs
        # เปลี่ยนเป็นการนับ unique orders จาก DB โดยตรง แทนการ sum จาก logs
        total_success = OrderLine.query.filter(
            OrderLine.allocation_status == 'BILL_EMPTY',
            OrderLine.import_date >= d_from,
            OrderLine.import_date <= d_to
        ).count()
        
        total_failed = 0
        failed_old = 0
        failed_new = 0

        # คำนวณยอดแยกเก่า/ใหม่
        today_date = now_thai().date()

        # คำนวณ failed จาก logs (รวมทุกครั้ง import)
        for log in logs:
            total_failed += log.failed_count
            if log.import_date < today_date:
                failed_old += log.failed_count
            else:
                failed_new += log.failed_count
        success_old = OrderLine.query.filter(
            OrderLine.allocation_status == 'BILL_EMPTY',
            OrderLine.import_date >= d_from,
            OrderLine.import_date < today_date
        ).count()
        success_new = OrderLine.query.filter(
            OrderLine.allocation_status == 'BILL_EMPTY',
            OrderLine.import_date >= today_date,
            OrderLine.import_date <= d_to
        ).count()
        
        total_duplicates = 0  # ไม่ใช้แล้วแต่เก็บไว้เผื่ออ้างอิง
        duplicates_old = 0
        duplicates_new = 0
        total_logs = len(logs)

        # [เพิ่ม] คำนวณจำนวนบิลเปล่าจริงจาก DB (นับ Order ที่ไม่ซ้ำกัน ไม่ใช่นับแถว)
        bill_empty_total = db.session.query(func.count(func.distinct(OrderLine.order_id))).filter(
            OrderLine.allocation_status == 'BILL_EMPTY'
        ).scalar() or 0
        bill_empty_old = db.session.query(func.count(func.distinct(OrderLine.order_id))).filter(
            OrderLine.allocation_status == 'BILL_EMPTY',
            OrderLine.import_date < today_date
        ).scalar() or 0
        bill_empty_new = db.session.query(func.count(func.distinct(OrderLine.order_id))).filter(
            OrderLine.allocation_status == 'BILL_EMPTY',
            OrderLine.import_date >= today_date
        ).scalar() or 0

        # [เพิ่ม] คำนวณจำนวน Packed จาก DB ตามช่วงวันที่ import
        packed_total = OrderLine.query.filter(
            OrderLine.allocation_status == 'PACKED',
            OrderLine.import_date >= d_from,
            OrderLine.import_date <= d_to
        ).count()
        packed_old = OrderLine.query.filter(
            OrderLine.allocation_status == 'PACKED',
            OrderLine.import_date >= d_from,
            OrderLine.import_date < today_date
        ).count()
        packed_new = OrderLine.query.filter(
            OrderLine.allocation_status == 'PACKED',
            OrderLine.import_date >= today_date,
            OrderLine.import_date <= d_to
        ).count()

        stats = {
            "total": total_logs,
            "success": total_success,
            "success_old": success_old,
            "success_new": success_new,
            "duplicates": total_duplicates,
            "duplicates_old": duplicates_old,
            "duplicates_new": duplicates_new,
            "failed": total_failed,
            "failed_old": failed_old,
            "failed_new": failed_new,
            "bill_empty_total": bill_empty_total,
            "bill_empty_old": bill_empty_old,
            "bill_empty_new": bill_empty_new,
            "packed": packed_total,
            "packed_old": packed_old,
            "packed_new": packed_new,
            "grouped_errors": []  # สำหรับแสดง error details
        }

        # จัดกลุ่ม errors ตามไฟล์ (แสดงเฉพาะ 10 ไฟล์ล่าสุด)
        for log in logs[:10]:
            if log.error_details:
                try:
                    errors = json.loads(log.error_details) if log.error_details else []
                    if errors:
                        stats["grouped_errors"].append({
                            "filename": log.filename or "Unknown",
                            "time": log.created_at.strftime("%d/%m/%Y %H:%M") if log.created_at else "",
                            "errors": errors[:5]  # แสดงแค่ 5 errors ต่อไฟล์
                        })
                except:
                    pass

        # ดึง URL ที่บันทึกไว้ (Config) จาก Shop
        saved_url = ""
        try:
            config_row = db.session.execute(
                text("SELECT google_sheet_url FROM shops WHERE platform = 'EMPTY_BILL_SYSTEM' AND name = 'GoogleSheet' LIMIT 1")
            ).fetchone()
            if config_row and config_row[0]:
                saved_url = config_row[0]
        except Exception:
            pass

        # เตรียมวันที่ภาษาไทยสำหรับหัวข้อ Modal
        date_from_thai = to_be_date_str(d_from) if d_from else ""
        date_to_thai = to_be_date_str(d_to) if d_to else ""

        return render_template(
            "import_bill_empty.html",
            date_from=date_from_str,
            date_to=date_to_str,
            date_from_thai=date_from_thai,
            date_to_thai=date_to_thai,
            stats=stats,
            logs=logs,
            saved_url=saved_url
        )

    # =========[ NEW ]=========  ล้างประวัติ Import Log สำหรับบิลเปล่า
    @app.route("/import/bill_empty/clear_logs", methods=["POST"])
    @login_required
    def clear_bill_empty_logs():
        """
        ล้างประวัติ Import Logs สำหรับบิลเปล่า
        รองรับการลบทั้งหมด หรือตามช่วงวันที่
        """
        try:
            # รับพารามิเตอร์
            clear_all = request.form.get("clear_all") == "true"
            date_from_str = request.form.get("date_from")
            date_to_str = request.form.get("date_to")

            if clear_all:
                # ลบทั้งหมด
                deleted_count = ImportLog.query.filter_by(platform='EMPTY_BILL_SYSTEM').delete()
                db.session.commit()
                flash(f"✅ ล้างประวัติทั้งหมดเรียบร้อยแล้ว ({deleted_count} records)", "success")
            elif date_from_str and date_to_str:
                # ลบตามช่วงวันที่
                d_from = parse_date_any(date_from_str)
                d_to = parse_date_any(date_to_str)

                deleted_count = ImportLog.query.filter(
                    ImportLog.platform == 'EMPTY_BILL_SYSTEM',
                    ImportLog.import_date >= d_from,
                    ImportLog.import_date <= d_to
                ).delete()
                db.session.commit()

                flash(f"✅ ล้างประวัติช่วงวันที่ {to_be_date_str(d_from)} - {to_be_date_str(d_to)} เรียบร้อยแล้ว ({deleted_count} records)", "success")
            else:
                flash("⚠️ กรุณาระบุช่วงวันที่หรือเลือกลบทั้งหมด", "warning")

        except Exception as e:
            db.session.rollback()
            flash(f"❌ เกิดข้อผิดพลาด: {str(e)}", "danger")
            app.logger.exception("Clear Bill Empty Logs Error")

        return redirect(url_for("import_bill_empty_view"))

    # =========[ NEW ]=========  API ดึงรายละเอียด Import Log
    @app.route("/api/import/bill_empty/log/<int:log_id>", methods=["GET"])
    @login_required
    def get_bill_empty_log_details(log_id):
        """
        API ดึงรายละเอียด Order IDs จาก Import Log
        สำหรับแสดงใน Modal View
        """
        try:
            # ดึง log ที่ต้องการ
            log = ImportLog.query.get(log_id)
            
            if not log:
                return jsonify({
                    "success": False,
                    "message": "ไม่พบข้อมูล Log"
                }), 404
            
            # ดึง batch_data จาก log
            new_ids = []
            duplicate_ids = []
            failed_ids = []
            packed_ids = []
            
            if log.batch_data:
                try:
                    batch_data = json.loads(log.batch_data)
                    new_ids = batch_data.get('new_ids', [])
                    duplicate_ids = batch_data.get('duplicate_ids', [])
                    failed_ids = batch_data.get('failed_ids', [])
                except:
                    pass

            # สร้างรายการ failed IDs ทั้งหมด (รวมซ้ำ) สำหรับแสดงใน Modal
            failed_ids_list = failed_ids.copy()
            # ถ้าจำนวนใน batch_data น้อยกว่า failed_count ให้สร้างรายการจำลองเพิ่ม
            if len(failed_ids_list) < log.failed_count:
                for i in range(len(failed_ids_list), log.failed_count):
                    failed_ids_list.append(f"Failed Order #{i + 1} (ไม่พบข้อมูล)")

            # ดึง Packed IDs จากฐานข้อมูล (ตาม import_date ของ log)
            if log.import_date:
                packed_orders = OrderLine.query.filter(
                    OrderLine.allocation_status == 'PACKED',
                    OrderLine.import_date >= log.import_date,
                    OrderLine.import_date <= log.import_date
                ).all()
                packed_ids = [order.order_id for order in packed_orders if order.order_id]

            return jsonify({
                "success": True,
                "log_id": log.id,
                "filename": log.filename,
                "import_date": log.import_date.strftime('%d/%m/%Y') if log.import_date else None,
                "new_ids": new_ids,
                "duplicate_ids": duplicate_ids,
                "failed_ids": failed_ids,  # เก่า: unique IDs
                "failed_ids_list": failed_ids_list,  # ใหม่: รายการทั้งหมด (รวมซ้ำ)
                "packed_ids": packed_ids,
                "added_count": log.added_count,
                "duplicates_count": log.duplicates_count or 0,
                "failed_count": log.failed_count
            })
            
        except Exception as e:
            app.logger.exception("Get Bill Empty Log Details Error")
            return jsonify({
                "success": False,
                "message": f"เกิดข้อผิดพลาด: {str(e)}"
            }), 500
    # =========[ /NEW ]=========

    # =========[ DEPRECATED: ยกเลิก - รวมไว้ใน Dashboard หลักแล้ว ]=========
    # @app.route("/dashboard/cancelled")
    # @login_required
    # def dashboard_cancelled():
    # def dashboard_cancelled():
    #     # สร้างตารางถ้ายังไม่มี (จากแพตช์ Import Orders ยกเลิก)
    #     ... (code commented out)
    #     return render_template(
    #         "dashboard_cancelled.html",
    #         rows=rows,
    #         q=q,
    #         platforms=platforms,
    #         shops=shops,
    #         platform_sel=platform_sel,
    #         shop_sel=shop_sel,
    #     )
    # =========[ /DEPRECATED ]=========

    # =========[ NEW ]=========  Dashboard: Order จ่ายแล้ว
    @app.route("/dashboard/issued")
    @login_required
    def dashboard_issued():
        if not current_user():
            return redirect(url_for("login"))

        q = (request.args.get("q") or "").strip()
        platform_sel = normalize_platform(request.args.get("platform"))
        shop_sel = request.args.get("shop_id")
        shop_sel = int(shop_sel) if shop_sel and str(shop_sel).isdigit() else None

        # Date range filter
        date_from_str = request.args.get("date_from") or ""
        date_to_str = request.args.get("date_to") or ""
        date_from_dt = None
        date_to_dt = None
        if date_from_str:
            try:
                date_from_dt = datetime.strptime(date_from_str, "%Y-%m-%d").replace(tzinfo=TH_TZ)
            except:
                pass
        if date_to_str:
            try:
                date_to_dt = datetime.strptime(date_to_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=TH_TZ)
            except:
                pass

        # สำหรับ dropdown เลือกแพลตฟอร์ม/ร้าน
        platforms = [p for (p,) in db.session.query(Shop.platform).filter(Shop.platform.isnot(None)).distinct().order_by(Shop.platform.asc()).all()]
        shop_query = Shop.query
        if platform_sel:
            shop_query = shop_query.filter(Shop.platform == platform_sel)
        shops = shop_query.order_by(Shop.name.asc()).all()

        # subquery map order_id -> (platform, shop_name, shop_id)
        sub = (
            db.session.query(
                OrderLine.order_id.label("oid"),
                func.min(OrderLine.shop_id).label("shop_id"),
                func.min(Shop.platform).label("platform"),
                func.min(Shop.name).label("shop_name"),
                func.min(OrderLine.logistic_type).label("logistic"),
            )
            .outerjoin(Shop, Shop.id == OrderLine.shop_id)
            .group_by(OrderLine.order_id)
            .subquery()
        )

        qry = (
            db.session.query(
                IssuedOrder.order_id,
                IssuedOrder.issued_at,
                sub.c.platform,
                sub.c.shop_name,
                sub.c.shop_id,
                sub.c.logistic,
            )
            .outerjoin(sub, sub.c.oid == IssuedOrder.order_id)
        )

        if q:
            qry = qry.filter(IssuedOrder.order_id.contains(q))
        if platform_sel:
            qry = qry.filter(sub.c.platform == platform_sel)
        if shop_sel:
            qry = qry.filter(sub.c.shop_id == shop_sel)
        if date_from_dt:
            qry = qry.filter(IssuedOrder.issued_at >= date_from_dt)
        if date_to_dt:
            qry = qry.filter(IssuedOrder.issued_at <= date_to_dt)

        rows = qry.order_by(IssuedOrder.issued_at.desc()).all()

        return render_template(
            "dashboard_issued.html",
            rows=rows, q=q, platforms=platforms, shops=shops,
            platform_sel=platform_sel, shop_sel=shop_sel,
            date_from_sel=date_from_str, date_to_sel=date_to_str
        )

    @app.post("/issued/unissue")
    @login_required
    def issued_unissue():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ต้องเป็นผู้ดูแลระบบหรือพนักงานเท่านั้น", "danger")
            return redirect(url_for("dashboard_issued"))

        ids = request.form.getlist("order_ids[]")
        if not ids:
            oid = request.form.get("order_id")
            if oid:
                ids = [oid]
        n = _unissue(ids or [])
        if n > 0:
            flash(f"ยกเลิกจ่ายงานแล้ว {n} ออเดอร์", "success")
        else:
            flash("ไม่พบออเดอร์ที่จะยกเลิกจ่ายงาน", "warning")
        return redirect(url_for("dashboard_issued"))
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  Dashboard: Order ที่ถูกลบ (Recycle Bin)
    @app.route("/dashboard/deleted")
    @login_required
    def dashboard_deleted():
        if not current_user():
            return redirect(url_for("login"))

        q = (request.args.get("q") or "").strip()
        platform_sel = normalize_platform(request.args.get("platform"))
        shop_sel = request.args.get("shop_id")
        shop_sel = int(shop_sel) if shop_sel and str(shop_sel).isdigit() else None

        # Date range filter
        date_from_str = request.args.get("date_from") or ""
        date_to_str = request.args.get("date_to") or ""
        date_from_dt = None
        date_to_dt = None
        if date_from_str:
            try:
                date_from_dt = datetime.strptime(date_from_str, "%Y-%m-%d").replace(tzinfo=TH_TZ)
            except:
                pass
        if date_to_str:
            try:
                date_to_dt = datetime.strptime(date_to_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=TH_TZ)
            except:
                pass

        # สำหรับ dropdown เลือกแพลตฟอร์ม/ร้าน
        platforms = [p for (p,) in db.session.query(Shop.platform).filter(Shop.platform.isnot(None)).distinct().order_by(Shop.platform.asc()).all()]
        shop_query = Shop.query
        if platform_sel:
            shop_query = shop_query.filter(Shop.platform == platform_sel)
        shops = shop_query.order_by(Shop.name.asc()).all()

        # subquery map order_id -> (platform, shop_name, shop_id, logistic)
        sub = (
            db.session.query(
                OrderLine.order_id.label("oid"),
                func.min(OrderLine.shop_id).label("shop_id"),
                func.min(Shop.platform).label("platform"),
                func.min(Shop.name).label("shop_name"),
                func.min(OrderLine.logistic_type).label("logistic"),
            )
            .outerjoin(Shop, Shop.id == OrderLine.shop_id)
            .group_by(OrderLine.order_id)
            .subquery()
        )

        qry = (
            db.session.query(
                DeletedOrder.order_id,
                DeletedOrder.deleted_at,
                sub.c.platform,
                sub.c.shop_name,
                sub.c.shop_id,
                sub.c.logistic,
                User.username.label("deleted_by")
            )
            .outerjoin(sub, sub.c.oid == DeletedOrder.order_id)
            .outerjoin(User, User.id == DeletedOrder.deleted_by_user_id)
        )

        if q:
            qry = qry.filter(DeletedOrder.order_id.contains(q))
        if platform_sel:
            qry = qry.filter(sub.c.platform == platform_sel)
        if shop_sel:
            qry = qry.filter(sub.c.shop_id == shop_sel)
        if date_from_dt:
            qry = qry.filter(DeletedOrder.deleted_at >= date_from_dt)
        if date_to_dt:
            qry = qry.filter(DeletedOrder.deleted_at <= date_to_dt)

        rows = qry.order_by(DeletedOrder.deleted_at.desc()).all()

        return render_template(
            "dashboard_deleted.html",
            rows=rows, q=q, platforms=platforms, shops=shops,
            platform_sel=platform_sel, shop_sel=shop_sel,
            date_from_sel=date_from_str, date_to_sel=date_to_str
        )

    @app.post("/deleted/restore")
    @login_required
    def deleted_restore():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ต้องเป็นผู้ดูแลระบบหรือพนักงานเท่านั้น", "danger")
            return redirect(url_for("dashboard_deleted"))

        ids = request.form.getlist("order_ids[]")
        if not ids:
            oid = request.form.get("order_id")
            if oid:
                ids = [oid]
        
        if not ids:
            flash("ไม่พบรายการที่จะกู้คืน", "warning")
            return redirect(url_for("dashboard_deleted"))

        # ลบออกจาก DeletedOrder = กู้คืนกลับหน้าหลัก
        n = db.session.query(DeletedOrder).filter(DeletedOrder.order_id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        
        if n > 0:
            flash(f"✅ กู้คืน {n} ออเดอร์ เรียบร้อยแล้ว", "success")
        else:
            flash("ไม่พบออเดอร์ที่จะกู้คืน", "warning")
        return redirect(url_for("dashboard_deleted"))
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  System Status Page (ตรวจสอบการเชื่อมต่อ Railway Volume)
    @app.route("/system-status")
    @login_required
    def system_status():
        """หน้าตรวจสอบสถานะระบบและการเชื่อมต่อ Database"""
        import sys
        import platform
        import flask
        import sqlalchemy

        # 1. Database Paths Information
        volume_path = os.environ.get("RAILWAY_VOLUME_MOUNT_PATH")
        db_location = "Railway Volume (Production)" if volume_path else "Local Filesystem (Development)"

        # Get all 3 database paths
        db_paths = get_db_paths()

        # 2. Database Info for all 3 DBs
        databases = {
            "data": {
                "name": "Main Database (data.db)",
                "path": db_paths["data"],
                **get_db_info(db_paths["data"])
            },
            "price": {
                "name": "Price Database (price.db)",
                "path": db_paths["price"],
                **get_db_info(db_paths["price"])
            },
            "supplier": {
                "name": "Supplier Database (supplier_stock.db)",
                "path": db_paths["supplier"],
                **get_db_info(db_paths["supplier"])
            }
        }

        # 3. Count Records from main database
        try:
            total_orders = db.session.query(func.count(func.distinct(OrderLine.order_id))).scalar() or 0
            total_products = db.session.query(func.count(Product.id)).scalar() or 0
            total_shops = db.session.query(func.count(Shop.id)).scalar() or 0
            total_users = db.session.query(func.count(User.id)).scalar() or 0
        except:
            total_orders = total_products = total_shops = total_users = 0

        # 4. System Information
        python_version = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
        flask_version = flask.__version__
        sqlalchemy_version = sqlalchemy.__version__
        os_info = f"{platform.system()} {platform.release()}"

        # 5. Environment Variables
        env_vars = {
            "RAILWAY_VOLUME_MOUNT_PATH": os.environ.get("RAILWAY_VOLUME_MOUNT_PATH", "Not Set"),
            "SECRET_KEY": "***" if os.environ.get("SECRET_KEY") else "Default (vnix-secret)",
            "APP_NAME": os.environ.get("APP_NAME", "VNIX ERP"),
        }

        status_info = {
            "databases": databases,
            "db_location": db_location,
            "volume_path": volume_path or "Not Configured",
            "total_orders": total_orders,
            "total_products": total_products,
            "total_shops": total_shops,
            "total_users": total_users,
            "python_version": python_version,
            "flask_version": flask_version,
            "sqlalchemy_version": sqlalchemy_version,
            "os_info": os_info,
            "env_vars": env_vars,
        }

        return render_template("system_status.html", status=status_info)

    # =========[ Database Download Routes ]=========
    @app.route("/api/database/download/<db_type>")
    @login_required
    def download_database(db_type):
        """Download individual database file"""
        user = current_user()
        if not user or user.role != 'admin':
            return jsonify({'error': 'ไม่มีสิทธิ์ดาวน์โหลดฐานข้อมูล (Admin only)'}), 403

        db_paths = get_db_paths()

        if db_type not in db_paths:
            return jsonify({'error': 'ประเภทฐานข้อมูลไม่ถูกต้อง'}), 400

        db_path = db_paths[db_type]

        if not os.path.exists(db_path):
            return jsonify({'error': f'ไม่พบไฟล์ฐานข้อมูล {db_type}.db'}), 404

        filename = os.path.basename(db_path)
        return send_file(db_path, as_attachment=True, download_name=filename)

    @app.route("/api/database/download-all")
    @login_required
    def download_all_databases():
        """Download all databases as ZIP file"""
        import zipfile

        user = current_user()
        if not user or user.role != 'admin':
            return jsonify({'error': 'ไม่มีสิทธิ์ดาวน์โหลดฐานข้อมูล (Admin only)'}), 403

        db_paths = get_db_paths()

        # Create temporary ZIP file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_filename = f'vnix_databases_{timestamp}.zip'
        zip_buffer = BytesIO()

        try:
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for db_type, db_path in db_paths.items():
                    if os.path.exists(db_path):
                        arcname = os.path.basename(db_path)
                        zip_file.write(db_path, arcname=arcname)

            zip_buffer.seek(0)
            return send_file(
                zip_buffer,
                mimetype='application/zip',
                as_attachment=True,
                download_name=zip_filename
            )
        except Exception as e:
            app.logger.exception("Download all databases failed")
            return jsonify({'error': str(e)}), 500

    # =========[ Database Backup Route ]=========
    @app.route("/api/database/backup/<db_type>", methods=['POST'])
    @login_required
    def backup_database(db_type):
        """Create manual backup of database file"""
        user = current_user()
        if not user or user.role != 'admin':
            return jsonify({'error': 'ไม่มีสิทธิ์สร้าง backup (Admin only)'}), 403

        db_paths = get_db_paths()

        if db_type not in db_paths:
            return jsonify({'error': 'ประเภทฐานข้อมูลไม่ถูกต้อง'}), 400

        db_path = db_paths[db_type]

        if not os.path.exists(db_path):
            return jsonify({'error': f'ไม่พบไฟล์ฐานข้อมูล {db_type}.db'}), 404

        try:
            db_dir = os.path.dirname(db_path)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = f'{db_type}_manual_backup_{timestamp}.db'
            backup_path = os.path.join(db_dir, backup_filename)

            import shutil
            shutil.copy2(db_path, backup_path)

            # Get backup file size
            size_mb = os.path.getsize(backup_path) / (1024 * 1024)

            app.logger.info(f"Manual backup created by {user.username}: {backup_filename}")

            return jsonify({
                'success': True,
                'message': 'Backup created successfully',
                'backup_file': backup_filename,
                'size': f'{size_mb:.2f} MB'
            })

        except Exception as e:
            app.logger.exception(f"Backup creation failed for {db_type}")
            return jsonify({'error': str(e)}), 500

    # =========[ Database Upload Routes ]=========
    @app.route("/api/database/upload/<db_type>", methods=['POST'])
    @login_required
    def upload_database(db_type):
        """Upload and replace individual database file"""
        user = current_user()
        if not user or user.role != 'admin':
            return jsonify({'error': 'ไม่มีสิทธิ์อัปโหลดฐานข้อมูล (Admin only)'}), 403

        db_paths = get_db_paths()

        if db_type not in db_paths:
            return jsonify({'error': 'ประเภทฐานข้อมูลไม่ถูกต้อง'}), 400

        # Check if file is in request
        if 'database' not in request.files:
            return jsonify({'error': 'ไม่พบไฟล์ที่อัปโหลด'}), 400

        file = request.files['database']

        # Check filename
        if file.filename == '':
            return jsonify({'error': 'ไม่ได้เลือกไฟล์'}), 400

        # Check file extension
        if not file.filename.endswith('.db'):
            return jsonify({'error': 'กรุณาอัปโหลดไฟล์ .db เท่านั้น'}), 400

        try:
            db_path = db_paths[db_type]
            db_dir = os.path.dirname(db_path)

            # Ensure directory exists
            os.makedirs(db_dir, exist_ok=True)

            # Create backup before replacing
            backup_filename = None
            if os.path.exists(db_path):
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_filename = f'{db_type}_backup_{timestamp}.db'
                backup_path = os.path.join(db_dir, backup_filename)

                import shutil
                shutil.copy2(db_path, backup_path)
                app.logger.info(f"Backup created: {backup_filename}")

            # Save uploaded file
            file.save(db_path)

            # Get file size
            size_mb = os.path.getsize(db_path) / (1024 * 1024)

            app.logger.info(f"Database {db_type}.db uploaded by {user.username}: {size_mb:.2f} MB")

            return jsonify({
                'success': True,
                'message': f'อัปโหลดและแทนที่ {db_type}.db สำเร็จ',
                'size': f'{size_mb:.2f} MB',
                'backup': backup_filename
            })

        except Exception as e:
            app.logger.exception(f"Database upload failed for {db_type}")
            return jsonify({'error': str(e)}), 500

    # =========[ /NEW ]=========

    @app.route("/import/products", methods=["GET", "POST"])
    @login_required
    def import_products_view():
        # --- ส่วนที่ 1: ดึง URL ที่บันทึกไว้ ---
        saved_url = ""
        CONFIG_SHOP_NAME = "GoogleSheet_Products"

        try:
            config_row = db.session.execute(
                text("SELECT google_sheet_url FROM shops WHERE platform = 'PRODUCTS_SYSTEM' AND name = :name LIMIT 1"),
                {"name": CONFIG_SHOP_NAME}
            ).fetchone()
            if config_row and config_row[0]:
                saved_url = config_row[0]
        except Exception:
            db.session.rollback()

        # --- ส่วนที่ 2: จัดการนำเข้า (POST) ---
        if request.method == "POST":
            mode = request.form.get("mode")
            df = None
            source_name = "Unknown"

            try:
                # >>>> Case 1: Google Sheet
                if mode == "gsheet":
                    sheet_url = request.form.get("sheet_url")
                    if not sheet_url:
                        flash("กรุณาระบุ Google Sheet URL", "danger")
                        return redirect(url_for("import_products_view"))

                    creds = get_google_credentials()
                    client = gspread.authorize(creds)

                    try:
                        sh = client.open_by_url(sheet_url)
                        worksheet = sh.worksheet("Import_product_master")
                    except gspread.WorksheetNotFound:
                        flash("ไม่พบ Tab ชื่อ 'Import_product_master'", "danger")
                        return redirect(url_for("import_products_view"))

                    data = worksheet.get_all_records()
                    if not data:
                        flash("ไม่มีข้อมูลใน Tab", "warning")
                        return redirect(url_for("import_products_view"))

                    df = pd.DataFrame(data)
                    source_name = "Google Sheet"

                    # Auto-save URL
                    try:
                        s = Shop.query.filter_by(platform='PRODUCTS_SYSTEM', name=CONFIG_SHOP_NAME).first()
                        if not s:
                            s = Shop(platform='PRODUCTS_SYSTEM', name=CONFIG_SHOP_NAME, is_system_config=True)
                            db.session.add(s)
                            db.session.commit()
                        db.session.execute(
                            text("UPDATE shops SET google_sheet_url = :u WHERE id = :id"),
                            {"u": sheet_url, "id": s.id}
                        )
                        db.session.commit()
                    except Exception as e_save:
                        db.session.rollback()
                        app.logger.error(f"Auto-save URL failed: {e_save}")

                # >>>> Case 2: File Upload
                else:
                    f = request.files.get("file")
                    if not f:
                        flash("กรุณาเลือกไฟล์", "danger")
                        return redirect(url_for("import_products_view"))
                    df = pd.read_excel(f)
                    source_name = f.filename

                # >>>> Process Import
                if df is not None:
                    # ลบแถวว่างท้ายไฟล์ทิ้ง
                    df.dropna(how='all', inplace=True)

                    cnt = import_products(df)

                    flash(f"✅ นำเข้าสินค้าสำเร็จ {cnt} รายการ (จาก {source_name})", "success")
                    return redirect(url_for("import_products_view"))

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import Products Error")
                flash(f"เกิดข้อผิดพลาด: {e}", "danger")
                return redirect(url_for("import_products_view"))

        # --- ส่วนที่ 3: นับจำนวน SKU ทั้งหมด ---
        total_skus = 0
        try:
            total_skus = Product.query.count()
        except Exception:
            pass

        return render_template("import_products.html", saved_url=saved_url, total_skus=total_skus)

    @app.route("/import/stock", methods=["GET", "POST"])
    @login_required
    def import_stock_view():
        # --- ส่วนที่ 1: ดึง URL ที่บันทึกไว้มาแสดง (GET) ---
        saved_url = ""
        try:
            # ใช้ platform='STOCK_SYSTEM' และ name='SabuySoft' เพื่อเก็บ URL
            config_row = db.session.execute(
                text("SELECT google_sheet_url FROM shops WHERE platform = 'STOCK_SYSTEM' AND name = 'SabuySoft' LIMIT 1")
            ).fetchone()
            if config_row and config_row[0]:
                saved_url = config_row[0]
        except Exception:
            pass

        # --- ส่วนที่ 2: จัดการการนำเข้า (POST) ---
        if request.method == "POST":
            mode = request.form.get("mode")  # 'file' หรือ 'gsheet'
            
            try:
                df = None
                
                # ==== กรณีนำเข้าผ่าน Google Sheet ====
                if mode == "gsheet":
                    sheet_url = request.form.get("sheet_url")
                    if not sheet_url:
                        flash("กรุณาระบุ Google Sheet URL", "danger")
                        return redirect(url_for("import_stock_view"))
                    
                    # 1. เชื่อมต่อ Google API
                    creds = get_google_credentials()
                    client = gspread.authorize(creds)
                    
                    # 2. เปิด Sheet และ Tab
                    try:
                        sh = client.open_by_url(sheet_url)
                        worksheet = sh.worksheet("Import_sabuysoft_stock")  # ตามชื่อที่ระบุ
                    except gspread.WorksheetNotFound:
                        flash("ไม่พบ Tab ชื่อ 'Import_sabuysoft_stock'", "danger")
                        return redirect(url_for("import_stock_view"))
                    except Exception as e:
                        flash(f"เข้าถึง Google Sheet ไม่ได้: {e}", "danger")
                        return redirect(url_for("import_stock_view"))
                        
                    # 3. ดึงข้อมูลแปลงเป็น DataFrame
                    data = worksheet.get_all_records()
                    if not data:
                        flash("ไม่พบข้อมูลใน Tab Import_sabuysoft_stock", "warning")
                        return redirect(url_for("import_stock_view"))
                    
                    df = pd.DataFrame(data)
                    
                    # (Optional) บันทึก URL ล่าสุดอัตโนมัติถ้าทำรายการสำเร็จ
                    try:
                        # หา Shop หรือสร้างใหม่
                        s = Shop.query.filter_by(platform='STOCK_SYSTEM', name='SabuySoft').first()
                        if not s:
                            s = Shop(platform='STOCK_SYSTEM', name='SabuySoft', is_system_config=True)
                            db.session.add(s)
                            db.session.commit()
                        
                        # Update URL
                        db.session.execute(
                            text("UPDATE shops SET google_sheet_url = :u WHERE id = :id"),
                            {"u": sheet_url, "id": s.id}
                        )
                        db.session.commit()
                    except Exception:
                        pass  # ถ้าบันทึก URL อัตโนมัติไม่ได้ ก็ไม่เป็นไร ให้ Import ต่อไป

                # ==== กรณีนำเข้าผ่านไฟล์ Excel ====
                else:
                    f = request.files.get("file")
                    if not f:
                        flash("กรุณาเลือกไฟล์สต็อก", "danger")
                        return redirect(url_for("import_stock_view"))
                    df = pd.read_excel(f)

                # ==== ส่ง DataFrame ไปเข้าฟังก์ชัน import_stock (Full Sync Mode) ====
                if df is not None:
                    cnt = import_stock(df, full_replace=True)
                    source_text = "Google Sheet" if mode == "gsheet" else "ไฟล์"
                    flash(f"✅ นำเข้าสต็อกสำเร็จ {cnt} SKU (Full Sync: SKU ที่ไม่อยู่ในไฟล์จะถูกตั้งเป็น 0) [จาก {source_text}]", "success")
                    return redirect(url_for("import_stock_view"))

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import Stock Error")
                flash(f"เกิดข้อผิดพลาดในการนำเข้าสต็อก: {e}", "danger")
                return redirect(url_for("import_stock_view"))

        return render_template("import_stock.html", saved_url=saved_url)

    @app.route("/import/sales", methods=["GET", "POST"])
    @login_required
    def import_sales_view():
        # --- ส่วนที่ 1: ดึง URL ที่บันทึกไว้ ---
        saved_url = ""
        CONFIG_SHOP_NAME = "GoogleSheet_Sales" 
        
        try:
            config_row = db.session.execute(
                text("SELECT google_sheet_url FROM shops WHERE platform = 'SALES_SYSTEM' AND name = :name LIMIT 1"),
                {"name": CONFIG_SHOP_NAME}
            ).fetchone()
            if config_row and config_row[0]:
                saved_url = config_row[0]
        except Exception:
            db.session.rollback()

        # ตัวแปร Filter วันที่
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        if not date_from_str: date_from_str = now_thai().date().isoformat()
        if not date_to_str: date_to_str = now_thai().date().isoformat()
        d_from = parse_date_any(date_from_str)
        d_to = parse_date_any(date_to_str)

        # --- ส่วนที่ 2: จัดการนำเข้า (POST) ---
        if request.method == "POST":
            mode = request.form.get("mode")
            df = None
            source_name = "Unknown"
            
            try:
                # >>>> Case 1: Google Sheet
                if mode == "gsheet":
                    sheet_url = request.form.get("sheet_url")
                    if not sheet_url:
                        flash("กรุณาระบุ URL", "danger")
                        return redirect(url_for("import_sales_view"))
                    
                    creds = get_google_credentials()
                    client = gspread.authorize(creds)
                    
                    try:
                        sh = client.open_by_url(sheet_url)
                        worksheet = sh.worksheet("Import_sabuysoft_sales_10d")
                    except gspread.WorksheetNotFound:
                        flash("ไม่พบ Tab ชื่อ 'Import_sabuysoft_sales_10d'", "danger")
                        return redirect(url_for("import_sales_view"))
                    
                    data = worksheet.get_all_records()
                    if not data:
                        flash("ไม่มีข้อมูลใน Tab", "warning")
                        return redirect(url_for("import_sales_view"))
                    
                    df = pd.DataFrame(data)
                    source_name = "Google Sheet"
                    
                    # Auto-save URL
                    try:
                        s = Shop.query.filter_by(platform='SALES_SYSTEM', name=CONFIG_SHOP_NAME).first()
                        if not s:
                            s = Shop(platform='SALES_SYSTEM', name=CONFIG_SHOP_NAME, is_system_config=True)
                            db.session.add(s)
                            db.session.commit()
                        db.session.execute(
                            text("UPDATE shops SET google_sheet_url = :u WHERE id = :id"),
                            {"u": sheet_url, "id": s.id}
                        )
                        db.session.commit()
                    except Exception as e_save:
                        db.session.rollback()
                        app.logger.error(f"Auto-save URL failed: {e_save}")

                # >>>> Case 2: File Upload
                else:
                    f = request.files.get("file")
                    if not f:
                        flash("กรุณาเลือกไฟล์", "danger")
                        return redirect(url_for("import_sales_view"))
                    df = pd.read_excel(f)
                    source_name = f.filename

                # >>>> Process Import
                if df is not None:
                    # [แก้จุด A] ลบแถวว่างท้ายไฟล์ทิ้ง (Clean Empty Rows)
                    df.dropna(how='all', inplace=True)

                    # นับจำนวนบรรทัดที่มีข้อมูล (Total)
                    total_rows = len(df)

                    # [แก้จุด B] เรียก Importer และรับ Dict กลับมา
                    result = import_sales(df)
                    success_ids = result.get('ids', [])
                    skipped_rows = result.get('skipped', [])

                    cnt = len(success_ids)
                    failed_cnt = len(skipped_rows)

                    # Logging: บันทึกรายละเอียด error ลง application log
                    if failed_cnt > 0:
                        app.logger.warning(f"[Import Sales] พบ {failed_cnt} รายการที่ไม่สำเร็จจากทั้งหมด {total_rows} รายการ")

                        # แสดง error สูงสุด 5 รายการแรกใน log
                        error_summary = {}
                        for skip in skipped_rows[:5]:
                            reason = skip.get('reason', 'ไม่ทราบสาเหตุ')
                            error_summary[reason] = error_summary.get(reason, 0) + 1
                            app.logger.warning(
                                f"  - Row {skip.get('row_number', 'N/A')}: {reason} "
                                f"(Order ID: {skip.get('order_id', 'N/A')}, PO: {skip.get('po_no', 'N/A')})"
                            )

                        if failed_cnt > 5:
                            app.logger.warning(f"  ... และอีก {failed_cnt - 5} รายการ (ดูรายละเอียดใน batch_data)")

                    # [แก้จุด C] บันทึก Log พร้อมรายชื่อ ID และข้อมูล Failed
                    log = ImportLog(
                        import_date=now_thai().date(),
                        platform="SALES_SYSTEM",
                        shop_name="-",
                        filename=source_name,
                        added_count=cnt,
                        duplicates_count=0,
                        failed_count=failed_cnt,
                        batch_data=json.dumps({
                            "ids": success_ids,
                            "skipped": skipped_rows  # เก็บรายละเอียด failed items
                        })
                    )
                    db.session.add(log)
                    db.session.commit()

                    # สร้างข้อความแจ้งเตือนที่มีรายละเอียดมากขึ้น
                    msg = f"✅ อัปเดตข้อมูลสั่งขายสำเร็จ {cnt} รายการ"
                    if failed_cnt > 0:
                        # นับประเภทของ error
                        error_reasons = {}
                        for skip in skipped_rows:
                            reason = skip.get('reason', 'ไม่ทราบสาเหตุ')
                            error_reasons[reason] = error_reasons.get(reason, 0) + 1

                        msg += f" (ไม่สำเร็จ {failed_cnt} รายการ)"

                        # แสดงสาเหตุหลักๆ (สูงสุด 3 ประเภท)
                        if error_reasons:
                            top_errors = sorted(error_reasons.items(), key=lambda x: x[1], reverse=True)[:3]
                            error_detail = ", ".join([f"{reason}: {count}" for reason, count in top_errors])
                            msg += f" | สาเหตุหลัก: {error_detail}"

                    flash(msg, "success" if failed_cnt == 0 else "warning")
                    return redirect(url_for("import_sales_view"))

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import Sales Error")
                flash(f"เกิดข้อผิดพลาด: {e}", "danger")
                return redirect(url_for("import_sales_view"))

        # --- ส่วนที่ 3: คำนวณ Dashboard Stats (นับ Unique) ---
        total_unique_success = 0
        total_failed = 0
        
        try:
            logs = ImportLog.query.filter(
                ImportLog.import_date >= d_from,
                ImportLog.import_date <= d_to,
                ImportLog.platform == 'SALES_SYSTEM'
            ).all()
            
            # [แก้จุด D] ใช้ Set เพื่อตัด Order ID ที่ซ้ำกันออก (สำหรับการ์ด Success)
            unique_ids_set = set()
            
            for l in logs:
                total_failed += (l.failed_count or 0)
                # แกะ batch_data เพื่อเอา ID มารวมใน Set
                if l.batch_data:
                    try:
                        data = json.loads(l.batch_data)
                        if "ids" in data:
                            unique_ids_set.update(data["ids"])
                    except:
                        pass
                else:
                    # Fallback สำหรับ Log เก่าที่ไม่มี batch_data
                    pass 

            # ถ้ายอด Set มีค่า (Log ใหม่) ให้ใช้ยอดนั้น
            if len(unique_ids_set) > 0:
                total_unique_success = len(unique_ids_set)
            else:
                # ถ้าเป็น Log เก่า (ก่อนแก้โค้ด) ให้ใช้การบวกยอดเอา
                total_unique_success = sum(l.added_count for l in logs)

        except:
            total_unique_success = 0
            total_failed = 0
        
        stats = {"success": total_unique_success, "failed": total_failed}
        date_from_thai = to_be_date_str(d_from) if d_from else ""
        date_to_thai = to_be_date_str(d_to) if d_to else ""

        return render_template(
            "import_sales.html",
            stats=stats,
            saved_url=saved_url,
            date_from=date_from_str,
            date_to=date_to_str,
            date_from_thai=date_from_thai,
            date_to_thai=date_to_thai
        )

    # =========[ NEW ]========= ล้างประวัติ Sales
    @app.route("/import/sales/clear_log", methods=["POST"])
    @login_required
    def clear_sales_log():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ลบประวัติ", "danger")
            return redirect(url_for("import_sales_view"))
            
        mode = request.form.get("mode") # 'range' or 'all'
        delete_data = request.form.get("delete_data") # 'yes' check
        
        try:
            log_deleted = 0
            data_deleted = 0
            
            if mode == 'all':
                log_deleted = ImportLog.query.filter_by(platform='SALES_SYSTEM').delete()
                if delete_data == 'yes':
                    data_deleted = db.session.query(Sales).delete()
            else:
                d_from = datetime.strptime(request.form.get("date_from"), "%Y-%m-%d").date()
                d_to = datetime.strptime(request.form.get("date_to"), "%Y-%m-%d").date()
                
                log_deleted = ImportLog.query.filter(
                    ImportLog.platform == 'SALES_SYSTEM',
                    ImportLog.import_date >= d_from,
                    ImportLog.import_date <= d_to
                ).delete(synchronize_session=False)
                
                # Sales ไม่มี import_date ใน Model มาตรฐาน (ปกติมีแค่ created_at)
                # ถ้าต้องการลบ data แบบช่วงเวลา อาจต้องใช้ created_at แต่เพื่อความปลอดภัย
                # ถ้าเป็น range จะลบแค่ log เว้นแต่ Sales มี import_date
                if delete_data == 'yes' and hasattr(Sales, 'created_at'):
                     dt_start = datetime.combine(d_from, datetime.min.time())
                     dt_end = datetime.combine(d_to, datetime.max.time())
                     data_deleted = Sales.query.filter(
                         Sales.created_at >= dt_start,
                         Sales.created_at <= dt_end
                     ).delete(synchronize_session=False)

            db.session.commit()
            msg = f"ล้างประวัติเรียบร้อย (Log {log_deleted})"
            if data_deleted: msg += f" (Data {data_deleted})"
            flash(msg, "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error: {e}", "danger")
            
        return redirect(url_for("import_sales_view"))
    # =========[ /NEW ]=========

    # -----------------------
    # Accept / Cancel / Bulk
    # -----------------------

    def _reserved_qty_for_sku(sku: str, exclude_line_id: int | None = None) -> int:
        """ยอดที่ถูก 'จองแล้ว' สำหรับ SKU นี้

        Rules:
        - นับ: Accepted + Issued
        - ไม่นับ: Cancelled
        - ไม่นับ: Packed (ถือว่าจบกระบวนการ/ถูกหักออกจากสต็อกหลักแล้ว)
        """
        issued_subq = db.session.query(IssuedOrder.order_id)
        cancelled_subq = db.session.query(CancelledOrder.order_id)

        sales_status_norm = func.lower(func.coalesce(Sales.status, ""))
        packed_oids_subq = (
            db.session.query(Sales.order_id)
            .filter(
                or_(
                    func.upper(func.coalesce(Sales.status, "")) == "PACKED",
                    sales_status_norm.like("%packed%"),
                    sales_status_norm.like("%แพ็ค%"),
                    sales_status_norm.like("%ครบตามจำนวน%"),
                    sales_status_norm.like("%เปิดใบขายครบตามจำนวนแล้ว%"),
                    sales_status_norm.like("%opened_full%"),
                )
            )
            .distinct()
        )

        q = db.session.query(func.coalesce(func.sum(OrderLine.qty), 0)).filter(OrderLine.sku == sku)
        if exclude_line_id is not None:
            q = q.filter(OrderLine.id != exclude_line_id)

        q = (
            q.filter(OrderLine.order_id.notin_(cancelled_subq))
            .filter(or_(OrderLine.accepted.is_(True), OrderLine.order_id.in_(issued_subq)))
            .filter(not_(OrderLine.order_id.in_(packed_oids_subq)))
        )

        return int(q.scalar() or 0)

    @app.route("/accept/<int:order_line_id>", methods=["POST"])
    @login_required
    def accept_order(order_line_id):
        ol = OrderLine.query.get_or_404(order_line_id)
        # ห้ามกดรับถ้าเลข Order ถูกทำเป็นยกเลิก
        if db.session.query(CancelledOrder.id).filter_by(order_id=ol.order_id).first():
            flash(f"Order {ol.order_id} ถูกทำเป็น 'ยกเลิก' แล้ว — ไม่สามารถกดรับได้", "warning")
            return redirect(url_for("dashboard", **request.args))

        cu = current_user()
        sales_status = (getattr(ol, "sales_status", "") or "").upper()
        if sales_status == "PACKED" or bool(getattr(ol, "packed", False)):
            flash("รายการนี้ถูกแพ็คแล้ว (PACKED) — ไม่สามารถกดรับได้", "warning")
            return redirect(url_for("dashboard", **request.args))

        stock_qty = _calc_stock_qty_for_line(ol)
        if stock_qty <= 0:
            flash("สต็อกหมด — ไม่สามารถกดรับได้", "warning")
            return redirect(url_for("dashboard", **request.args))

        sku = _get_line_sku(ol)
        if not sku:
            flash("ไม่พบ SKU ของรายการนี้ — ไม่สามารถกดรับได้", "warning")
            return redirect(url_for("dashboard", **request.args))

        used_qty = _reserved_qty_for_sku(sku, exclude_line_id=ol.id)
        proposed_total = int(used_qty) + int(ol.qty or 0)
        if proposed_total > int(stock_qty):
            remain_real = max(0, int(stock_qty) - int(used_qty))
            over = proposed_total - int(stock_qty)
            flash(
                f"สต็อกไม่พอ! (มี {stock_qty}, ถูกจองแล้ว {used_qty}, เหลือให้รับได้ {remain_real}) — ต้องการรับ {int(ol.qty or 0)} (เกิน {over})",
                "warning",
            )
            return redirect(url_for("dashboard", **request.args))

        ol.accepted = True
        ol.accepted_at = now_thai()
        ol.accepted_by_user_id = cu.id if cu else None
        ol.accepted_by_username = cu.username if cu else None
        db.session.commit()
        flash(f"ทำเครื่องหมายกดรับ Order {ol.order_id} • SKU {sku} แล้ว", "success")
        return redirect(url_for("dashboard", **request.args))

    @app.route("/cancel_accept/<int:order_line_id>", methods=["POST"])
    @login_required
    def cancel_accept(order_line_id):
        ol = OrderLine.query.get_or_404(order_line_id)
        ol.accepted = False
        ol.accepted_at = None
        ol.accepted_by_user_id = None
        ol.accepted_by_username = None
        db.session.commit()
        flash(f"ยกเลิกการกดรับ Order {ol.order_id} • SKU {getattr(ol, 'sku', '')}", "warning")
        return redirect(url_for("dashboard", **request.args))

    # =========[ NEW ]========= ฟังก์ชั่นยกเลิก Order ถาวร (พร้อมเหตุผล)
    @app.post("/cancel_order_permanent")
    @login_required
    def cancel_order_permanent():
        """ยกเลิก Order ถาวร พร้อมบันทึกเหตุผล - ใช้ได้ทุกเวลา ทั้งก่อน/หลังจ่ายงาน"""
        cu = current_user()
        order_id = (request.form.get("order_id") or "").strip()
        reason = (request.form.get("reason") or "").strip()

        if not order_id:
            flash("ไม่พบเลข Order", "danger")
            return redirect(url_for("dashboard", **request.args))
        
        if not reason:
            flash("กรุณาระบุเหตุผลการยกเลิก", "warning")
            return redirect(url_for("dashboard", **request.args))

        # ตรวจสอบว่ามีอยู่แล้วหรือไม่ (อัปเดตเหตุผลใหม่ได้)
        existing = CancelledOrder.query.filter_by(order_id=order_id).first()
        if existing:
            existing.note = reason
            existing.imported_by_user_id = cu.id if cu else None
            existing.imported_at = datetime.utcnow()
            flash(f"อัปเดตข้อมูลการยกเลิก Order {order_id} แล้ว (เหตุผล: {reason})", "info")
        else:
            new_cancel = CancelledOrder(
                order_id=order_id, 
                note=reason, 
                imported_by_user_id=cu.id if cu else None,
                imported_at=datetime.utcnow()
            )
            db.session.add(new_cancel)
            flash(f"ยกเลิก Order {order_id} สำเร็จ (เหตุผล: {reason})", "success")

        db.session.commit()
        return redirect(url_for("dashboard", **request.args))
    # =========[ /NEW ]=========

    @app.route("/bulk_accept", methods=["POST"])
    @login_required
    def bulk_accept():
        cu = current_user()
        order_line_ids = request.form.getlist("order_line_ids[]")
        if not order_line_ids:
            flash("กรุณาเลือกรายการที่ต้องการกดรับ", "warning")
            return redirect(url_for("dashboard", **request.args))
        success_count = 0
        error_messages = []
        for ol_id in order_line_ids:
            try:
                ol = db.session.get(OrderLine, int(ol_id))
                if not ol:
                    continue
                # [NEW] block ถ้าจ่ายงานแล้ว
                if db.session.query(IssuedOrder.id).filter_by(order_id=ol.order_id).first():
                    error_messages.append(f"Order {ol.order_id} จ่ายงานแล้ว")
                    continue
                # block ถ้ายกเลิก
                if db.session.query(CancelledOrder.id).filter_by(order_id=ol.order_id).first():
                    error_messages.append(f"Order {ol.order_id} ถูกยกเลิก")
                    continue
                sales_status = (getattr(ol, "sales_status", "") or "").upper()
                if sales_status == "PACKED" or bool(getattr(ol, "packed", False)):
                    error_messages.append(f"Order {ol.order_id} ถูกแพ็คแล้ว")
                    continue
                stock_qty = _calc_stock_qty_for_line(ol)
                if stock_qty <= 0:
                    error_messages.append(f"Order {ol.order_id} สต็อกหมด")
                    continue
                # [NEW] ป้องกัน Low Stock (สินค้าน้อย <= 3 ชิ้น) ห้ามกดรับแบบกลุ่ม
                if stock_qty <= 3:
                    error_messages.append(f"Order {ol.order_id} สินค้าน้อย (Low Stock) - กรุณาตรวจสอบและกดรับรายออเดอร์")
                    continue
                sku = _get_line_sku(ol)
                if not sku:
                    error_messages.append(f"Order {ol.order_id} ไม่พบ SKU")
                    continue

                used_qty = _reserved_qty_for_sku(sku, exclude_line_id=ol.id)
                proposed_total = int(used_qty) + int(ol.qty or 0)
                if proposed_total > int(stock_qty):
                    remain_real = max(0, int(stock_qty) - int(used_qty))
                    error_messages.append(
                        f"Order {ol.order_id} สินค้าไม่พอส่ง (มี {stock_qty}, จองแล้ว {used_qty}, เหลือ {remain_real})"
                    )
                    continue
                ol.accepted = True
                ol.accepted_at = now_thai()
                ol.accepted_by_user_id = cu.id if cu else None
                ol.accepted_by_username = cu.username if cu else None
                success_count += 1
            except Exception as e:
                error_messages.append(f"Order ID {ol_id}: {str(e)}")
                continue
        db.session.commit()
        if success_count > 0:
            flash(f"✅ กดรับสำเร็จ {success_count} รายการ", "success")
        if error_messages:
            for msg in error_messages[:5]:
                flash(f"⚠️ {msg}", "warning")
            if len(error_messages) > 5:
                flash(f"... และอีก {len(error_messages) - 5} รายการที่ไม่สามารถกดรับได้", "warning")
        return redirect(url_for("dashboard", **request.args))

    @app.route("/bulk_cancel", methods=["POST"])
    @login_required
    def bulk_cancel():
        order_line_ids = request.form.getlist("order_line_ids[]")
        if not order_line_ids:
            flash("กรุณาเลือกรายการที่ต้องการยกเลิก", "warning")
            return redirect(url_for("dashboard", **request.args))
        success_count = 0
        for ol_id in order_line_ids:
            try:
                ol = db.session.get(OrderLine, int(ol_id))
                if ol:
                    ol.accepted = False
                    ol.accepted_at = None
                    ol.accepted_by_user_id = None
                    ol.accepted_by_username = None
                    success_count += 1
            except Exception:
                continue
        db.session.commit()
        if success_count > 0:
            flash(f"✅ ยกเลิกสำเร็จ {success_count} รายการ", "success")
        return redirect(url_for("dashboard", **request.args))

    # ================== NEW: Bulk Delete Orders (เปลี่ยนเป็น Soft Delete) ==================
    @app.route("/bulk_delete_orders", methods=["POST"])
    @login_required
    def bulk_delete_orders():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("เฉพาะแอดมินหรือพนักงานเท่านั้นที่ลบได้", "danger")
            return redirect(url_for("dashboard", **request.args))

        ids = request.form.getlist("order_line_ids[]")
        if not ids:
            flash("กรุณาเลือกรายการที่ต้องการลบ", "warning")
            return redirect(url_for("dashboard", **request.args))

        # แปลง id -> set ของ order_id
        id_ints = [int(i) for i in ids if str(i).isdigit()]
        lines = OrderLine.query.filter(OrderLine.id.in_(id_ints)).all()
        oids = { (l.order_id or "").strip() for l in lines if l and l.order_id }
        if not oids:
            flash("ไม่พบเลข Order สำหรับลบ", "warning")
            return redirect(url_for("dashboard", **request.args))

        # [NEW] ย้ายไปถังขยะ (Soft Delete) แทนการลบจริง
        existing_deleted = _deleted_oids_set()
        inserted = 0
        
        for oid in oids:
            oid = (oid or "").strip()
            if not oid or oid in existing_deleted:
                # มีข้อมูลอยู่ในถังขยะแล้ว ข้ามไป
                continue
            db.session.add(DeletedOrder(
                order_id=oid,
                deleted_at=now_thai(),
                deleted_by_user_id=cu.id if cu else None
            ))
            inserted += 1
        
        db.session.commit()
        
        if inserted > 0:
            flash(f"🗑️ ย้าย {inserted} ออเดอร์ ไปที่ 'Order ที่ถูกลบ' เรียบร้อยแล้ว", "success")
        else:
            flash("ออเดอร์ที่เลือกถูกย้ายไปถังขยะแล้วก่อนหน้านี้", "info")
            
        return redirect(url_for("dashboard", **request.args))
    # ================== /NEW ==================

    # ================== NEW: Update Dispatch Round ==================
    @app.route("/update_dispatch_round", methods=["POST"])
    @login_required
    def update_dispatch_round():
        """Update dispatch_round for selected orders"""
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "error": "Unauthorized"}), 401
        
        try:
            data = request.get_json()
            order_ids = data.get("order_ids", [])
            dispatch_round = data.get("dispatch_round")
            
            if not order_ids:
                return jsonify({"success": False, "error": "ไม่มีออเดอร์ที่เลือก"}), 400
            
            if dispatch_round is None or dispatch_round == "":
                return jsonify({"success": False, "error": "กรุณาระบุรอบการจ่ายงาน"}), 400
            
            # Convert to integer
            try:
                dispatch_round = int(dispatch_round)
            except (ValueError, TypeError):
                return jsonify({"success": False, "error": "รอบการจ่ายงานต้องเป็นตัวเลข"}), 400
            
            # Update all OrderLine records matching the order_ids
            updated = db.session.query(OrderLine).filter(
                OrderLine.order_id.in_(order_ids)
            ).update(
                {"dispatch_round": dispatch_round},
                synchronize_session=False
            )
            
            db.session.commit()
            
            return jsonify({
                "success": True,
                "message": f"อัปเดตรอบการจ่ายงานเป็น {dispatch_round} สำเร็จ {updated} รายการ",
                "updated": updated
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 500
    # ================== /NEW ==================

    # ================== NEW: Barcode Scan API ==================
    @app.route("/api/scan_order", methods=["POST"])
    @login_required
    def api_scan_order():
        """บันทึกการสแกนบาร์โค้ดลง Database"""
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "error": "Unauthorized"}), 401
        
        try:
            data = request.get_json() or {}
            order_id = data.get("order_id")
            if not order_id:
                return jsonify({"success": False, "error": "Missing order_id"}), 400
            
            # อัปเดตเวลาที่สแกนลงในฐานข้อมูล
            tbl = _ol_table_name()
            sql = text(f"UPDATE {tbl} SET scanned_at=:now, scanned_by=:u WHERE order_id=:oid")
            db.session.execute(sql, {
                "now": now_thai().isoformat(),
                "u": cu.username,
                "oid": order_id
            })
            db.session.commit()
            
            return jsonify({"success": True})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 500

    @app.route("/api/reset_scans", methods=["POST"])
    @login_required
    def api_reset_scans():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        try:
            data = request.get_json() or {}
            order_ids = data.get("order_ids", [])
            if not order_ids:
                return jsonify({"success": False, "error": "Missing order_ids"}), 400

            tbl = _ol_table_name()
            reset_count = 0

            for order_id in order_ids:
                sql = text(f"UPDATE {tbl} SET scanned_at=NULL, scanned_by=NULL WHERE order_id=:oid")
                db.session.execute(sql, {"oid": order_id})
                reset_count += 1

            db.session.commit()

            return jsonify({"success": True, "message": f"Reset {reset_count} scans"})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 500
    # ================== /NEW ==================

    # ================== NEW: Check Order Status API (สำหรับสแกนแยกงาน) ==================
    @app.route("/api/check_order_status", methods=["POST"])
    @login_required
    def api_check_order_status():
        """เช็คสถานะ Order อย่างละเอียด - รองรับหลายสถานะพร้อมกัน (Multi-Status)"""
        cu = current_user()
        if not cu:
            return jsonify({"found": False, "message": "Unauthorized"}), 401
        
        try:
            data = request.get_json() or {}
            oid = (data.get("order_id") or "").strip()
            
            if not oid:
                return jsonify({"found": False, "message": "ไม่ระบุเลข Order"})

            # --- เริ่มเก็บสถานะ (ใช้ List) ---
            found_statuses = []
            
            # 1. เช็คสถานะหลัก (Cancelled / Issued)
            if db.session.query(CancelledOrder).filter_by(order_id=oid).first():
                found_statuses.append("CANCELLED")
            
            if db.session.query(IssuedOrder).filter_by(order_id=oid).first():
                found_statuses.append("ISSUED")

            # 2. ดึงรายการสินค้าเพื่อเช็คสถานะอื่นๆ
            lines = OrderLine.query.filter_by(order_id=oid).all()
            if not lines:
                return jsonify({"found": False, "message": f"❌ ไม่พบ Order {oid} ในระบบ"})

            # 3. เช็ค Sales Status (SBS / Packed)
            sale = Sales.query.filter_by(order_id=oid).first()
            if not sale:
                found_statuses.append("NOT_IN_SBS")
            else:
                s_status = (sale.status or "").upper()
                if "PACKED" in s_status or "แพ็คแล้ว" in s_status or "ครบตามจำนวน" in s_status:
                    found_statuses.append("PACKED")

            # 4. เช็ค Stock รายสินค้า
            stock_statuses = []
            for line in lines:
                sku = (line.sku or "").strip()
                qty = int(line.qty or 0)
                stock_qty = 0
                
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except: stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        if st and st.qty is not None:
                            stock_qty = int(st.qty)
                
                # Logic คำนวณสถานะ Stock
                if stock_qty <= 0:
                    stock_statuses.append("SHORTAGE")
                elif stock_qty < qty:
                    stock_statuses.append("NOT_ENOUGH")
                elif stock_qty - qty <= 3:
                    stock_statuses.append("LOW_STOCK")
                else:
                    stock_statuses.append("READY")

            # สรุปสถานะ Stock (เอาที่แย่ที่สุดอันเดียวพอ)
            if "SHORTAGE" in stock_statuses:
                found_statuses.append("SHORTAGE")
            elif "NOT_ENOUGH" in stock_statuses:
                found_statuses.append("NOT_ENOUGH")
            elif "LOW_STOCK" in stock_statuses:
                found_statuses.append("LOW_STOCK")
            else:
                found_statuses.append("READY")

            # --- กำหนดสีตามความรุนแรง ---
            color = "success"
            if "CANCELLED" in found_statuses or "SHORTAGE" in found_statuses or "NOT_ENOUGH" in found_statuses:
                color = "danger"
            elif "NOT_IN_SBS" in found_statuses or "LOW_STOCK" in found_statuses:
                color = "warning"
            elif "PACKED" in found_statuses:
                color = "dark"
            elif "ISSUED" in found_statuses:
                color = "info"

            # สร้างข้อความรวม (Fallback)
            msg = f"สถานะ: {', '.join(found_statuses)}"

            return jsonify({
                "found": True, 
                "statuses": found_statuses,  # ส่งกลับเป็น List
                "status": found_statuses[0] if found_statuses else "UNKNOWN",  # รองรับโค้ดเก่า
                "message": msg, 
                "color": color
            })
            
        except Exception as e:
            return jsonify({"found": False, "message": f"เกิดข้อผิดพลาด: {str(e)}"}), 500
    # ================== /NEW ==================

    # ================== NEW: Update Low Stock Round (ข้อ 1) ==================
    @app.route("/report/lowstock/update_round", methods=["POST"])
    @login_required
    def update_lowstock_round():
        """อัปเดต lowstock_round สำหรับออเดอร์ในรายงานสินค้าน้อย (ข้อ 1)"""
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "message": "Unauthorized"}), 401

        data = request.get_json(silent=True) or {}
        order_ids = [str(s).strip() for s in (data.get("order_ids") or []) if str(s).strip()]
        round_raw = data.get("round")

        if not order_ids:
            return jsonify({"success": False, "message": "ไม่พบออเดอร์ในรายงานนี้"}), 400
        try:
            round_no = int(round_raw)
        except Exception:
            return jsonify({"success": False, "message": "รอบที่ต้องเป็นตัวเลข"}), 400

        # อัปเดตทุกบรรทัดของออเดอร์ที่เลือก (ใช้ raw SQL เพราะ lowstock_round ไม่มีในโมเดล)
        try:
            tbl = _ol_table_name()
            sql = text(f"""
                UPDATE {tbl}
                   SET lowstock_round = :r
                 WHERE order_id IN :oids
            """).bindparams(bindparam("oids", expanding=True))
            result = db.session.execute(sql, {"r": round_no, "oids": order_ids})
            db.session.commit()
            
            return jsonify({
                "success": True,
                "message": f"อัปเดตรอบเป็น {round_no} ให้ {result.rowcount} รายการ",
                "updated": result.rowcount
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({
                "success": False,
                "message": f"เกิดข้อผิดพลาด: {str(e)}"
            }), 500
    # ================== /NEW ==================

    # -----------------------
    # Export dashboard
    # -----------------------
    @app.route("/export.xlsx")
    @login_required
    def export_excel():
        # รับค่าทั้งหมดเหมือน Dashboard
        platform = normalize_platform(request.args.get("platform"))
        shop_id = request.args.get("shop_id")
        
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        status = request.args.get("status")
        
        q = (request.args.get("q") or "").strip()       # [NEW] รับคำค้นหา
        all_time = request.args.get("all_time")         # [NEW] รับ All Time
        mode = request.args.get("mode")                 # [NEW] รับ Mode (Today)

        # แปลงวันที่
        def _p(s): return parse_date_any(s)
        imp_from = _p(import_from_str)
        imp_to = _p(import_to_str)
        d_from = datetime.combine(_p(date_from), datetime.min.time(), tzinfo=TH_TZ) if date_from else None
        d_to = datetime.combine(_p(date_to) + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if date_to else None

        has_date_filter = bool(imp_from or imp_to or d_from or d_to)
        is_all_time = bool(all_time)

        # --- 1. Logic การดึงข้อมูล (เหมือน Dashboard เป๊ะ) ---
        rows = []
        base_filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
        }

        if is_all_time:
            # All Time
            filters = base_filters.copy()
            filters["active_only"] = False 
            filters["all_time"] = True
            rows, _ = compute_allocation(db.session, filters)

        elif mode == 'today':
            # Order ปัจจุบัน (วันนี้) + Order ที่ยกเลิกวันนี้
            today = now_thai().date()
            
            # 1. ดึง Order ที่นำเข้าวันนี้
            filters = base_filters.copy()
            filters["active_only"] = False
            filters["import_from"] = today
            filters["import_to"] = today
            rows_import, _ = compute_allocation(db.session, filters)
            
            # 2. ดึง Order ที่ "ยกเลิกวันนี้" (บวก 7 ชม. เพื่อให้ตรงกับเวลาไทย)
            cancel_today_oids = [
                r[0] for r in db.session.query(CancelledOrder.order_id)
                .filter(func.date(CancelledOrder.imported_at, '+7 hours') == today).all()
            ]
            
            rows_cancel = []
            if cancel_today_oids:
                f_cancel = base_filters.copy()
                f_cancel["all_time"] = True
                f_cancel["active_only"] = False
                temp_rows, _ = compute_allocation(db.session, f_cancel)
                rows_cancel = [r for r in temp_rows if r.get("order_id") in cancel_today_oids]
            
            # 3. รวมรายการ (ตัดตัวซ้ำด้วย id)
            seen_ids = set()
            rows = []
            for r in (rows_import + rows_cancel):
                rid = r.get("id")
                if rid not in seen_ids:
                    rows.append(r)
                    seen_ids.add(rid)

        elif has_date_filter:
            # กรองตามวันที่
            filters = base_filters.copy()
            filters["active_only"] = False
            filters["import_from"] = imp_from
            filters["import_to"] = imp_to
            filters["date_from"] = d_from
            filters["date_to"] = d_to
            rows, _ = compute_allocation(db.session, filters)
            
        else:
            # Default View (Order ค้าง + จบงานวันนี้)
            f_active = base_filters.copy()
            f_active["active_only"] = True
            rows_active, _ = compute_allocation(db.session, f_active)
            
            today = now_thai().date()
            f_inactive = base_filters.copy()
            f_inactive["active_only"] = False
            f_inactive["import_from"] = today
            f_inactive["import_to"] = today
            
            rows_today_all, _ = compute_allocation(db.session, f_inactive)
            
            existing_ids = set(r["id"] for r in rows_active)
            rows = list(rows_active)
            for r in rows_today_all:
                if r["id"] not in existing_ids:
                    if r.get("is_packed") or r.get("is_cancelled"):
                         rows.append(r)

        # --- 2. Post-Processing Rows ---
        # [แก้ไข] ใช้ _cancelled_oids_map แทน set เพื่อดึงเหตุผล (note) มาด้วย
        cancelled_map = _cancelled_oids_map()
        packed_oids = _orders_packed_set(rows)
        orders_not_in_sbs = _orders_not_in_sbs_set(rows)
        orders_no_sales = _orders_no_sales_set(rows)

        # ให้ Export เหมือนตาราง Dashboard: กรอง Order ที่ถูกลบออก + Inject scan status
        deleted_oids = _deleted_oids_set()
        rows = [r for r in rows if (r.get("order_id") or "").strip() not in deleted_oids]
        _inject_scan_status(rows)
        
        # เตรียม Stock/AllQty
        totals = _build_allqty_map(rows)
        
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            
            # Stock Logic
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except: stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty

            r["allqty"] = int(totals.get((r.get("sku") or "").strip(), r.get("qty", 0)) or 0)
            r["accepted"] = bool(r.get("accepted", False))
            r["sales_status"] = r.get("sales_status", None)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            
            r["is_cancelled"] = False
            r["is_not_in_sbs"] = False
            r["packed"] = False
            r["cancel_reason"] = ""  # [NEW] เตรียมตัวแปรเก็บเหตุผล
            r["cancel_str"] = ""     # [NEW] ข้อความรวมสำหรับ Excel
            
            # [NEW] เช็คว่า Order นี้เคยแพ็คแล้วหรือยัง (ก่อนถูกยกเลิก)
            r["was_packed"] = (oid in packed_oids)

            # [แก้ไข] เช็คจาก map แทน set
            if oid in cancelled_map:
                r["allocation_status"] = "CANCELLED"
                r["is_cancelled"] = True
                
                # [NEW] แกะข้อมูล Note และ Time จาก dict ซ้อน
                c_info = cancelled_map[oid]
                note_txt = c_info.get('note', '')
                time_obj = c_info.get('at')
                
                # จัด Format เวลา (แปลงเป็น พ.ศ.)
                time_str = ""
                if time_obj:
                    try:
                        if time_obj.year < 2400:
                            time_obj_be = time_obj.replace(year=time_obj.year + 543)
                        else:
                            time_obj_be = time_obj
                        time_str = time_obj_be.strftime("%d/%m/%Y %H:%M")
                    except Exception:
                        pass
                
                r["cancel_reason"] = note_txt
                r["cancel_str"] = f"{note_txt} [เมื่อ: {time_str}]" if time_str else note_txt
            elif oid in packed_oids:
                r["allocation_status"] = "PACKED"
                r["packed"] = True
            else:
                if oid in orders_not_in_sbs:
                    r["is_not_in_sbs"] = True

        # --- 3. คำนวณ KPI Sets (ต้องใช้สำหรับการกรองสถานะแบบกลุ่ม) ---
        kpi_orders_ready = _orders_ready_set(rows)
        kpi_orders_low = _orders_lowstock_order_set(rows)
        
        kpi_orders_problem = set()
        for r in rows:
            # [แก้ไข] เพิ่มเงื่อนไข: ต้องยังไม่จ่ายงาน (is_issued) ด้วย ถึงจะนับเข้ากอง 3
            if not r.get("packed") and not r.get("is_cancelled") and not r.get("is_issued"):
                status_alloc = (r.get("allocation_status") or "").strip().upper()
                if status_alloc in ("SHORTAGE", "NOT_ENOUGH"):
                    oid = (r.get("order_id") or "").strip()
                    if oid:
                        kpi_orders_problem.add(oid)

        # ===== Scan (Barcode) KPI Sets (สำหรับการกรองแบบกลุ่ม) =====
        def _active_oids(source_rows: list[dict]) -> set[str]:
            return {
                (r.get("order_id") or "").strip()
                for r in source_rows
                if r.get("order_id") and not r.get("packed") and not r.get("is_cancelled")
            }

        kpi_active_oids = _active_oids(rows)
        kpi_orders_scanned = {
            (r.get("order_id") or "").strip()
            for r in rows
            if r.get("order_id")
            and not r.get("packed")
            and not r.get("is_cancelled")
            and r.get("scanned_at")
        }
        kpi_orders_not_scanned = kpi_active_oids - kpi_orders_scanned

        # ===== Warehouse Receive (Issued but Not Packed) KPI Sets =====
        def _compute_wh_receive_sets(source_rows: list[dict]):
            issued_active_oids = {
                (r.get("order_id") or "").strip()
                for r in source_rows
                if r.get("order_id")
                and r.get("is_issued")
                and not r.get("packed")
                and not r.get("is_cancelled")
            }
            if not issued_active_oids:
                return {"total": set(), "g1": set(), "g2": set(), "g3": set()}

            issued_rows = (
                db.session.query(IssuedOrder.order_id, IssuedOrder.source)
                .filter(IssuedOrder.order_id.in_(issued_active_oids))
                .all()
            )
            src_map: dict[str, str] = {str(r[0]): (r[1] or "") for r in issued_rows}

            g1, g2, g3 = set(), set(), set()
            for oid in issued_active_oids:
                src = (src_map.get(oid) or "").strip().lower()
                if src == "print:lowstock":
                    g2.add(oid)
                elif src in {"print:nostock", "print:notenough"}:
                    g3.add(oid)
                else:
                    g1.add(oid)

            total = set().union(g1, g2, g3)
            return {"total": total, "g1": g1, "g2": g2, "g3": g3}

        wh_sets = _compute_wh_receive_sets(rows)
        wh_total_oids = wh_sets["total"]
        wh_g1_oids = wh_sets["g1"]
        wh_g2_oids = wh_sets["g2"]
        wh_g3_oids = wh_sets["g3"]

        # --- 4. กรองข้อมูล (Filtering) ---
        
        # 4.1 กรองด้วย Search Q (ถ้ามี)
        if q:
            q_lower = q.lower()
            rows = [
                r for r in rows
                if q_lower in (
                    str(r.get("order_id") or "") + " " +
                    str(r.get("sku") or "") + " " +
                    str(r.get("brand") or "") + " " +
                    str(r.get("model") or "") + " " +
                    str(r.get("shop") or "") + " " +
                    str(r.get("sales_status") or "")
                ).lower()
            ]

        # 4.2 กรองด้วย Status (เหมือนหน้า Dashboard: ถ้ามี q ให้ถือว่าเป็น Search Mode และไม่ใช้ status)
        if not q:
            status_norm = (status or "").strip().upper()
            if status_norm == "TOTAL":
                # รวมทั้งหมดใน scope (ไม่ซ่อน Packed/Cancelled)
                pass
            elif status_norm == "ORDER_CANCELLED":
                # [แก้ไข] กรองเฉพาะยกเลิกที่ยังไม่เคยแพ็ค (ก่อนแพ็ค)
                rows = [r for r in rows if r.get("is_cancelled") and not r.get("was_packed")]
            elif status_norm == "ORDER_CANCELLED_PACKED":
                # [NEW] กรองเฉพาะยกเลิกหลังแพ็ค (เคยแพ็คแล้ว)
                rows = [r for r in rows if r.get("is_cancelled") and r.get("was_packed")]
            elif status_norm == "ORDER_NOT_IN_SBS":
                rows = [r for r in rows if r.get("is_not_in_sbs")]
            elif status_norm in {"ORDER_NOT_SCANNED", "ORDER_SCAN_BARCODE"}:
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_not_scanned]
            elif status_norm == "WH_RECEIVE_TOTAL":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in wh_total_oids]
            elif status_norm == "WH_RECEIVE_G1":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in wh_g1_oids]
            elif status_norm == "WH_RECEIVE_G2":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in wh_g2_oids]
            elif status_norm == "WH_RECEIVE_G3":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in wh_g3_oids]
            elif status_norm == "ORDER_PROBLEM":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_problem]
            elif status_norm == "PACKED":
                rows = [r for r in rows if r.get("packed")]
            elif status_norm == "ORDER_READY":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_ready]
            elif status_norm in {"ORDER_LOW_STOCK", "ORDER_LOW"}:
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_low]
            elif status_norm == "ORDER_NO_SALES":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in orders_no_sales]
            elif status_norm:
                # กรองรายบรรทัด (Ready, Accepted, etc.)
                rows = [r for r in rows if (r.get("allocation_status") or "").strip().upper() == status_norm]
            else:
                # Default Table View (ซ่อน Packed/Cancelled) ยกเว้น All Time หรือ Mode Today
                if not is_all_time and mode != 'today':
                    rows = [r for r in rows if not r.get("packed") and not r.get("is_cancelled")]

        # --- 5. จัดคอลัมน์ให้ตรงกับตาราง Dashboard ---
        rows = _annotate_order_spans(rows)

        data = []
        columns = [
            "Platform",
            "ร้าน",
            "เลข Order",
            "สินค้า (SKU)",
            "Brand",
            "Stock",
            "Qty",
            "AllQty",
            "เวลาสั่ง",
            "SLA",
            "ขนส่ง",
            "สถานะ",
            "Scan Order",
            "ผู้รับ",
            "หมายเหตุ",
        ]
        for r in rows:
            # แปลง Status เป็นภาษาไทย/คำที่เข้าใจง่าย
            st = r.get("allocation_status")
            if r.get("is_issued"): st_display = "จ่ายแล้ว"
            elif st == "READY_ACCEPT": st_display = "พร้อมรับ"
            elif st == "ACCEPTED": st_display = "รับแล้ว"
            elif st == "PACKED": st_display = "แพ็คแล้ว"
            elif st == "CANCELLED": st_display = "ยกเลิก"
            elif st == "LOW_STOCK": st_display = "สินค้าน้อย"
            elif st == "SHORTAGE": st_display = "ไม่มีสินค้า"
            elif st == "NOT_ENOUGH": st_display = "สินค้าไม่พอส่ง"
            else: st_display = st

            data.append({
                "Platform": r.get("platform"),
                "ร้าน": r.get("shop"),
                "เลข Order": r.get("order_id"),
                "สินค้า (SKU)": r.get("sku"),
                "Brand": r.get("brand"),
                "Stock": r.get("stock_qty"),
                "Qty": r.get("qty"),
                "AllQty": r.get("allqty"),
                "เวลาสั่ง": r.get("order_time"),
                "SLA": r.get("sla"),
                "ขนส่ง": r.get("logistic"),
                "สถานะ": st_display,
                "Scan Order": "✓" if r.get("scanned_at") else "",
                "ผู้รับ": r.get("accepted_by"),
                "หมายเหตุ": r.get("cancel_str") or r.get("cancel_reason") or r.get("note") or "",
            })

        df = pd.DataFrame(data, columns=columns)

        out = BytesIO()
        with pd.ExcelWriter(out, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="Dashboard")
            
            # จัดความกว้างคอลัมน์
            worksheet = w.sheets['Dashboard']
            worksheet.set_column('A:A', 12)  # Platform
            worksheet.set_column('B:B', 18)  # ร้าน
            worksheet.set_column('C:C', 22)  # เลข Order
            worksheet.set_column('D:D', 18)  # สินค้า (SKU)
            worksheet.set_column('E:E', 15)  # Brand
            worksheet.set_column('F:F', 8)   # Stock
            worksheet.set_column('G:G', 8)   # Qty
            worksheet.set_column('H:H', 8)   # AllQty
            worksheet.set_column('I:I', 18)  # เวลาสั่ง
            worksheet.set_column('J:J', 18)  # SLA
            worksheet.set_column('K:K', 20)  # ขนส่ง
            worksheet.set_column('L:L', 15)  # สถานะ
            worksheet.set_column('M:M', 10)  # Scan Order
            worksheet.set_column('N:N', 12)  # ผู้รับ
            worksheet.set_column('O:O', 30)  # หมายเหตุ
            
        out.seek(0)
        filename = f"Dashboard_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(out, as_attachment=True, download_name=filename)

    # -----------------------
    # ใบงานคลัง (Warehouse Job Sheet)
    # -----------------------
    @app.route("/report/warehouse", methods=["GET"])
    @login_required
    def print_warehouse():
        # รับค่า reset mode และ search query
        reset_mode = request.args.get("reset")
        q = (request.args.get("q") or "").strip()  # [NEW] รับค่าคำค้นหา
        
        if reset_mode == 'all':
            # ถ้ากดรีเฟรช: เคลียร์ตัวกรองทุกอย่างให้เป็น None
            platform = None
            shop_id = None
            logistic = None
            acc_from = None
            acc_to = None
            acc_from_str = ""
            acc_to_str = ""
            q = ""  # เคลียร์คำค้นหาด้วย
            round_sel = None
            print_count_sel = None
        else:
            # ถ้าไม่ได้กดรีเฟรช: รับค่าจากฟอร์มปกติ
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            acc_from_str = request.args.get("accepted_from")
            acc_to_str = request.args.get("accepted_to")
            acc_from = parse_date_any(acc_from_str)
            acc_to = parse_date_any(acc_to_str)
            round_sel = request.args.get("round")
            print_count_sel = request.args.get("print_count")
        
        # [NEW] ถ้ามีคำค้นหา ให้ล้าง filter วันที่ (ค้นหาทั้งหมด)
        if q:
            acc_from = None
            acc_to = None
            acc_from_str = ""
            acc_to_str = ""

        filters = {
            "platform": platform, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]

        # *** กรองออเดอร์ที่พิมพ์แล้วออก - แสดงเฉพาะที่ยังไม่ได้พิมพ์ ***
        # ดึง count จาก DB จริงแทนที่จะใช้ r.get("printed_warehouse") ที่เป็น 0 ตลอด
        oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        counts = _get_print_counts_local(oids, kind="warehouse")
        rows = [r for r in rows if int(counts.get((r.get("order_id") or "").strip(), 0)) == 0]

        if logistic:
            rows = [r for r in rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # [NEW] กรอง rows ตามคำค้นหา q (ค้นหาใน order_id, sku, shop, logistic)
        if q:
            q_lower = q.lower()
            rows = [
                r for r in rows 
                if q_lower in (
                    str(r.get("order_id") or "") + 
                    str(r.get("sku") or "") + 
                    str(r.get("shop") or "") + 
                    str(r.get("logistic") or "")
                ).lower()
            ]

        _inject_print_counts_to_rows(rows, kind="warehouse")
        _inject_scan_status(rows)  # Inject scan data before grouping
        rows = _group_rows_for_warehouse_report(rows)  # Use warehouse-specific grouping

        # [NEW] กรอง Round และ Print Count หลังจากจัดกลุ่มแล้ว
        if round_sel and round_sel.strip():
            filtered_rows = []
            for r in rows:
                try:
                    if str(r.get("dispatch_round") or "") == str(round_sel):
                        filtered_rows.append(r)
                except:
                    pass
            rows = filtered_rows
        
        if print_count_sel and print_count_sel.strip():
            filtered_rows = []
            for r in rows:
                try:
                    p_val = int(r.get("printed_warehouse") or r.get("printed_count") or 0)
                    if p_val == int(print_count_sel):
                        filtered_rows.append(r)
                except:
                    pass
            rows = filtered_rows
        # [/NEW]

        total_orders = len(rows)  # Now 1 row = 1 order
        shops = Shop.query.all()
        logistics = sorted(set(r.get("logistic") for r in rows if r.get("logistic")))
        return render_template(
            "report.html",
            rows=rows,
            count_orders=total_orders,
            shops=shops,
            logistics=logistics,
            platform_sel=platform if reset_mode != 'all' else None,
            shop_sel=shop_id if reset_mode != 'all' else None,
            logistic_sel=logistic if reset_mode != 'all' else None,
            official_print=False,
            printed_meta=None,
            accepted_from=acc_from_str if reset_mode != 'all' else "",
            accepted_to=acc_to_str if reset_mode != 'all' else "",
            q=q,  # [NEW] ส่งค่าคำค้นหากลับไป template
            round_sel=round_sel if reset_mode != 'all' else None,
            print_count_sel=print_count_sel if reset_mode != 'all' else None,
        )

    @app.route("/report/warehouse/print", methods=["POST"])
    @login_required
    def print_warehouse_commit():
        cu = current_user()
        platform = normalize_platform(request.form.get("platform"))
        shop_id = request.form.get("shop_id")
        logistic = request.form.get("logistic")
        override = request.form.get("override") in ("1", "true", "yes")
        
        # Get selected order IDs from form
        selected_order_ids = request.form.get("order_ids", "")
        selected_order_ids = [oid.strip() for oid in selected_order_ids.split(",") if oid.strip()]

        filters = {"platform": platform, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]

        if logistic:
            rows = [r for r in rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # If specific order IDs were selected, filter to only those orders
        if selected_order_ids:
            rows = [r for r in rows if (r.get("order_id") or "").strip() in selected_order_ids]
            oids = sorted(selected_order_ids)
        else:
            oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        
        if not oids:
            flash("ไม่พบออเดอร์สำหรับพิมพ์", "warning")
            return redirect(url_for("print_warehouse", platform=platform, shop_id=shop_id, logistic=logistic))

        already = _detect_already_printed(oids, kind="warehouse")
        if already and not (override and cu and cu.role == "admin"):
            head = ", ".join(list(already)[:10])
            more = "" if len(already) <= 10 else f" ... (+{len(already)-10})"
            flash(f"มีบางออเดอร์เคยพิมพ์ใบงานคลังไปแล้ว: {head}{more}", "danger")
            flash("ถ้าจำเป็นต้องพิมพ์ซ้ำ โปรดให้แอดมินกดยืนยัน 'อนุญาตพิมพ์ซ้ำ' แล้วพิมพ์อีกครั้ง", "warning")
            return redirect(url_for("print_warehouse", platform=platform, shop_id=shop_id, logistic=logistic))

        now_iso = now_thai().isoformat()
        _mark_printed(oids, kind="warehouse", user_id=(cu.id if cu else None), when_iso=now_iso)
        
        # [แก้ไข] ปิดการจบงาน (Issued) ณ จุดนี้ เพื่อให้ Order ไปรอที่หน้า Picking ก่อน
        # _mark_issued(oids, user_id=(cu.id if cu else None), source="print:warehouse", when_dt=now_thai())
        
        db.session.commit()  # Ensure changes are committed
        db.session.expire_all()  # Force refresh to get updated print counts

        _inject_print_counts_to_rows(rows, kind="warehouse")
        _inject_scan_status(rows)  # Inject scan data to preserve in print view
        rows = _group_rows_for_warehouse_report(rows)  # Use warehouse-specific grouping

        total_orders = len(rows)  # Now 1 row = 1 order
        shops = Shop.query.all()
        logistics = sorted(set(r.get("logistic") for r in rows if r.get("logistic")))
        printed_meta = {"by": (cu.username if cu else "-"), "at": now_thai(), "orders": total_orders, "override": bool(already)}
        return render_template(
            "report.html",
            rows=rows,
            count_orders=total_orders,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            official_print=True,
            printed_meta=printed_meta
        )

    # ================== NEW: View Printed Warehouse Jobs ==================
    @app.route("/report/warehouse/printed", methods=["GET"])
    @login_required
    def warehouse_printed_history():
        """ดูใบงานคลังที่พิมพ์แล้ว - สามารถเลือกวันที่และพิมพ์ซ้ำได้"""
        # รับค่า reset mode และ search query
        reset_mode = request.args.get("reset")
        q = (request.args.get("q") or "").strip()  # [NEW] รับค่าคำค้นหา
        
        # [NEW] ถ้ามีคำค้นหา ให้ข้ามการตั้งค่าวันที่ไปเลย (ค้นหาทั้งหมด)
        if q:
            target_date = None
            platform = None
            shop_id = None
            logistic = None
            print_date = None
            print_date_from = None
            print_date_to = None
            raw_from = None
            raw_to = None
            round_sel = None
            print_count_sel = None
        elif reset_mode == 'today':
            # ถ้ากดรีเฟรช: แสดงเฉพาะของวันนี้
            target_date = now_thai().date()
            platform = None
            shop_id = None
            logistic = None
            print_date = None
            print_date_from = None
            print_date_to = None
            raw_from = None
            raw_to = None
            round_sel = None
            print_count_sel = None
        else:
            # กรณีปกติ: รับค่าจากฟอร์ม
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            print_date = request.args.get("print_date")  # วันที่พิมพ์ (YYYY-MM-DD) - เก็บไว้สำหรับ backward compatible
            
            # [NEW] รับค่า Date Range สำหรับวันที่พิมพ์
            print_date_from = request.args.get("print_date_from")
            print_date_to = request.args.get("print_date_to")
            
            raw_from = request.args.get("accepted_from")
            raw_to = request.args.get("accepted_to")
            round_sel = request.args.get("round")
            print_count_sel = request.args.get("print_count")
            
            # ============================================================
            # [แก้ไข] ถ้าเข้าหน้านี้ครั้งแรก (ไม่มี Params วันที่)
            # ให้ Default เป็น "วันนี้" ทันที
            # ============================================================
            if print_date_from is None and print_date_to is None and print_date is None:
                today_str = now_thai().date().isoformat()
                print_date_from = today_str
                print_date_to = today_str
            # ============================================================
            
            # ถ้าเลือกวันที่พิมพ์ (ระบบเก่า - single date)
            if print_date:
                try:
                    target_date = datetime.strptime(print_date, "%Y-%m-%d").date()
                except:
                    target_date = None
            else:
                target_date = None
        
        # ไม่ตั้งค่า default - ให้เป็นค่าว่าง (mm/dd/yyyy)
        acc_from = parse_date_any(raw_from)
        acc_to = parse_date_any(raw_to)
        
        # Get all orders that have been printed
        tbl = _ol_table_name()
        
        # Build query to get orders with print history
        if q:
            # [NEW] กรณีค้นหา: หาจากประวัติทั้งหมด (printed_warehouse > 0) ที่เลข Order ตรงกัน
            # ไม่สนวันที่พิมพ์ (Global Search in History)
            sql = text(f"""
                SELECT DISTINCT order_id 
                FROM {tbl} 
                WHERE printed_warehouse > 0 
                AND order_id LIKE :q
            """)
            result = db.session.execute(sql, {"q": f"%{q}%"}).fetchall()
        elif target_date:
            # Filter by specific print date (หรือวันนี้ถ้า reset)
            # หมายเหตุ: printed_warehouse_at ถูกบันทึกเป็นเวลาไทยอยู่แล้ว (ไม่ต้อง +7)
            sql = text(f"""
                SELECT DISTINCT order_id 
                FROM {tbl} 
                WHERE printed_warehouse > 0 
                AND DATE(printed_warehouse_at) = :target_date
            """)
            result = db.session.execute(sql, {"target_date": target_date.isoformat()}).fetchall()
        elif print_date_from or print_date_to:
            # [NEW] Filter by date range (เริ่ม - ถึง)
            # หมายเหตุ: printed_warehouse_at ถูกบันทึกเป็นเวลาไทยอยู่แล้ว (ไม่ต้อง +7)
            sql_where = "WHERE printed_warehouse > 0"
            params = {}
            if print_date_from:
                sql_where += " AND DATE(printed_warehouse_at) >= :pf"
                params["pf"] = print_date_from
            if print_date_to:
                sql_where += " AND DATE(printed_warehouse_at) <= :pt"
                params["pt"] = print_date_to
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} {sql_where}")
            result = db.session.execute(sql, params).fetchall()
        else:
            # Get all printed orders
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_warehouse > 0")
            result = db.session.execute(sql).fetchall()
        
        printed_order_ids = [row[0] for row in result if row[0]]
        
        if not printed_order_ids:
            # No printed orders found
            shops = Shop.query.all()
            return render_template(
                "report.html",
                rows=[],
                count_orders=0,
                shops=shops,
                logistics=[],
                platform_sel=platform,
                shop_sel=shop_id,
                logistic_sel=logistic,
                official_print=False,
                printed_meta=None,
                is_history_view=True,
                print_date_sel=None if reset_mode == 'today' else print_date,
                print_date_from=print_date_from,
                print_date_to=print_date_to,
                accepted_from="" if reset_mode == 'today' else raw_from,
                accepted_to="" if reset_mode == 'today' else raw_to,
                q=q,  # [NEW] ส่งค่าคำค้นหากลับไป template
            )
        
        # Get full data for these orders
        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        
        # Filter to only printed orders
        rows = [r for r in rows if (r.get("order_id") or "").strip() in printed_order_ids]
        
        if logistic:
            rows = [r for r in rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        
        _inject_print_counts_to_rows(rows, kind="warehouse")
        _inject_scan_status(rows)  # Inject scan data before grouping
        rows = _group_rows_for_warehouse_report(rows)
        
        # [NEW] กรอง Round และ Print Count หลังจากจัดกลุ่มแล้ว
        if round_sel and round_sel.strip():
            filtered_rows = []
            for r in rows:
                try:
                    if str(r.get("dispatch_round") or "") == str(round_sel):
                        filtered_rows.append(r)
                except:
                    pass
            rows = filtered_rows
        
        if print_count_sel and print_count_sel.strip():
            filtered_rows = []
            for r in rows:
                try:
                    p_val = int(r.get("printed_warehouse") or r.get("printed_count") or 0)
                    if p_val == int(print_count_sel):
                        filtered_rows.append(r)
                except:
                    pass
            rows = filtered_rows
        # [/NEW]
        
        total_orders = len(rows)
        shops = Shop.query.all()
        logistics = sorted(set(r.get("logistic") for r in rows if r.get("logistic")))
        
        # Get available print dates for dropdown
        # หมายเหตุ: printed_warehouse_at ถูกบันทึกเป็นเวลาไทยอยู่แล้ว (ไม่ต้อง +7)
        sql_dates = text(f"""
            SELECT DISTINCT DATE(printed_warehouse_at) as print_date 
            FROM {tbl} 
            WHERE printed_warehouse > 0 AND printed_warehouse_at IS NOT NULL
            ORDER BY print_date DESC
        """)
        available_dates = [row[0] for row in db.session.execute(sql_dates).fetchall()]
        
        return render_template(
            "report.html",
            rows=rows,
            count_orders=total_orders,
            shops=shops,
            logistics=logistics,
            platform_sel=platform if reset_mode != 'today' else None,
            shop_sel=shop_id if reset_mode != 'today' else None,
            logistic_sel=logistic if reset_mode != 'today' else None,
            official_print=False,
            printed_meta=None,
            is_history_view=True,
            print_date_sel=None if reset_mode == 'today' else print_date,
            available_dates=available_dates,
            
            # [NEW] ส่งค่าวันที่พิมพ์กลับไปแสดงใน Input (Date Range)
            print_date_from=print_date_from,
            print_date_to=print_date_to,
            
            accepted_from="" if reset_mode == 'today' else raw_from,
            accepted_to="" if reset_mode == 'today' else raw_to,
            q=q,  # [NEW] ส่งค่าคำค้นหากลับไป template
            round_sel=round_sel if reset_mode != 'today' else None,
            print_count_sel=print_count_sel if reset_mode != 'today' else None,
        )

    # ================== NEW: Export Warehouse Excel ==================
    @app.route("/report/warehouse/export.xlsx")
    @login_required
    def export_warehouse_excel():
        """Export ใบงานคลัง (หน้าปัจจุบัน) - แสดงเฉพาะงานที่ยังไม่พิมพ์"""
        # รับค่า Filter เหมือนหน้า Warehouse
        reset_mode = request.args.get("reset")
        
        if reset_mode == 'all':
            platform = None
            shop_id = None
            logistic = None
            acc_from = None
            acc_to = None
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            acc_from = parse_date_any(request.args.get("accepted_from"))
            acc_to = parse_date_any(request.args.get("accepted_to"))

        filters = {
            "platform": platform, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]

        # กรองเฉพาะยังไม่พิมพ์
        oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        counts = _get_print_counts_local(oids, kind="warehouse")
        rows = [r for r in rows if int(counts.get((r.get("order_id") or "").strip(), 0)) == 0]

        if logistic:
            rows = [r for r in rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        _inject_print_counts_to_rows(rows, kind="warehouse")
        _inject_scan_status(rows)
        rows = _group_rows_for_warehouse_report(rows)

        # สร้าง DataFrame ให้ตรงกับคอลัมน์หน้าจอ
        data = []
        for r in rows:
            data.append({
                "แพลตฟอร์ม": r.get("platform", ""),
                "ร้าน": r.get("shop", ""),
                "เลข Order": r.get("order_id", ""),
                "ประเภทขนส่ง": r.get("logistic", ""),
                "ผู้กดรับ": r.get("accepted_by", ""),
                "Scan Order": "✓ แล้ว" if r.get("scanned_at") else "",
                "จ่ายงาน(รอบที่)": r.get("dispatch_round", ""),
                "พิมพ์แล้ว(ครั้ง)": r.get("printed_warehouse", 0),
                "วัน/เดือน/ปี/เวลา ที่พิมพ์": to_thai_be(r.get("printed_warehouse_at")) if r.get("printed_warehouse_at") else ""
            })

        df = pd.DataFrame(data)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="Warehouse")
        bio.seek(0)
        
        filename = f"ใบงานคลัง_Warehouse_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename)

    @app.route("/report/warehouse/history/export.xlsx")
    @login_required
    def export_warehouse_history_excel():
        """Export ใบงานคลังประวัติ - แสดงงานที่พิมพ์แล้ว"""
        # รับค่า Filter เหมือนหน้า History
        reset_mode = request.args.get("reset")
        
        if reset_mode == 'today':
            target_date = now_thai().date()
            platform = None
            shop_id = None
            logistic = None
            print_date = None
            raw_from = None
            raw_to = None
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            print_date = request.args.get("print_date")
            raw_from = request.args.get("accepted_from")
            raw_to = request.args.get("accepted_to")
            
            if print_date:
                try:
                    target_date = datetime.strptime(print_date, "%Y-%m-%d").date()
                except:
                    target_date = None
            else:
                target_date = None
        
        acc_from = parse_date_any(raw_from)
        acc_to = parse_date_any(raw_to)
        
        # Get printed orders
        tbl = _ol_table_name()
        
        if target_date:
            sql = text(f"""
                SELECT DISTINCT order_id 
                FROM {tbl} 
                WHERE printed_warehouse > 0 
                AND DATE(printed_warehouse_at) = :target_date
            """)
            result = db.session.execute(sql, {"target_date": target_date.isoformat()}).fetchall()
        else:
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_warehouse > 0")
            result = db.session.execute(sql).fetchall()
        
        printed_order_ids = [row[0] for row in result if row[0]]
        
        if not printed_order_ids:
            # Return empty Excel
            df = pd.DataFrame()
            bio = BytesIO()
            with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
                df.to_excel(w, index=False, sheet_name="History")
            bio.seek(0)
            filename = f"ใบงานคลังประวัติ_History_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            return send_file(bio, as_attachment=True, download_name=filename)
        
        # Get full data
        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if (r.get("order_id") or "").strip() in printed_order_ids]
        
        if logistic:
            rows = [r for r in rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        
        _inject_print_counts_to_rows(rows, kind="warehouse")
        _inject_scan_status(rows)
        rows = _group_rows_for_warehouse_report(rows)
        
        # สร้าง DataFrame
        data = []
        for r in rows:
            data.append({
                "แพลตฟอร์ม": r.get("platform", ""),
                "ร้าน": r.get("shop", ""),
                "เลข Order": r.get("order_id", ""),
                "ประเภทขนส่ง": r.get("logistic", ""),
                "ผู้กดรับ": r.get("accepted_by", ""),
                "Scan Order": "✓ แล้ว" if r.get("scanned_at") else "",
                "จ่ายงาน(รอบที่)": r.get("dispatch_round", ""),
                "พิมพ์แล้ว(ครั้ง)": r.get("printed_warehouse", 0),
                "วัน/เดือน/ปี/เวลา ที่พิมพ์": to_thai_be(r.get("printed_warehouse_at")) if r.get("printed_warehouse_at") else ""
            })

        df = pd.DataFrame(data)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="History")
        bio.seek(0)
        
        filename = f"ใบงานคลังประวัติ_History_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename)

    # ================== NEW: Low-Stock & No-Stock Reports ==================

    @app.route("/report/lowstock", methods=["GET"])
    @login_required
    def report_lowstock():
        """
        รายงานสินค้าน้อย — อ้างอิงชุด SKU/Order จาก Dashboard โดยตรง
        ข้อสำคัญตาม requirement:
          - ไม่ดึงออเดอร์ที่ PACKED แล้ว (ข้อ 1)
          - 'จ่ายงาน(รอบที่)' ใช้คอลัมน์ lowstock_round แยกจาก dispatch_round (ข้อ 2)
          - 'พิมพ์แล้ว(ครั้ง)' ใช้ printed_lowstock (ข้อ 3)
          - รองรับ filter ครบ (ข้อ 4)
          - รองรับ sort ทุกคอลัมน์ (ข้อ 5)
          - ดึงเฉพาะชุด Order สินค้าน้อยจาก Dashboard (ข้อ 6)
        """
        # ไม่ต้องใช้ services.lowstock_queue แล้ว - ใช้ compute_allocation โดยตรง

        # ---- รับตัวกรอง/เรียง ----
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        round_num = request.args.get("round")  # ข้อ 7: กรองรอบ
        q        = (request.args.get("q") or "").strip()
        sort_col = (request.args.get("sort") or "").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าวันที่กรอง
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")

        shops = Shop.query.order_by(Shop.name.asc()).all()

        # ---- 1) ดึง allocation rows เหมือน Dashboard ----
        filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
            "import_date": None
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)
        rows = _filter_out_lowstock_printed_rows(rows)  # <<<< NEW (ข้อ 2): ตัดออเดอร์ที่พิมพ์รายงานสินค้าน้อยออก

        # คำนวณออเดอร์ที่แพ็คแล้ว (เช็คจาก sales_status)
        packed_oids = _orders_packed_set(rows)

        # เติม stock_qty / logistic ให้ครบ + ไม่เอา PACKED (ข้อ 1)
        safe = []
        for r in rows:
            r = dict(r)
            # กรองออเดอร์ที่อยู่ในลิสต์แพ็คแล้วออก
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            sales_status = (str(r.get("sales_status") or "")).upper()
            if sales_status == "PACKED" or bool(r.get("packed", False)):
                continue
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            # ไม่ต้อง _recompute เพราะ allocation_status มาจาก compute_allocation แล้ว
            safe.append(r)

        # ---- 2) ให้ "Order สินค้าน้อย" เป็นตัวตั้ง (ข้อ 6) ----
        orders_low = _orders_lowstock_order_set(safe)
        safe = [r for r in safe if (r.get("order_id") or "").strip() in orders_low]

        # ---- 2.5) กรองตามวันที่สั่งซื้อและวันที่นำเข้า ----
        if date_from_str or date_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            date_from = parse_date_str(date_from_str)
            date_to = parse_date_str(date_to_str)
            filtered = []
            for r in safe:
                order_dt = r.get("order_time")
                if isinstance(order_dt, str):
                    try: order_dt = datetime.strptime(order_dt.split()[0], "%Y-%m-%d").date()
                    except: order_dt = None
                elif isinstance(order_dt, datetime):
                    order_dt = order_dt.date()
                if order_dt:
                    if date_from and order_dt < date_from: continue
                    if date_to and order_dt > date_to: continue
                elif date_from or date_to:
                    continue
                filtered.append(r)
            safe = filtered
        
        if import_from_str or import_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            import_from = parse_date_str(import_from_str)
            import_to = parse_date_str(import_to_str)
            filtered = []
            for r in safe:
                imp_dt = r.get("import_date")
                if isinstance(imp_dt, str):
                    try: imp_dt = datetime.strptime(imp_dt, "%Y-%m-%d").date()
                    except: imp_dt = None
                elif isinstance(imp_dt, datetime):
                    imp_dt = imp_dt.date()
                elif isinstance(imp_dt, date):
                    pass
                else:
                    imp_dt = None
                if imp_dt:
                    if import_from and imp_dt < import_from: continue
                    if import_to and imp_dt > import_to: continue
                elif import_from or import_to:
                    continue
                filtered.append(r)
            safe = filtered

        # ---- 3) กรองเฉพาะ allocation_status == "LOW_STOCK" ตาม compute_allocation ----
        # ใช้ allocation_status จาก compute_allocation โดยตรง (Single Source of Truth)
        lines = [r for r in safe if r.get("allocation_status") == "LOW_STOCK"]

        # ---- 4) กรองเพิ่มตามคำค้น/โลจิสติกส์ (ข้อ 4) ----
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        if q:
            ql = q.lower()
            def _hit(s):
                return ql in (str(s or "").lower())
            lines = [r for r in lines if (
                _hit(r.get("order_id")) or _hit(r.get("sku")) or _hit(r.get("brand")) or
                _hit(r.get("model")) or _hit(r.get("shop")) or _hit(r.get("platform")) or _hit(r.get("logistic"))
            )]

        # ---- NEW (ข้อ 1): อ่านค่า lowstock_round จาก DB เผื่อ compute_allocation ไม่ส่งฟิลด์มา ----
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        low_round_by_oid = {}
        if order_ids_for_round:
            # ใช้ raw SQL แทน ORM เพราะ lowstock_round ไม่มีในโมเดล
            tbl = _ol_table_name()
            sql = text(f"""
                SELECT order_id, MAX(lowstock_round) AS r
                  FROM {tbl}
                 WHERE order_id IN :oids
                 GROUP BY order_id
            """).bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                low_round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except Exception:
                # ถ้าคอลัมน์ยังไม่มี ให้ใช้ค่าว่าง
                low_round_by_oid = {}

        # ---- เตรียมข้อมูล Mixed Status ----
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "SHORTAGE": "ไม่มีของ",
            "NOT_ENOUGH": "ไม่พอส่ง",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "LOW_STOCK":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        # ---- 5) แปลงเป็นคอลัมน์ของรายงาน + AllQty ----
        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform":      r.get("platform"),
                "store":         r.get("shop"),
                "order_no":      oid,
                "sku":           r.get("sku"),
                "brand":         r.get("brand"),
                "product_name":  r.get("model"),
                "stock":         int(r.get("stock_qty", 0) or 0),
                "qty":           int(r.get("qty", 0) or 0),
                "order_time":    r.get("order_time"),
                "due_date":      r.get("due_date"),
                "sla":           r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round":  low_round_by_oid.get(oid, r.get("lowstock_round")),  # <<<< ใช้ค่าจาก DB (ข้อ 1)
                "printed_count": 0,
                "note":          mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        # ---- 6) เรียงลำดับ (ข้อ 5) ----
        sort_col = sort_col if sort_col in {"platform","store","order_no","sku","brand","product_name","stock","qty","allqty","order_time","due_date","sla","shipping_type","assign_round","printed_count"} else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            if sort_col in {"stock","qty","allqty","assign_round","printed_count"}:
                try: return int(v.get(sort_col) or 0)
                except: return 0
            elif sort_col in {"order_time","due_date"}:
                try: return datetime.fromisoformat(str(v.get(sort_col)))
                except: return str(v.get(sort_col) or "")
            else:
                return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        # ---- 7) นับ "พิมพ์แล้ว(ครั้ง)" (ข้อ 3) ----
        order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
        counts_low = _get_print_counts_local(order_ids, "lowstock")
        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts_low.get(oid, 0))

        # ---- 8) เตรียม context สำหรับ template ----
        # คำนวณจำนวน SKU ที่ไม่ซ้ำจาก out
        low_skus = {(r.get("sku") or "").strip() for r in out if r.get("sku")}
        summary = {"sku_count": len(low_skus), "orders_count": len(order_ids)}
        # ข้อ 1: ไม่ต้องแสดงเวลาพิมพ์ในหน้าปกติ (ยังไม่ได้พิมพ์จริง)
        for r in out:
            r["printed_at"] = None  # ไม่ใส่เวลา

        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))
        
        # ข้อ 7: หา available rounds สำหรับ dropdown
        available_rounds = sorted({r["assign_round"] for r in out if r["assign_round"] is not None})
        if not available_rounds:
            rs = db.session.execute(text("SELECT DISTINCT lowstock_round FROM order_lines WHERE lowstock_round IS NOT NULL ORDER BY lowstock_round")).fetchall()
            available_rounds = [x[0] for x in rs]

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if order_ids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": order_ids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_lowstock.html",
            rows=out,
            summary=summary,
            printed_at=None,  # ข้อ 1: ไม่แสดงเวลาพิมพ์ในหน้าปกติ
            order_ids=order_ids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            round_sel=round_num,
            available_rounds=available_rounds,
            sort_col=sort_col,
            sort_dir=("desc" if rev else "asc"),
            q=q,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str,
            mixed_status=mixed_info,
            is_history_view=False
        )

    @app.post("/report/lowstock/print")
    @login_required
    def report_lowstock_print():
        """บันทึกการพิมพ์รายงานสินค้าน้อย + ย้ายไปหน้าประวัติ (ข้อ 7)"""
        cu = current_user()
        order_ids_raw = (request.form.get("order_ids") or "").strip()
        order_ids = [s.strip() for s in order_ids_raw.split(",") if s.strip()]
        if not order_ids:
            flash("ไม่พบออเดอร์สำหรับพิมพ์", "warning")
            return redirect(url_for("report_lowstock"))

        # ตรวจสอบว่าทุกออเดอร์มีรอบหรือไม่ (STRICT MODE)
        orders = OrderLine.query.filter(OrderLine.order_id.in_(order_ids)).all()
        orders_without_round = [o.order_id for o in orders if not o.lowstock_round]

        if orders_without_round:
            head = ', '.join(orders_without_round[:5])
            more = f" และอีก {len(orders_without_round)-5} รายการ" if len(orders_without_round) > 5 else ""
            flash(f"ออเดอร์เหล่านี้ยังไม่ได้บันทึกรอบ: {head}{more}", "danger")
            flash("กรุณาบันทึกรอบให้ทุกออเดอร์ก่อนทำการพิมพ์", "warning")
            return redirect(url_for("report_lowstock"))

        now_iso = now_thai().isoformat()

        # 1. บันทึกว่าพิมพ์ Low Stock แล้ว
        _mark_lowstock_printed(order_ids, username=(cu.username if cu else None), when_iso=now_iso)
        
        # 2. ย้ายไป "Order จ่ายแล้ว" (Issued) ทันที
        _mark_issued(order_ids, user_id=(cu.id if cu else None), source="print:lowstock", when_dt=now_thai())
        
        db.session.commit()
        return redirect(url_for("report_lowstock_printed", auto_print="1"))

    @app.get("/report/lowstock/printed")
    @login_required
    def report_lowstock_printed():
        """ประวัติรายงานสินค้าน้อยที่พิมพ์แล้ว (ข้อ 7)"""
        # ไม่ต้องใช้ services.lowstock_queue แล้ว - ใช้ compute_allocation โดยตรง
        
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q        = (request.args.get("q") or "").strip()
        round_num = request.args.get("round")
        sort_col = (request.args.get("sort") or "order_no").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าตัวกรองวันที่สั่งซื้อและนำเข้า
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        # รับค่าช่วงวันที่พิมพ์
        print_date_from = request.args.get("print_date_from")
        print_date_to = request.args.get("print_date_to")
        reset_mode = request.args.get("reset")  # [NEW] รับค่า reset
        action = request.args.get("action")  # [NEW] รับค่า action (เพื่อแยกการกดปุ่มกรอง กับการเข้าหน้าเว็บครั้งแรก)
        
        # [SMART DEFAULT] ถ้าไม่มีวันที่ส่งมา AND ไม่มีคำค้นหา AND ไม่ได้ reset AND ไม่ใช่การกดปุ่มกรอง -> ให้กรอง "วันนี้"
        if not action and reset_mode != 'all' and not print_date_from and not print_date_to and not q:
            # เข้าหน้าเว็บครั้งแรก (ไม่มี action) = ดูงานวันนี้
            today = now_thai().date().isoformat()
            print_date_from = today
            print_date_to = today
        # ถ้ามี action (กดปุ่มกรอง) หรือ q หรือ reset='all' แต่ไม่มีวันที่ -> ค้นหาทั้งหมด

        tbl = _ol_table_name()
        
        # ========================================================
        # [FIX] ดึงข้อมูลเฉพาะเมื่อ: มีคำค้นหา หรือ มีการเลือกวันที่
        # ========================================================
        if q:
            # กรณี 1: มีคำค้นหา -> ค้นหาทั้งหมด (Global Search)
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_lowstock > 0")
            result = db.session.execute(sql).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        elif print_date_from or print_date_to:
            # กรณี 2: มีการเลือกวันที่ -> กรองตามวันที่
            sql_where = "printed_lowstock > 0"
            params = {}
            if print_date_from:
                sql_where += " AND DATE(printed_lowstock_at) >= :pf"
                params["pf"] = print_date_from
            if print_date_to:
                sql_where += " AND DATE(printed_lowstock_at) <= :pt"
                params["pt"] = print_date_to
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE {sql_where}")
            result = db.session.execute(sql, params).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        else:
            # กรณี 3: ไม่ค้นหา และ ไม่เลือกวัน (เช่น กด reset='all') -> ไม่แสดงอะไร
            printed_oids = []

        def _available_dates():
            sql = text(f"SELECT DISTINCT DATE(printed_lowstock_at) as d FROM {tbl} WHERE printed_lowstock > 0 AND printed_lowstock_at IS NOT NULL ORDER BY d DESC")
            return [r[0] for r in db.session.execute(sql).fetchall()]

        shops = Shop.query.order_by(Shop.name.asc()).all()
        
        if not printed_oids:
            return render_template(
                "report_lowstock.html",
                rows=[],
                summary={"sku_count": 0, "orders_count": 0},
                printed_at=None,
                order_ids=[],
                shops=shops,
                logistics=[],
                platform_sel=platform,
                shop_sel=shop_id,
                logistic_sel=logistic,
                is_history_view=True,
                available_dates=_available_dates(),
                print_date_from=print_date_from,
                print_date_to=print_date_to,
                sort_col=sort_col,
                sort_dir=sort_dir,
                q=q,
                round_sel=round_num,
                date_from=date_from_str,
                date_to=date_to_str,
                import_from=import_from_str,
                import_to=import_to_str
            )

        # เตรียมตัวกรองวันที่สั่งซื้อ
        date_from_dt = None
        date_to_dt = None
        if date_from_str:
            try:
                date_from_dt = datetime.combine(parse_date_any(date_from_str), datetime.min.time(), tzinfo=TH_TZ)
            except: pass
        if date_to_str:
            try:
                date_to_dt = datetime.combine(parse_date_any(date_to_str) + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ)
            except: pass

        filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
            "import_date": None,
            "date_from": date_from_dt,
            "date_to": date_to_dt
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if (r.get("order_id") or "").strip() in printed_oids]
        
        # กรองวันที่นำเข้า (Import Date)
        if import_from_str or import_to_str:
            imp_from = parse_date_any(import_from_str) if import_from_str else None
            imp_to = parse_date_any(import_to_str) if import_to_str else None
            filtered_rows = []
            for r in rows:
                d = r.get("import_date")
                if isinstance(d, str):
                    try:
                        d = datetime.strptime(d, "%Y-%m-%d").date()
                    except:
                        d = None
                elif isinstance(d, datetime):
                    d = d.date()
                
                if d:
                    if imp_from and d < imp_from:
                        continue
                    if imp_to and d > imp_to:
                        continue
                elif imp_from or imp_to:
                    continue
                filtered_rows.append(r)
            rows = filtered_rows

        safe = []
        for r in rows:
            r = dict(r)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except Exception: stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)

        # [CRITICAL FIX] Logic กรองสำหรับหน้าประวัติ
        # เพราะออเดอร์ถูก mark เป็น ISSUED แล้ว allocation_status อาจไม่ใช่ LOW_STOCK
        # ต้อง fallback เช็ค stock condition แทน
        def _is_low_for_history(r):
            # 1. ถ้า status เป็น LOW_STOCK อยู่แล้ว -> เอา
            if r.get("allocation_status") == "LOW_STOCK": return True
            # 2. Fallback: ถ้า stock <= 3 (เกณฑ์ Low Stock มาตรฐาน) -> เอา
            try: s = int(r.get("stock_qty") or 0)
            except: s = 0
            if s <= 3: return True
            return False

        low_skus = {(r.get("sku") or "").strip() for r in safe if _is_low_for_history(r)}
        lines = [r for r in safe if (r.get("sku") or "").strip() in low_skus]

        # เตรียมข้อมูล Mixed Status สำหรับหน้าประวัติ
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "SHORTAGE": "ไม่มีของ",
            "NOT_ENOUGH": "ไม่พอส่ง",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "LOW_STOCK":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # กรองตามคำค้นหา (q)
        if q:
            q_lower = q.lower()
            lines = [
                r for r in lines
                if q_lower in (
                    str(r.get("order_id") or "") +
                    str(r.get("sku") or "") +
                    str(r.get("brand") or "") +
                    str(r.get("model") or "") +
                    str(r.get("shop") or "") +
                    str(r.get("platform") or "") +
                    str(r.get("logistic") or "")
                ).lower()
            ]

        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()  # [FIX] เพิ่มการดึงค่า oid ในแต่ละรอบ
            out.append({
                "platform":      r.get("platform"),
                "store":         r.get("shop"),
                "order_no":      oid,
                "sku":           r.get("sku"),
                "brand":         r.get("brand"),
                "product_name":  r.get("model"),
                "stock":         int(r.get("stock_qty", 0) or 0),
                "qty":           int(r.get("qty", 0) or 0),
                "order_time":    r.get("order_time"),
                "due_date":      r.get("due_date"),
                "sla":           r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round":  r.get("lowstock_round"),
                "printed_count": 0,
                "note":          mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        # เรียง
        sort_col = sort_col if sort_col in {"platform","store","order_no","sku","brand","product_name","stock","qty","allqty","order_time","due_date","sla","shipping_type","assign_round","printed_count"} else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            if sort_col in {"stock","qty","allqty","assign_round","printed_count"}:
                try: return int(v.get(sort_col) or 0)
                except: return 0
            elif sort_col in {"order_time","due_date"}:
                try: return datetime.fromisoformat(str(v.get(sort_col)))
                except: return str(v.get(sort_col) or "")
            else:
                return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
        counts_low = _get_print_counts_local(order_ids, "lowstock")
        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts_low.get(oid, 0))

        # ข้อ 1: ดึงเวลา printed_lowstock_at ต่อ order_id จาก DB
        tbl = _ol_table_name()
        sql_ts = text(f"""
            SELECT order_id, MAX(printed_lowstock_at) AS ts
            FROM {tbl}
            WHERE order_id IN :oids AND printed_lowstock_at IS NOT NULL
            GROUP BY order_id
        """).bindparams(bindparam("oids", expanding=True))
        rows_ts = db.session.execute(sql_ts, {"oids": order_ids}).all()
        ts_map = {}
        for oid, ts in rows_ts:
            if not ts:
                continue
            try:
                dt = datetime.fromisoformat(ts)
                if dt.tzinfo is None:
                    dt = TH_TZ.localize(dt)
                ts_map[str(oid)] = dt
            except Exception:
                pass

        # ใส่ลงในแต่ละแถว
        for r in out:
            r["printed_at"] = ts_map.get((r.get("order_no") or "").strip())

        # เวลาพิมพ์บนหัวรายงาน (ล่าสุดสุดในชุด)
        meta_printed_at = max(ts_map.values()) if ts_map else None

        # ดึงค่า lowstock_round จาก DB เพื่อให้แน่ใจว่าหน้าประวัติแสดงเลขรอบ (แก้ปัญหาเลขหาย)
        if order_ids:
            tbl = _ol_table_name()
            sql = text(f"""
                SELECT order_id, MAX(lowstock_round) AS r
                  FROM {tbl}
                 WHERE order_id IN :oids
                 GROUP BY order_id
            """).bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids}).all()
                round_map = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
                for r in out:
                    oid = (r.get("order_no") or "").strip()
                    if oid in round_map and round_map[oid] is not None:
                        r["assign_round"] = round_map[oid]
            except Exception:
                pass  # ถ้าคอลัมน์ยังไม่มีก็ข้าม

        # กรองตามรอบ (หลังจากดึงค่าจาก DB แล้ว)
        if round_num and round_num != "all":
            try:
                r_int = int(round_num)
                out = [r for r in out if r.get("assign_round") == r_int]
                # อัปเดต order_ids หลังกรอง
                order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
            except:
                pass

        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if order_ids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": order_ids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_lowstock.html",
            rows=out,
            summary={"sku_count": len(low_skus), "orders_count": len(order_ids)},
            printed_at=meta_printed_at,  # ข้อ 1: ใช้เวลาจริงที่ถูกบันทึกไว้
            order_ids=order_ids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            is_history_view=True,
            available_dates=_available_dates(),
            print_date_from=print_date_from,
            print_date_to=print_date_to,
            sort_col=sort_col,
            sort_dir=sort_dir,
            q=q,
            round_sel=round_num,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str
        )

    @app.route("/report/lowstock.xlsx", methods=["GET"])
    @login_required
    def report_lowstock_export():
        """ส่งออกรายงานสินค้าน้อยเป็น Excel (ข้อ 2: ตรงกับตารางในหน้าเว็บ)"""
        # ไม่ต้องใช้ services.lowstock_queue แล้ว
        
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q        = (request.args.get("q") or "").strip()
        sort_col = (request.args.get("sort") or "order_no").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        round_num = request.args.get("round")
        
        # รับค่าวันที่กรอง (เพิ่มใหม่)
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")

        filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
            "import_date": None
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)
        
        # คำนวณออเดอร์ที่แพ็คแล้ว (เช็คจาก sales_status)
        packed_oids = _orders_packed_set(rows)
        
        # ข้อ 4: กรอง PACKED
        safe = []
        for r in rows:
            r = dict(r)
            # กรองออเดอร์ที่อยู่ในลิสต์แพ็คแล้วออก
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            sales_status = (str(r.get("sales_status") or "")).upper()
            if sales_status == "PACKED" or bool(r.get("packed", False)):
                continue
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except: stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)

        orders_low = _orders_lowstock_order_set(safe)
        safe = [r for r in safe if (r.get("order_id") or "").strip() in orders_low]
        
        # กรองเฉพาะ allocation_status == "LOW_STOCK" ตาม compute_allocation
        lines = [r for r in safe if r.get("allocation_status") == "LOW_STOCK"]

        # กรองเพิ่ม
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        if q:
            ql = q.lower()
            def _hit(s): return ql in (str(s or "").lower())
            lines = [r for r in lines if (
                _hit(r.get("order_id")) or _hit(r.get("sku")) or _hit(r.get("brand")) or
                _hit(r.get("model")) or _hit(r.get("shop")) or _hit(r.get("platform")) or _hit(r.get("logistic"))
            )]
        if round_num and round_num != "all":
            try:
                r_int = int(round_num)
                lines = [r for r in lines if r.get("lowstock_round") == r_int]
            except: pass

        # กรองวันที่ (เพิ่มใหม่)
        def _parse_date(s):
            if not s: return None
            try: return datetime.strptime(s, "%Y-%m-%d").date()
            except: return None
        
        d_from = _parse_date(date_from_str)
        d_to = _parse_date(date_to_str)
        i_from = _parse_date(import_from_str)
        i_to = _parse_date(import_to_str)
        
        if d_from or d_to or i_from or i_to:
            filtered_lines = []
            for r in lines:
                # กรองวันสั่งซื้อ
                pass_order = True
                if d_from or d_to:
                    od = r.get("order_time")
                    if isinstance(od, str):
                        try: od = datetime.strptime(od.split()[0], "%Y-%m-%d").date()
                        except: od = None
                    elif isinstance(od, datetime): od = od.date()
                    elif hasattr(od, 'date'): od = od.date()
                    else: od = None
                    
                    if od:
                        if d_from and od < d_from: pass_order = False
                        if d_to and od > d_to: pass_order = False
                    elif d_from or d_to:
                        pass_order = False
                
                # กรองวันนำเข้า
                pass_import = True
                if i_from or i_to:
                    id_ = r.get("import_date")
                    if isinstance(id_, str):
                        try: id_ = datetime.strptime(id_, "%Y-%m-%d").date()
                        except: id_ = None
                    elif isinstance(id_, datetime): id_ = id_.date()
                    elif hasattr(id_, 'date'): id_ = id_.date()
                    else: id_ = None
                    
                    if id_:
                        if i_from and id_ < i_from: pass_import = False
                        if i_to and id_ > i_to: pass_import = False
                    elif i_from or i_to:
                        pass_import = False

                if pass_order and pass_import:
                    filtered_lines.append(r)
            lines = filtered_lines

        # คำนวณ AllQty
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in lines:
            sum_by_sku[(r.get("sku") or "").strip()] += int(r.get("qty") or 0)

        # อ่านค่า lowstock_round จาก DB เหมือนหน้ารายงาน (ข้อ 1)
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        low_round_by_oid = {}
        if order_ids_for_round:
            # ใช้ raw SQL แทน ORM เพราะ lowstock_round ไม่มีในโมเดล
            tbl = _ol_table_name()
            sql = text(f"""
                SELECT order_id, MAX(lowstock_round) AS r
                  FROM {tbl}
                 WHERE order_id IN :oids
                 GROUP BY order_id
            """).bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                low_round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except Exception:
                low_round_by_oid = {}

        # สร้าง output rows
        out = []
        for r in lines:
            sku = (r.get("sku") or "").strip()
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform":      r.get("platform"),
                "store":         r.get("shop"),
                "order_no":      oid,
                "sku":           sku,
                "brand":         r.get("brand"),
                "product_name":  r.get("model"),
                "stock":         int(r.get("stock_qty", 0) or 0),
                "qty":           int(r.get("qty", 0) or 0),
                "allqty":        sum_by_sku[sku],
                "order_time":    r.get("order_time"),
                "due_date":      r.get("due_date"),
                "sla":           r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round":  low_round_by_oid.get(oid, r.get("lowstock_round")),  # <<<< ใช้ค่าจาก DB
            })

        # เรียง
        sort_col = sort_col if sort_col in {"platform","store","order_no","sku","brand","product_name","stock","qty","allqty","order_time","due_date","sla","shipping_type","assign_round","printed_count"} else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            if sort_col in {"stock","qty","allqty","assign_round","printed_count"}:
                try: return int(v.get(sort_col) or 0)
                except: return 0
            elif sort_col in {"order_time","due_date"}:
                try: return datetime.fromisoformat(str(v.get(sort_col)))
                except: return str(v.get(sort_col) or "")
            else:
                return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        # เพิ่มคอลัมน์ "พิมพ์แล้ว(ครั้ง)"
        order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
        counts_low = _get_print_counts_local(order_ids, "lowstock")
        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts_low.get(oid, 0))
        
        # สร้าง DataFrame
        df_data = []
        for r in out:
            df_data.append({
                "แพลตฟอร์ม": r["platform"],
                "ร้าน": r["store"],
                "เลข Order": r["order_no"],
                "SKU": r["sku"],
                "Brand": r["brand"],
                "ชื่อสินค้า": r["product_name"],
                "Stock": r["stock"],
                "Qty": r["qty"],
                "AllQty": r["allqty"],
                "เวลาที่ลูกค้าสั่ง": r["order_time"],
                "กำหนดส่ง": r["due_date"],
                "SLA (ชม.)": r["sla"],
                "ประเภทขนส่ง": r["shipping_type"],
                "จ่ายงาน(รอบที่)": r["assign_round"] if r["assign_round"] is not None else "",
                "พิมพ์แล้ว(ครั้ง)": r["printed_count"],
            })

        df = pd.DataFrame(df_data)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="LowStock")
        bio.seek(0)
        
        filename = f"lowstock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


    @app.route("/report/nostock", methods=["GET"])
    @login_required
    def report_nostock():
        """
        รายงานไม่มีสินค้า — กรองเฉพาะ SHORTAGE (stock = 0) เท่านั้น
        """
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        round_num = request.args.get("round")
        q        = (request.args.get("q") or "").strip()
        sort_col = (request.args.get("sort") or "").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าวันที่กรอง
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")

        shops = Shop.query.order_by(Shop.name.asc()).all()

        # 1) ดึง allocation rows
        filters = {"platform": platform or None, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)

        # คำนวณออเดอร์ที่แพ็คแล้ว (เช็คจาก sales_status)
        packed_oids = _orders_packed_set(rows)

        # เติม stock_qty/logistic
        safe = []
        for r in rows:
            r = dict(r)
            # กรองออเดอร์ที่อยู่ในลิสต์แพ็คแล้วออก
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            if (str(r.get("sales_status") or "")).upper() == "PACKED" or bool(r.get("packed", False)):
                continue
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)

        # กรองตามวันที่สั่งซื้อและวันที่นำเข้า
        if date_from_str or date_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            date_from = parse_date_str(date_from_str)
            date_to = parse_date_str(date_to_str)
            filtered = []
            for r in safe:
                order_dt = r.get("order_time")
                if isinstance(order_dt, str):
                    try: order_dt = datetime.strptime(order_dt.split()[0], "%Y-%m-%d").date()
                    except: order_dt = None
                elif isinstance(order_dt, datetime):
                    order_dt = order_dt.date()
                if order_dt:
                    if date_from and order_dt < date_from: continue
                    if date_to and order_dt > date_to: continue
                elif date_from or date_to:
                    continue
                filtered.append(r)
            safe = filtered
        
        if import_from_str or import_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            import_from = parse_date_str(import_from_str)
            import_to = parse_date_str(import_to_str)
            filtered = []
            for r in safe:
                imp_dt = r.get("import_date")
                if isinstance(imp_dt, str):
                    try: imp_dt = datetime.strptime(imp_dt, "%Y-%m-%d").date()
                    except: imp_dt = None
                elif isinstance(imp_dt, datetime):
                    imp_dt = imp_dt.date()
                elif isinstance(imp_dt, date):
                    pass
                else:
                    imp_dt = None
                if imp_dt:
                    if import_from and imp_dt < import_from: continue
                    if import_to and imp_dt > import_to: continue
                elif import_from or import_to:
                    continue
                filtered.append(r)
            safe = filtered

        # 2) กรองเฉพาะ allocation_status == "SHORTAGE" ตาม compute_allocation
        lines = [r for r in safe if r.get("allocation_status") == "SHORTAGE"]

        # 3) ฟิลเตอร์
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        if q:
            ql = q.lower()
            lines = [r for r in lines if ql in (str(r.get("order_id","")) + str(r.get("sku","")) + 
                    str(r.get("brand","")) + str(r.get("model","")) + str(r.get("shop",""))).lower()]

        # 4) ดึงค่า nostock_round จาก DB
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        nostock_round_by_oid = {}
        if order_ids_for_round:
            tbl = _ol_table_name()
            sql = text(f"SELECT order_id, MAX(nostock_round) AS r FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                nostock_round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except Exception:
                nostock_round_by_oid = {}

        # กรองตาม round ถ้ามีเลือก
        if round_num not in (None, "", "all"):
            try:
                round_filter = int(round_num)
                lines = [r for r in lines if nostock_round_by_oid.get((r.get("order_id") or "").strip()) == round_filter]
            except:
                pass

        # เตรียมข้อมูล Mixed Status
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "LOW_STOCK": "สินค้าน้อย",
            "NOT_ENOUGH": "ไม่พอส่ง",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "SHORTAGE":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        # 5) แปลงเป็นคอลัมน์รายงาน
        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform":      r.get("platform"),
                "store":         r.get("shop"),
                "order_no":      oid,
                "sku":           r.get("sku"),
                "brand":         r.get("brand"),
                "product_name":  r.get("model"),
                "stock":         int(r.get("stock_qty", 0) or 0),
                "qty":           int(r.get("qty", 0) or 0),
                "order_time":    r.get("order_time"),
                "due_date":      r.get("due_date"),
                "sla":           r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round":  nostock_round_by_oid.get(oid, r.get("nostock_round")),
                "printed_count": 0,
                "note":          mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        # 6) เรียงลำดับ
        sort_col = sort_col if sort_col in {"platform","store","order_no","sku","brand","product_name","stock","qty","allqty","order_time","due_date","sla","shipping_type","assign_round","printed_count"} else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            if sort_col in {"stock","qty","allqty","assign_round","printed_count"}:
                try: return int(v.get(sort_col) or 0)
                except: return 0
            elif sort_col in {"order_time","due_date"}:
                try: return datetime.fromisoformat(str(v.get(sort_col)))
                except: return str(v.get(sort_col) or "")
            else:
                return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        # 7) นับ "พิมพ์แล้ว(ครั้ง)"
        order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
        counts_nostock = _get_print_counts_local(order_ids, "nostock")
        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts_nostock.get(oid, 0))
            r["printed_at"] = None  # ไม่แสดงเวลาในหน้าปกติ

        # 8) กรองเฉพาะออเดอร์ที่ยังไม่พิมพ์
        out = [r for r in out if (r.get("printed_count") or 0) == 0]

        # 9) คำนวณสรุป + order_ids ใหม่หลังกรอง
        order_ids = sorted({(r.get("order_no") or "").strip() for r in out if r.get("order_no")})
        nostock_skus = {(r["sku"] or "").strip() for r in out if r.get("sku")}
        summary = {"sku_count": len(nostock_skus), "orders_count": len(order_ids)}

        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))
        available_rounds = sorted({r["assign_round"] for r in out if r["assign_round"] is not None})

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if order_ids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": order_ids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_nostock_READY.html",
            rows=out,
            summary=summary,
            printed_at=None,
            order_ids=order_ids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            round_sel=round_num,
            available_rounds=available_rounds,
            sort_col=sort_col,
            sort_dir=("desc" if rev else "asc"),
            q=q,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str,
            mixed_status=mixed_info,
            is_history_view=False
        )

    @app.post("/report/nostock/print")
    @login_required
    def report_nostock_print():
        """บันทึกการพิมพ์รายงานไม่มีสินค้า + ย้ายไปหน้าประวัติ"""
        cu = current_user()
        order_ids_raw = (request.form.get("order_ids") or "").strip()
        order_ids = [s.strip() for s in order_ids_raw.split(",") if s.strip()]
        if not order_ids:
            flash("ไม่พบออเดอร์สำหรับพิมพ์", "warning")
            return redirect(url_for("report_nostock"))

        # ตรวจสอบว่าทุกออเดอร์มีรอบหรือไม่ (STRICT MODE)
        orders = OrderLine.query.filter(OrderLine.order_id.in_(order_ids)).all()
        orders_without_round = [o.order_id for o in orders if not o.nostock_round]

        if orders_without_round:
            head = ', '.join(orders_without_round[:5])
            more = f" และอีก {len(orders_without_round)-5} รายการ" if len(orders_without_round) > 5 else ""
            flash(f"ออเดอร์เหล่านี้ยังไม่ได้บันทึกรอบ: {head}{more}", "danger")
            flash("กรุณาบันทึกรอบให้ทุกออเดอร์ก่อนทำการพิมพ์", "warning")
            return redirect(url_for("report_nostock"))

        now_iso = now_thai().isoformat()

        # 1. บันทึกว่าพิมพ์ No Stock แล้ว
        _mark_nostock_printed(order_ids, username=(cu.username if cu else None), when_iso=now_iso)
        
        # 2. ย้ายไป "Order จ่ายแล้ว" (Issued) ทันที
        _mark_issued(order_ids, user_id=(cu.id if cu else None), source="print:nostock", when_dt=now_thai())
        
        db.session.commit()
        return redirect(url_for("report_nostock_printed", auto_print="1"))

    @app.get("/report/nostock/printed")
    @login_required
    def report_nostock_printed():
        """ประวัติรายงานไม่มีสินค้าที่พิมพ์แล้ว"""
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q        = (request.args.get("q") or "").strip()
        round_num = request.args.get("round")
        sort_col = (request.args.get("sort") or "order_no").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าตัวกรองวันที่สั่งซื้อและนำเข้า
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        # รับค่าช่วงวันที่พิมพ์
        print_date_from = request.args.get("print_date_from")
        print_date_to = request.args.get("print_date_to")
        reset_mode = request.args.get("reset")  # [NEW] รับค่า reset
        action = request.args.get("action")  # [NEW] รับค่า action (เพื่อแยกการกดปุ่มกรอง กับการเข้าหน้าเว็บครั้งแรก)
        
        # [SMART DEFAULT] ถ้าไม่มีวันที่ส่งมา AND ไม่มีคำค้นหา AND ไม่ได้ reset AND ไม่ใช่การกดปุ่มกรอง -> ให้กรอง "วันนี้"
        if not action and reset_mode != 'all' and not print_date_from and not print_date_to and not q:
            # เข้าหน้าเว็บครั้งแรก (ไม่มี action) = ดูงานวันนี้
            today = now_thai().date().isoformat()
            print_date_from = today
            print_date_to = today
        # ถ้ามี action (กดปุ่มกรอง) หรือ q หรือ reset='all' แต่ไม่มีวันที่ -> ค้นหาทั้งหมด

        tbl = _ol_table_name()
        
        # ========================================================
        # [FIX] ดึงข้อมูลเฉพาะเมื่อ: มีคำค้นหา หรือ มีการเลือกวันที่
        # ========================================================
        if q:
            # กรณี 1: มีคำค้นหา -> ค้นหาทั้งหมด (Global Search)
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_nostock > 0")
            result = db.session.execute(sql).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        elif print_date_from or print_date_to:
            # กรณี 2: มีการเลือกวันที่ -> กรองตามวันที่
            sql_where = "printed_nostock > 0"
            params = {}
            if print_date_from:
                sql_where += " AND DATE(printed_nostock_at) >= :pf"
                params["pf"] = print_date_from
            if print_date_to:
                sql_where += " AND DATE(printed_nostock_at) <= :pt"
                params["pt"] = print_date_to
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE {sql_where}")
            result = db.session.execute(sql, params).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        else:
            # กรณี 3: ไม่ค้นหา และ ไม่เลือกวัน (เช่น กด reset='all') -> ไม่แสดงอะไร
            printed_oids = []

        def _available_dates():
            sql = text(f"SELECT DISTINCT DATE(printed_nostock_at) as d FROM {tbl} WHERE printed_nostock > 0 AND printed_nostock_at IS NOT NULL ORDER BY d DESC")
            return [r[0] for r in db.session.execute(sql).fetchall()]

        shops = Shop.query.order_by(Shop.name.asc()).all()
        
        if not printed_oids:
            return render_template(
                "report_nostock_READY.html",
                rows=[],
                summary={"sku_count": 0, "orders_count": 0},
                printed_at=None,
                order_ids=[],
                shops=shops,
                logistics=[],
                platform_sel=platform,
                shop_sel=shop_id,
                logistic_sel=logistic,
                is_history_view=True,
                available_dates=_available_dates(),
                print_date_from=print_date_from,
                print_date_to=print_date_to,
                sort_col=sort_col,
                sort_dir=sort_dir,
                q=q,
                round_sel=round_num,
                date_from=date_from_str,
                date_to=date_to_str,
                import_from=import_from_str,
                import_to=import_to_str
            )

        # เตรียมตัวกรองวันที่สั่งซื้อ
        date_from_dt = None
        date_to_dt = None
        if date_from_str:
            try:
                date_from_dt = datetime.combine(parse_date_any(date_from_str), datetime.min.time(), tzinfo=TH_TZ)
            except: pass
        if date_to_str:
            try:
                date_to_dt = datetime.combine(parse_date_any(date_to_str) + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ)
            except: pass

        filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
            "import_date": None,
            "date_from": date_from_dt,
            "date_to": date_to_dt
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if (r.get("order_id") or "").strip() in printed_oids]
        
        # กรองวันที่นำเข้า (Import Date) - [แก้ไข] ถ้าไม่มีวันที่ก็เอามาด้วย
        if import_from_str or import_to_str:
            from datetime import date as date_type
            imp_from = parse_date_any(import_from_str) if import_from_str else None
            imp_to = parse_date_any(import_to_str) if import_to_str else None
            filtered_rows = []
            for r in rows:
                raw_d = r.get("import_date")
                d_obj = None
                if isinstance(raw_d, str):
                    try: d_obj = datetime.strptime(raw_d, "%Y-%m-%d").date()
                    except: pass
                elif isinstance(raw_d, datetime):
                    d_obj = raw_d.date()
                elif isinstance(raw_d, date_type):
                    d_obj = raw_d
                
                if d_obj:
                    if imp_from and d_obj < imp_from: continue
                    if imp_to and d_obj > imp_to: continue
                    filtered_rows.append(r)
                else:
                    # ข้อมูลไม่มีวันที่นำเข้า -> เอามาด้วย
                    filtered_rows.append(r)
            rows = filtered_rows

        safe = []
        for r in rows:
            r = dict(r)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except Exception: stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            # ไม่ต้อง _recompute เพราะ allocation_status มาจาก compute_allocation แล้ว
            safe.append(r)

        # กรองเฉพาะ SHORTAGE (stock = 0)
        def is_nostock(r):
            try:
                stk = int(r.get("stock_qty") or 0)
            except:
                stk = 0
            return (r.get("allocation_status") == "SHORTAGE") or (stk <= 0)
        
        lines = [r for r in safe if is_nostock(r)]

        # เตรียมข้อมูล Mixed Status สำหรับหน้าประวัติ
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "LOW_STOCK": "สินค้าน้อย",
            "NOT_ENOUGH": "ไม่พอส่ง",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "SHORTAGE":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # กรองตามคำค้นหา (q)
        if q:
            q_lower = q.lower()
            lines = [
                r for r in lines
                if q_lower in (
                    str(r.get("order_id") or "") +
                    str(r.get("sku") or "") +
                    str(r.get("brand") or "") +
                    str(r.get("model") or "") +
                    str(r.get("shop") or "") +
                    str(r.get("platform") or "") +
                    str(r.get("logistic") or "")
                ).lower()
            ]

        # ดึงค่า nostock_round จาก DB
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        nostock_round_by_oid = {}
        if order_ids_for_round:
            sql = text(f"SELECT order_id, MAX(nostock_round) AS r FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                nostock_round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except Exception:
                nostock_round_by_oid = {}

        # กรองตาม round ถ้ามี
        if round_num not in (None, "", "all"):
            try:
                round_filter = int(round_num)
                lines = [r for r in lines if nostock_round_by_oid.get((r.get("order_id") or "").strip()) == round_filter]
            except:
                pass

        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform":      r.get("platform"),
                "store":         r.get("shop"),
                "order_no":      oid,
                "sku":           r.get("sku"),
                "brand":         r.get("brand"),
                "product_name":  r.get("model"),
                "stock":         int(r.get("stock_qty", 0) or 0),
                "qty":           int(r.get("qty", 0) or 0),
                "order_time":    r.get("order_time"),
                "due_date":      r.get("due_date"),
                "sla":           r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round":  nostock_round_by_oid.get(oid, r.get("nostock_round")),
                "printed_count": 0,
                "note":          mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        # เรียง
        sort_col = sort_col if sort_col in {"platform","store","order_no","sku","brand","product_name","stock","qty","allqty","order_time","due_date","sla","shipping_type","assign_round","printed_count"} else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            if sort_col in {"stock","qty","allqty","assign_round","printed_count"}:
                try: return int(v.get(sort_col) or 0)
                except: return 0
            elif sort_col in {"order_time","due_date"}:
                try: return datetime.fromisoformat(str(v.get(sort_col)))
                except: return str(v.get(sort_col) or "")
            else:
                return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
        counts_nostock = _get_print_counts_local(order_ids, "nostock")
        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts_nostock.get(oid, 0))

        # ดึงเวลาพิมพ์จาก DB
        sql_ts = text(f"""
            SELECT order_id, MAX(printed_nostock_at) AS ts
            FROM {tbl}
            WHERE printed_nostock > 0
              AND order_id IN :oids
            GROUP BY order_id
        """).bindparams(bindparam("oids", expanding=True))
        rows_ts = db.session.execute(sql_ts, {"oids": order_ids}).all() if order_ids else []
        ts_map = {}
        for row_ts in rows_ts:
            if not row_ts or not row_ts[0] or not row_ts[1]:
                continue
            oid_str = str(row_ts[0]).strip()
            ts_str = row_ts[1]
            try:
                dt = datetime.fromisoformat(ts_str)
                if dt.tzinfo is None:
                    dt = TH_TZ.localize(dt)
                ts_map[oid_str] = dt
            except Exception:
                pass

        for r in out:
            r["printed_at"] = ts_map.get((r.get("order_no") or "").strip())

        meta_printed_at = max(ts_map.values()) if ts_map else None

        # ดึงค่า nostock_round จาก DB
        if order_ids:
            sql = text(f"SELECT order_id, MAX(nostock_round) AS r FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids}).all()
                round_map = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
                for r in out:
                    oid = (r.get("order_no") or "").strip()
                    if oid in round_map and round_map[oid] is not None:
                        r["assign_round"] = round_map[oid]
            except Exception:
                pass

        if round_num and round_num != "all":
            try:
                r_int = int(round_num)
                out = [r for r in out if r.get("assign_round") == r_int]
                order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
            except:
                pass

        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))
        nostock_skus = {(r["sku"] or "").strip() for r in out if r.get("sku")}

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if order_ids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": order_ids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_nostock_READY.html",
            rows=out,
            summary={"sku_count": len(nostock_skus), "orders_count": len(order_ids)},
            printed_at=meta_printed_at,
            order_ids=order_ids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            is_history_view=True,
            available_dates=_available_dates(),
            print_date_from=print_date_from,
            print_date_to=print_date_to,
            sort_col=sort_col,
            sort_dir=sort_dir,
            q=q,
            round_sel=round_num,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str
        )

    @app.route("/report/nostock.xlsx", methods=["GET"])
    @login_required
    def report_nostock_export():
        """Export Excel รายงานไม่มีสินค้า"""
        # ไม่ต้องใช้ services.lowstock แล้ว
        import pandas as pd
        
        platform = normalize_platform(request.args.get("platform"))
        shop_id = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q = (request.args.get("q") or "").strip()
        round_num = request.args.get("round")
        
        # รับค่าวันที่กรอง (เพิ่มใหม่)
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        filters = {"platform": platform or None, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)
        
        # คำนวณออเดอร์ที่แพ็คแล้ว (เช็คจาก sales_status)
        packed_oids = _orders_packed_set(rows)
        
        safe = []
        for r in rows:
            r = dict(r)
            # กรองออเดอร์ที่อยู่ในลิสต์แพ็คแล้วออก
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            if (str(r.get("sales_status") or "")).upper() == "PACKED":
                continue
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except: stock_qty = 0
                r["stock_qty"] = stock_qty
            safe.append(r)
        
        # กรองเฉพาะ allocation_status == "SHORTAGE"
        lines = [r for r in safe if r.get("allocation_status") == "SHORTAGE"]
        
        # ---------- กรอง logistic ----------
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").strip().upper() == logistic.strip().upper()]
        
        # ---------- กรอง round ----------
        if round_num:
            try:
                round_int = int(round_num)
                lines = [r for r in lines if r.get("nostock_round") == round_int]
            except:
                pass
        
        # ---------- กรองด้วย date filters (เพิ่มใหม่) ----------
        def _parse_date(d_str):
            if not d_str:
                return None
            from utils import parse_datetime_guess
            dt = parse_datetime_guess(d_str)
            if dt:
                return dt.date() if hasattr(dt, 'date') else dt
            return None
        
        date_from = _parse_date(date_from_str)
        date_to = _parse_date(date_to_str)
        import_from = _parse_date(import_from_str)
        import_to = _parse_date(import_to_str)
        
        # กรองด้วยวันที่สั่ง (order_time / due_date)
        if date_from or date_to:
            filtered = []
            for r in lines:
                order_time = r.get("order_time") or r.get("due_date")
                if not order_time:
                    continue
                try:
                    from utils import parse_datetime_guess
                    dt = parse_datetime_guess(order_time)
                    if dt:
                        dt_date = dt.date() if hasattr(dt, 'date') else dt
                        if date_from and dt_date < date_from:
                            continue
                        if date_to and dt_date > date_to:
                            continue
                        filtered.append(r)
                except:
                    continue
            lines = filtered
        
        # กรองด้วยวันที่นำเข้า (import_date)
        if import_from or import_to:
            filtered = []
            for r in lines:
                imp_date = r.get("import_date")
                if not imp_date:
                    continue
                try:
                    from utils import parse_datetime_guess
                    dt = parse_datetime_guess(imp_date)
                    if dt:
                        dt_date = dt.date() if hasattr(dt, 'date') else dt
                        if import_from and dt_date < import_from:
                            continue
                        if import_to and dt_date > import_to:
                            continue
                        filtered.append(r)
                except:
                    continue
            lines = filtered
        
        # ---------- กรองด้วยคำค้นหา q ----------
        if q:
            q_lower = q.lower()
            lines = [r for r in lines if q_lower in (r.get("sku") or "").lower() 
                     or q_lower in (r.get("model") or "").lower() 
                     or q_lower in (r.get("order_id") or "").lower()]
        
        df = pd.DataFrame([{
            "แพลตฟอร์ม": r.get("platform"),
            "ร้าน": r.get("shop"),
            "เลข Order": r.get("order_id"),
            "SKU": r.get("sku"),
            "Brand": r.get("brand"),
            "ชื่อสินค้า": r.get("model"),
            "Stock": int(r.get("stock_qty", 0) or 0),
            "Qty": int(r.get("qty", 0) or 0),
            "เวลาที่ลูกค้าสั่ง": r.get("order_time"),
            "กำหนดส่ง": r.get("due_date"),
            "ประเภทขนส่ง": r.get("logistic"),
        } for r in lines])
        
        out = BytesIO()
        with pd.ExcelWriter(out, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="NoStock")
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="report_nostock.xlsx")

    # ================== NEW: Update No Stock Round ==================
    @app.route("/report/nostock/update_round", methods=["POST"])
    @login_required
    def update_nostock_round():
        """อัปเดตรอบสำหรับรายงานไม่มีสินค้า"""
        data = request.get_json() or {}
        order_ids = data.get("order_ids", [])
        round_num = data.get("round")
        
        if not order_ids or round_num is None:
            return jsonify({"success": False, "message": "ข้อมูลไม่ครบ"})
        
        try:
            round_int = int(round_num)
        except:
            return jsonify({"success": False, "message": "รอบต้องเป็นตัวเลข"})
        
        tbl = _ol_table_name()
        sql = text(f"UPDATE {tbl} SET nostock_round = :r WHERE order_id IN :oids")
        sql = sql.bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"r": round_int, "oids": order_ids})
        db.session.commit()
        
        return jsonify({"success": True, "message": f"อัปเดตรอบเป็น {round_int} สำเร็จ ({len(order_ids)} ออเดอร์)"})
    # ================== /NEW ==================

    # ================== NEW: Report Not Enough (สินค้าไม่พอส่ง) ==================
    @app.route("/report/notenough", methods=["GET"])
    @login_required
    def report_notenough():
        """รายงานสินค้าไม่พอส่ง (NOT_ENOUGH) — กรองเฉพาะสินค้าไม่พอส่ง"""
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        round_num = request.args.get("round")
        q        = (request.args.get("q") or "").strip()
        sort_col = (request.args.get("sort") or "").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าวันที่กรอง
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")

        shops = Shop.query.order_by(Shop.name.asc()).all()

        # 1) ดึง allocation rows
        filters = {"platform": platform or None, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)

        # คำนวณออเดอร์ที่แพ็คแล้ว (เช็คจาก sales_status)
        packed_oids = _orders_packed_set(rows)
        
        safe = []
        for r in rows:
            r = dict(r)
            # กรองออเดอร์ที่อยู่ในลิสต์แพ็คแล้วออก
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            # หรือถ้า sales_status เป็น 'PACKED' ก็ข้ามไป
            if (str(r.get("sales_status") or "")).upper() == "PACKED":
                continue
            if bool(r.get("packed", False)):
                continue
            
            # ตรวจ stock_qty (ถ้า compute_allocation ไม่ได้เติมให้)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except:
                            stock_qty = 0
                    if not prod:
                        st = Stock.query.filter_by(sku=sku).first()
                        if st and st.qty is not None:
                            stock_qty = int(st.qty)
                r["stock_qty"] = stock_qty
            
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)

        # กรองตามวันที่สั่งซื้อและวันที่นำเข้า
        if date_from_str or date_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            date_from = parse_date_str(date_from_str)
            date_to = parse_date_str(date_to_str)
            filtered = []
            for r in safe:
                order_dt = r.get("order_time")
                if isinstance(order_dt, str):
                    try: order_dt = datetime.strptime(order_dt.split()[0], "%Y-%m-%d").date()
                    except: order_dt = None
                elif isinstance(order_dt, datetime):
                    order_dt = order_dt.date()
                if order_dt:
                    if date_from and order_dt < date_from: continue
                    if date_to and order_dt > date_to: continue
                elif date_from or date_to:
                    continue
                filtered.append(r)
            safe = filtered
        
        if import_from_str or import_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            import_from = parse_date_str(import_from_str)
            import_to = parse_date_str(import_to_str)
            filtered = []
            for r in safe:
                imp_dt = r.get("import_date")
                if isinstance(imp_dt, str):
                    try: imp_dt = datetime.strptime(imp_dt, "%Y-%m-%d").date()
                    except: imp_dt = None
                elif isinstance(imp_dt, datetime):
                    imp_dt = imp_dt.date()
                elif isinstance(imp_dt, date):
                    pass
                else:
                    imp_dt = None
                if imp_dt:
                    if import_from and imp_dt < import_from: continue
                    if import_to and imp_dt > import_to: continue
                elif import_from or import_to:
                    continue
                filtered.append(r)
            safe = filtered

        # กรองเฉพาะ allocation_status == "NOT_ENOUGH" ตาม compute_allocation
        lines = [r for r in safe if r.get("allocation_status") == "NOT_ENOUGH"]

        # Filter ตามขนส่ง
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # Search
        if q:
            ql = q.lower()
            lines = [
                r for r in lines 
                if ql in (
                    str(r.get("order_id","")) + 
                    str(r.get("sku","")) + 
                    str(r.get("model","")) +
                    str(r.get("brand","")) +
                    str(r.get("shop","")) +
                    str(r.get("logistic",""))
                ).lower()
            ]

        # ดึง Round
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        round_by_oid = {}
        if order_ids_for_round:
            tbl = _ol_table_name()
            sql = text(f"SELECT order_id, MAX(notenough_round) AS r FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except:
                pass

        # Filter by round
        if round_num not in (None, "", "all"):
            try:
                r_int = int(round_num)
                lines = [r for r in lines if round_by_oid.get((r.get("order_id") or "").strip()) == r_int]
            except:
                pass

        # เตรียมข้อมูล Mixed Status
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "LOW_STOCK": "สินค้าน้อย",
            "SHORTAGE": "ไม่มีของ",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "NOT_ENOUGH":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        # Map output
        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform": r.get("platform"),
                "store": r.get("shop"),
                "order_no": oid,
                "sku": r.get("sku"),
                "brand": r.get("brand"),
                "product_name": r.get("model"),
                "stock": int(r.get("stock_qty", 0) or 0),
                "qty": int(r.get("qty", 0) or 0),
                "order_time": r.get("order_time"),
                "due_date": r.get("due_date"),
                "sla": r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round": round_by_oid.get(oid),
                "printed_count": 0,
                "printed_at": None,
                "note": mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        
        # AllQty
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        # Sort
        sort_col = sort_col if sort_col else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        # Print Count
        oids = sorted({(r["order_no"] or "").strip() for r in out if r["order_no"]})
        counts = _get_print_counts_local(oids, "notenough")
        
        # [เพิ่ม] ดึงเวลาพิมพ์ล่าสุด (printed_notenough_at) จาก DB
        ts_map = {}
        if oids:
            tbl = _ol_table_name()
            sql_ts = text(f"""
                SELECT order_id, MAX(printed_notenough_at) 
                FROM {tbl} 
                WHERE order_id IN :oids 
                GROUP BY order_id
            """).bindparams(bindparam("oids", expanding=True))
            try:
                res_ts = db.session.execute(sql_ts, {"oids": oids}).fetchall()
                for row in res_ts:
                    if row[1]:
                        dt = datetime.fromisoformat(row[1])
                        if dt.tzinfo is None: dt = TH_TZ.localize(dt)
                        ts_map[str(row[0])] = dt
            except: pass

        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts.get(oid, 0))
            r["printed_at"] = ts_map.get(oid)  # ใส่เวลาจริงแทน None

        # กรองที่พิมพ์แล้วออก (ไม่แสดงในรายงานหลัก)
        out = [r for r in out if r["printed_count"] == 0]
        
        # Summary
        final_oids = sorted({(r["order_no"] or "").strip() for r in out if r["order_no"]})
        skus = {(r["sku"] or "").strip() for r in out if r["sku"]}
        summary = {
            "sku_count": len(skus),
            "orders_count": len(final_oids),
        }
        
        # ดึงรายการขนส่ง
        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))
        
        # ดึงรอบที่มี
        available_rounds = sorted({r["assign_round"] for r in out if r["assign_round"] is not None})

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if final_oids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": final_oids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_notenough.html",
            rows=out,
            summary=summary,
            printed_at=None,
            order_ids=final_oids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            round_sel=round_num,
            available_rounds=available_rounds,
            sort_col=sort_col,
            sort_dir=sort_dir,
            q=q,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str,
            mixed_status=mixed_info,
            is_history_view=False
        )

    @app.post("/report/notenough/print")
    @login_required
    def report_notenough_print():
        """บันทึกการพิมพ์รายงานสินค้าไม่พอส่ง + ย้ายไปหน้าประวัติ"""
        cu = current_user()
        order_ids_raw = (request.form.get("order_ids") or "").strip()
        order_ids = [s.strip() for s in order_ids_raw.split(",") if s.strip()]
        if not order_ids:
            flash("ไม่พบออเดอร์สำหรับพิมพ์", "warning")
            return redirect(url_for("report_notenough"))

        # ตรวจสอบว่าทุกออเดอร์มีรอบหรือไม่ (STRICT MODE)
        orders = OrderLine.query.filter(OrderLine.order_id.in_(order_ids)).all()
        orders_without_round = [o.order_id for o in orders if not o.notenough_round]

        if orders_without_round:
            head = ', '.join(orders_without_round[:5])
            more = f" และอีก {len(orders_without_round)-5} รายการ" if len(orders_without_round) > 5 else ""
            flash(f"ออเดอร์เหล่านี้ยังไม่ได้บันทึกรอบ: {head}{more}", "danger")
            flash("กรุณาบันทึกรอบให้ทุกออเดอร์ก่อนทำการพิมพ์", "warning")
            return redirect(url_for("report_notenough"))

        now_iso = now_thai().isoformat()

        # 1. บันทึกว่าพิมพ์ Not Enough แล้ว
        _mark_notenough_printed(order_ids, username=(cu.username if cu else None), when_iso=now_iso)
        
        # 2. ย้ายไป "Order จ่ายแล้ว" (Issued) ทันที
        _mark_issued(order_ids, user_id=(cu.id if cu else None), source="print:notenough", when_dt=now_thai())
        
        db.session.commit()
        return redirect(url_for("report_notenough_printed", auto_print="1"))

    @app.get("/report/notenough/printed")
    @login_required
    def report_notenough_printed():
        """ประวัติรายงานสินค้าไม่พอส่งที่พิมพ์แล้ว"""
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q        = (request.args.get("q") or "").strip()
        round_num = request.args.get("round")
        sort_col = (request.args.get("sort") or "order_no").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าตัวกรองวันที่สั่งซื้อและนำเข้า
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        # รับค่าช่วงวันที่พิมพ์
        print_date_from = request.args.get("print_date_from")
        print_date_to = request.args.get("print_date_to")
        reset_mode = request.args.get("reset")  # [NEW] รับค่า reset
        action = request.args.get("action")  # [NEW] รับค่า action (เพื่อแยกการกดปุ่มกรอง กับการเข้าหน้าเว็บครั้งแรก)
        
        # [SMART DEFAULT] ถ้าไม่มีวันที่ส่งมา AND ไม่มีคำค้นหา AND ไม่ได้ reset AND ไม่ใช่การกดปุ่มกรอง -> ให้กรอง "วันนี้"
        if not action and reset_mode != 'all' and not print_date_from and not print_date_to and not q:
            # เข้าหน้าเว็บครั้งแรก (ไม่มี action) = ดูงานวันนี้
            today = now_thai().date().isoformat()
            print_date_from = today
            print_date_to = today
        # ถ้ามี action (กดปุ่มกรอง) หรือ q หรือ reset='all' แต่ไม่มีวันที่ -> ค้นหาทั้งหมด

        tbl = _ol_table_name()
        
        # ========================================================
        # [FIX] ดึงข้อมูลเฉพาะเมื่อ: มีคำค้นหา หรือ มีการเลือกวันที่
        # ========================================================
        if q:
            # กรณี 1: มีคำค้นหา -> ค้นหาทั้งหมด (Global Search)
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_notenough > 0")
            result = db.session.execute(sql).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        elif print_date_from or print_date_to:
            # กรณี 2: มีการเลือกวันที่ -> กรองตามวันที่
            sql_where = "printed_notenough > 0"
            params = {}
            if print_date_from:
                sql_where += " AND DATE(printed_notenough_at) >= :pf"
                params["pf"] = print_date_from
            if print_date_to:
                sql_where += " AND DATE(printed_notenough_at) <= :pt"
                params["pt"] = print_date_to
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE {sql_where}")
            result = db.session.execute(sql, params).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        else:
            # กรณี 3: ไม่ค้นหา และ ไม่เลือกวัน (เช่น กด reset='all') -> ไม่แสดงอะไร
            printed_oids = []

        def _available_dates():
            sql = text(f"SELECT DISTINCT DATE(printed_notenough_at) as d FROM {tbl} WHERE printed_notenough > 0 AND printed_notenough_at IS NOT NULL ORDER BY d DESC")
            return [r[0] for r in db.session.execute(sql).fetchall()]

        shops = Shop.query.order_by(Shop.name.asc()).all()
        
        if not printed_oids:
            return render_template(
                "report_notenough.html",
                rows=[],
                summary={"sku_count": 0, "orders_count": 0},
                printed_at=None,
                order_ids=[],
                shops=shops,
                logistics=[],
                platform_sel=platform,
                shop_sel=shop_id,
                logistic_sel=logistic,
                is_history_view=True,
                available_dates=_available_dates(),
                print_date_from=print_date_from,
                print_date_to=print_date_to,
                sort_col=sort_col,
                sort_dir=sort_dir,
                q=q,
                round_sel=round_num,
                date_from=date_from_str,
                date_to=date_to_str,
                import_from=import_from_str,
                import_to=import_to_str
            )

        # เตรียมตัวกรองวันที่สั่งซื้อ
        date_from_dt = None
        date_to_dt = None
        if date_from_str:
            try:
                date_from_dt = datetime.combine(parse_date_any(date_from_str), datetime.min.time(), tzinfo=TH_TZ)
            except: pass
        if date_to_str:
            try:
                date_to_dt = datetime.combine(parse_date_any(date_to_str) + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ)
            except: pass

        # ดึงข้อมูลจริง
        filters = {
            "platform": platform or None,
            "shop_id": int(shop_id) if shop_id else None,
            "import_date": None,
            "date_from": date_from_dt,
            "date_to": date_to_dt
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        # [FIX] ในหน้าประวัติ (printed) ไม่กรอง Issued ออก เพราะเราเพิ่ง mark Issued ไป
        rows = [r for r in rows if (r.get("order_id") or "").strip() in printed_oids]
        
        # กรองวันที่นำเข้า (Import Date) - [แก้ไข] ถ้าไม่มีวันที่ก็เอามาด้วย
        if import_from_str or import_to_str:
            from datetime import date as date_type
            imp_from = parse_date_any(import_from_str) if import_from_str else None
            imp_to = parse_date_any(import_to_str) if import_to_str else None
            filtered_rows = []
            for r in rows:
                raw_d = r.get("import_date")
                d_obj = None
                if isinstance(raw_d, str):
                    try: d_obj = datetime.strptime(raw_d, "%Y-%m-%d").date()
                    except: pass
                elif isinstance(raw_d, datetime):
                    d_obj = raw_d.date()
                elif isinstance(raw_d, date_type):
                    d_obj = raw_d
                
                if d_obj:
                    if imp_from and d_obj < imp_from: continue
                    if imp_to and d_obj > imp_to: continue
                    filtered_rows.append(r)
                else:
                    # ข้อมูลไม่มีวันที่นำเข้า -> เอามาด้วย
                    filtered_rows.append(r)
            rows = filtered_rows

        packed_oids = _orders_packed_set(rows)
        
        safe = []
        for r in rows:
            r = dict(r)
            oid = (r.get("order_id") or "").strip()
            if oid not in printed_oids:
                continue
            if oid in packed_oids:
                continue
            if (str(r.get("sales_status") or "")).upper() == "PACKED":
                continue
            if bool(r.get("packed", False)):
                continue
            
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except:
                            stock_qty = 0
                r["stock_qty"] = stock_qty
            
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)

        # [CRITICAL FIX] Fallback Logic สำหรับ Not Enough
        # เพราะออเดอร์ถูก mark เป็น ISSUED แล้ว allocation_status อาจไม่ใช่ NOT_ENOUGH
        def _is_not_enough_for_history(r):
            # 1. ถ้า status เป็น NOT_ENOUGH อยู่แล้ว -> เอา
            if r.get("allocation_status") == "NOT_ENOUGH": return True
            # 2. Fallback: ถ้า stock < qty (เกณฑ์ Not Enough) -> เอา
            try:
                s = int(r.get("stock_qty") or 0)
                q = int(r.get("qty") or 0)
                return s < q and s > 0  # stock มีแต่ไม่พอ
            except: return False

        lines = [r for r in safe if _is_not_enough_for_history(r)]

        # เตรียมข้อมูล Mixed Status สำหรับหน้าประวัติ
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "LOW_STOCK": "สินค้าน้อย",
            "SHORTAGE": "ไม่มีของ",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "NOT_ENOUGH":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # กรองตามคำค้นหา (q)
        if q:
            q_lower = q.lower()
            lines = [
                r for r in lines
                if q_lower in (
                    str(r.get("order_id") or "") +
                    str(r.get("sku") or "") +
                    str(r.get("brand") or "") +
                    str(r.get("model") or "") +
                    str(r.get("shop") or "") +
                    str(r.get("platform") or "") +
                    str(r.get("logistic") or "")
                ).lower()
            ]

        # ดึง Round
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        round_by_oid = {}
        if order_ids_for_round:
            sql = text(f"SELECT order_id, MAX(notenough_round) AS r FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except:
                pass

        if round_num not in (None, "", "all"):
            try:
                r_int = int(round_num)
                lines = [r for r in lines if round_by_oid.get((r.get("order_id") or "").strip()) == r_int]
            except:
                pass

        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform": r.get("platform"),
                "store": r.get("shop"),
                "order_no": oid,
                "sku": r.get("sku"),
                "brand": r.get("brand"),
                "product_name": r.get("model"),
                "stock": int(r.get("stock_qty", 0) or 0),
                "qty": int(r.get("qty", 0) or 0),
                "order_time": r.get("order_time"),
                "due_date": r.get("due_date"),
                "sla": r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round": round_by_oid.get(oid),
                "printed_count": 0,
                "printed_at": None,
                "note": mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        sort_col = sort_col if sort_col else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        oids = sorted({(r["order_no"] or "").strip() for r in out if r["order_no"]})
        counts = _get_print_counts_local(oids, "notenough")
        for r in out:
            r["printed_count"] = int(counts.get(r["order_no"], 0))

        # [เพิ่ม] ดึงเวลาพิมพ์จาก DB (ใช้ printed_notenough_at ที่ถูกต้อง)
        ts_map = {}
        if oids:
            tbl_ts = _ol_table_name()
            sql_ts = text(f"""
                SELECT order_id, MAX(printed_notenough_at) AS ts 
                FROM {tbl_ts}
                WHERE printed_notenough > 0
                  AND order_id IN :oids
                GROUP BY order_id
            """).bindparams(bindparam("oids", expanding=True))
            try:
                rows_ts = db.session.execute(sql_ts, {"oids": oids}).all()
                for row_ts in rows_ts:
                    if not row_ts or not row_ts[0] or not row_ts[1]:
                        continue
                    oid_str = str(row_ts[0]).strip()
                    ts_str = row_ts[1]
                    try:
                        dt = datetime.fromisoformat(ts_str)
                        if dt.tzinfo is None:
                            dt = TH_TZ.localize(dt)
                        ts_map[oid_str] = dt
                    except:
                        pass
            except:
                pass

        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_at"] = ts_map.get(oid)

        final_oids = sorted({(r["order_no"] or "").strip() for r in out if r["order_no"]})
        skus = {(r["sku"] or "").strip() for r in out if r["sku"]}
        summary = {
            "sku_count": len(skus),
            "orders_count": len(final_oids),
        }
        
        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))
        available_rounds = sorted({r["assign_round"] for r in out if r["assign_round"] is not None})

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if final_oids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": final_oids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_notenough.html",
            rows=out,
            summary=summary,
            printed_at=None,
            order_ids=final_oids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            is_history_view=True,
            available_dates=_available_dates(),
            print_date_from=print_date_from,
            print_date_to=print_date_to,
            sort_col=sort_col,
            sort_dir=sort_dir,
            q=q,
            round_sel=round_num,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str
        )

    @app.route("/report/notenough/update_round", methods=["POST"])
    @login_required
    def update_notenough_round():
        """อัปเดตรอบสำหรับรายงานสินค้าไม่พอส่ง"""
        data = request.get_json() or {}
        order_ids = data.get("order_ids", [])
        round_num = data.get("round")
        
        if not order_ids or round_num is None:
            return jsonify({"success": False, "message": "ข้อมูลไม่ครบ"})
        
        try:
            round_int = int(round_num)
        except:
            return jsonify({"success": False, "message": "รอบต้องเป็นตัวเลข"})
        
        tbl = _ol_table_name()
        sql = text(f"UPDATE {tbl} SET notenough_round = :r WHERE order_id IN :oids")
        sql = sql.bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"r": round_int, "oids": order_ids})
        db.session.commit()
        
        return jsonify({"success": True, "message": f"อัปเดตรอบเป็น {round_int} สำเร็จ ({len(order_ids)} ออเดอร์)"})

    @app.route("/report/notenough.xlsx", methods=["GET"])
    @login_required
    def report_notenough_export():
        """Export Excel รายงานสินค้าไม่พอส่ง"""
        # ไม่ต้องใช้ services.lowstock แล้ว
        import pandas as pd
        
        platform = normalize_platform(request.args.get("platform"))
        shop_id = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q = (request.args.get("q") or "").strip()
        round_num = request.args.get("round")
        
        # รับค่าวันที่กรอง (เพิ่มใหม่)
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        filters = {"platform": platform or None, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)
        
        packed_oids = _orders_packed_set(rows)
        
        safe = []
        for r in rows:
            r = dict(r)
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            if (str(r.get("sales_status") or "")).upper() == "PACKED":
                continue
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except:
                            stock_qty = 0
                r["stock_qty"] = stock_qty
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)
        
        # กรองเฉพาะ allocation_status == "NOT_ENOUGH"
        lines = [r for r in safe if r.get("allocation_status") == "NOT_ENOUGH"]
        
        # ---------- กรอง logistic ----------
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").strip().upper() == logistic.strip().upper()]
        
        # ---------- กรอง round ----------
        if round_num:
            try:
                round_int = int(round_num)
                lines = [r for r in lines if r.get("notenough_round") == round_int]
            except:
                pass
        
        # ---------- กรองด้วย date filters (เพิ่มใหม่) ----------
        def _parse_date(d_str):
            if not d_str:
                return None
            from utils import parse_datetime_guess
            dt = parse_datetime_guess(d_str)
            if dt:
                return dt.date() if hasattr(dt, 'date') else dt
            return None
        
        date_from = _parse_date(date_from_str)
        date_to = _parse_date(date_to_str)
        import_from = _parse_date(import_from_str)
        import_to = _parse_date(import_to_str)
        
        # กรองด้วยวันที่สั่ง (order_time / due_date)
        if date_from or date_to:
            filtered = []
            for r in lines:
                order_time = r.get("order_time") or r.get("due_date")
                if not order_time:
                    continue
                try:
                    from utils import parse_datetime_guess
                    dt = parse_datetime_guess(order_time)
                    if dt:
                        dt_date = dt.date() if hasattr(dt, 'date') else dt
                        if date_from and dt_date < date_from:
                            continue
                        if date_to and dt_date > date_to:
                            continue
                        filtered.append(r)
                except:
                    continue
            lines = filtered
        
        # กรองด้วยวันที่นำเข้า (import_date)
        if import_from or import_to:
            filtered = []
            for r in lines:
                imp_date = r.get("import_date")
                if not imp_date:
                    continue
                try:
                    from utils import parse_datetime_guess
                    dt = parse_datetime_guess(imp_date)
                    if dt:
                        dt_date = dt.date() if hasattr(dt, 'date') else dt
                        if import_from and dt_date < import_from:
                            continue
                        if import_to and dt_date > import_to:
                            continue
                        filtered.append(r)
                except:
                    continue
            lines = filtered
        
        # ---------- กรองด้วยคำค้นหา q ----------
        if q:
            q_lower = q.lower()
            lines = [r for r in lines if q_lower in (r.get("sku") or "").lower() 
                     or q_lower in (r.get("model") or "").lower() 
                     or q_lower in (r.get("order_id") or "").lower()]
        
        df = pd.DataFrame([{
            "แพลตฟอร์ม": r.get("platform"),
            "ร้าน": r.get("shop"),
            "เลข Order": r.get("order_id"),
            "SKU": r.get("sku"),
            "Brand": r.get("brand"),
            "ชื่อสินค้า": r.get("model"),
            "Stock": int(r.get("stock_qty", 0) or 0),
            "Qty": int(r.get("qty", 0) or 0),
            "เวลาที่ลูกค้าสั่ง": r.get("order_time"),
            "กำหนดส่ง": r.get("due_date"),
            "ประเภทขนส่ง": r.get("logistic"),
        } for r in lines])
        
        out = BytesIO()
        with pd.ExcelWriter(out, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="NotEnough")
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="report_notenough.xlsx")
    # ================== /NEW: Report Not Enough ==================

    # -----------------------
    # Picking (รวมยอดหยิบ)
    # -----------------------
    def _aggregate_picking(rows: list[dict], group_by_round: bool = False) -> list[dict]:
        """
        รวมยอดหยิบตาม SKU
        - group_by_round=False: รวมทุกรอบเข้าด้วยกัน (default เดิม)
        - group_by_round=True: แยก key ตาม SKU+dispatch_round (ไม่รวมข้ามรอบ)
        """
        rows = rows or []
        agg: dict[str, dict] = {}
        for r in rows:
            if not bool(r.get("accepted")):
                continue
            # [แก้ไข] เพิ่ม "ISSUED" เพื่อให้หน้าประวัติ (ที่จ่ายงานแล้ว) แสดงข้อมูลได้
            if (r.get("allocation_status") or "") not in ("ACCEPTED", "READY_ACCEPT", "ISSUED"):
                continue
            sku = str(r.get("sku") or "").strip()
            if not sku:
                continue
            brand = str(r.get("brand") or "").strip()
            model = str(r.get("model") or "").strip()
            qty = int(r.get("qty", 0) or 0)
            stock_qty = int(r.get("stock_qty", 0) or 0)
            dispatch_round = r.get("dispatch_round")
            
            # [NEW] สร้าง key ที่รวม dispatch_round ด้วย (ถ้า group_by_round=True)
            if group_by_round and dispatch_round is not None:
                agg_key = f"{sku}__round_{dispatch_round}"
            else:
                agg_key = sku
            
            a = agg.setdefault(agg_key, {
                "sku": sku, 
                "brand": brand, 
                "model": model, 
                "need_qty": 0, 
                "stock_qty": 0,
                "dispatch_rounds": set(),
                "dispatch_round_single": dispatch_round  # เก็บค่าเดี่ยวไว้
            })
            a["need_qty"] += qty
            if stock_qty > a["stock_qty"]:
                a["stock_qty"] = stock_qty
            if dispatch_round is not None:
                a["dispatch_rounds"].add(dispatch_round)

        items = []
        for _, a in agg.items():
            need = int(a["need_qty"])
            stock = int(a["stock_qty"])
            shortage = max(0, need - stock)
            remain = stock - need
            
            # Handle dispatch_round display
            dispatch_rounds = sorted(a["dispatch_rounds"])
            if len(dispatch_rounds) == 0:
                dispatch_round_display = None
            elif len(dispatch_rounds) == 1:
                dispatch_round_display = dispatch_rounds[0]
            else:
                # ถ้า group_by_round=True ไม่ควรมีกรณีนี้ แต่ fallback ไว้
                dispatch_round_display = f"{dispatch_rounds[0]}-{dispatch_rounds[-1]}"
            
            items.append({
                "sku": a["sku"], 
                "brand": a["brand"], 
                "model": a["model"],
                "need_qty": need, 
                "stock_qty": stock, 
                "shortage": shortage, 
                "remain_after_pick": remain,
                "dispatch_round": dispatch_round_display,
            })
        items.sort(key=lambda x: (x["brand"].lower(), x["model"].lower(), x["sku"].lower()))
        return items

    @app.route("/report/picking", methods=["GET"])
    @login_required
    def picking_list():
        # Check for reset mode
        reset_mode = request.args.get("reset")
        
        if reset_mode == 'all':
            # Clear all filters and show all pending orders
            platform = None
            shop_id = None
            logistic = None
            acc_from = None
            acc_to = None
            acc_from_str = ""
            acc_to_str = ""
            round_sel = None
            print_count_sel = None
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            
            # รับค่าวันที่กดพร้อมรับ (accepted_at)
            acc_from_str = request.args.get("accepted_from")
            acc_to_str = request.args.get("accepted_to")
            acc_from = parse_date_any(acc_from_str)
            acc_to = parse_date_any(acc_to_str)
            
            # [NEW] รับค่ารอบจ่ายงาน และจำนวนครั้งที่พิมพ์
            round_sel = request.args.get("round")
            print_count_sel = request.args.get("print_count")

        # [แก้ไข] ไม่กรอง accepted date ใน compute_allocation
        # เราจะกรองด้วย printed_warehouse_at เองทีหลัง
        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": None,  # ไม่กรองตรงนี้
            "accepted_to": None,    # ไม่กรองตรงนี้
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)

        # ========================================================
        # [FIXED LOGIC] กรอง Order ที่ "พิมพ์คลังแล้ว" แต่ "ยังไม่พิมพ์หยิบ"
        # + กรองด้วยวันที่พิมพ์ใบงานคลัง (printed_warehouse_at)
        # ========================================================
        
        # 1. รวบรวม Order ID ทั้งหมดในหน้านี้
        all_oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        
        valid_rows = []
        
        if all_oids:
            tbl = _ol_table_name()
            # 2. Query เช็คสถานะการพิมพ์จาก DB โดยตรง (แม่นยำกว่า)
            # ดึงจำนวนครั้งที่พิมพ์ Warehouse และ Picking + เวลาที่พิมพ์ Warehouse
            sql = text(f"""
                SELECT order_id, 
                       MAX(COALESCE(printed_warehouse, 0)) as wh_count, 
                       MAX(COALESCE(printed_picking, 0)) as pk_count,
                       MAX(printed_warehouse_at) as wh_at
                FROM {tbl} 
                WHERE order_id IN :oids 
                GROUP BY order_id
            """)
            sql = sql.bindparams(bindparam("oids", expanding=True))
            
            print_status = db.session.execute(sql, {"oids": all_oids}).fetchall()
            
            # สร้าง Map {order_id: (wh_count, pk_count, wh_at_str)}
            status_map = {}
            for row in print_status:
                status_map[row[0]] = (int(row[1] or 0), int(row[2] or 0), row[3])
            
            # แปลงวันที่กรองเป็น datetime เพื่อเปรียบเทียบ
            f_start = datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None
            f_end = datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None
            
            for r in rows:
                oid = (r.get("order_id") or "").strip()
                wh, pk, wh_at_str = status_map.get(oid, (0, 0, None))
                
                # เงื่อนไขสำคัญ: 
                # 1. ต้องพิมพ์คลังแล้ว (wh > 0)
                # 2. ต้องยังไม่พิมพ์หยิบ (pk == 0)
                if wh > 0 and pk == 0:
                    # เงื่อนไข 3: กรองวันที่พิมพ์ใบงานคลัง (ถ้ามีการกรอง)
                    pass_date = True
                    if f_start or f_end:
                        if not wh_at_str:
                            pass_date = False  # ไม่มีวันที่พิมพ์ = ไม่ผ่าน
                        else:
                            try:
                                dt_print = datetime.fromisoformat(wh_at_str)
                                if dt_print.tzinfo is None:
                                    dt_print = TH_TZ.localize(dt_print)
                                if f_start and dt_print < f_start:
                                    pass_date = False
                                if f_end and dt_print >= f_end:
                                    pass_date = False
                            except Exception:
                                pass_date = False
                    
                    if pass_date:
                        valid_rows.append(r)
            
        rows = valid_rows

        # เตรียมข้อมูลปลอดภัย + ใส่ stock_qty ให้ครบ
        safe_rows = []
        for r in rows:
            r = dict(r)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["accepted"] = bool(r.get("accepted", False))
            r["sales_status"] = r.get("sales_status", None)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            safe_rows.append(r)

        if logistic:
            safe_rows = [r for r in safe_rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        
        # [NEW] กรองตามรอบจ่ายงาน (dispatch_round)
        if round_sel:
            try:
                round_int = int(round_sel)
                safe_rows = [r for r in safe_rows if r.get("dispatch_round") == round_int]
            except (ValueError, TypeError):
                pass
        
        # [NEW] กรองตามจำนวนครั้งที่พิมพ์ - ต้องดึงข้อมูล print count ของแต่ละ order ก่อน
        if print_count_sel:
            try:
                target_pc = int(print_count_sel)
                # ดึง order_ids ที่มี print count ตามเงื่อนไข
                temp_oids = sorted({(r.get("order_id") or "").strip() for r in safe_rows if r.get("order_id")})
                if temp_oids:
                    pc_map = _get_print_counts_local(temp_oids, "picking")
                    # กรองเฉพาะ order ที่มี print count ตรงกับที่ระบุ
                    valid_oids = {oid for oid, cnt in pc_map.items() if cnt == target_pc}
                    safe_rows = [r for r in safe_rows if (r.get("order_id") or "").strip() in valid_oids]
            except (ValueError, TypeError):
                pass

        # รวมต่อ SKU
        items = _aggregate_picking(safe_rows)

        # ===== นับจำนวนครั้งที่พิมพ์ Picking (รวมทั้งชุดงาน) — ใช้ MAX ไม่ใช่ SUM =====
        valid_rows = [r for r in safe_rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]
        order_ids = sorted({(r.get("order_id") or "").strip() for r in valid_rows if r.get("order_id")})
        print_counts_pick = _get_print_counts_local(order_ids, "picking")
        print_count_overall = max(print_counts_pick.values()) if print_counts_pick else 0
        
        # Get the latest print timestamp and user
        print_timestamp_overall = None
        print_user_overall = None
        if order_ids:
            tbl = _ol_table_name()
            sql = text(f"SELECT printed_picking_at, printed_picking_by FROM {tbl} WHERE order_id IN :oids AND printed_picking_at IS NOT NULL ORDER BY printed_picking_at DESC LIMIT 1")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            result = db.session.execute(sql, {"oids": order_ids}).first()
            if result:
                try:
                    dt = datetime.fromisoformat(result[0])
                    if dt.tzinfo is None:
                        dt = TH_TZ.localize(dt)
                    print_timestamp_overall = dt
                    print_user_overall = result[1]  # username
                except Exception:
                    pass

        # ชื่อร้านสำหรับแสดงในคอลัมน์ใหม่
        shop_sel_name = None
        if shop_id:
            s = Shop.query.get(int(shop_id))
            if s:
                shop_sel_name = f"{s.platform} • {s.name}"

        # เติมแพลตฟอร์ม/ร้าน/ประเภทขนส่งให้แต่ละ item เพื่อไม่ให้ขึ้น '-'
        for it in items:
            it["platform"] = platform or "-"
            it["shop"] = shop_sel_name or "-"
            it["logistic"] = logistic or "-"
        
        # ===== ดึงข้อมูลการเชื่อมโยงกับใบงานคลัง (Warehouse) =====
        # เพื่อแสดงว่า Picking ใบนี้ผูกกับใบงานคลังที่พิมพ์เมื่อไหร่
        warehouse_print_info = None
        if order_ids:
            tbl = _ol_table_name()
            sql = text(f"""
                SELECT printed_warehouse_at, printed_warehouse_by, printed_warehouse
                FROM {tbl} 
                WHERE order_id IN :oids 
                AND printed_warehouse > 0 
                ORDER BY printed_warehouse_at DESC 
                LIMIT 1
            """)
            sql = sql.bindparams(bindparam("oids", expanding=True))
            result = db.session.execute(sql, {"oids": order_ids}).first()
            if result and result[0]:
                try:
                    dt = datetime.fromisoformat(result[0])
                    if dt.tzinfo is None:
                        dt = TH_TZ.localize(dt)
                    warehouse_print_info = {
                        "printed_at": dt,
                        "printed_by": result[1],
                        "print_count": result[2]
                    }
                except Exception:
                    pass

        totals = {
            "total_skus": len(items),
            "total_need_qty": sum(i["need_qty"] for i in items),
            "total_shortage": sum(i["shortage"] for i in items),
        }
        shops = Shop.query.order_by(Shop.name.asc()).all()
        logistics = sorted(set(r.get("logistic") for r in safe_rows if r.get("logistic")))

        return render_template(
            "picking.html",
            items=items,
            totals=totals,
            shops=shops,
            logistics=logistics,
            platform_sel=platform if reset_mode != 'all' else None,
            shop_sel=shop_id if reset_mode != 'all' else None,
            shop_sel_name=shop_sel_name if reset_mode != 'all' else None,
            logistic_sel=logistic if reset_mode != 'all' else None,
            official_print=False,
            printed_meta=None,
            print_count_overall=print_count_overall,
            print_timestamp_overall=print_timestamp_overall,
            print_user_overall=print_user_overall,
            order_ids=order_ids,  # Pass order IDs for dispatch round update
            accepted_from=acc_from_str if reset_mode != 'all' else "",
            accepted_to=acc_to_str if reset_mode != 'all' else "",
            is_history_view=False,
            warehouse_print_info=warehouse_print_info,  # เชื่อมโยงกับใบงานคลัง
            round_sel=round_sel if reset_mode != 'all' else None,  # [NEW] ส่งค่ารอบจ่ายงาน
            print_count_sel=print_count_sel if reset_mode != 'all' else None,  # [NEW] ส่งค่าจำนวนครั้งที่พิมพ์
            available_rounds=[],  # [NEW] สำหรับหน้าปัจจุบันไม่มี dropdown รอบ
        )

    @app.route("/report/picking/print", methods=["POST"])
    @login_required
    def picking_list_commit():
        cu = current_user()
        platform = normalize_platform(request.form.get("platform"))
        shop_id = request.form.get("shop_id")
        logistic = request.form.get("logistic")
        override = request.form.get("override") in ("1", "true", "yes")

        # Idempotency token: กัน request ซ้ำ (double-submit / retry)
        print_token = (request.form.get("print_token") or "").strip()
        
        # Get selected order IDs from form (comma-separated)
        # ถ้าเป็น '', 'all', 'ALL' ให้ถือว่า "ไม่ระบุ"
        order_ids_raw = (request.form.get("order_ids") or "").strip()
        selected_order_ids = [] if order_ids_raw.lower() in ("", "all") else \
            [oid.strip() for oid in order_ids_raw.split(",") if oid.strip()]

        filters = {"platform": platform if platform else None, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)

        safe_rows = []
        for r in rows:
            r = dict(r)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["accepted"] = bool(r.get("accepted", False))
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            safe_rows.append(r)

        if logistic:
            safe_rows = [r for r in safe_rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        valid_rows = [r for r in safe_rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]
        
        # If specific order IDs were selected, filter to only those
        if selected_order_ids:
            valid_rows = [r for r in valid_rows if (r.get("order_id") or "").strip() in selected_order_ids]
            oids = sorted(selected_order_ids)
        else:
            oids = sorted({(r.get("order_id") or "").strip() for r in valid_rows if r.get("order_id")})

        if not oids:
            flash("ไม่พบออเดอร์สำหรับพิมพ์ Picking", "warning")
            return redirect(url_for("picking_list", platform=platform, shop_id=shop_id, logistic=logistic))

        already = _detect_already_printed(oids, kind="picking")
        if already and not (override and cu and cu.role == "admin"):
            head = ", ".join(list(already)[:10])
            more = "" if len(already) <= 10 else f" ... (+{len(already)-10})"
            flash(f"มีบางออเดอร์เคยพิมพ์ Picking ไปแล้ว: {head}{more}", "danger")
            flash("ถ้าจำเป็นต้องพิมพ์ซ้ำ โปรดให้แอดมินติ๊ก 'อนุญาตพิมพ์ซ้ำ' แล้วพิมพ์อีกครั้ง", "warning")
            return redirect(url_for("picking_list", platform=platform, shop_id=shop_id, logistic=logistic))

        deduped = False
        if print_token:
            try:
                ins = text(
                    """
                    INSERT OR IGNORE INTO action_dedupe(token, kind, created_at, user_id)
                    VALUES (:t, :k, :at, :uid)
                    """
                )
                res = db.session.execute(
                    ins,
                    {
                        "t": print_token,
                        "k": "picking_print",
                        "at": now_thai().isoformat(),
                        "uid": (cu.id if cu else None),
                    },
                )
                if int(getattr(res, "rowcount", 0) or 0) == 0:
                    deduped = True
                    try:
                        db.session.rollback()
                    except Exception:
                        pass
            except Exception as e:
                # ถ้า insert token พัง ให้เดินต่อแบบเดิม (ไม่บล็อกการพิมพ์)
                app.logger.warning(f"[action_dedupe] insert failed: {e}")
                try:
                    db.session.rollback()
                except Exception:
                    pass

        if not deduped:
            now_dt = now_thai()
            now_iso = now_dt.isoformat()
            _mark_printed(oids, kind="picking", user_id=(cu.id if cu else None), when_iso=now_iso, commit=False)
            # >>> NEW: ย้ายไป Orderจ่ายแล้ว (บันทึกเวลาตอนพิมพ์)
            _mark_issued(oids, user_id=(cu.id if cu else None), source="print:picking", when_dt=now_dt, commit=False)
            db.session.commit()
        else:
            flash("ตรวจพบการส่งซ้ำ ระบบจึงไม่บวกจำนวนครั้งพิมพ์เพิ่ม", "warning")

        db.session.expire_all()  # Force refresh to get updated print counts

        items = _aggregate_picking(safe_rows)
        for it in items:
            it["platform"] = platform or "-"
            if shop_id:
                s = Shop.query.get(int(shop_id))
                it["shop"] = (f"{s.platform} • {s.name}") if s else "-"
            else:
                it["shop"] = "-"
            it["logistic"] = logistic or "-"

        totals = {
            "total_skus": len(items),
            "total_need_qty": sum(i["need_qty"] for i in items),
            "total_shortage": sum(i["shortage"] for i in items),
        }
        shops = Shop.query.order_by(Shop.name.asc()).all()
        logistics = sorted(set(r.get("logistic") for r in safe_rows if r.get("logistic")))
        printed_meta = {"by": (cu.username if cu else "-"), "at": now_thai(), "orders": len(oids), "override": bool(already)}

        print_counts_pick = _get_print_counts_local(oids, "picking")
        print_count_overall = max(print_counts_pick.values()) if print_counts_pick else 0
        
        # Use current timestamp and user
        print_timestamp_overall = now_thai()
        print_user_overall = cu.username if cu else None

        shop_sel_name = None
        if shop_id:
            s = Shop.query.get(int(shop_id))
            if s:
                shop_sel_name = f"{s.platform} • {s.name}"

        return render_template(
            "picking.html",
            items=items,
            totals=totals,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            shop_sel_name=shop_sel_name,
            logistic_sel=logistic,
            official_print=True,
            printed_meta=printed_meta,
            print_count_overall=print_count_overall,
            print_timestamp_overall=print_timestamp_overall,
            print_user_overall=print_user_overall,
            order_ids=oids,  # Pass order IDs for dispatch round update
        )

    # ================== NEW: Update Dispatch Round from Picking ==================
    @app.route("/picking/update_dispatch", methods=["POST"])
    @login_required
    def picking_update_dispatch():
        """อัปเดตเลขรอบจ่ายงานจากหน้า Picking List และทำเครื่องหมายจ่ายงานแล้ว"""
        cu = current_user()
        if not cu:
            flash("กรุณาเข้าสู่ระบบก่อน", "warning")
            return redirect(url_for("login"))
        
        try:
            # รับค่าจาก Form
            order_ids = request.form.getlist("order_ids[]")
            dispatch_round = request.form.get("dispatch_round", type=int) or 1
            
            if not order_ids:
                flash("ไม่พบรายการที่เลือก", "warning")
                return redirect(request.referrer or url_for("picking_list"))
            
            # 1. บันทึกสถานะ Issued (จ่ายงานแล้ว)
            source = f"picking:round_{dispatch_round}"
            _mark_issued(order_ids, cu.id, source=source)
            
            # 2. อัปเดตเลขรอบ (dispatch_round) ลง DB
            tbl = _ol_table_name()
            sql = text(f"UPDATE {tbl} SET dispatch_round = :r WHERE order_id IN :oids")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            db.session.execute(sql, {"r": dispatch_round, "oids": order_ids})
            db.session.commit()
            
            flash(f"✅ จ่ายงาน {len(order_ids)} ออเดอร์ เป็นรอบที่ {dispatch_round} เรียบร้อยแล้ว", "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"เกิดข้อผิดพลาด: {e}", "danger")
        
        # Redirect กลับหน้าเดิมพร้อมฟิลเตอร์
        platform = request.form.get("platform") or ""
        shop_id = request.form.get("shop_id") or ""
        logistic = request.form.get("logistic") or ""
        return redirect(url_for("picking_list", platform=platform, shop_id=shop_id, logistic=logistic))
    # ================== /NEW ==================

    # ================== NEW: View Printed Picking Lists ==================
    @app.route("/report/picking/printed", methods=["GET"])
    @login_required
    def picking_printed_history():
        """ดู Picking List ที่พิมพ์แล้ว - สามารถเลือกวันที่และพิมพ์ซ้ำได้"""
        # Check for reset mode
        reset_mode = request.args.get("reset")
        target_date = None
        
        # ตรวจสอบว่ามีการส่งพารามิเตอร์มาบ้างไหม (เพื่อดูว่าเป็น First Load หรือไม่)
        has_params = any([
            request.args.get("platform"),
            request.args.get("shop_id"),
            request.args.get("logistic"),
            request.args.get("print_date"),
            request.args.get("accepted_from"),
            request.args.get("accepted_to"),
            request.args.get("round"),
            request.args.get("print_count"),  # [NEW]
            request.args.get("reset")
        ])
        
        if reset_mode == 'today' or not has_params:
            # Reset หรือ เข้ามาครั้งแรก (ไม่ส่ง param) -> เอาของ "วันนี้"
            target_date = now_thai().date()
            platform = None
            shop_id = None
            logistic = None
            print_date = None
            raw_from = None
            raw_to = None
            round_sel = None  # [NEW]
            print_count_sel = None  # [NEW]
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            print_date = request.args.get("print_date")  # วันที่พิมพ์ (YYYY-MM-DD)
            
            # รับค่าวันที่กดพร้อมรับ - ไม่ตั้งค่า default
            raw_from = request.args.get("accepted_from")
            raw_to = request.args.get("accepted_to")
            
            # [NEW] รับค่ารอบจ่ายงาน และจำนวนครั้งที่พิมพ์
            round_sel = request.args.get("round")
            print_count_sel = request.args.get("print_count")
            
            if print_date:
                try:
                    target_date = datetime.strptime(print_date, "%Y-%m-%d").date()
                except:
                    target_date = None
        
        # ไม่ตั้งค่า default - ให้เป็นค่าว่าง (mm/dd/yyyy)
        acc_from = parse_date_any(raw_from)
        acc_to = parse_date_any(raw_to)
        
        # Get all orders that have been printed for picking
        tbl = _ol_table_name()
        
        # Build query to get orders with print history
        if target_date:
            # Filter by specific print date (or today if reset)
            # หมายเหตุ: printed_picking_at ถูกบันทึกเป็นเวลาไทยอยู่แล้ว (ไม่ต้อง +7)
            sql = text(f"""
                SELECT DISTINCT order_id 
                FROM {tbl} 
                WHERE printed_picking > 0 
                AND DATE(printed_picking_at) = :target_date
            """)
            result = db.session.execute(sql, {"target_date": target_date.isoformat()}).fetchall()
        else:
            # Get all printed orders
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_picking > 0")
            result = db.session.execute(sql).fetchall()
        
        printed_order_ids = [row[0] for row in result if row[0]]
        
        if not printed_order_ids:
            # No printed orders found
            shops = Shop.query.order_by(Shop.name.asc()).all()
            return render_template(
                "picking.html",
                items=[],
                totals={"total_skus": 0, "total_need_qty": 0, "total_shortage": 0},
                shops=shops,
                logistics=[],
                platform_sel=platform,
                shop_sel=shop_id,
                shop_sel_name=None,
                logistic_sel=logistic,
                official_print=False,
                printed_meta=None,
                print_count_overall=0,
                print_timestamp_overall=None,
                order_ids=[],
                is_history_view=True,
                print_date_sel=print_date,
                available_dates=[],
                available_rounds=[],
                round_sel=round_sel,
                print_count_sel=print_count_sel,  # [NEW]
                accepted_from=raw_from,
                accepted_to=raw_to,
            )
        
        # Get full data for these orders
        # [แก้ไข] ไม่กรอง accepted date ใน compute_allocation - เราจะกรองด้วย printed_warehouse_at เอง
        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": None,  # ไม่กรองตรงนี้
            "accepted_to": None,    # ไม่กรองตรงนี้
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        
        # [แก้ไข] ดึงเวลาพิมพ์ Warehouse มาเพื่อกรองด้วย printed_warehouse_at
        all_oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        wh_print_map = {}
        dispatch_round_map = {}  # [NEW] เก็บ dispatch_round ของแต่ละ order+sku
        if all_oids:
            sql = text(f"SELECT order_id, MAX(printed_warehouse_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            res = db.session.execute(sql, {"oids": all_oids}).fetchall()
            for row in res:
                wh_print_map[row[0]] = row[1]
            
            # [NEW] ดึง dispatch_round แยกตาม order_id + sku (ระดับบรรทัด)
            sql_dr = text(f"SELECT order_id, sku, dispatch_round FROM {tbl} WHERE order_id IN :oids AND dispatch_round IS NOT NULL")
            sql_dr = sql_dr.bindparams(bindparam("oids", expanding=True))
            res_dr = db.session.execute(sql_dr, {"oids": all_oids}).fetchall()
            for row_dr in res_dr:
                key = (row_dr[0], row_dr[1])  # (order_id, sku)
                dispatch_round_map[key] = row_dr[2]
        
        # แปลงวันที่กรองเป็น datetime เพื่อเปรียบเทียบ
        f_start = datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None
        f_end = datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None
        
        # Filter to only printed orders + กรองด้วยวันที่พิมพ์ Warehouse
        safe_rows = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if oid not in printed_order_ids:
                continue
            
            # [แก้ไข] กรองด้วยวันที่พิมพ์ Warehouse (ถ้ามีการกรอง)
            if f_start or f_end:
                wh_at_str = wh_print_map.get(oid)
                if not wh_at_str:
                    continue  # ไม่มีวันที่พิมพ์ Warehouse -> ข้าม
                try:
                    dt_print = datetime.fromisoformat(wh_at_str)
                    if dt_print.tzinfo is None:
                        dt_print = TH_TZ.localize(dt_print)
                    if f_start and dt_print < f_start:
                        continue
                    if f_end and dt_print >= f_end:
                        continue
                except Exception:
                    continue
            
            r = dict(r)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["accepted"] = bool(r.get("accepted", False))
            r["sales_status"] = r.get("sales_status", None)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            
            # [NEW] ใส่ dispatch_round จาก DB (ระดับบรรทัด order_id + sku)
            sku_key = (oid, (r.get("sku") or "").strip())
            if sku_key in dispatch_round_map:
                r["dispatch_round"] = dispatch_round_map[sku_key]
            
            safe_rows.append(r)
        
        if logistic:
            safe_rows = [r for r in safe_rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        
        # [NEW] กรองตามรอบจ่ายงาน (dispatch_round) ถ้ามีการเลือก
        if round_sel:
            try:
                round_int = int(round_sel)
                safe_rows = [r for r in safe_rows if r.get("dispatch_round") == round_int]
            except (ValueError, TypeError):
                pass
        
        # [NEW] กรองตามจำนวนครั้งที่พิมพ์ - ต้องดึงข้อมูล print count ของแต่ละ order ก่อน
        if print_count_sel:
            try:
                target_pc = int(print_count_sel)
                # ดึง order_ids ที่มี print count ตามเงื่อนไข
                temp_oids = sorted({(r.get("order_id") or "").strip() for r in safe_rows if r.get("order_id")})
                if temp_oids:
                    pc_map = _get_print_counts_local(temp_oids, "picking")
                    # กรองเฉพาะ order ที่มี print count ตรงกับที่ระบุ
                    valid_oids = {oid for oid, cnt in pc_map.items() if cnt == target_pc}
                    safe_rows = [r for r in safe_rows if (r.get("order_id") or "").strip() in valid_oids]
            except (ValueError, TypeError):
                pass
        
        # [NEW] ดึงรายการรอบที่มีทั้งหมด (สำหรับ Dropdown)
        available_rounds = []
        try:
            rounds_sql = text(f"""
                SELECT DISTINCT dispatch_round 
                FROM {tbl} 
                WHERE printed_picking > 0 
                  AND dispatch_round IS NOT NULL
                ORDER BY dispatch_round ASC
            """)
            rounds_result = db.session.execute(rounds_sql).fetchall()
            available_rounds = [r[0] for r in rounds_result if r[0] is not None]
        except Exception:
            pass
        
        # Aggregate by SKU
        # [แก้ไข] ถ้ามีการเลือกรอบ ให้แยก aggregate ตามรอบด้วย (ไม่รวมข้ามรอบ)
        items = _aggregate_picking(safe_rows, group_by_round=bool(round_sel))
        
        # Get print counts
        # [แก้ไข] เพิ่ม "ISSUED" เพื่อให้หน้าประวัติ (ที่จ่ายงานแล้ว) นับ order ได้ถูกต้อง
        valid_rows = [r for r in safe_rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT", "ISSUED")]
        order_ids = sorted({(r.get("order_id") or "").strip() for r in valid_rows if r.get("order_id")})
        print_counts_pick = _get_print_counts_local(order_ids, "picking")
        print_count_overall = max(print_counts_pick.values()) if print_counts_pick else 0
        
        # Get the latest print timestamp and user
        print_timestamp_overall = None
        print_user_overall = None
        if order_ids:
            sql = text(f"SELECT printed_picking_at, printed_picking_by FROM {tbl} WHERE order_id IN :oids AND printed_picking_at IS NOT NULL ORDER BY printed_picking_at DESC LIMIT 1")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            result = db.session.execute(sql, {"oids": order_ids}).first()
            if result:
                try:
                    dt = datetime.fromisoformat(result[0])
                    if dt.tzinfo is None:
                        dt = TH_TZ.localize(dt)
                    print_timestamp_overall = dt
                    print_user_overall = result[1]
                except Exception:
                    pass
        
        # Shop name
        shop_sel_name = None
        if shop_id:
            s = Shop.query.get(int(shop_id))
            if s:
                shop_sel_name = f"{s.platform} • {s.name}"
        
        # Fill in platform/shop/logistic for each item
        for it in items:
            it["platform"] = platform or "-"
            it["shop"] = shop_sel_name or "-"
            it["logistic"] = logistic or "-"
        
        totals = {
            "total_skus": len(items),
            "total_need_qty": sum(i["need_qty"] for i in items),
            "total_shortage": sum(i["shortage"] for i in items),
        }
        shops = Shop.query.order_by(Shop.name.asc()).all()
        logistics = sorted(set(r.get("logistic") for r in safe_rows if r.get("logistic")))
        
        # Get available print dates for dropdown
        # หมายเหตุ: printed_picking_at ถูกบันทึกเป็นเวลาไทยอยู่แล้ว (ไม่ต้อง +7)
        sql_dates = text(f"""
            SELECT DISTINCT DATE(printed_picking_at) as print_date 
            FROM {tbl} 
            WHERE printed_picking > 0 AND printed_picking_at IS NOT NULL
            ORDER BY print_date DESC
        """)
        available_dates = [row[0] for row in db.session.execute(sql_dates).fetchall()]
        
        return render_template(
            "picking.html",
            items=items,
            totals=totals,
            shops=shops,
            logistics=logistics,
            platform_sel=platform if reset_mode != 'today' else None,
            shop_sel=shop_id if reset_mode != 'today' else None,
            shop_sel_name=shop_sel_name if reset_mode != 'today' else None,
            logistic_sel=logistic if reset_mode != 'today' else None,
            official_print=False,
            printed_meta=None,
            print_count_overall=print_count_overall,
            print_timestamp_overall=print_timestamp_overall,
            print_user_overall=print_user_overall,
            order_ids=order_ids,
            is_history_view=True,
            # ถ้าเป็น Default/Today ให้ส่งค่า print_date_sel เป็นวันนี้ เพื่อให้ Dropdown เลือกถูก
            print_date_sel=print_date if print_date else (target_date.isoformat() if target_date else None),
            available_dates=available_dates,
            available_rounds=available_rounds,
            round_sel=round_sel if reset_mode != 'today' else None,
            print_count_sel=print_count_sel if reset_mode != 'today' else None,  # [NEW] ส่งค่าจำนวนครั้งที่พิมพ์
            accepted_from=raw_from if reset_mode != 'today' else "",
            accepted_to=raw_to if reset_mode != 'today' else "",
        )

    @app.route("/export_picking.xlsx")
    @login_required
    def export_picking_excel():
        """Export ใบงานหยิบสินค้าปัจจุบัน - แสดงงานที่ยังไม่ได้พิมพ์"""
        # Check for reset mode
        reset_mode = request.args.get("reset")
        
        if reset_mode == 'all':
            platform = None
            shop_id = None
            logistic = None
            acc_from = None
            acc_to = None
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            acc_from = parse_date_any(request.args.get("accepted_from"))
            acc_to = parse_date_any(request.args.get("accepted_to"))

        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        
        # *** [NEW LOGIC] กรองเฉพาะ Order ที่พิมพ์คลังแล้ว แต่ยังไม่พิมพ์หยิบ ***
        oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        warehouse_counts = _get_print_counts_local(oids, kind="warehouse")
        picking_counts = _get_print_counts_local(oids, kind="picking")
        
        valid_rows = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            wh_count = int(warehouse_counts.get(oid, 0))
            pk_count = int(picking_counts.get(oid, 0))
            if wh_count > 0 and pk_count == 0:
                valid_rows.append(r)
        
        rows = valid_rows

        safe_rows = []
        for r in rows:
            r = dict(r)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["accepted"] = bool(r.get("accepted", False))
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            safe_rows.append(r)

        if logistic:
            safe_rows = [r for r in safe_rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        items = _aggregate_picking(safe_rows)

        valid_rows = [r for r in safe_rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]
        order_ids = sorted({(r.get("order_id") or "").strip() for r in valid_rows if r.get("order_id")})
        print_counts_pick = _get_print_counts_local(order_ids, "picking")
        print_count_overall = max(print_counts_pick.values()) if print_counts_pick else 0

        shop_name = ""
        if shop_id:
            s = Shop.query.get(int(shop_id))
            if s:
                shop_name = f"{s.platform} • {s.name}"

        for it in items:
            it["platform"] = platform or ""
            it["shop_name"] = shop_name or ""
            it["logistic"] = logistic or ""

        # Get dispatch_round data for items
        dispatch_rounds = {}
        if order_ids:
            tbl = _ol_table_name()
            sql = text(f"SELECT DISTINCT order_id, dispatch_round FROM {tbl} WHERE order_id IN :oids")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            for row in db.session.execute(sql, {"oids": order_ids}).fetchall():
                if row[1] is not None:
                    dispatch_rounds[row[0]] = row[1]
        
        df = pd.DataFrame([{
            "แพลตฟอร์ม": it["platform"],
            "ร้าน": it["shop_name"],
            "SKU": it["sku"],
            "Brand": it["brand"],
            "สินค้า": it["model"],
            "ต้องหยิบ": it["need_qty"],
            "สต็อก": it["stock_qty"],
            "ขาด": it["shortage"],
            "คงเหลือหลังหยิบ": it["remain_after_pick"],
            "ประเภทขนส่ง": it["logistic"],
            "จ่ายงาน(รอบที่)": it.get("dispatch_round", ""),
            "พิมพ์แล้ว (ครั้ง)": 0,  # Current page: not printed yet
        } for it in items])

        out = BytesIO()
        with pd.ExcelWriter(out, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="Picking List")
        out.seek(0)
        
        filename = f"ใบงานหยิบสินค้า_Picking_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(out, as_attachment=True, download_name=filename)

    @app.route("/report/picking/history/export.xlsx")
    @login_required
    def export_picking_history_excel():
        """Export ใบงานหยิบสินค้าประวัติ - แสดงงานที่พิมพ์แล้ว"""
        # Check for reset mode
        reset_mode = request.args.get("reset")
        
        if reset_mode == 'today':
            target_date = now_thai().date()
            platform = None
            shop_id = None
            logistic = None
            print_date = None
            raw_from = None
            raw_to = None
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            print_date = request.args.get("print_date")
            raw_from = request.args.get("accepted_from")
            raw_to = request.args.get("accepted_to")
            
            if print_date:
                try:
                    target_date = datetime.strptime(print_date, "%Y-%m-%d").date()
                except:
                    target_date = None
            else:
                target_date = None
        
        acc_from = parse_date_any(raw_from)
        acc_to = parse_date_any(raw_to)
        
        # Get printed orders
        tbl = _ol_table_name()
        
        if target_date:
            sql = text(f"""
                SELECT DISTINCT order_id 
                FROM {tbl} 
                WHERE printed_picking > 0 
                AND DATE(printed_picking_at) = :target_date
            """)
            result = db.session.execute(sql, {"target_date": target_date.isoformat()}).fetchall()
        else:
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_picking > 0")
            result = db.session.execute(sql).fetchall()
        
        printed_order_ids = [row[0] for row in result if row[0]]
        
        if not printed_order_ids:
            # Return empty file if no data
            df = pd.DataFrame(columns=["แพลตฟอร์ม", "ร้าน", "SKU", "Brand", "สินค้า", "ต้องหยิบ", "สต็อก", "ขาด", "คงเหลือหลังหยิบ", "ประเภทขนส่ง", "จ่ายงาน(รอบที่)", "พิมพ์แล้ว (ครั้ง)"])
            bio = BytesIO()
            with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
                df.to_excel(w, index=False, sheet_name="Picking History")
            bio.seek(0)
            return send_file(bio, as_attachment=True, download_name="ใบงานหยิบสินค้าประวัติ_Empty.xlsx")
        
        # Get full data for printed orders
        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        
        # Filter to only printed orders
        safe_rows = []
        for r in rows:
            if (r.get("order_id") or "").strip() not in printed_order_ids:
                continue
            r = dict(r)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["accepted"] = bool(r.get("accepted", False))
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            safe_rows.append(r)
        
        if logistic:
            safe_rows = [r for r in safe_rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        
        # Aggregate by SKU
        items = _aggregate_picking(safe_rows)
        
        # Get print counts
        valid_rows = [r for r in safe_rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]
        order_ids = sorted({(r.get("order_id") or "").strip() for r in valid_rows if r.get("order_id")})
        print_counts_pick = _get_print_counts_local(order_ids, "picking")
        print_count_overall = max(print_counts_pick.values()) if print_counts_pick else 0
        
        # Shop name
        shop_name = ""
        if shop_id and reset_mode != 'today':
            s = Shop.query.get(int(shop_id))
            if s:
                shop_name = f"{s.platform} • {s.name}"
        
        # Fill in platform/shop/logistic for each item
        for it in items:
            it["platform"] = platform or "-"
            it["shop_name"] = shop_name or "-"
            it["logistic"] = logistic or "-"
        
        # Get dispatch_round data
        dispatch_rounds = {}
        if order_ids:
            sql = text(f"SELECT DISTINCT order_id, dispatch_round FROM {tbl} WHERE order_id IN :oids")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            for row in db.session.execute(sql, {"oids": order_ids}).fetchall():
                if row[1] is not None:
                    dispatch_rounds[row[0]] = row[1]
        
        # Create DataFrame
        df = pd.DataFrame([{
            "แพลตฟอร์ม": it["platform"],
            "ร้าน": it["shop_name"],
            "SKU": it["sku"],
            "Brand": it["brand"],
            "สินค้า": it["model"],
            "ต้องหยิบ": it["need_qty"],
            "สต็อก": it["stock_qty"],
            "ขาด": it["shortage"],
            "คงเหลือหลังหยิบ": it["remain_after_pick"],
            "ประเภทขนส่ง": it["logistic"],
            "จ่ายงาน(รอบที่)": it.get("dispatch_round", ""),
            "พิมพ์แล้ว (ครั้ง)": print_count_overall,
        } for it in items])
        
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="Picking History")
        bio.seek(0)
        
        filename = f"ใบงานหยิบสินค้าประวัติ_History_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename)

    # -----------------------
    # ดาวน์โหลด Orders Excel Template (เดิม)
    # -----------------------
    @app.route("/download/orders-template")
    @login_required
    def download_orders_template():
        platform = normalize_platform(request.args.get("platform") or "Shopee")
        cols = ["ชื่อร้าน", "Order ID", "SKU", "Item Name", "Qty", "Order Time", "Logistics"]

        sample = pd.DataFrame(columns=cols)
        sample.loc[0] = ["Your Shop", "ORDER123", "SKU-001", "สินค้าทดลอง", 1, "2025-01-01 12:00", "J&T"]

        out = BytesIO()
        with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
            sample.to_excel(writer, index=False, sheet_name=f"{platform} Orders")
        out.seek(0)
        return send_file(out, as_attachment=True, download_name=f"{platform}_Orders_Template.xlsx")

    # -----------------------
    # Admin clear
    # -----------------------
    @app.route("/admin/clear", methods=["GET","POST"])
    @login_required
    def admin_clear():
        cu = current_user()
        if not cu or cu.role != "admin":
            flash("เฉพาะแอดมินเท่านั้นที่สามารถล้างข้อมูลได้", "danger")
            return redirect(url_for("dashboard"))
        
        if request.method == "POST":
            scope = request.form.get("scope")
            
            if scope == "today":
                today = now_thai().date()
                
                # [แก้ไข] หา order_id ที่จะถูกลบก่อน เพื่อเอาไปลบในตาราง DeletedOrder ด้วย
                today_lines = db.session.query(OrderLine.order_id).filter(OrderLine.import_date == today).all()
                today_oids = list(set(r[0] for r in today_lines if r[0]))
                
                # ลบข้อมูลจริง
                deleted = OrderLine.query.filter(OrderLine.import_date == today).delete(synchronize_session=False)
                
                # [เพิ่ม] ลบข้อมูลในถังขยะที่เกี่ยวข้องกับ ID พวกนี้
                del_bin = 0
                if today_oids:
                    del_bin = db.session.query(DeletedOrder).filter(DeletedOrder.order_id.in_(today_oids)).delete(synchronize_session=False)

                db.session.commit()
                flash(f"ลบข้อมูลของวันนี้แล้ว ({deleted} รายการ, ถังขยะ {del_bin} รายการ)", "warning")
            
            elif scope == "date_range":
                d_from_str = request.form.get("date_from")
                d_to_str = request.form.get("date_to")
                
                # รับค่า Checkbox เป็น list (เช่น ['orders', 'sales'])
                targets = request.form.getlist("targets")
                
                if not d_from_str or not d_to_str:
                    flash("กรุณาระบุวันที่เริ่มต้นและสิ้นสุด", "danger")
                elif not targets:
                    flash("กรุณาติ๊กเลือกประเภทข้อมูลที่ต้องการลบอย่างน้อย 1 อย่าง", "warning")
                else:
                    try:
                        # 1. แปลงวันที่จาก String เป็น Date Object
                        d_from = datetime.strptime(d_from_str, "%Y-%m-%d").date()
                        d_to = datetime.strptime(d_to_str, "%Y-%m-%d").date()
                        
                        # สร้างตัวแปร DateTime สำหรับฟิลด์ที่เป็น timestamp (เริ่ม 00:00:00 ถึง 23:59:59)
                        dt_start = datetime.combine(d_from, datetime.min.time())
                        dt_end = datetime.combine(d_to, datetime.max.time())
                        
                        msg_parts = []

                        # 2. เช็คว่าติ๊ก "ออเดอร์" ไหม
                        if "orders" in targets:
                            # [แก้ไข] หา order_id ก่อนลบ เพื่อตามไปลบในถังขยะด้วย
                            lines_q = db.session.query(OrderLine.order_id).filter(
                                OrderLine.import_date >= d_from,
                                OrderLine.import_date <= d_to
                            )
                            target_oids = list(set(r[0] for r in lines_q.all() if r[0]))
                            
                            del_orders = OrderLine.query.filter(
                                OrderLine.import_date >= d_from,
                                OrderLine.import_date <= d_to
                            ).delete(synchronize_session=False)
                            
                            # ลบในถังขยะด้วย (Cascading delete logic)
                            del_bin = 0
                            if target_oids:
                                del_bin = db.session.query(DeletedOrder).filter(DeletedOrder.order_id.in_(target_oids)).delete(synchronize_session=False)

                            msg_parts.append(f"ออเดอร์ {del_orders} รายการ (ถังขยะ {del_bin})")

                        # 3. เช็คว่าติ๊ก "ใบสั่งขาย" ไหม
                        if "sales" in targets:
                            try:
                                if hasattr(Sales, 'import_date'):
                                    del_sales = Sales.query.filter(
                                        Sales.import_date >= d_from,
                                        Sales.import_date <= d_to
                                    ).delete(synchronize_session=False)
                                    msg_parts.append(f"ใบขาย {del_sales} รายการ")
                            except Exception:
                                pass

                        # 4. เช็คว่าติ๊ก "จ่ายงานแล้ว" ไหม
                        if "issued" in targets:
                            del_issued = IssuedOrder.query.filter(
                                IssuedOrder.issued_at >= dt_start,
                                IssuedOrder.issued_at <= dt_end
                            ).delete(synchronize_session=False)
                            msg_parts.append(f"จ่ายแล้ว {del_issued} รายการ")

                        # 5. เช็คว่าติ๊ก "ยกเลิก" ไหม
                        if "cancelled" in targets:
                            # [แก้ไข] แปลงเวลาไทย (Input) -> UTC (Database) เพื่อให้ลบได้ตรงช่วง
                            dt_start_utc = dt_start - timedelta(hours=7)
                            dt_end_utc = dt_end - timedelta(hours=7)
                            
                            del_cancelled = CancelledOrder.query.filter(
                                CancelledOrder.imported_at >= dt_start_utc,
                                CancelledOrder.imported_at <= dt_end_utc
                            ).delete(synchronize_session=False)
                            msg_parts.append(f"ยกเลิก {del_cancelled} รายการ")
                        
                        # [เพิ่ม] 6. เช็คว่าติ๊ก "ประวัติการลบ" ไหม
                        if "deleted" in targets:
                            del_deleted_log = db.session.query(DeletedOrder).filter(
                                DeletedOrder.deleted_at >= dt_start,
                                DeletedOrder.deleted_at <= dt_end
                            ).delete(synchronize_session=False)
                            msg_parts.append(f"ประวัติลบ {del_deleted_log} รายการ")

                        db.session.commit()
                        
                        if msg_parts:
                            flash(f"ลบข้อมูลช่วง {d_from_str} - {d_to_str} เรียบร้อย: " + ", ".join(msg_parts), "success")
                        else:
                            flash("ไม่ได้ลบข้อมูลใดๆ", "info")
                              
                    except Exception as e:
                        db.session.rollback()
                        app.logger.exception("Clear date range failed")
                        flash(f"เกิดข้อผิดพลาดในการลบ: {e}", "danger")
                
            elif scope == "all":
                # 1. ลบรายการสินค้า
                deleted = OrderLine.query.delete()
                # 2. ลบถังขยะ
                del_bin = db.session.query(DeletedOrder).delete()
                # 3. [เพิ่ม] ลบประวัติการจ่ายงาน (Issued)
                del_issued = db.session.query(IssuedOrder).delete()
                # 4. [เพิ่ม] ลบประวัติการยกเลิก (Cancelled)
                del_cancel = db.session.query(CancelledOrder).delete()
                
                db.session.commit()
                flash(f"ล้างระบบใหม่หมดแล้ว (ออเดอร์ {deleted}, จ่ายแล้ว {del_issued}, ยกเลิก {del_cancel}, ถังขยะ {del_bin})", "success")
            
            # --- [เพิ่ม] CASE: ล้างถังขยะอย่างเดียว ---
            elif scope == "deleted_bin":
                n = db.session.query(DeletedOrder).delete()
                db.session.commit()
                flash(f"ล้างถังขยะเรียบร้อย ({n} รายการ)", "success")
                
            elif scope == "cancelled":
                # Get all cancelled order IDs
                cancelled_orders = CancelledOrder.query.all()
                cancelled_order_ids = [co.order_id for co in cancelled_orders]
                
                if cancelled_order_ids:
                    # Delete OrderLine records
                    deleted_lines = OrderLine.query.filter(
                        OrderLine.order_id.in_(cancelled_order_ids)
                    ).delete(synchronize_session=False)
                    
                    # Delete CancelledOrder records
                    deleted_cancelled = CancelledOrder.query.delete()
                    
                    db.session.commit()
                    flash(f"ลบ Order ยกเลิกทั้งหมดแล้ว ({len(cancelled_order_ids)} ออเดอร์, {deleted_lines} รายการ)", "warning")
                else:
                    flash("ไม่พบ Order ยกเลิก", "info")
                    
            elif scope == "issued":
                # Get all issued order IDs
                issued_orders = IssuedOrder.query.all()
                issued_order_ids = [io.order_id for io in issued_orders]
                
                if issued_order_ids:
                    # Delete OrderLine records
                    deleted_lines = OrderLine.query.filter(
                        OrderLine.order_id.in_(issued_order_ids)
                    ).delete(synchronize_session=False)
                    
                    # Delete IssuedOrder records
                    deleted_issued = IssuedOrder.query.delete()
                    
                    db.session.commit()
                    flash(f"ลบ Order จ่ายแล้วทั้งหมดแล้ว ({len(issued_order_ids)} ออเดอร์, {deleted_lines} รายการ)", "warning")
                else:
                    flash("ไม่พบ Order จ่ายแล้ว", "info")
                    
            elif scope == "sales":
                # ลบข้อมูลในตาราง Sales ทั้งหมด
                deleted = db.session.query(Sales).delete()
                db.session.commit()
                flash(f"ลบข้อมูลใบสั่งขาย (Sales) ทั้งหมดแล้ว ({deleted} รายการ)", "danger")

            elif scope == "bill_empty_today":
                # ล้างบิลเปล่าเฉพาะวันนี้
                today = now_thai().date()
                deleted = OrderLine.query.filter(
                    OrderLine.allocation_status == 'BILL_EMPTY',
                    OrderLine.import_date == today
                ).delete(synchronize_session=False)
                db.session.commit()
                flash(f"ลบบิลเปล่าของวันนี้แล้ว ({deleted} รายการ)", "warning")

            elif scope == "bill_empty_all":
                # ล้างบิลเปล่าทั้งหมด
                deleted = OrderLine.query.filter(
                    OrderLine.allocation_status == 'BILL_EMPTY'
                ).delete(synchronize_session=False)
                db.session.commit()
                flash(f"ลบบิลเปล่าทั้งหมดแล้ว ({deleted} รายการ)", "danger")

            return redirect(url_for("admin_clear"))
        
        # GET request - show stats
        today = now_thai().date()
        stats = {
            "total_orders": db.session.query(func.count(func.distinct(OrderLine.order_id))).scalar() or 0,
            "cancelled_orders": CancelledOrder.query.count(),
            "issued_orders": IssuedOrder.query.count(),
            "deleted_orders": DeletedOrder.query.count(),  # [เพิ่ม] นับถังขยะ
            "today_orders": db.session.query(func.count(func.distinct(OrderLine.order_id))).filter(
                OrderLine.import_date == today
            ).scalar() or 0,
            "total_sales": Sales.query.count(),
            "bill_empty_all": OrderLine.query.filter(OrderLine.allocation_status == 'BILL_EMPTY').count(),
            "bill_empty_today": OrderLine.query.filter(
                OrderLine.allocation_status == 'BILL_EMPTY',
                OrderLine.import_date == today
            ).count(),
        }
        
        return render_template("clear_confirm.html", stats=stats)

    return app


app = create_app()

if __name__ == "__main__":
    from waitress import serve
    port = int(os.environ.get("PORT", "8000"))
    print(f"[Vnix_ERP_OMS] Serving from {os.path.abspath(__file__)} on port {port}")
    serve(app, host="0.0.0.0", port=port)